"""
Analyse des temps par collaborateur — Export Excel
====================================================
Génère un tableur avec :
  - Onglet "Synthèse"       : vue consolidée (Projet / Rework / Maintenance / Gratuit / Support / Interne / Non saisi)
  - Onglet "Détail worklogs": toutes les saisies individuelles
  - Un onglet par collaborateur : récap + détail Interne + détail tickets

Catégorisation :
  Projet      → cat brute "Projet" + tickets PSC (clé PSC-*)
  Rework      → cat brute "Rework"
  Maintenance → cat brute "Maintenance" + "COORDIN : Suivi hors projet"
  Gratuit     → cat brute "Prestation offerte" + "Commande sans prestation"
  Support     → cat brute "Support N2"
  Interne     → cat brute "Interne" (détaillé par ticket fixe sur onglets individuels)
  Hors périmètre → tout le reste (tickets hors PDMDEP/PSC ou catégorie inconnue)
"""

import os
import sys
from datetime import datetime, timedelta
from types import SimpleNamespace

import holidays
import pandas as pd
from jira import JIRA
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# ══════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════
JIRA_SERVER = "https://sogelink.atlassian.net"
JIRA_EMAIL = "mathilde.panier@sogelink.com"
JIRA_TOKEN = os.environ["JIRA_API_TOKEN"]
ABSENCES_FILE = "absences.xlsx"
MODE          = "mois_precedent"   # "mois_courant" ou "mois_precedent"
MODE          = os.environ.get("EXPORT_MODE", MODE)  # override via workflow

today           = datetime.now()
_first_of_month = today.replace(day=1)
if MODE == "mois_precedent":
    REFERENCE_DATE = (_first_of_month.replace(month=_first_of_month.month - 1)
                      if _first_of_month.month > 1
                      else _first_of_month.replace(year=_first_of_month.year - 1, month=12))
else:
    REFERENCE_DATE = _first_of_month

month_start = REFERENCE_DATE
month_end = (month_start.replace(month=month_start.month + 1)
             if month_start.month < 12
             else month_start.replace(year=month_start.year + 1, month=1))

MOIS_FR    = ["Janvier","Février","Mars","Avril","Mai","Juin",
              "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
mois_label = MOIS_FR[month_start.month - 1]
EXPORT_FILE = f"analyse_temps_{month_start:%Y-%m}.xlsx"

all_days         = pd.date_range(month_start, month_end, freq="D")
fr_holidays      = holidays.France(years=month_start.year)
working_days     = [d for d in all_days if d.weekday() < 5 and d not in fr_holidays]
num_working_days = len(working_days)
HEURES_PAR_JOUR  = 7

print(f"📅 Période : {month_start:%Y-%m-%d} → {month_end:%Y-%m-%d}  |  {num_working_days} jours ouvrés")

# ══════════════════════════════════════════════════════════════════
# COLLABORATEURS (Timon Bos, Valentin Caujolle, Alizée Margot exclus)
# ══════════════════════════════════════════════════════════════════
ACCOUNT_TO_COLLAB = {
    "63fe26c90a4a47fb8d213c54"                    : "Quentin Bordillon",
    "712020:bb1c3fc2-7106-4c8b-9955-2e4917706722" : "Fabien Reutenauer",
    "63fe26c7f00d095406f2590c"                    : "Bérénice Bossard",
    "712020:f59016f5-1528-45f5-a182-d91eb6f69cad" : "Flavie Bardin",
    "712020:e3996f09-39b9-4957-8f83-d925d9543c98" : "Maxime Pontonnier",
    "712020:30146e03-1c8d-42f4-a105-c6987f9b2e2c" : "Marine Masingarbe",
    "712020:ab30e040-78a7-40e1-aef4-ff6781645c8e" : "Duncan Hamelin",
    "712020:a40869a4-1333-4240-ac8a-670ed2314e0b" : "Rémy Vincent",
}
ALLOWED_ACCOUNT_IDS = set(ACCOUNT_TO_COLLAB.keys())

TAUX_DISPO = {
    "Quentin Bordillon" : 1.00,
    "Fabien Reutenauer" : 1.00,
    "Bérénice Bossard"  : 1.00,
    "Flavie Bardin"     : 1.00,
    "Maxime Pontonnier" : 0.80,
    "Marine Masingarbe" : 1.00,
    "Duncan Hamelin"    : 1.00,
    "Rémy Vincent"      : 1.00,
}

# ══════════════════════════════════════════════════════════════════
# CATÉGORISATION
# ══════════════════════════════════════════════════════════════════
SOLUTION_MAPPING = {
    "GEODP1"             : "GEODP",
    "GEODP2 new"         : "GEODP",
    "GEODP2 migration"   : "GEODP",
    "LITTERALIS"         : "LITTERALIS",
    "LITTERALIS STANDARD": "LITTERALIS",
    "SHERPA"             : "LITTERALIS",
}

SUPPORT_LITT_EPICS  = {"PDMDEP-2171", "PDMDEP-12707"}
SUPPORT_GEODP_EPICS = {"PDMDEP-2172"}

# Epics "Divers" : comptés en Projet avec la solution forcée
DIVERS_EPICS = {
    "PDMDEP-31498": "LITTERALIS",  # DIVERS PROJET LITTERALIS
    "PDMDEP-31499": "GEODP",       # DIVERS PROJET GEODP
}

# Tickets fixes Interne avec leur libellé métier
INTERNE_TICKETS = {
    "PDMDEP-22900": "Skill development",
    "PDMDEP-22901": "Group Meeting",
    "PDMDEP-22903": "Internal process improvement",
    "PDMDEP-22904": "Travelling",
    "PDMDEP-24398": "Pre-sales",
    "PDMDEP-24399": "COORDIN Meeting",
    "PDMDEP-24400": "Career management",
    "PDMDEP-24401": "Administrative Tasks",
}
INTERNE_LABELS = list(dict.fromkeys(INTERNE_TICKETS.values()))

# Catégories brutes → catégorie synthèse
CAT_MAP = {
    "Projet"                      : "Projet",
    "Rework"                      : "Rework",
    "Maintenance"                 : "Maintenance",
    "COORDIN : Suivi hors projet" : "Maintenance",
    "Prestation offerte"          : "Gratuit",
    "Commande sans prestation"    : "Gratuit",
    "Support N2"                  : "Support",
    "Interne"                     : "Interne",
}

def categorize(cat_raw, ticket_key):
    if ticket_key.startswith("PSC-"):
        return "Projet"
    return CAT_MAP.get(cat_raw, "Hors périmètre")

def interne_label(ticket_key):
    """Remonte la hiérarchie pour trouver le libellé du ticket Interne."""
    key = ticket_key
    for _ in range(4):
        lbl = INTERNE_TICKETS.get(key)
        if lbl:
            return lbl
        issue = issue_map.get(key)
        if not issue:
            break
        parent = getattr(issue.fields, "parent", None)
        if not parent:
            break
        key = parent.key
    return "Interne (autre)"

# ══════════════════════════════════════════════════════════════════
# STYLES
# ══════════════════════════════════════════════════════════════════
BLUE_DARK  = "1F3864"
BLUE_MID   = "2E75B6"
BLUE_LIGHT = "D6E4F0"
BLUE_XL    = "EBF3FB"
GREY_LINE  = "BDD7EE"
WHITE      = "FFFFFF"
GREEN_FILL = "C6EFCE"
RED_FILL   = "FFCCCC"

def _font(bold=False, size=11, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)
def _border(color=GREY_LINE):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

STYLE_HDR_DARK = {"font": _font(bold=True, color=WHITE), "fill": _fill(BLUE_DARK),
                  "alignment": _align("center"), "border": _border(BLUE_DARK)}
STYLE_HDR_MID  = {"font": _font(bold=True, color=WHITE), "fill": _fill(BLUE_MID),
                  "alignment": _align("center"), "border": _border(BLUE_MID)}

def _apply(cell, style):
    for k, v in style.items():
        setattr(cell, k, v)

def _style_header_row(ws, row, ncols, dark=True):
    for c in range(1, ncols + 1):
        _apply(ws.cell(row=row, column=c), STYLE_HDR_DARK if dark else STYLE_HDR_MID)

def _style_row(ws, row, ncols, even=False, total=False, bg=None):
    if bg is None:
        bg = BLUE_LIGHT if total else (BLUE_XL if even else WHITE)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = _fill(bg)
        cell.border    = _border()
        cell.font      = _font(bold=total)
        cell.alignment = _align("left")

def _autofit(ws, min_w=10, max_w=50):
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 3, min_w), max_w)

def _sheet_title(ws, title, subtitle=None):
    ws["A1"] = title
    ws["A1"].font = _font(bold=True, size=14, color=BLUE_DARK)
    ws.row_dimensions[1].height = 24
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = _font(italic=True, size=10, color="595959")
    ws.append([])

# ══════════════════════════════════════════════════════════════════
# CONNEXION JIRA
# ══════════════════════════════════════════════════════════════════
try:
    jira = JIRA(options={"server": JIRA_SERVER}, basic_auth=(JIRA_EMAIL, JIRA_TOKEN))
    print("✅ Connexion Jira établie.")
except Exception as e:
    print(f"❌ Connexion impossible : {e}")
    sys.exit(1)

# ══════════════════════════════════════════════════════════════════
# RÉCUPÉRATION DES TICKETS
# ══════════════════════════════════════════════════════════════════
def get_all_issues(jql, expand="worklog,subtasks"):
    issues, token = [], None
    fields = ("summary,issuetype,parent,assignee,"
              "customfield_20514,customfield_23103,"
              "status,timespent,worklog,subtasks")
    while True:
        batch = jira.enhanced_search_issues(
            jql, expand=expand, maxResults=50, fields=fields, nextPageToken=token)
        issues.extend(batch)
        token = getattr(batch, "nextPageToken", None)
        if not token:
            break
    return issues

print("📥 Récupération des tickets PDMDEP…")
all_issues = get_all_issues("project = PDMDEP")
print(f"   → {len(all_issues)} tickets")

print("📥 Récupération des tickets PSC…")
psc_epics    = get_all_issues('project = PSC AND issuetype = "New delivery"')
psc_children = get_all_issues('project = PSC AND issuetype != "New delivery"')
print(f"   → {len(psc_epics)} epics PSC / {len(psc_children)} enfants PSC")

issue_map   = {i.key: i for i in all_issues}
worklog_map = {i.key: getattr(i.fields.worklog, "worklogs", []) for i in all_issues}
for issue in psc_epics + psc_children:
    issue_map[issue.key]   = issue
    worklog_map[issue.key] = getattr(issue.fields.worklog, "worklogs", [])
_all_pool = list(all_issues) + list(psc_epics) + list(psc_children)

# ══════════════════════════════════════════════════════════════════
# UTILITAIRES
# ══════════════════════════════════════════════════════════════════
def _make_wl(w_dict):
    author = SimpleNamespace(accountId=w_dict["author"]["accountId"],
                             displayName=w_dict["author"]["displayName"])
    return SimpleNamespace(author=author,
                           timeSpentSeconds=w_dict["timeSpentSeconds"],
                           started=w_dict["started"])

def fetch_worklogs(ticket):
    existing   = worklog_map.get(ticket.key, [])
    time_spent = getattr(ticket.fields, "timespent", None) or 0
    if len(existing) < 20 and not (len(existing) == 0 and time_spent > 0):
        return existing
    all_wl, start_at = [], 0
    url = f"{JIRA_SERVER}/rest/api/3/issue/{ticket.key}/worklog"
    while True:
        resp  = jira._session.get(url, params={"startAt": start_at, "maxResults": 50})
        resp.raise_for_status()
        data  = resp.json()
        batch = data.get("worklogs", [])
        all_wl.extend(_make_wl(w) for w in batch)
        if len(batch) < 50:
            break
        start_at += 50
    return all_wl

def get_category(issue):
    cur = issue
    for _ in range(5):
        if cur.fields.issuetype.name.lower() == "epic":
            cf = getattr(cur.fields, "customfield_23103", None)
            return (cf.value if hasattr(cf, "value") else "") or ""
        parent = getattr(cur.fields, "parent", None)
        if not parent:
            return ""
        cur = issue_map.get(parent.key)
        if not cur:
            return ""
    return ""

def get_solution(issue):
    cur = issue
    for _ in range(5):
        if cur.fields.issuetype.name.lower() == "epic":
            cf  = getattr(cur.fields, "customfield_20514", None)
            raw = (cf.value if hasattr(cf, "value") else "") or ""
            return SOLUTION_MAPPING.get(raw, raw or "Autre")
        parent = getattr(cur.fields, "parent", None)
        if not parent:
            return "Autre"
        cur = issue_map.get(parent.key)
        if not cur:
            return "Autre"
    return "Autre"

def get_support_solution(ticket_key):
    key = ticket_key
    for _ in range(4):
        issue = issue_map.get(key)
        if not issue:
            break
        if key in SUPPORT_LITT_EPICS:
            return "LITTERALIS"
        if key in SUPPORT_GEODP_EPICS:
            return "GEODP"
        parent = getattr(issue.fields, "parent", None)
        if not parent:
            break
        key = parent.key
    return ""

def get_divers_solution(ticket_key):
    """Retourne la solution si le ticket est sous un epic Divers, sinon ''."""
    key = ticket_key
    for _ in range(4):
        sol = DIVERS_EPICS.get(key)
        if sol:
            return sol
        issue = issue_map.get(key)
        if not issue:
            break
        parent = getattr(issue.fields, "parent", None)
        if not parent:
            break
        key = parent.key
    return ""

# ══════════════════════════════════════════════════════════════════
# COLLECTE DES WORKLOGS
# ══════════════════════════════════════════════════════════════════
print("📊 Collecte des worklogs du mois…")
rows = []
for issue in _all_pool:
    try:
        wl_list = fetch_worklogs(issue)
    except Exception as e:
        print(f"⚠️  {issue.key} : {e}")
        continue
    for w in wl_list:
        if w.author.accountId not in ALLOWED_ACCOUNT_IDS:
            continue
        started = pd.to_datetime(w.started).tz_convert("Europe/Paris").date()
        if started < month_start.date() or started >= month_end.date():
            continue

        collab   = ACCOUNT_TO_COLLAB[w.author.accountId]
        cat_raw  = get_category(issue)
        cat      = categorize(cat_raw, issue.key)
        solution = get_solution(issue)

        # Tickets Divers → forcé en Projet avec la bonne solution
        divers_sol = get_divers_solution(issue.key)
        if divers_sol:
            cat      = "Projet"
            solution = divers_sol
        elif cat_raw == "Support N2":
            sol_sup = get_support_solution(issue.key)
            if sol_sup:
                solution = sol_sup
        lbl_interne = interne_label(issue.key) if cat == "Interne" else ""

        rows.append({
            "Collaborateur"  : collab,
            "Ticket"         : issue.key,
            "Résumé"         : getattr(issue.fields, "summary", ""),
            "Catégorie brute": cat_raw,
            "Catégorie"      : cat,
            "Interne détail" : lbl_interne,
            "Solution"       : solution,
            "Temps (h)"      : round(w.timeSpentSeconds / 3600, 3),
            "Date"           : started,
        })

df = pd.DataFrame(rows)
print(f"   → {len(df)} entrées worklogs")

# ══════════════════════════════════════════════════════════════════
# ABSENCES
# ══════════════════════════════════════════════════════════════════
def load_absences(filepath):
    empty = {}, {}, {}
    if not os.path.exists(filepath):
        print(f"⚠️  Fichier absences introuvable : {filepath} → absences à 0.")
        return empty
    try:
        df_abs      = pd.read_excel(filepath, header=3).fillna(0)
        col_conges  = next((c for c in df_abs.columns if "cong"   in c.lower()), None)
        col_maladie = next((c for c in df_abs.columns if "malad"  in c.lower()), None)
        col_collab  = next((c for c in df_abs.columns if "collab" in c.lower()), None)
        if not all([col_collab, col_conges, col_maladie]):
            print(f"⚠️  Colonnes attendues : Collaborateur | Congés (jours) | Maladie (jours)")
            print(f"   Colonnes trouvées : {list(df_abs.columns)}")
            return empty
        df_abs["_total"] = df_abs[col_conges] + df_abs[col_maladie]
        total   = dict(zip(df_abs[col_collab], df_abs["_total"]))
        conges  = dict(zip(df_abs[col_collab], df_abs[col_conges]))
        maladie = dict(zip(df_abs[col_collab], df_abs[col_maladie]))
        print(f"✅ Absences chargées pour {len(total)} collaborateurs.")
        return total, conges, maladie
    except Exception as e:
        print(f"⚠️  Impossible de lire {filepath} : {e} → absences à 0.")
        return empty

absences, absences_conges, absences_maladie = load_absences(ABSENCES_FILE)

def capacite(name):
    taux  = TAUX_DISPO.get(name, 1.0)
    abs_j = absences.get(name, 0)
    return round(max(0, num_working_days - abs_j) * HEURES_PAR_JOUR * taux, 2)

def absences_detail(name):
    return absences_conges.get(name, 0), absences_maladie.get(name, 0), absences.get(name, 0)

# ══════════════════════════════════════════════════════════════════
# SYNTHÈSE PAR COLLABORATEUR
# ══════════════════════════════════════════════════════════════════
def pivot_collab(df_in, collab):
    df_c = df_in[df_in["Collaborateur"] == collab] if not df_in.empty else pd.DataFrame()
    cap  = capacite(collab)
    j_cg, j_ml, j_abs = absences_detail(collab)

    def _h(cat):
        return df_c[df_c["Catégorie"] == cat]["Temps (h)"].sum() if not df_c.empty else 0

    h_projet  = _h("Projet")
    h_rework  = _h("Rework")
    h_maint   = _h("Maintenance")
    h_gratuit = _h("Gratuit")
    h_support = _h("Support")
    h_interne = _h("Interne")
    h_hors    = _h("Hors périmètre")
    h_total   = h_projet + h_rework + h_maint + h_gratuit + h_support + h_interne + h_hors
    non_saisi = max(0, cap - h_total)

    return {
        "Collaborateur"      : collab,
        "Congés (j)"         : j_cg,
        "Maladie (j)"        : j_ml,
        "Absences (j)"       : j_abs,
        "Capacité (h)"       : cap,
        "Projet (h)"         : round(h_projet,  2),
        "Rework (h)"         : round(h_rework,  2),
        "Maintenance (h)"    : round(h_maint,   2),
        "Gratuit (h)"        : round(h_gratuit, 2),
        "Support (h)"        : round(h_support, 2),
        "Interne (h)"        : round(h_interne, 2),
        "Hors périmètre (h)" : round(h_hors,    2),
        "Total saisi (h)"    : round(h_total,   2),
        "Non saisi (h)"      : round(non_saisi, 2),
        "% Saisie"           : round(h_total   / cap * 100, 1) if cap else 0,
        "% Projet"           : round(h_projet  / cap * 100, 1) if cap else 0,
        "% Rework"           : round(h_rework  / cap * 100, 1) if cap else 0,
        "% Maintenance"      : round(h_maint   / cap * 100, 1) if cap else 0,
        "% Gratuit"          : round(h_gratuit / cap * 100, 1) if cap else 0,
        "% Support"          : round(h_support / cap * 100, 1) if cap else 0,
        "% Interne"          : round(h_interne / cap * 100, 1) if cap else 0,
    }

collabs  = sorted(ACCOUNT_TO_COLLAB.values())
synth    = [pivot_collab(df, c) for c in collabs]
df_synth = pd.DataFrame(synth)

# ══════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════
wb = Workbook()

# ─────────────────────────────────────────────────────────────────
# ONGLET 1 : SYNTHÈSE
# ─────────────────────────────────────────────────────────────────
ws_s = wb.active
ws_s.title = "Synthèse"
ws_s.sheet_view.showGridLines = False
_sheet_title(ws_s,
    f"Analyse des temps — {mois_label} {month_start.year}",
    f"Capacité : {num_working_days} jours ouvrés × {HEURES_PAR_JOUR}h  |  Export le {today:%d/%m/%Y}")

COLS_SYNTH = [
    "Collaborateur",
    "Congés (j)", "Maladie (j)", "Absences (j)", "Capacité (h)",
    "Projet (h)", "Rework (h)", "Maintenance (h)", "Gratuit (h)",
    "Support (h)", "Interne (h)", "Hors périmètre (h)",
    "Total saisi (h)", "Non saisi (h)", "% Saisie",
    "% Projet", "% Rework", "% Maintenance", "% Gratuit", "% Support", "% Interne",
]
ncols_s = len(COLS_SYNTH)
hdr_row = 4
ws_s.append(COLS_SYNTH)
_style_header_row(ws_s, hdr_row, ncols_s)

HDR_COMMENTS = {
    "Congés (j)"         : "Jours de congés posés — lu depuis absences.xlsx",
    "Maladie (j)"        : "Jours d'arrêt maladie — lu depuis absences.xlsx",
    "Absences (j)"       : "Congés + Maladie",
    "Capacité (h)"       : f"(Jours ouvrés − Absences) × {HEURES_PAR_JOUR}h × taux dispo",
    "Gratuit (h)"        : "Prestation offerte + Commande sans prestation",
    "Hors périmètre (h)" : "Tickets hors PDMDEP/PSC ou catégorie non reconnue",
    "Non saisi (h)"      : "Capacité − Total saisi",
    "% Saisie"           : "Total saisi / Capacité",
}
for cell in ws_s[hdr_row]:
    if cell.value in HDR_COMMENTS:
        cell.comment = Comment(HDR_COMMENTS[cell.value], "Info")

COL_PROJET   = COLS_SYNTH.index("Projet (h)")        + 1
COL_REWORK   = COLS_SYNTH.index("Rework (h)")        + 1
COL_MAINT    = COLS_SYNTH.index("Maintenance (h)")   + 1
COL_GRATUIT  = COLS_SYNTH.index("Gratuit (h)")       + 1
COL_SUPPORT  = COLS_SYNTH.index("Support (h)")       + 1
COL_INTERNE  = COLS_SYNTH.index("Interne (h)")       + 1
COL_NONSAISI = COLS_SYNTH.index("Non saisi (h)")     + 1
COL_PCT      = COLS_SYNTH.index("% Saisie")          + 1

for i, row in enumerate(df_synth[COLS_SYNTH].itertuples(index=False)):
    ws_s.append(list(row))
    ri       = hdr_row + 1 + i
    _style_row(ws_s, ri, ncols_s, even=(i % 2 == 0))
    pct_cell = ws_s.cell(row=ri, column=COL_PCT)
    pct_val  = pct_cell.value or 0
    if pct_val < 80:
        pct_cell.fill = _fill(RED_FILL);   pct_cell.font = _font(bold=True, color="C00000")
    elif pct_val >= 95:
        pct_cell.fill = _fill(GREEN_FILL); pct_cell.font = _font(bold=True, color="375623")

# Ligne totaux
total_row_idx = hdr_row + 1 + len(df_synth)
cap_total = df_synth["Capacité (h)"].sum()
totals = ["TOTAL ÉQUIPE"]
for col in COLS_SYNTH[1:]:
    if col.startswith("%"):
        key_h = col.replace("% ", "") + " (h)"
        if key_h in df_synth.columns:
            totals.append(round(df_synth[key_h].sum() / cap_total * 100, 1) if cap_total else 0)
        elif col == "% Saisie":
            totals.append(round(df_synth["Total saisi (h)"].sum() / cap_total * 100, 1) if cap_total else 0)
        else:
            totals.append("")
    else:
        totals.append(round(df_synth[col].sum(), 2))
ws_s.append(totals)
_style_row(ws_s, total_row_idx, ncols_s, total=True)

# Graphique barre empilée
chart_start = hdr_row + 1
chart_end   = chart_start + len(df_synth) - 1
bar = BarChart()
bar.type = "bar"; bar.title = "Répartition des temps par collaborateur"
bar.grouping = "stacked"; bar.y_axis.title = "Heures"
bar.width = 32; bar.height = 20

for col_idx, color in [
    (COL_PROJET,   "70AD47"),
    (COL_REWORK,   "C55A11"),
    (COL_MAINT,    "FFD966"),
    (COL_GRATUIT,  "9DC3E6"),
    (COL_SUPPORT,  "2E75B6"),
    (COL_INTERNE,  "808080"),
    (COL_NONSAISI, "FFCCCC"),
]:
    data  = Reference(ws_s, min_col=col_idx, min_row=hdr_row, max_row=chart_end)
    serie = Series(data, title_from_data=True)
    serie.graphicalProperties.solidFill      = color
    serie.graphicalProperties.line.solidFill = color
    bar.series.append(serie)

bar.set_categories(Reference(ws_s, min_col=1, min_row=chart_start, max_row=chart_end))
bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True
bar.legend.position = "b"
ws_s.add_chart(bar, f"V{hdr_row}")
_autofit(ws_s)

# ─────────────────────────────────────────────────────────────────
# ONGLET 2 : DÉTAIL WORKLOGS
# ─────────────────────────────────────────────────────────────────
ws_d = wb.create_sheet("Détail worklogs")
ws_d.sheet_view.showGridLines = False
_sheet_title(ws_d, "Détail des saisies de temps",
             f"Toutes les entrées Clockwork — {mois_label} {month_start.year}")

COLS_DET = ["Collaborateur", "Date", "Ticket", "Résumé", "Solution",
            "Catégorie brute", "Catégorie", "Interne détail", "Temps (h)"]
hdr_d = 4
ws_d.append(COLS_DET)
_style_header_row(ws_d, hdr_d, len(COLS_DET))

df_sorted = df[COLS_DET].sort_values(["Collaborateur", "Date", "Ticket"]) if not df.empty else df
for i, row in enumerate(df_sorted.itertuples(index=False)):
    ws_d.append(list(row))
    _style_row(ws_d, hdr_d + 1 + i, len(COLS_DET), even=(i % 2 == 0))

last_col = get_column_letter(len(COLS_DET))
ws_d.auto_filter.ref = f"A{hdr_d}:{last_col}{hdr_d}"
ws_d.freeze_panes    = f"A{hdr_d + 1}"
_autofit(ws_d)

# ─────────────────────────────────────────────────────────────────
# ONGLETS PAR COLLABORATEUR
# ─────────────────────────────────────────────────────────────────
for collab in collabs:
    ws_c = wb.create_sheet(collab[:28])
    ws_c.sheet_view.showGridLines = False
    cap = capacite(collab)
    j_cg, j_ml, j_abs = absences_detail(collab)
    taux_str = f"{int(TAUX_DISPO.get(collab, 1) * 100)}%"
    _sheet_title(ws_c,
        f"{collab} — {mois_label} {month_start.year}",
        f"Capacité : {cap}h  |  Dispo : {taux_str}  |  {num_working_days} jours ouvrés  |  "
        f"Absences : {j_abs}j (congés {j_cg}j / maladie {j_ml}j)")

    row_s = next((s for s in synth if s["Collaborateur"] == collab), {})
    df_c  = df[df["Collaborateur"] == collab].sort_values(["Date", "Ticket"]) if not df.empty else pd.DataFrame()

    # ── Bloc récap ────────────────────────────────────────────────
    ws_c.append(["Catégorie", "Heures", "% Capacité"])
    hdr_c = ws_c.max_row
    _style_header_row(ws_c, hdr_c, 3)

    def _pct_str(h):
        return f"{round(h / cap * 100, 1)}%" if cap else "0%"

    recap = [
        # Absences
        ("── Absences ──",  "",                              ""),
        ("  Congés",        f"{j_cg}j",                     ""),
        ("  Maladie",       f"{j_ml}j",                     ""),
        ("  Total absences",f"{j_abs}j",                    ""),
        # Productif — on n'affiche que les catégories avec des heures
        ("── Productif ──", "",                              ""),
    ]
    for cat_lbl, key_h, key_pct in [
        ("  Projet",      "Projet (h)",      "% Projet"),
        ("  Rework",      "Rework (h)",      "% Rework"),
        ("  Maintenance", "Maintenance (h)", "% Maintenance"),
        ("  Gratuit",     "Gratuit (h)",     "% Gratuit"),
    ]:
        h_val = row_s.get(key_h, 0)
        if h_val > 0:
            recap.append((cat_lbl, h_val, f"{row_s.get(key_pct, 0)}%"))

    # Support
    h_sup = row_s.get("Support (h)", 0)
    if h_sup > 0:
        recap += [
            ("── Support ──",  "",      ""),
            ("  Support",      h_sup,   f"{row_s.get('% Support', 0)}%"),
        ]

    # Interne détaillé — on n'affiche que les sous-catégories avec des heures
    interne_rows = []
    for lbl in INTERNE_LABELS:
        h_lbl = df_c[df_c["Interne détail"] == lbl]["Temps (h)"].sum() if not df_c.empty else 0
        if h_lbl > 0:
            interne_rows.append((f"  {lbl}", round(h_lbl, 2), _pct_str(h_lbl)))
    h_int_autre = df_c[df_c["Interne détail"] == "Interne (autre)"]["Temps (h)"].sum() if not df_c.empty else 0
    if h_int_autre > 0:
        interne_rows.append(("  Interne (autre)", round(h_int_autre, 2), _pct_str(h_int_autre)))
    if interne_rows:
        recap.append(("── Interne ──", "", ""))
        recap.extend(interne_rows)

    recap += [
        # Totaux
        ("── Totaux ──",    "",                                   ""),
        ("Total saisi",     row_s.get("Total saisi (h)", 0),     f"{row_s.get('% Saisie', 0)}%"),
        ("Non saisi",       row_s.get("Non saisi (h)",   0),     _pct_str(row_s.get("Non saisi (h)", 0))),
        ("Capacité",        f"{cap}h",                            "100%"),
    ]

    for j, (cat, val, pct) in enumerate(recap):
        ws_c.append([cat, val, pct])
        ri          = ws_c.max_row
        is_section  = cat.startswith("──")
        is_total    = cat in ("Total saisi", "Capacité", "Total absences")
        bg = BLUE_MID if is_section else None
        _style_row(ws_c, ri, 3, even=(j % 2 == 0), total=is_total, bg=bg)
        if is_section:
            for ci in range(1, 4):
                ws_c.cell(row=ri, column=ci).font = _font(bold=True, color=WHITE)
        if cat == "Non saisi" and row_s.get("Non saisi (h)", 0) > 0:
            for ci in range(1, 4):
                ws_c.cell(row=ri, column=ci).fill = _fill(RED_FILL)

    ws_c.append([])

    # ── Détail tickets ────────────────────────────────────────────
    if not df_c.empty:
        ticket_headers = ["Date", "Ticket", "Résumé", "Solution",
                          "Catégorie brute", "Catégorie", "Interne détail", "Temps (h)"]
        ws_c.append(ticket_headers)
        hdr_t = ws_c.max_row
        _style_header_row(ws_c, hdr_t, len(ticket_headers), dark=False)

        for k, row in enumerate(df_c[ticket_headers].itertuples(index=False)):
            ws_c.append(list(row))
            _style_row(ws_c, ws_c.max_row, len(ticket_headers), even=(k % 2 == 0))

        ws_c.append(["", "", "", "", "", "", "TOTAL", round(df_c["Temps (h)"].sum(), 2)])
        _style_row(ws_c, ws_c.max_row, len(ticket_headers), total=True)

        last_col_c = get_column_letter(len(ticket_headers))
        ws_c.auto_filter.ref = f"A{hdr_t}:{last_col_c}{hdr_t}"
        ws_c.freeze_panes    = f"A{hdr_t + 1}"
    else:
        ws_c["A20"] = "⚠️  Aucune saisie ce mois."
        ws_c["A20"].font = _font(italic=True, color="C00000")

    _autofit(ws_c)

# ─────────────────────────────────────────────────────────────────
# SAUVEGARDE
# ─────────────────────────────────────────────────────────────────
wb.save(EXPORT_FILE)
print(f"✅ Export terminé : {EXPORT_FILE}")
print(f"   Onglets : Synthèse, Détail worklogs, + {len(collabs)} onglets collaborateurs")