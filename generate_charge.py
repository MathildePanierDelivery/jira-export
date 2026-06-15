"""
generate_charge.py
------------------
Lit le dernier fichier jira_export_*.xlsx et génère charge.html
"""
import glob, os, sys
from datetime import datetime
from openpyxl import load_workbook
import design_system as ds

def find_file(path=None):
    if path and os.path.exists(path): return path
    files = sorted(glob.glob("jira_export_*.xlsx"), reverse=True)
    if not files:
        print("❌ Aucun fichier jira_export_*.xlsx trouvé"); sys.exit(1)
    print(f"📂 {files[0]}"); return files[0]

def safe_float(v):
    try: return float(str(v).replace("h","").strip() or 0)
    except: return 0

def load_data(fp):
    wb = load_workbook(fp, read_only=True, data_only=True)

    def read_sheet(name):
        ws = next((wb[s] for s in wb.sheetnames if name.lower() in s.lower()), None)
        if not ws: return []
        return list(ws.iter_rows(values_only=True))

    def parse_charge(rows):
        d = {"productif":0,"support":0,"interne":0,"rework":0,"projet":0,"maintenance":0,"offerte":0,"bloque":0,"historique":[]}
        for row in rows:
            vals = [v for v in row if v is not None]
            if not vals: continue
            lbl = str(vals[0]).strip()
            if lbl == "Productif" and len(vals) >= 2:        d["productif"]   = safe_float(vals[1])
            elif lbl == "  dont Projet" and len(vals) >= 2:  d["projet"]      = safe_float(vals[1])
            elif lbl == "  dont Rework" and len(vals) >= 2:  d["rework"]      = safe_float(vals[1])
            elif lbl == "  dont Maintenance" and len(vals) >= 2: d["maintenance"] = safe_float(vals[1])
            elif lbl == "  dont Prestation offerte" and len(vals) >= 2: d["offerte"] = safe_float(vals[1])
            elif lbl == "  dont Bloqué" and len(vals) >= 2:  d["bloque"]      = safe_float(vals[1])
            elif lbl == "Support" and len(vals) >= 2:         d["support"]     = safe_float(vals[1])
            elif lbl == "Interne" and len(vals) >= 2:         d["interne"]     = safe_float(vals[1])
        return d

    global_rows = read_sheet("Charge globale")
    litt_rows   = read_sheet("Charge Litt")
    geodp_rows  = read_sheet("Charge GEODP")

    global_d = parse_charge(global_rows)
    litt_d   = parse_charge(litt_rows)
    geodp_d  = parse_charge(geodp_rows)

    # Historique mensuel Littéralis
    hist_litt = []
    for row in litt_rows:
        vals = [v for v in row if v is not None]
        if not vals: continue
        lbl = str(vals[0]).strip()
        mois_fr = ["Janvier","Février","Mars","Avril","Mai","Juin","Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
        for m in mois_fr:
            if lbl.startswith(m) and len(vals) >= 2:
                hist_litt.append({"mois": m, "productif": safe_float(vals[1]), "rework": safe_float(vals[2]) if len(vals) > 2 else 0})

    # Heures par collaborateur
    collabs = []
    in_collab = False
    for row in global_rows:
        vals = [v for v in row if v is not None]
        if not vals: continue
        lbl = str(vals[0]).strip()
        if lbl == "Collaborateur": in_collab = True; continue
        if in_collab and len(vals) >= 3:
            try:
                collabs.append({"nom": lbl, "realise": safe_float(vals[1]), "theorique": safe_float(vals[2]), "delta": safe_float(vals[3]) if len(vals) > 3 else 0})
            except: pass

    wb.close()
    return {"global": global_d, "litt": litt_d, "geodp": geodp_d, "hist_litt": hist_litt, "collabs": collabs}

def charge_panel(d, label, color):
    total = d["productif"] + d["support"] + d["interne"]
    items = [
        ("Productif",   d["productif"], color),
        ("Support",     d["support"],   "#854F0B"),
        ("Interne",     d["interne"],   "#6B6B67"),
    ]
    bars = ""
    for lbl, val, col in items:
        pct = (val / total * 100) if total else 0
        bars += f"""
        <div class="bar-row">
          <div class="bar-label">{lbl}</div>
          <div class="bar-track"><div class="bar-fill" style="width:{min(pct,100):.1f}%;background:{col}"></div></div>
          <div class="bar-value">{ds.fmt_h(val)}</div>
        </div>"""

    detail_rows = ""
    for lbl, val in [("dont Projet", d["projet"]), ("dont Rework", d["rework"]), ("dont Maintenance", d["maintenance"]), ("dont Offerte", d["offerte"]), ("dont Bloqué", d["bloque"])]:
        if val > 0:
            detail_rows += f'<tr><td style="color:var(--text-2);padding-left:20px">{lbl}</td><td class="r">{ds.fmt_h(val)}</td></tr>'

    return f"""
    <div class="kpi-row">
      <div class="kpi">
        <div class="kpi-label">Total réalisé</div>
        <div class="kpi-value">{ds.fmt_h(total)}</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">Productif</div>
        <div class="kpi-value" style="color:{color}">{ds.fmt_h(d['productif'])}</div>
        <div class="kpi-trend flat">{(d['productif']/total*100):.1f}% du total</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">Rework</div>
        <div class="kpi-value amber">{ds.fmt_h(d['rework'])}</div>
        <div class="kpi-trend flat">{(d['rework']/d['productif']*100 if d['productif'] else 0):.1f}% du productif</div>
      </div>
    </div>

    <div class="g2">
      <div>
        <div class="section-label">Répartition par poste</div>
        <div class="card"><div class="card-body">{bars}</div></div>
      </div>
      <div>
        <div class="section-label">Détail du productif</div>
        <div class="card">
          <div class="card-body" style="padding:0">
            <table class="data-table">
              <thead><tr><th>Catégorie</th><th class="r">Heures</th></tr></thead>
              <tbody>
                <tr><td style="font-weight:500">Productif total</td><td class="r">{ds.fmt_h(d['productif'])}</td></tr>
                {detail_rows}
              </tbody>
            </table>
          </div>
        </div>
      </div>
    </div>"""

def generate_html(data, today_str):
    # Panels par scope
    panels = [
        ("GLOBAL",    "Global",     charge_panel(data["global"], "Global", "#185FA5")),
        ("LITTERALIS","Littéralis", charge_panel(data["litt"],   "Littéralis", "#534AB7")),
        ("GEODP",     "GEODP",      charge_panel(data["geodp"],  "GEODP", "#0F6E56")),
    ]

    scope_html = "".join(
        f'<div id="panel-{k}" class="scope-panel" style="display:{"block" if i==0 else "none"}">{html}</div>'
        for i, (k, _, html) in enumerate(panels)
    )

    # Historique mensuel (barres)
    hist = data["hist_litt"]
    max_val = max((h["productif"] for h in hist), default=1) or 1
    hist_bars = ""
    for h in hist:
        pct = h["productif"] / max_val * 100
        pct_rw = h["rework"] / max_val * 100
        hist_bars += f"""
        <div class="bar-row">
          <div class="bar-label">{h['mois'][:4]}</div>
          <div class="bar-track" style="height:16px;position:relative">
            <div class="bar-fill" style="width:{pct:.1f}%;background:#534AB7;opacity:.7;position:absolute"></div>
            <div class="bar-fill" style="width:{pct_rw:.1f}%;background:#854F0B;position:absolute"></div>
          </div>
          <div class="bar-value">{ds.fmt_h(h['productif'])}</div>
        </div>"""

    # Tableau collaborateurs
    collab_rows = ""
    for c in sorted(data["collabs"], key=lambda x: x["realise"], reverse=True):
        delta = c["delta"]
        delta_css = "red" if delta < -20 else ("green" if delta > 0 else "")
        delta_str = f'+{delta:.1f}' if delta >= 0 else f'{delta:.1f}'
        pct_saisie = (c["realise"] / c["theorique"] * 100) if c["theorique"] else 0
        collab_rows += f"""
        <tr>
          <td>{c['nom']}</td>
          <td class="r">{ds.fmt_h(c['realise'])}</td>
          <td class="r">{ds.fmt_h(c['theorique'])}</td>
          <td class="r" style="color:var(--{delta_css or 'text-2'})">{delta_str} h</td>
          <td class="r">
            <div style="display:flex;align-items:center;gap:6px;justify-content:flex-end">
              <div style="width:60px;height:6px;background:var(--bg);border-radius:3px;overflow:hidden;border:1px solid var(--border)">
                <div style="height:100%;width:{min(pct_saisie,100):.0f}%;background:{'#3B6D11' if pct_saisie>=80 else '#854F0B' if pct_saisie>=50 else '#A32D2D'};border-radius:3px"></div>
              </div>
              <span>{pct_saisie:.0f}%</span>
            </div>
          </td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Charge équipe — PS France</title>
{ds.DESIGN_CSS}
</head>
<body>
{ds.nav("charge", today_str)}
<div class="page">
  <div class="page-header">
    <div class="page-title">Charge équipe</div>
    <div class="page-sub">Heures réalisées — données Jira du mois en cours</div>
  </div>

  <div class="scope-wrap">
    <div class="scope-tabs">
      <button class="scope-tab active" onclick="setScope('GLOBAL',this)">Global</button>
      <button class="scope-tab" onclick="setScope('LITTERALIS',this)">Littéralis</button>
      <button class="scope-tab" onclick="setScope('GEODP',this)">GEODP</button>
    </div>
  </div>

  {scope_html}

  <div class="section-label" style="margin-top:8px">Évolution mensuelle — productif Littéralis</div>
  <div class="card">
    <div class="card-head">
      <span class="card-title">Heures productives vs Rework</span>
      <div style="display:flex;gap:12px;align-items:center;font-size:11px">
        <span style="display:flex;align-items:center;gap:4px"><span style="width:10px;height:10px;background:#534AB7;border-radius:2px;display:inline-block;opacity:.7"></span>Productif</span>
        <span style="display:flex;align-items:center;gap:4px"><span style="width:10px;height:10px;background:#854F0B;border-radius:2px;display:inline-block"></span>Rework</span>
      </div>
    </div>
    <div class="card-body">{hist_bars if hist_bars else '<p style="color:var(--text-3);text-align:center;padding:16px">Pas de données historiques disponibles</p>'}</div>
  </div>

  <div class="section-label">Heures par collaborateur</div>
  <div class="card">
    <div class="card-head">
      <span class="card-title">Réalisé vs théorique — mois en cours</span>
    </div>
    <div class="card-body" style="padding:0">
      <table class="data-table">
        <thead><tr><th>Collaborateur</th><th class="r">Réalisé</th><th class="r">Théorique</th><th class="r">Delta</th><th class="r">% saisie</th></tr></thead>
        <tbody>{collab_rows}</tbody>
      </table>
    </div>
  </div>
</div>

<script>
function setScope(key, btn) {{
  document.querySelectorAll(".scope-panel").forEach(p => p.style.display="none");
  document.querySelectorAll(".scope-tab").forEach(b => b.classList.remove("active"));
  document.getElementById("panel-"+key).style.display="block";
  btn.classList.add("active");
}}
</script>
</body></html>"""

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--file")
    parser.add_argument("--output", default="charge.html")
    args = parser.parse_args()
    fp = find_file(args.file)
    today_str = datetime.now().strftime("%d/%m/%Y")
    print("📊 Lecture des données...")
    data = load_data(fp)
    print(f"   {len(data['collabs'])} collaborateurs trouvés")
    print("🖊️  Génération HTML...")
    html = generate_html(data, today_str)
    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ {args.output}")
