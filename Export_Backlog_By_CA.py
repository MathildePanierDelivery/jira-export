import os
"""
Export Jira vers Excel - Backlog BDC (tickets Suivi CA)
=======================================================
- Unité d'analyse : chaque ticket "COORDIN : Suivi CA"
- Blocage & To Recognized : lus sur le ticket CA
- Solution cible & Type Epic : remontés depuis l'Epic parent
- Design moderne : palette Indigo/Coral, graphiques donut, sans section Rework

Pré-requis :
    pip install jira openpyxl pandas
"""

import sys, io
# Force UTF-8 sur stdout/stderr pour éviter les erreurs d'encodage sur Windows (cp1252)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from jira import JIRA
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, Reference
from openpyxl.chart import DoughnutChart
from openpyxl.utils import column_index_from_string
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from datetime import date
import pandas as pd

# ─────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────
JIRA_URL = "https://sogelink.atlassian.net"
JIRA_EMAIL = "mathilde.panier@sogelink.com"
JIRA_TOKEN = os.environ["JIRA_API_TOKEN"]
JIRA_PROJECT  = "PDMDEP"
today         = date.today().strftime("%Y%m%d")
OUTPUT_FILE   = f"Backlog_projets_by_CA_{today}.xlsx"

FIELD_TO_RECOGNIZED   = "customfield_22042"   # Montant restant à date
FIELD_MONTANT_TOTAL   = "customfield_20701"   # Montant total
FIELD_REMISE          = "customfield_21447"   # Remise
FIELD_SOLUTION_CIBLE = "customfield_20514"   # Solution cible (Epic)
FIELD_BLOCAGE        = "customfield_24269"   # Blocage
FIELD_BLOCAGE_DETAIL = "customfield_26672"   # Détail blocage
FIELD_BLOCAGE_DATE   = "customfield_26845"   # Date de déblocage estimée
FIELD_DEPASSEMENT        = "customfield_26606"   # Dépassement
FIELD_SERVICE_RESPONSABLE = "customfield_26878"  # Service responsable
FIELD_TYPE_EPIC      = "customfield_23103"   # Type d'Epic
FIELD_BDC            = "customfield_24438"   # N° BDC / Order
FIELD_CLIENT         = "customfield_10070"   # Client
FIELD_NOM_PROJET     = "customfield_23608"   # Nom du projet (Epic)
FIELD_JALON_NOM      = "customfield_23645"   # Jalon actuel (nom)
FIELD_JALON_PCT      = "customfield_21883"   # Jalon actuel (%)
FIELD_TEMPS_RESTANT  = "customfield_10384"   # Temps restant (somme sous-tâches)
FIELD_PREV = {                               # Prévisionnel mensuel
    "Juin":       "customfield_23610",
    "Juillet":      "customfield_23602",
    "Août":   "customfield_23603",
    "Septembre":      "customfield_23604",
    "Octobre": "customfield_23605",
    "Novembre":   "customfield_23606",
    "Décembre":  "customfield_23607",
}
ISSUE_TYPE           = "COORDIN : Suivi CA"

LITTERALIS_SOLUTIONS = ["LITTERALIS", "LITTERALIS STANDARD", "SHERPA"]
GEODP_SOLUTIONS      = ["GEODP1", "GEODP2 new", "GEODP migration", "GEODP2 migration"]
# ─────────────────────────────────────────────────────────────

# ── Palette (fidèle au fichier de référence Mars26.xlsx) ─────
CLR_BAND    = "1F3864"   # Bleu marine foncé — bandeaux de section
CLR_VALUE   = "EBF3FB"   # Bleu très pâle   — fond des grandes valeurs
CLR_WHITE   = "FFFFFF"
TXT_BAND    = "FFFFFF"   # Blanc sur bandeau
TXT_LABEL   = "595959"   # Gris foncé — labels
TXT_VALUE   = "2E75B6"   # Bleu — grandes valeurs positives
TXT_RED     = "C00000"   # Rouge — valeurs négatives / alertes
TXT_GREEN   = "375623"   # Vert foncé — taux bons
TXT_TITLE   = "1F3864"   # Bleu marine — titre principal
TXT_SUB     = "595959"   # Gris — sous-titre

# Couleurs des tranches de donut
DONUT_COLORS = ["2E75B6", "C00000", "ED7D31", "70AD47"]


# ── 1. Connexion & extraction ─────────────────────────────────

def connect_jira() -> JIRA:
    return JIRA(server=JIRA_URL, basic_auth=(JIRA_EMAIL, JIRA_TOKEN))


# Mapping normalisation solution cible
SOLUTION_MAP = {
    "GEODP2 migration":  "GEODP",
    "GEODP2 new":        "GEODP",
    "GEODP1":            "GEODP",
    "GEODP migration":   "GEODP",
    "LITTERALIS":            "LITTERALIS",
    "LITTERALIS STANDARD":   "LITTERALIS",
    "SHERPA":                "LITTERALIS",
}


def fetch_tickets(jira: JIRA) -> pd.DataFrame:

    def opt_val(fd, fid):
        v = fd.get(fid)
        if v is None:
            return None
        if isinstance(v, dict):
            val = v.get("value") or v.get("name") or v.get("displayName")
        elif isinstance(v, list):
            val = ", ".join(
                (i.get("value") or i.get("name") or i.get("displayName") or str(i))
                for i in v if i
            )
        else:
            val = str(v)
        return val if val and val.lower() not in ("aucun", "none", "") else None

    # ── 1. Index des Epics PDMDEP ──────────────────────────────
    print("  Récupération des Epics PDMDEP...")
    jql_epics = (
        f'project = "{JIRA_PROJECT}" AND issuetype = Epic '
        f'AND statusCategory != Done '
        f'AND "Type d\'Epic[Dropdown]" NOT IN (Interne, "Support N2", "Suivi client hors projet") '
        f'ORDER BY created DESC'
    )
    epic_fields = ",".join([
        FIELD_TYPE_EPIC, FIELD_SOLUTION_CIBLE, FIELD_NOM_PROJET, FIELD_CLIENT, "status", "created",
    ])
    epics = jira.search_issues(jql_epics, maxResults=False, fields=epic_fields)
    print(f"  → {len(epics)} Epics")

    epic_index = {}
    for e in epics:
        ef = e.raw["fields"]
        client_raw = ef.get(FIELD_CLIENT)
        if isinstance(client_raw, list):
            client = ", ".join(
                (i.get("displayName") or i.get("name") or "") for i in client_raw if i
            ) or None
        else:
            client = opt_val(ef, FIELD_CLIENT)
        epic_index[e.key] = {
            "solution_cible": opt_val(ef, FIELD_SOLUTION_CIBLE),
            "type_epic":      opt_val(ef, FIELD_TYPE_EPIC),
            "nom_projet":     opt_val(ef, FIELD_NOM_PROJET),
            "client":         client,
            "epic_status":    (ef.get("status") or {}).get("name"),
            "epic_created":   (ef.get("created") or "")[:10],
        }

    # ── 2. Tickets Suivi CA — avec issuelinks, jalon, prévisionnel ──
    print("  Récupération des tickets Suivi CA...")
    jql_ca = (
        f'project = "{JIRA_PROJECT}" AND issuetype = "{ISSUE_TYPE}" '
        f'AND statusCategory != Done ORDER BY created DESC'
    )
    ca_fields = ",".join([
        FIELD_TO_RECOGNIZED, FIELD_BLOCAGE, FIELD_BLOCAGE_DETAIL, FIELD_BDC,
        FIELD_JALON_PCT, FIELD_BLOCAGE_DATE, FIELD_MONTANT_TOTAL, FIELD_REMISE,
        *FIELD_PREV.values(),
        "parent", "summary", "status", "issuelinks", "created", FIELD_CLIENT, FIELD_NOM_PROJET,
        "assignee", FIELD_SERVICE_RESPONSABLE,
    ])
    ca_issues = jira.search_issues(jql_ca, maxResults=False, fields=ca_fields)
    print(f"  → {len(ca_issues)} tickets Suivi CA")

    # ── 3. Temps restant : somme timeestimate de tous les enfants de chaque Epic ──
    print("  Calcul du temps restant par Epic...")

    epic_keys = list(epic_index.keys())
    temps_by_epic  = {}
    estime_by_epic = {}   # somme timeoriginalestimate
    depas_by_epic  = {}   # somme customfield_26606
    batch_size = 50

    # Passe 1 : tickets directement liés à l'Epic (children)
    # On stocke (epic_link, child_key) pour récupérer leurs sous-tâches en passe 2
    child_keys_by_epic = []   # [(epic_link, child_key), ...]

    for i in range(0, len(epic_keys), batch_size):
        batch = epic_keys[i:i + batch_size]
        keys_str = ", ".join(f'"{k}"' for k in batch)
        try:
            children = jira.search_issues(
                f'project = "{JIRA_PROJECT}" AND "Epic Link" in ({keys_str})',
                maxResults=False, fields=f"timeestimate,timeoriginalestimate,{FIELD_DEPASSEMENT},customfield_10014,subtasks"
            )
            for child in children:
                cf = child.raw["fields"]
                epic_link = cf.get("customfield_10014")
                if not epic_link:
                    continue
                temps_by_epic[epic_link]  = temps_by_epic.get(epic_link, 0)  + float(cf.get("timeestimate") or 0)
                estime_by_epic[epic_link] = estime_by_epic.get(epic_link, 0) + float(cf.get("timeoriginalestimate") or 0)
                depas_by_epic[epic_link]  = depas_by_epic.get(epic_link, 0)  + float(cf.get(FIELD_DEPASSEMENT) or 0)
                # Mémoriser ce child pour récupérer ses sous-tâches en passe 2
                for st in (cf.get("subtasks") or []):
                    child_keys_by_epic.append((epic_link, child.key, st["key"]))
        except Exception as ex:
            print(f"    ⚠ Erreur lot {i//batch_size}: {ex}")

    # Passe 2 : sous-tâches directes des children
    if child_keys_by_epic:
        all_st_keys = [st_key for _, _, st_key in child_keys_by_epic]
        # Table de correspondance st_key → epic_link pour attribution rapide
        st_to_epic = {st_key: epic_link for epic_link, _, st_key in child_keys_by_epic}

        print(f"  Récupération des sous-tâches ({len(all_st_keys)} tickets)...")
        grand_children_keys = []   # sous-tâches des sous-tâches éventuelles

        for i in range(0, len(all_st_keys), batch_size):
            batch = all_st_keys[i:i + batch_size]
            keys_str = ", ".join(f'"{k}"' for k in batch)
            try:
                subtasks = jira.search_issues(
                    f'key in ({keys_str})',
                    maxResults=False, fields=f"timeestimate,timeoriginalestimate,{FIELD_DEPASSEMENT},subtasks"
                )
                for st in subtasks:
                    sf = st.raw["fields"]
                    epic_link = st_to_epic.get(st.key)
                    if epic_link:
                        temps_by_epic[epic_link]  = temps_by_epic.get(epic_link, 0)  + float(sf.get("timeestimate") or 0)
                        estime_by_epic[epic_link] = estime_by_epic.get(epic_link, 0) + float(sf.get("timeoriginalestimate") or 0)
                        depas_by_epic[epic_link]  = depas_by_epic.get(epic_link, 0)  + float(sf.get(FIELD_DEPASSEMENT) or 0)
                        # Collecter les sous-sous-tâches (niveau 3) si elles existent
                        for gst in (sf.get("subtasks") or []):
                            grand_children_keys.append((epic_link, gst["key"]))
            except Exception as ex:
                print(f"    ⚠ Erreur sous-tâches lot {i//batch_size}: {ex}")

        # Passe 3 : sous-tâches de niveau 3 (si présentes)
        if grand_children_keys:
            all_gst_keys = [k for _, k in grand_children_keys]
            gst_to_epic = {k: epic_link for epic_link, k in grand_children_keys}
            print(f"  Récupération sous-tâches niveau 3 ({len(all_gst_keys)} tickets)...")
            for i in range(0, len(all_gst_keys), batch_size):
                batch = all_gst_keys[i:i + batch_size]
                keys_str = ", ".join(f'"{k}"' for k in batch)
                try:
                    gsubtasks = jira.search_issues(
                        f'key in ({keys_str})',
                        maxResults=False, fields=f"timeestimate,timeoriginalestimate,{FIELD_DEPASSEMENT}"
                    )
                    for gst in gsubtasks:
                        gf = gst.raw["fields"]
                        epic_link = gst_to_epic.get(gst.key)
                        if epic_link:
                            temps_by_epic[epic_link]  = temps_by_epic.get(epic_link, 0)  + float(gf.get("timeestimate") or 0)
                            estime_by_epic[epic_link] = estime_by_epic.get(epic_link, 0) + float(gf.get("timeoriginalestimate") or 0)
                            depas_by_epic[epic_link]  = depas_by_epic.get(epic_link, 0)  + float(gf.get(FIELD_DEPASSEMENT) or 0)
                except Exception as ex:
                    print(f"    ⚠ Erreur sous-tâches niv.3 lot {i//batch_size}: {ex}")

    print(f"  → Temps restant calculé pour {len(temps_by_epic)} Epics (enfants + sous-tâches)")

    # ── 3b. Temps restant PSC : enfants du ticket "New delivery" lié ──
    # Pour les tickets CA sans Epic parent mais liés à un ticket PSC "New delivery",
    # le temps restant = somme timeestimate des enfants (et sous-tâches) du ticket PSC.
    print("  Calcul du temps restant pour les tickets PSC...")
    # Collecter tous les ps_pdm_key distincts nécessaires
    psc_keys_needed = set()
    for issue in ca_issues:
        f_tmp = issue.raw["fields"]
        parent_tmp = f_tmp.get("parent")
        if parent_tmp:
            continue  # a un Epic parent → pas PSC
        for link in (f_tmp.get("issuelinks") or []):
            for direction in ("inwardIssue", "outwardIssue"):
                linked = link.get(direction)
                if not linked:
                    continue
                linked_key = linked.get("key", "")
                issue_type_name = linked.get("fields", {}).get("issuetype", {}).get("name", "")
                if linked_key.startswith("PSC") and issue_type_name.lower() == "new delivery":
                    psc_keys_needed.add(linked_key)
                    break

    temps_by_psc  = {}
    estime_by_psc = {}
    depas_by_psc  = {}
    if psc_keys_needed:
        psc_keys_list = list(psc_keys_needed)
        print(f"  → {len(psc_keys_list)} tickets PSC New delivery trouvés")
        psc_child_keys = []   # [(psc_key, child_key, st_key), ...]

        for i in range(0, len(psc_keys_list), batch_size):
            batch = psc_keys_list[i:i + batch_size]
            keys_str = ", ".join(f'"{k}"' for k in batch)
            try:
                psc_children = jira.search_issues(
                    f'parent in ({keys_str})',
                    maxResults=False, fields=f"timeestimate,timeoriginalestimate,{FIELD_DEPASSEMENT},parent,subtasks"
                )
                for child in psc_children:
                    cf = child.raw["fields"]
                    psc_key = (cf.get("parent") or {}).get("key")
                    if not psc_key:
                        continue
                    temps_by_psc[psc_key]  = temps_by_psc.get(psc_key, 0)  + float(cf.get("timeestimate") or 0)
                    estime_by_psc[psc_key] = estime_by_psc.get(psc_key, 0) + float(cf.get("timeoriginalestimate") or 0)
                    depas_by_psc[psc_key]  = depas_by_psc.get(psc_key, 0)  + float(cf.get(FIELD_DEPASSEMENT) or 0)
                    for st in (cf.get("subtasks") or []):
                        psc_child_keys.append((psc_key, child.key, st["key"]))
            except Exception as ex:
                print(f"    ⚠ Erreur enfants PSC lot {i//batch_size}: {ex}")

        # Passe 2 : sous-tâches des enfants PSC
        if psc_child_keys:
            all_psc_st = [st_key for _, _, st_key in psc_child_keys]
            psc_st_to_psc = {st_key: psc_key for psc_key, _, st_key in psc_child_keys}
            psc_grand_children = []

            for i in range(0, len(all_psc_st), batch_size):
                batch = all_psc_st[i:i + batch_size]
                keys_str = ", ".join(f'"{k}"' for k in batch)
                try:
                    subtasks = jira.search_issues(
                        f'key in ({keys_str})',
                        maxResults=False, fields=f"timeestimate,timeoriginalestimate,{FIELD_DEPASSEMENT},subtasks"
                    )
                    for st in subtasks:
                        sf = st.raw["fields"]
                        psc_key = psc_st_to_psc.get(st.key)
                        if psc_key:
                            temps_by_psc[psc_key]  = temps_by_psc.get(psc_key, 0)  + float(sf.get("timeestimate") or 0)
                            estime_by_psc[psc_key] = estime_by_psc.get(psc_key, 0) + float(sf.get("timeoriginalestimate") or 0)
                            depas_by_psc[psc_key]  = depas_by_psc.get(psc_key, 0)  + float(sf.get(FIELD_DEPASSEMENT) or 0)
                            for gst in (sf.get("subtasks") or []):
                                psc_grand_children.append((psc_key, gst["key"]))
                except Exception as ex:
                    print(f"    ⚠ Erreur sous-tâches PSC: {ex}")

            # Passe 3 : niveau 3
            if psc_grand_children:
                all_pgst = [k for _, k in psc_grand_children]
                pgst_to_psc = {k: psc_key for psc_key, k in psc_grand_children}
                for i in range(0, len(all_pgst), batch_size):
                    batch = all_pgst[i:i + batch_size]
                    keys_str = ", ".join(f'"{k}"' for k in batch)
                    try:
                        gsubtasks = jira.search_issues(
                            f'key in ({keys_str})',
                            maxResults=False, fields=f"timeestimate,timeoriginalestimate,{FIELD_DEPASSEMENT}"
                        )
                        for gst in gsubtasks:
                            gf = gst.raw["fields"]
                            psc_key = pgst_to_psc.get(gst.key)
                            if psc_key:
                                temps_by_psc[psc_key]  = temps_by_psc.get(psc_key, 0)  + float(gf.get("timeestimate") or 0)
                                estime_by_psc[psc_key] = estime_by_psc.get(psc_key, 0) + float(gf.get("timeoriginalestimate") or 0)
                                depas_by_psc[psc_key]  = depas_by_psc.get(psc_key, 0)  + float(gf.get(FIELD_DEPASSEMENT) or 0)
                    except Exception as ex:
                        print(f"    ⚠ Erreur sous-tâches PSC niv.3: {ex}")

        print(f"  → Temps restant PSC calculé pour {len(temps_by_psc)} tickets")

    def action_attendue(pct):
        """Règle issue du modèle Suivi_projets.xlsx (valeur Jira 0-100) :
          0        → Lancement
          10       → Mise à disposition PF
          20–89    → Paramétrages
          90       → PV
          autres   → None (non renseigné)
        Note : jalon_pct est déjà divisé par 100 pour Excel (0.0–1.0),
        donc on compare sur la valeur originale * 100.
        """
        if pct is None:
            return "Lancement"   # non renseigné = pas encore démarré
        pct_orig = round(pct * 100)   # repasse en 0-100 pour la comparaison
        if pct_orig == 0:             return "Lancement"
        if pct_orig == 10:            return "Mise à disposition PF"
        if 20 <= pct_orig <= 89:      return "Paramétrages"
        if pct_orig >= 90:            return "PV"
        return ""

    # ── 4. Construction du DataFrame ──────────────────────────
    # Diagnostic issuelinks — tous les tickets ayant au moins un lien PSC
    print("  Diagnostic issuelinks (tickets avec liens PSC) :")
    found = 0
    for issue in ca_issues:
        links = issue.raw["fields"].get("issuelinks") or []
        psc_links = []
        for lk in links:
            for d in ("inwardIssue", "outwardIssue"):
                tgt = lk.get(d)
                if tgt and tgt.get("key", "").startswith("PSC"):
                    psc_links.append((d, lk, tgt))
        if psc_links:
            print(f"    {issue.key}:")
            for d, lk, tgt in psc_links:
                print(f"      [{d}] key={tgt.get('key')}")
                print(f"        type_lien name={lk.get('type',{}).get('name')!r}")
                print(f"        type_lien inward={lk.get('type',{}).get('inward')!r}")
                print(f"        type_lien outward={lk.get('type',{}).get('outward')!r}")
                print(f"        issuetype={tgt.get('fields',{}).get('issuetype',{}).get('name')!r}")
                print(f"        summary={tgt.get('fields',{}).get('summary','')!r}")
            found += 1
            if found >= 5:
                break
    if found == 0:
        print("    ⚠ Aucun lien PSC trouvé dans les issuelinks")

    rows = []
    for issue in ca_issues:
        f = issue.raw["fields"]
        to_recognized = float(f.get(FIELD_TO_RECOGNIZED) or 0)
        montant_total = f.get(FIELD_MONTANT_TOTAL)
        remise        = f.get(FIELD_REMISE)

        # Blocage — "Aucun" si pas de blocage, "NC" pour détail/date si pas de blocage
        blocage_raw = f.get(FIELD_BLOCAGE)
        blocage_val = None
        if isinstance(blocage_raw, dict):
            val = blocage_raw.get("value") or blocage_raw.get("name")
            blocage_val = val if val and val.lower() != "aucun" else None
        elif blocage_raw:
            blocage_val = str(blocage_raw)
        blocage          = blocage_val or "Aucun"   # colonne Blocage
        blocage_detail   = "" if blocage_val else "NC"   # colonne Détail blocage
        blocage_date_def = None if blocage_val else "NC" # colonne Date déblocage

        # Parent Epic PDMDEP
        parent   = f.get("parent")
        epic_key = parent["key"] if parent else None
        epic_info = epic_index.get(epic_key, {}) if epic_key else {}

        # Ticket PSC lié : uniquement si pas d'epic parent
        # type_lien="Relates", issuetype="New delivery" (casse exacte Jira)
        ps_pdm_key = None
        if not epic_key:
            for link in (f.get("issuelinks") or []):
                for direction in ("inwardIssue", "outwardIssue"):
                    linked = link.get(direction)
                    if not linked:
                        continue
                    linked_key = linked.get("key", "")
                    issue_type_name = linked.get("fields", {}).get("issuetype", {}).get("name", "")
                    if (linked_key.startswith("PSC")
                            and issue_type_name.lower() == "new delivery"):
                        ps_pdm_key = linked_key
                        break
                if ps_pdm_key:
                    break

        # Jalon = statut de l'Epic ; % lu sur le ticket CA
        jalon_nom = epic_info.get("epic_status")
        jalon_pct_raw = f.get(FIELD_JALON_PCT)
        # Jira stocke 0-100, Excel format 0% attend 0.0-1.0 → diviser par 100
        # 0 ou absent → affiché comme 0% (pas None)
        jalon_pct = (float(jalon_pct_raw) / 100) if jalon_pct_raw not in (None, "") else 0.0

        # Temps restant : via Epic pour les tickets PDMDEP, via PSC New delivery pour les tickets PSC
        if epic_key:
            _temps_sec   = temps_by_epic.get(epic_key, 0)
            _estime_sec  = estime_by_epic.get(epic_key, 0)
            _depas_sec   = depas_by_epic.get(epic_key, 0)
        elif ps_pdm_key:
            _temps_sec   = temps_by_psc.get(ps_pdm_key, 0)
            _estime_sec  = estime_by_psc.get(ps_pdm_key, 0)
            _depas_sec   = depas_by_psc.get(ps_pdm_key, 0)
        else:
            _temps_sec  = 0
            _estime_sec = 0
            _depas_sec  = 0

        # Prévisionnel mensuel
        prev = {mois: f.get(fid) for mois, fid in FIELD_PREV.items()}

        rows.append({
            "key":            issue.key,
            "epic_key":       epic_key,
            "ps_pdm_key":     ps_pdm_key,
            "bdc":            f.get(FIELD_BDC) or "",
            "date_reception":  epic_info.get("epic_created") or (f.get("created") or "")[:10],
            "summary":        f.get("summary", ""),
            "status":         f.get("status", {}).get("name", ""),
            "to_recognized":  to_recognized,
            "montant_total":  montant_total,
            "remise":         remise,
            "blocage":          blocage,
            "blocage_detail":   f.get(FIELD_BLOCAGE_DETAIL) or blocage_detail,
            "blocage_date_def": f.get(FIELD_BLOCAGE_DATE) or blocage_date_def,
            "solution_cible": SOLUTION_MAP.get(epic_info.get("solution_cible"), epic_info.get("solution_cible")) or ("LITTERALIS" if ps_pdm_key else None),
            "type_epic":      epic_info.get("type_epic"),
            "client":         epic_info.get("client") or opt_val(f, FIELD_CLIENT),
            "chef_de_projet":      (f.get("assignee") or {}).get("displayName"),
            "service_responsable": opt_val(f, FIELD_SERVICE_RESPONSABLE) or "Delivery",
            "nom_projet":     epic_info.get("nom_projet") or opt_val(f, FIELD_NOM_PROJET),
            "jalon_nom":      jalon_nom,
            "jalon_pct":      jalon_pct,
            "action_attendue":action_attendue(jalon_pct),
            "temps_restant":  _temps_sec,   # en secondes, réparti après
            "temps_estime":   _estime_sec,  # en secondes
            "depassement":    _depas_sec,   # en secondes
            **{f"prev_{m}": v for m, v in prev.items()},
        })

    base_cols = ["key", "epic_key", "ps_pdm_key", "bdc", "date_reception", "summary", "status",
                 "to_recognized", "montant_total", "remise", "blocage", "blocage_detail", "blocage_date_def", "solution_cible", "type_epic",
                 "client", "chef_de_projet", "service_responsable", "nom_projet", "jalon_nom", "jalon_pct",
                 "action_attendue", "temps_restant", "temps_estime", "depassement"]
    prev_cols = [f"prev_{m}" for m in FIELD_PREV]
    cols = base_cols + prev_cols
    df = pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)

    # Répartition uniforme du temps restant entre tous les CA d'un même Epic
    if not df.empty:
        ca_per_epic = df[df["epic_key"].notna()].groupby("epic_key")["key"].transform("count")
        df["temps_restant"] = df.apply(
            lambda r: round((r["temps_restant"] / 3600) / ca_per_epic[r.name], 2)
                      if pd.notna(r["epic_key"]) and r["temps_restant"] > 0
                      else (round(r["temps_restant"] / 3600, 2) if r["temps_restant"] else 0),
            axis=1
        )
        df["temps_estime"] = df.apply(
            lambda r: round((r["temps_estime"] / 3600) / ca_per_epic[r.name], 2)
                      if pd.notna(r["epic_key"]) and r["temps_estime"] > 0
                      else (round(r["temps_estime"] / 3600, 2) if r["temps_estime"] else 0),
            axis=1
        )
        df["depassement"] = df.apply(
            lambda r: round((r["depassement"] / 3600) / ca_per_epic[r.name], 2)
                      if pd.notna(r["epic_key"]) and r["depassement"] > 0
                      else (round(r["depassement"] / 3600, 2) if r["depassement"] else 0),
            axis=1
        )

        # Ajustement selon service_responsable :
        # - hors Delivery (et hors Multi-pôles) → temps = 0
        # - Multi-pôles → temps / 2
        def _ajuster_temps(val, service):
            if service == "Delivery":
                return val
            elif service == "Multi-pôles":
                return round(val / 2, 2) if val else 0
            else:
                return 0

        df["temps_restant"] = df.apply(
            lambda r: _ajuster_temps(r["temps_restant"], r["service_responsable"]), axis=1
        )
        df["temps_estime"] = df.apply(
            lambda r: _ajuster_temps(r["temps_estime"], r["service_responsable"]), axis=1
        )
    return df


# ── 2. Métriques ──────────────────────────────────────────────

def compute_metrics(df: pd.DataFrame) -> dict:
    total_backlog     = df["to_recognized"].sum()
    blocage_client    = df[df["blocage"] == "Blocage client"]["to_recognized"].sum()
    blocage_commerce  = df[df["blocage"] == "Blocage commerce"]["to_recognized"].sum()
    blocage_produit   = df[df["blocage"] == "Blocage produit"]["to_recognized"].sum()
    total_bloque      = blocage_client + blocage_commerce + blocage_produit
    total_mobilisable = total_backlog - total_bloque

    nb_tickets = len(df)
    bloque_df  = df[df["blocage"].notna() & (df["blocage"] != "Aucun")]
    nb_bloques = len(bloque_df)

    def nb_b(label):
        return len(bloque_df[bloque_df["blocage"] == label])

    return {
        "total_backlog":     total_backlog,
        "blocage_client":    blocage_client,
        "blocage_commerce":  blocage_commerce,
        "blocage_produit":   blocage_produit,
        "total_bloque":      total_bloque,
        "total_mobilisable": total_mobilisable,
        "nb_tickets":        nb_tickets,
        "nb_bloques":        nb_bloques,
        "nb_b_client":       nb_b("Blocage client"),
        "nb_b_commerce":     nb_b("Blocage commerce"),
        "nb_b_produit":      nb_b("Blocage produit"),
        "nb_b_juridique":    nb_b("Blocage juridique"),
    }


# ── 3. Helpers Excel ──────────────────────────────────────────

def _no_border():
    s = Side(style=None)
    return Border(left=s, right=s, top=s, bottom=s)

def _thin():
    s = Side(style="thin", color="D9D9D9")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, coord_or_row, col=None, value=None, bold=False, bg=None,
       num_fmt=None, h_align="left", font_color=TXT_LABEL, size=10,
       border=None, italic=False, merge_to=None):
    """Set Cell — helper principal."""
    if col is not None:
        c = ws.cell(row=coord_or_row, column=col, value=value)
    else:
        c = ws[coord_or_row]
        c.value = value
    c.font      = Font(name="Calibri", bold=bold, italic=italic, color=font_color, size=size)
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    c.border    = border if border else _no_border()
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        c.number_format = num_fmt
    if merge_to and col:
        ws.merge_cells(start_row=coord_or_row, start_column=col,
                       end_row=coord_or_row, end_column=merge_to)
    return c


def add_donut(ws, data_col, data_row_start, data_row_end,
              lbl_row_start, lbl_row_end, title, anchor, invisible_last=False):
    """Graphique donut avec DoughnutChart natif openpyxl."""
    chart = DoughnutChart()
    chart.title    = title
    chart.style    = 2
    chart.width    = 13
    chart.height   = 10
    chart.holeSize = 50

    col_idx    = column_index_from_string(data_col)
    data_ref   = Reference(ws, min_col=col_idx,
                           min_row=data_row_start, max_row=data_row_end)
    labels_ref = Reference(ws, min_col=1,
                           min_row=lbl_row_start, max_row=lbl_row_end)
    chart.add_data(data_ref)
    chart.set_categories(labels_ref)

    ser = chart.series[0]
    n_points = data_row_end - data_row_start + 1

    # Couleurs des 4 tranches de blocage + gris clair pour "Non bloqués"
    colors = DONUT_COLORS[:4] + ["D9D9D9"]
    for i, hex_color in enumerate(colors[:n_points]):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = hex_color
        ser.dPt.append(pt)

    dl = DataLabelList()
    dl.showPercent  = True
    dl.showVal      = False
    dl.showCatName  = False
    dl.showSerName  = False
    chart.dataLabels = dl

    from openpyxl.chart.legend import Legend
    chart.legend          = Legend()
    chart.legend.position = "b"

    ws.add_chart(chart, anchor)


# ── 4. Écriture d'un onglet ───────────────────────────────────
#
# Layout inspiré de Mars26.xlsx :
#  Col A (col 1)  : labels (fusionnés sur A:C, soit cols 1–3)
#  Col D (col 4)  : grandes valeurs (fusionnées sur D:F)
#  Col G (col 7)  : 2e métrique (fusionnée G:I)
#  Col J (col 10) : 3e métrique (fusionnée J:L) — onglet TOTAL seulement
#
#  Ligne 1  → Titre de l'onglet (A:L)
#  Ligne 2  → Sous-titre / date
#  Ligne 3  → séparateur (hauteur 8)
#  Ligne 4  → bandeau "💰 Montants backlog"
#  Ligne 5  → labels  (Total backlog | Total bloqué | Total mobilisable)
#  Ligne 6  → valeurs (grandes)
#  Ligne 7  → séparateur
#  Ligne 8  → sous-labels blocage (Blocage client | commerce | produit)
#  Ligne 9  → valeurs montants
#  Ligne 10 → séparateur
#  Ligne 11 → bandeau "📋 BDC"
#  Ligne 12 → labels (Total BDC ouverts | BDC bloqués)
#  Ligne 13 → valeurs (grandes)
#  Ligne 14 → séparateur
#  Ligne 15 → bandeau "🔒 Répartition des blocages"
#  Ligne 16 → labels (Blocage client | commerce | produit | juridique)
#  Ligne 17 → valeurs nb
#  Ligne 18 → % (onglet TOTAL seulement)

def write_sheet(ws, m: dict, tab_label: str, has_pct_col: bool):
    ws.sheet_view.showGridLines     = True
    ws.sheet_view.showRowColHeaders = True

    # Colonnes
    col_widths = {"A": 0.8, "B": 0.8, "C": 22, "D": 2,
                  "E": 22, "F": 2, "G": 22, "H": 2, "I": 22,
                  "J": 2,  "K": 22, "L": 2}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Hauteurs
    ws.row_dimensions[1].height  = 24
    ws.row_dimensions[2].height  = 14
    ws.row_dimensions[3].height  = 8
    ws.row_dimensions[4].height  = 22
    ws.row_dimensions[5].height  = 16
    ws.row_dimensions[6].height  = 28
    ws.row_dimensions[7].height  = 8
    ws.row_dimensions[8].height  = 16
    ws.row_dimensions[9].height  = 28
    ws.row_dimensions[10].height = 8
    ws.row_dimensions[11].height = 22
    ws.row_dimensions[12].height = 16
    ws.row_dimensions[13].height = 28
    ws.row_dimensions[14].height = 8
    ws.row_dimensions[15].height = 22
    ws.row_dimensions[16].height = 16
    ws.row_dimensions[17].height = 28
    ws.row_dimensions[18].height = 16

    # ── Ligne 1 : titre principal ──
    ws.merge_cells("C1:L1")
    sc(ws, 1, 3, tab_label, bold=True, size=14, font_color=TXT_TITLE, h_align="left")

    # ── Ligne 2 : sous-titre ──
    ws.merge_cells("C2:L2")
    sc(ws, 1, 3)  # déjà posé
    c2 = ws.cell(row=2, column=3,
                 value=f"Export du {date.today().strftime('%d/%m/%Y')}")
    c2.font      = Font(name="Calibri", size=10, color=TXT_SUB)
    c2.alignment = Alignment(horizontal="left", vertical="center")

    def fmt_eur(v):
        """Formate un montant pour affichage en grande valeur."""
        return v  # On laisse Excel formater via num_fmt

    # ── Ligne 4 : bandeau Montants ──
    ws.merge_cells("C4:L4")
    c4 = ws.cell(row=4, column=3, value="💰  Montants backlog")
    c4.font      = Font(name="Calibri", bold=True, size=12, color=TXT_BAND)
    c4.fill      = PatternFill("solid", fgColor=CLR_BAND)
    c4.alignment = Alignment(horizontal="left", vertical="center")

    # Ligne 5 : labels financiers (3 colonnes : Total backlog | Total bloqué | Total mobilisable)
    for col, label in [(3, "Total backlog"), (5, "Total bloqué"), (7, "Total mobilisable")]:
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        sc(ws, 5, col, label, bold=True, size=10, font_color=TXT_LABEL, h_align="left")

    # Ligne 6 : grandes valeurs financières
    FMT = '#,##0.000 "€"'
    val_color_map = [
        (3, m["total_backlog"],     TXT_VALUE),
        (5, m["total_bloque"],      TXT_RED),
        (7, m["total_mobilisable"], TXT_GREEN),
    ]
    for col, val, color in val_color_map:
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
        c = ws.cell(row=6, column=col, value=val)
        c.font         = Font(name="Calibri", bold=True, size=20, color=color)
        c.fill         = PatternFill("solid", fgColor=CLR_VALUE)
        c.alignment    = Alignment(horizontal="left", vertical="center")
        c.number_format = FMT

    # Ligne 8 : sous-labels blocage
    for col, label in [(3, "Blocage client"), (5, "Blocage commerce"), (7, "Blocage produit")]:
        ws.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)
        sc(ws, 8, col, label, bold=True, size=10, font_color=TXT_LABEL, h_align="left")

    # Ligne 9 : valeurs montants blocage
    for col, val in [(3, m["blocage_client"]), (5, m["blocage_commerce"]), (7, m["blocage_produit"])]:
        ws.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col+1)
        c = ws.cell(row=9, column=col, value=val)
        c.font          = Font(name="Calibri", bold=True, size=20, color=TXT_RED)
        c.fill          = PatternFill("solid", fgColor=CLR_VALUE)
        c.alignment     = Alignment(horizontal="left", vertical="center")
        c.number_format = FMT

    # ── Ligne 11 : bandeau BDC ──
    ws.merge_cells("C11:L11")
    c11 = ws.cell(row=11, column=3, value="📋  BDC ouverts")
    c11.font      = Font(name="Calibri", bold=True, size=12, color=TXT_BAND)
    c11.fill      = PatternFill("solid", fgColor=CLR_BAND)
    c11.alignment = Alignment(horizontal="left", vertical="center")

    # Ligne 12 : labels BDC
    for col, label in [(3, "Total BDC ouverts"), (5, "BDC bloqués")]:
        ws.merge_cells(start_row=12, start_column=col, end_row=12, end_column=col+1)
        sc(ws, 12, col, label, bold=True, size=10, font_color=TXT_LABEL, h_align="left")

    # Ligne 13 : grandes valeurs BDC
    for col, val, color in [(3, m["nb_tickets"], TXT_VALUE), (5, m["nb_bloques"], TXT_RED)]:
        ws.merge_cells(start_row=13, start_column=col, end_row=13, end_column=col+1)
        c = ws.cell(row=13, column=col, value=val)
        c.font          = Font(name="Calibri", bold=True, size=20, color=color)
        c.fill          = PatternFill("solid", fgColor=CLR_VALUE)
        c.alignment     = Alignment(horizontal="left", vertical="center")
        c.number_format = "#,##0"

    # ── Ligne 15 : bandeau répartition blocages ──
    ws.merge_cells("C15:L15")
    c15 = ws.cell(row=15, column=3, value="🔒  Répartition des blocages")
    c15.font      = Font(name="Calibri", bold=True, size=12, color=TXT_BAND)
    c15.fill      = PatternFill("solid", fgColor=CLR_BAND)
    c15.alignment = Alignment(horizontal="left", vertical="center")

    # Ligne 16 : labels répartition (4 colonnes)
    blocage_items = [
        (3, "Blocage client",    m["nb_b_client"]),
        (5, "Blocage commerce",  m["nb_b_commerce"]),
        (7, "Blocage produit",   m["nb_b_produit"]),
        (9, "Blocage juridique", m["nb_b_juridique"]),
    ]
    for col, label, val in blocage_items:
        ws.merge_cells(start_row=16, start_column=col, end_row=16, end_column=col+1)
        sc(ws, 16, col, label, bold=True, size=10, font_color=TXT_LABEL, h_align="left")

    # Ligne 17 : valeurs nb blocage
    for col, label, val in blocage_items:
        ws.merge_cells(start_row=17, start_column=col, end_row=17, end_column=col+1)
        c = ws.cell(row=17, column=col, value=val)
        c.font          = Font(name="Calibri", bold=True, size=20, color=TXT_RED)
        c.fill          = PatternFill("solid", fgColor=CLR_VALUE)
        c.alignment     = Alignment(horizontal="left", vertical="center")
        c.number_format = "#,##0"

    # Ligne 18 : % (onglet TOTAL uniquement)
    if has_pct_col:
        for col, label, val in blocage_items:
            ws.merge_cells(start_row=18, start_column=col, end_row=18, end_column=col+1)
            sc(ws, 18, col,
               f"={chr(64+col)}17/{chr(64+3)}13" if m["nb_tickets"] else 0,
               size=10, font_color=TXT_LABEL, h_align="left",
               num_fmt="0.0%")

    # ── Colonne auxiliaire masquée (col 13 = M) pour le donut ──
    # On ajoute une 5e tranche "Non bloqués" pour que la somme = nb_tickets,
    # ce qui force Excel à calculer les % sur le total BDC (cohérent avec ligne 18).
    # Cette tranche est rendue invisible (couleur blanche).
    aux_labels = ["Blocage client", "Blocage commerce", "Blocage produit", "Blocage juridique", "Non bloqués"]
    nb_non_bloque = m["nb_tickets"] - m["nb_bloques"]
    aux_vals   = [m["nb_b_client"], m["nb_b_commerce"], m["nb_b_produit"], m["nb_b_juridique"], nb_non_bloque]
    for i, (lbl, val) in enumerate(zip(aux_labels, aux_vals)):
        ws.cell(row=20 + i, column=12, value=lbl)
        ws.cell(row=20 + i, column=13, value=val)
    ws.column_dimensions["L"].width = 0.1
    ws.column_dimensions["M"].width = 0.1

    # Graphique donut pointant sur la colonne auxiliaire (5 tranches, la 5e invisible)
    add_donut(ws,
              data_col="M", data_row_start=20, data_row_end=24,
              lbl_row_start=20, lbl_row_end=24,
              title="Répartition des BDC bloqués",
              anchor="C20",
              invisible_last=True)



# ── 7. Onglet détail ──────────────────────────────────────────

def write_detail_sheet(ws, df: pd.DataFrame):
    """
    Onglet détail — colonnes conformes au modèle Suivi_projets.xlsx :

    ADMINISTRATIF
      BDC / Order | Ticket CA | Ticket Jira (Projet/Epic) | Solution
      Client | Nom du projet | Montant restant

    AVANCEMENT
      Jalon actuel (nom) | Jalon actuel (%) | Action attendue | Temps restant

    BLOCAGE
      Blocage | Détail blocage | Date de déblocage estimée

    PRÉVISIONNEL
      Juin | Juillet | Août | Septembre | Octobre | Novembre | Décembre

    MANUEL (vide à remplir)
      Type de prestation | Date de prochaine reconnaissance
    """
    from openpyxl.utils import get_column_letter

    ws.sheet_view.showGridLines     = True
    ws.sheet_view.showRowColHeaders = True

    # ── Définition des colonnes ──
    # (label, largeur, clé_df_ou_None, num_fmt, section)
    COLS = [
        # --- ADMINISTRATIF ---
        ("BDC / Order",                  18, "bdc",            None,         "Administratif"),
        ("Date de réception",            16, "date_reception", "DD/MM/YYYY", "Administratif"),
        ("Ticket CA",                    16, "key",            None,         "Administratif"),
        ("Ticket Jira (Projet)",         18, "ps_pdm_key",     None,         "Administratif"),
        ("Solution",                     16, "solution_cible", None,         "Administratif"),
        ("Client",                       22, "client",         None,         "Administratif"),
        ("Chef de projet",               20, "chef_de_projet",      None,         "Administratif"),
        ("Nom du projet",                28, "nom_projet",     None,         "Administratif"),
        ("Montant restant",              18, "to_recognized",  '#,##0.000 "€"', "Administratif"),
        ("Montant total",                18, "montant_total",  '#,##0.000 "€"', "Administratif"),
        ("Remise",                       18, "remise",         '#,##0.000 "€"', "Administratif"),
        # --- AVANCEMENT ---
        ("Jalon actuel (nom)",           22, "jalon_nom",      None,         "Avancement"),
        ("Jalon actuel (%)",             16, "jalon_pct",      "0%",         "Avancement"),
        ("Action attendue",              20, "action_attendue",     None,         "Avancement"),
        ("Service responsable",          20, "service_responsable", None,         "Avancement"),
        ("Temps restant (h)",            18, "temps_restant",  "#,##0.00",   "Avancement"),
        ("Temps estimé (h)",             18, "temps_estime",   "#,##0.00",   "Avancement"),
        ("Dépassement (h)",              18, "depassement",    "#,##0.00",   "Avancement"),
        # --- BLOCAGE ---
        ("Blocage",                      22, "blocage",        None,         "Blocage"),
        ("Détail blocage",               30, "blocage_detail", None,         "Blocage"),
        ("Date de déblocage estimée",    24, "blocage_date_def","MMM-YYYY",   "Blocage"),
        # --- PRÉVISIONNEL ---
        ("Juin",                          14, "prev_Juin",       '#,##0.000 "€"', "Prévisionnel"),
        ("Juillet",                         14, "prev_Juillet",      '#,##0.000 "€"', "Prévisionnel"),
    ("Août",                      14, "prev_Août",   '#,##0.000 "€"', "Prévisionnel"),
        ("Septembre",                         14, "prev_Septembre",      '#,##0.000 "€"', "Prévisionnel"),
        ("Octobre",                    14, "prev_Octobre", '#,##0.000 "€"', "Prévisionnel"),
        ("Novembre",                      14, "prev_Novembre",   '#,##0.000 "€"', "Prévisionnel"),
        ("Décembre",                     14, "prev_Décembre",  '#,##0.000 "€"', "Prévisionnel"),

    ]

    # Couleurs par section
    SECTION_COLORS = {
        "Administratif": ("1F3864", "BDD7EE"),   # (bandeau, en-tête)
        "Avancement":    ("375623", "C6EFCE"),
        "Blocage":       ("843C0C", "FCE4D6"),
        "Prévisionnel":  ("4B2D83", "E2CFED"),
    }

    n_cols = len(COLS)

    # Largeurs colonnes
    for i, (_, w, *_rest) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Ligne 1 : bandeau titre ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1,
                value=f"Suivi des projets — {date.today().strftime('%d/%m/%Y')}")
    c.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Ligne 2 : bandeaux de section (fusion par section) ──
    ws.row_dimensions[2].height = 18
    section_ranges = {}
    for i, (*_, section) in enumerate(COLS, start=1):
        if section not in section_ranges:
            section_ranges[section] = [i, i]
        else:
            section_ranges[section][1] = i

    for section, (c_start, c_end) in section_ranges.items():
        band_clr, _ = SECTION_COLORS[section]
        if c_start < c_end:
            ws.merge_cells(start_row=2, start_column=c_start,
                           end_row=2, end_column=c_end)
        c = ws.cell(row=2, column=c_start, value=section)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=band_clr)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Ligne 3 : en-têtes de colonnes ──
    ws.row_dimensions[3].height = 30
    for i, (label, _, _key, _fmt, section) in enumerate(COLS, start=1):
        _, hdr_clr = SECTION_COLORS[section]
        c = ws.cell(row=3, column=i, value=label)
        c.font      = Font(name="Calibri", bold=True, size=9, color="1F1F1F")
        c.fill      = PatternFill("solid", fgColor=hdr_clr)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border    = Border(
            right=Side(style="thin", color="FFFFFF"),
            bottom=Side(style="medium", color="FFFFFF")
        )

    # ── Lignes de données ──
    for r_idx, (_, row) in enumerate(df.iterrows(), start=4):
        bg = "F7FBFF" if r_idx % 2 == 0 else "FFFFFF"
        ws.row_dimensions[r_idx].height = 16

        for c_idx, (label, _, key, num_fmt, section) in enumerate(COLS, start=1):
            val = row[key] if key and key in row.index else None
            _, hdr_clr = SECTION_COLORS[section]

            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = Border(
                right=Side(style="thin", color="E0E0E0"),
                bottom=Side(style="thin", color="E0E0E0")
            )

            # Styles spécifiques
            if label == "Blocage" and val == "Aucun":
                c.font = Font(name="Calibri", size=9, color="375623")   # vert
            elif label == "Blocage" and val:
                c.font = Font(name="Calibri", size=9, color="C00000", bold=True)
            elif label in ("Détail blocage", "Date de déblocage estimée") and val == "NC":
                c.font = Font(name="Calibri", size=9, color="595959")   # noir normal
            elif label == "Montant restant":
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.font = Font(name="Calibri", size=9,
                              color="2E75B6" if val and float(val) > 0 else "595959")
            elif label == "Temps restant (h)":
                is_zero = not val  # 0, 0.0, None ou "" → rouge gras
                c.value = val if val else 0
                c.font = Font(name="Calibri", size=9,
                              color="C00000" if is_zero else "595959",
                              bold=is_zero)
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif label in ("Temps estimé (h)", "Dépassement (h)"):
                c.value = val if val else 0
                c.font = Font(name="Calibri", size=9, color="595959")
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif label == "Action attendue":
                color_map = {"Lancement": "2E75B6", "En cours": "ED7D31",
                             "Reconnaissance": "375623"}
                c.font = Font(name="Calibri", size=9, bold=True,
                              color=color_map.get(str(val), "595959"))
            elif label == "Jalon actuel (%)":
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Calibri", size=9, color="595959")
            elif num_fmt and num_fmt.endswith('"€"') and val:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.font = Font(name="Calibri", size=9, color="595959")
            else:
                c.font = Font(name="Calibri", size=9, color="595959")

            if num_fmt:
                c.number_format = num_fmt



    # ── Hyperlinks sur Ticket CA (col B) et Ticket Projet (col C) ──
    col_ca    = next(i for i, (lbl, *_) in enumerate(COLS, 1) if lbl == "Ticket CA")
    col_epic  = next(i for i, (lbl, *_) in enumerate(COLS, 1) if lbl == "Ticket Jira (Projet)")
    for r_idx, (_, row) in enumerate(df.iterrows(), start=4):
        # Ticket CA — clé clicable
        ca_key = row["key"]
        if ca_key:
            c = ws.cell(row=r_idx, column=col_ca)
            c.value     = ca_key                          # clé seule, pas l'URL
            c.hyperlink = f"{JIRA_URL}/browse/{ca_key}"
            c.font = Font(name="Calibri", size=9, color="0563C1", underline="single")
        # Ticket Projet : PS-PDM si dispo, sinon Epic PDMDEP — clé seule
        proj_key = row.get("ps_pdm_key") or row.get("epic_key")
        if proj_key:
            c = ws.cell(row=r_idx, column=col_epic)
            c.value     = proj_key                        # clé seule, pas l'URL
            c.hyperlink = f"{JIRA_URL}/browse/{proj_key}"
            c.font = Font(name="Calibri", size=9, color="0563C1", underline="single")

    # ── Filtre + freeze ──
    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"
    ws.freeze_panes    = "A4"



# ── 8. Fetch & onglet Rework ──────────────────────────────────

def fetch_prestations_gratuites(jira: JIRA) -> pd.DataFrame:
    """Récupère les Epics Maintenance, Prestation offerte et Commande sans prestation."""
    print("  Récupération des Epics Prestations gratuites...")
    jql = (
        f'project = "{JIRA_PROJECT}" AND issuetype = Epic '
        f'AND statusCategory != Done '
        f'AND "Type d\'Epic[Dropdown]" in ("Maintenance", "Prestation offerte", "Commande sans prestation") '
        f'ORDER BY created DESC'
    )
    fields = ",".join([
        FIELD_SOLUTION_CIBLE, FIELD_NOM_PROJET, FIELD_CLIENT,
        FIELD_BLOCAGE, FIELD_BLOCAGE_DETAIL,
        "status", "summary", "subtasks",
    ])
    epics = jira.search_issues(jql, maxResults=False, fields=fields)
    print(f"  → {len(epics)} Epics Prestations gratuites")

    def opt_val(fd, fid):
        v = fd.get(fid)
        if v is None:
            return None
        if isinstance(v, dict):
            val = v.get("value") or v.get("name") or v.get("displayName")
        elif isinstance(v, list):
            val = ", ".join(
                (i.get("value") or i.get("name") or i.get("displayName") or str(i))
                for i in v if i
            )
        else:
            val = str(v)
        return val if val and val.lower() not in ("aucun", "none", "") else None

    # Temps restant : somme timeestimate sur 3 niveaux
    epic_keys  = [e.key for e in epics]
    temps_by_epic = {}
    batch_size = 50

    child_keys_by_epic = []
    for i in range(0, len(epic_keys), batch_size):
        batch    = epic_keys[i:i + batch_size]
        keys_str = ", ".join(f'"{k}"' for k in batch)
        try:
            children = jira.search_issues(
                f'project = "{JIRA_PROJECT}" AND "Epic Link" in ({keys_str})',
                maxResults=False, fields="timeestimate,customfield_10014,subtasks"
            )
            for child in children:
                cf         = child.raw["fields"]
                epic_link  = cf.get("customfield_10014")
                if not epic_link:
                    continue
                temps_by_epic[epic_link] = temps_by_epic.get(epic_link, 0) + float(cf.get("timeestimate") or 0)
                for st in (cf.get("subtasks") or []):
                    child_keys_by_epic.append((epic_link, child.key, st["key"]))
        except Exception as ex:
            print(f"    ⚠ Erreur lot {i//batch_size}: {ex}")

    if child_keys_by_epic:
        all_st_keys  = [st_key for _, _, st_key in child_keys_by_epic]
        st_to_epic   = {st_key: epic_link for epic_link, _, st_key in child_keys_by_epic}
        grand_children_keys = []

        for i in range(0, len(all_st_keys), batch_size):
            batch    = all_st_keys[i:i + batch_size]
            keys_str = ", ".join(f'"{k}"' for k in batch)
            try:
                subtasks = jira.search_issues(
                    f'key in ({keys_str})',
                    maxResults=False, fields="timeestimate,subtasks"
                )
                for st in subtasks:
                    sf        = st.raw["fields"]
                    epic_link = st_to_epic.get(st.key)
                    if epic_link:
                        temps_by_epic[epic_link] = temps_by_epic.get(epic_link, 0) + float(sf.get("timeestimate") or 0)
                        for gst in (sf.get("subtasks") or []):
                            grand_children_keys.append((epic_link, gst["key"]))
            except Exception as ex:
                print(f"    ⚠ Erreur sous-tâches: {ex}")

        if grand_children_keys:
            all_gst_keys = [k for _, k in grand_children_keys]
            gst_to_epic  = {k: epic_link for epic_link, k in grand_children_keys}
            for i in range(0, len(all_gst_keys), batch_size):
                batch    = all_gst_keys[i:i + batch_size]
                keys_str = ", ".join(f'"{k}"' for k in batch)
                try:
                    gsubtasks = jira.search_issues(
                        f'key in ({keys_str})',
                        maxResults=False, fields="timeestimate"
                    )
                    for gst in gsubtasks:
                        epic_link = gst_to_epic.get(gst.key)
                        if epic_link:
                            temps_by_epic[epic_link] = temps_by_epic.get(epic_link, 0) + float(gst.raw["fields"].get("timeestimate") or 0)
                except Exception as ex:
                    print(f"    ⚠ Erreur sous-tâches niv.3: {ex}")

    rows = []
    for epic in epics:
        ef = epic.raw["fields"]

        blocage_raw = ef.get(FIELD_BLOCAGE)
        blocage_val = None
        if isinstance(blocage_raw, dict):
            val = blocage_raw.get("value") or blocage_raw.get("name")
            blocage_val = val if val and val.lower() != "aucun" else None
        elif blocage_raw:
            blocage_val = str(blocage_raw)

        client_raw = ef.get(FIELD_CLIENT)
        if isinstance(client_raw, list):
            client = ", ".join(
                (i.get("displayName") or i.get("name") or "") for i in client_raw if i
            ) or None
        else:
            client = opt_val(ef, FIELD_CLIENT)

        solution_raw = opt_val(ef, FIELD_SOLUTION_CIBLE)
        solution     = SOLUTION_MAP.get(solution_raw, solution_raw)

        temps_sec = temps_by_epic.get(epic.key, 0)
        temps_h   = round(temps_sec / 3600, 2) if temps_sec else 0

        rows.append({
            "key":            epic.key,
            "summary":        ef.get("summary", ""),
            "solution":       solution,
            "client":         client,
            "nom_projet":     opt_val(ef, FIELD_NOM_PROJET),
            "jalon":          (ef.get("status") or {}).get("name"),
            "blocage":        blocage_val or "Aucun",
            "blocage_detail": ef.get(FIELD_BLOCAGE_DETAIL) or ("NC" if not blocage_val else ""),
            "temps_restant":  temps_h,
        })

    cols = ["key", "summary", "solution", "client", "nom_projet",
            "jalon", "blocage", "blocage_detail", "temps_restant"]
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def fetch_rework(jira: JIRA) -> pd.DataFrame:
    """Récupère les Epics de type Rework encore ouverts."""
    print("  Récupération des Epics Rework...")
    jql = (
        f'project = "{JIRA_PROJECT}" AND issuetype = Epic '
        f'AND statusCategory != Done '
        f'AND "Type d\'Epic[Dropdown]" = Rework '
        f'ORDER BY created DESC'
    )
    fields = ",".join([
        FIELD_SOLUTION_CIBLE, FIELD_NOM_PROJET, FIELD_CLIENT,
        FIELD_BLOCAGE, FIELD_BLOCAGE_DETAIL,
        "status", "summary", "subtasks",
    ])
    epics = jira.search_issues(jql, maxResults=False, fields=fields)
    print(f"  → {len(epics)} Epics Rework")

    def opt_val(fd, fid):
        v = fd.get(fid)
        if v is None:
            return None
        if isinstance(v, dict):
            val = v.get("value") or v.get("name") or v.get("displayName")
        elif isinstance(v, list):
            val = ", ".join(
                (i.get("value") or i.get("name") or i.get("displayName") or str(i))
                for i in v if i
            )
        else:
            val = str(v)
        return val if val and val.lower() not in ("aucun", "none", "") else None

    # Temps restant : somme timeestimate de tous les enfants (sous-tâches)
    epic_keys = [e.key for e in epics]
    temps_by_epic = {}
    batch_size = 50

    # Passe 1 : children directement liés à l'Epic
    child_keys_by_epic = []   # [(epic_link, child_key, st_key), ...]
    for i in range(0, len(epic_keys), batch_size):
        batch = epic_keys[i:i + batch_size]
        keys_str = ", ".join(f'"{k}"' for k in batch)
        try:
            children = jira.search_issues(
                f'project = "{JIRA_PROJECT}" AND "Epic Link" in ({keys_str})',
                maxResults=False, fields="timeestimate,customfield_10014,subtasks"
            )
            for child in children:
                cf = child.raw["fields"]
                epic_link = cf.get("customfield_10014")
                if not epic_link:
                    continue
                temps_by_epic[epic_link] = temps_by_epic.get(epic_link, 0) + float(cf.get("timeestimate") or 0)
                for st in (cf.get("subtasks") or []):
                    child_keys_by_epic.append((epic_link, child.key, st["key"]))
        except Exception as ex:
            print(f"    ⚠ Erreur lot {i//batch_size}: {ex}")

    # Passe 2 : sous-tâches des children
    if child_keys_by_epic:
        all_st_keys = [st_key for _, _, st_key in child_keys_by_epic]
        st_to_epic = {st_key: epic_link for epic_link, _, st_key in child_keys_by_epic}
        grand_children_keys = []

        for i in range(0, len(all_st_keys), batch_size):
            batch = all_st_keys[i:i + batch_size]
            keys_str = ", ".join(f'"{k}"' for k in batch)
            try:
                subtasks = jira.search_issues(
                    f'key in ({keys_str})',
                    maxResults=False, fields="timeestimate,subtasks"
                )
                for st in subtasks:
                    sf = st.raw["fields"]
                    epic_link = st_to_epic.get(st.key)
                    if epic_link:
                        temps_by_epic[epic_link] = temps_by_epic.get(epic_link, 0) + float(sf.get("timeestimate") or 0)
                        for gst in (sf.get("subtasks") or []):
                            grand_children_keys.append((epic_link, gst["key"]))
            except Exception as ex:
                print(f"    ⚠ Erreur sous-tâches: {ex}")

        # Passe 3 : sous-tâches niveau 3
        if grand_children_keys:
            all_gst_keys = [k for _, k in grand_children_keys]
            gst_to_epic = {k: epic_link for epic_link, k in grand_children_keys}
            for i in range(0, len(all_gst_keys), batch_size):
                batch = all_gst_keys[i:i + batch_size]
                keys_str = ", ".join(f'"{k}"' for k in batch)
                try:
                    gsubtasks = jira.search_issues(
                        f'key in ({keys_str})',
                        maxResults=False, fields="timeestimate"
                    )
                    for gst in gsubtasks:
                        epic_link = gst_to_epic.get(gst.key)
                        if epic_link:
                            temps_by_epic[epic_link] = temps_by_epic.get(epic_link, 0) + float(gst.raw["fields"].get("timeestimate") or 0)
                except Exception as ex:
                    print(f"    ⚠ Erreur sous-tâches niv.3: {ex}")

    rows = []
    for epic in epics:
        ef = epic.raw["fields"]

        blocage_raw = ef.get(FIELD_BLOCAGE)
        blocage_val = None
        if isinstance(blocage_raw, dict):
            val = blocage_raw.get("value") or blocage_raw.get("name")
            blocage_val = val if val and val.lower() != "aucun" else None
        elif blocage_raw:
            blocage_val = str(blocage_raw)

        client_raw = ef.get(FIELD_CLIENT)
        if isinstance(client_raw, list):
            client = ", ".join(
                (i.get("displayName") or i.get("name") or "") for i in client_raw if i
            ) or None
        else:
            client = opt_val(ef, FIELD_CLIENT)

        solution_raw = opt_val(ef, FIELD_SOLUTION_CIBLE)
        solution     = SOLUTION_MAP.get(solution_raw, solution_raw)

        temps_sec    = temps_by_epic.get(epic.key, 0)
        temps_h      = round(temps_sec / 3600, 2) if temps_sec else None

        rows.append({
            "key":           epic.key,
            "summary":       ef.get("summary", ""),
            "solution":      solution,
            "client":        client,
            "nom_projet":    opt_val(ef, FIELD_NOM_PROJET),
            "jalon":         (ef.get("status") or {}).get("name"),
            "blocage":       blocage_val or "Aucun",
            "blocage_detail":ef.get(FIELD_BLOCAGE_DETAIL) or ("NC" if not blocage_val else ""),
            "temps_restant": temps_h,
        })

    cols = ["key", "summary", "solution", "client", "nom_projet",
            "jalon", "blocage", "blocage_detail", "temps_restant"]
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def write_rework_sheet(ws, df: pd.DataFrame):
    from openpyxl.utils import get_column_letter

    ws.sheet_view.showGridLines     = True
    ws.sheet_view.showRowColHeaders = True

    COLS = [
        ("Ticket Epic",       16, "key",           None,          "Administratif"),
        ("Solution",          16, "solution",       None,          "Administratif"),
        ("Client",            22, "client",         None,          "Administratif"),
        ("Nom du projet",     30, "nom_projet",     None,          "Administratif"),
        ("Jalon",             20, "jalon",          None,          "Avancement"),
        ("Temps restant (h)", 18, "temps_restant",  "#,##0.00",    "Avancement"),
        ("Blocage",           22, "blocage",        None,          "Blocage"),
        ("Détail blocage",    30, "blocage_detail", None,          "Blocage"),
    ]
    SECTION_COLORS = {
        "Administratif": ("1F3864", "BDD7EE"),
        "Avancement":    ("375623", "C6EFCE"),
        "Blocage":       ("843C0C", "FCE4D6"),
    }
    n_cols = len(COLS)

    for i, (_, w, *_r) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Ligne 1 : bandeau titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1,
                value=f"Analyse Rework — {date.today().strftime('%d/%m/%Y')}")
    c.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Ligne 2 : bandeaux section
    ws.row_dimensions[2].height = 16
    section_ranges = {}
    for i, (*_, section) in enumerate(COLS, start=1):
        if section not in section_ranges:
            section_ranges[section] = [i, i]
        else:
            section_ranges[section][1] = i
    for section, (c_start, c_end) in section_ranges.items():
        band_clr, _ = SECTION_COLORS[section]
        if c_start < c_end:
            ws.merge_cells(start_row=2, start_column=c_start, end_row=2, end_column=c_end)
        c = ws.cell(row=2, column=c_start, value=section)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=band_clr)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Ligne 3 : en-têtes
    ws.row_dimensions[3].height = 28
    for i, (label, _, _key, _fmt, section) in enumerate(COLS, start=1):
        _, hdr_clr = SECTION_COLORS[section]
        c = ws.cell(row=3, column=i, value=label)
        c.font      = Font(name="Calibri", bold=True, size=9, color="1F1F1F")
        c.fill      = PatternFill("solid", fgColor=hdr_clr)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = Border(
            right=Side(style="thin", color="FFFFFF"),
            bottom=Side(style="medium", color="FFFFFF")
        )

    # Données
    for r_idx, (_, row) in enumerate(df.iterrows(), start=4):
        bg = "F7FBFF" if r_idx % 2 == 0 else "FFFFFF"
        ws.row_dimensions[r_idx].height = 16
        for c_idx, (label, _, key, num_fmt, section) in enumerate(COLS, start=1):
            val = row[key] if key and key in row.index else None
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = Border(
                right=Side(style="thin", color="E0E0E0"),
                bottom=Side(style="thin", color="E0E0E0")
            )
            if label == "Blocage" and val == "Aucun":
                c.font = Font(name="Calibri", size=9, color="375623")
            elif label == "Blocage" and val:
                c.font = Font(name="Calibri", size=9, color="C00000", bold=True)
            elif label == "Détail blocage" and val == "NC":
                c.font = Font(name="Calibri", size=9, color="595959")
            elif label == "Temps restant (h)":
                is_zero = not val  # 0, 0.0, None ou "" → rouge gras
                c.value = val if val else 0
                c.font = Font(name="Calibri", size=9,
                              color="C00000" if is_zero else "595959",
                              bold=is_zero)
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.font = Font(name="Calibri", size=9, color="595959")
            if num_fmt:
                c.number_format = num_fmt

        # Hyperlink ticket Epic
        key_val = row["key"]
        if key_val:
            c = ws.cell(row=r_idx, column=1)
            c.value     = key_val
            c.hyperlink = f"{JIRA_URL}/browse/{key_val}"
            c.font = Font(name="Calibri", size=9, color="0563C1", underline="single")

    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"
    ws.freeze_panes    = "A4"



# ── 6. Main ───────────────────────────────────────────────────

def main():
    print("Connexion à Jira...")
    jira = connect_jira()
    print("Connecté ✓")

    print("\nRécupération des tickets...")
    df = fetch_tickets(jira)

    print("\nAperçu :")
    print(f"  blocage : {df['blocage'].value_counts(dropna=False).to_dict()}")
    print(f"  → {len(df)} tickets CA au total")

    if df.empty:
        print("\n⚠ Aucun ticket trouvé.")
        return

    df_litt  = df[df["solution_cible"] == "LITTERALIS"]
    df_geodp = df[df["solution_cible"] == "GEODP"]
    print(f"  → {len(df_litt)} LITTERALIS / {len(df_geodp)} GEODP")

    m_total = compute_metrics(df)
    m_litt  = compute_metrics(df_litt)
    m_geodp = compute_metrics(df_geodp)

    print("\nGénération du fichier Excel...")
    wb = openpyxl.Workbook()

    ws_total = wb.active
    ws_total.title = f"TOTAL_{today}"
    write_sheet(ws_total, m_total,
                tab_label=f"Vue globale — {date.today().strftime('%d/%m/%Y')}",
                has_pct_col=True)

    ws_litt = wb.create_sheet(f"LITTERALIS_{today}")
    write_sheet(ws_litt, m_litt,
                tab_label=f"LITTERALIS — {date.today().strftime('%d/%m/%Y')}",
                has_pct_col=False)

    ws_geodp = wb.create_sheet(f"GEODP_{today}")
    write_sheet(ws_geodp, m_geodp,
                tab_label=f"GEODP — {date.today().strftime('%d/%m/%Y')}",
                has_pct_col=False)

    ws_detail = wb.create_sheet(f"DETAIL_{today}")
    write_detail_sheet(ws_detail, df)

    # ── Onglet Rework unique (TOTAL / LITTERALIS / GEODP) ──
    print("\nRécupération des Epics Rework...")
    df_rework = fetch_rework(jira)
    print(f"  → {len(df_rework)} Epics Rework")

    df_rw_litt  = df_rework[df_rework["solution"] == "LITTERALIS"]
    df_rw_geodp = df_rework[df_rework["solution"] == "GEODP"]

    ws_rework = wb.create_sheet(f"REWORK_{today}")
    write_rework_sheet(ws_rework, df_rework)

    # ── Onglet Prestations gratuites ──
    print("\nRécupération des Epics Prestations gratuites...")
    df_pg = fetch_prestations_gratuites(jira)
    ws_pg = wb.create_sheet(f"PRESTATIONS_GRATUITES_{today}")
    write_rework_sheet(ws_pg, df_pg)

    wb.save(OUTPUT_FILE)
    print(f"\n✓ Fichier sauvegardé : {OUTPUT_FILE}")


if __name__ == "__main__":
    main()