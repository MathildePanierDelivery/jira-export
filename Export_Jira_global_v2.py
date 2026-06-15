"""
Export Worklog — V3
-------------------
Améliorations vs V2 :
  - Token Jira lu depuis variable d'environnement (JIRA_TOKEN)
  - REFERENCE_DATE automatique (mois courant)
  - Absences lues depuis un fichier Excel d'entrée (absences.xlsx)
  - Gestion d'erreurs API avec try/except + messages clairs
  - Correction du bug de filtre df_wl (week_end → month_end)
  - Classe WL définie une seule fois (SimpleNamespace)
  - solution_mapping défini en global (non répété dans la boucle)
  - issue_map construit une seule fois et enrichi en place
  - Correction du bug de scope sur type_presta
  - get_all_children optimisé (pool précalculé)
  - Imports dédupliqués et regroupés
"""

import os
import sys
from datetime import datetime, timedelta
from types import SimpleNamespace

import holidays
import pandas as pd
from jira import JIRA
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ══════════════════════════════════════════════════════════════════
# PALETTE & STYLES GLOBAUX
# ══════════════════════════════════════════════════════════════════
BLUE_DARK   = "1F3864"   # entêtes principaux
BLUE_MID    = "2E75B6"   # entêtes secondaires / titres
BLUE_LIGHT  = "D6E4F0"   # alternance lignes paires
BLUE_XLIGHT = "EBF3FB"   # fond léger
GREY_LINE   = "BDD7EE"   # bordures
WHITE       = "FFFFFF"
ORANGE      = "C55A11"   # alerte / rework
GREEN       = "375623"   # positif
RED_LIGHT   = "FFCCCC"   # alerte rouge

FONT_DEFAULT = "Calibri"
FONT_SIZE    = 11

def _font(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name=FONT_DEFAULT, bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _border(color=GREY_LINE):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

STYLE_HEADER_DARK = {
    "font"      : _font(bold=True, color=WHITE),
    "fill"      : _fill(BLUE_DARK),
    "alignment" : _align("center"),
    "border"    : _border(BLUE_DARK),
}
STYLE_HEADER_MID = {
    "font"      : _font(bold=True, color=WHITE),
    "fill"      : _fill(BLUE_MID),
    "alignment" : _align("center"),
    "border"    : _border(BLUE_MID),
}
STYLE_TOTAL = {
    "font"      : _font(bold=True),
    "fill"      : _fill(BLUE_LIGHT),
    "alignment" : _align("center"),
    "border"    : _border(),
}

def apply_style(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)

def style_row(ws, row_idx, ncols, even=False, bold=False, total=False):
    """Applique alternance + bordures sur une ligne entière."""
    fill_color = BLUE_XLIGHT if even else WHITE
    if total:
        fill_color = BLUE_LIGHT
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill      = _fill(fill_color)
        cell.border    = _border()
        cell.font      = _font(bold=bold or total)
        cell.alignment = _align("left")

def style_header_row(ws, row_idx, ncols, dark=True):
    style = STYLE_HEADER_DARK if dark else STYLE_HEADER_MID
    for col in range(1, ncols + 1):
        apply_style(ws.cell(row=row_idx, column=col), style)

def autofit(ws, min_w=10, max_w=50, padding=3):
    """Ajuste la largeur de chaque colonne au contenu."""
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + padding, min_w), max_w)

def write_sheet_title(ws, title, subtitle=None):
    """Écrit le titre en A1, le sous-titre en A2, et laisse A3 vide.
    Les sections qui suivent n'ont besoin que d'un ws.append([]) pour
    atterrir en ligne 4 (en-tête de tableau)."""
    ws["A1"] = title
    ws["A1"].font = _font(bold=True, size=14, color=BLUE_DARK)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = _font(italic=True, size=10, color="595959")
    ws.row_dimensions[1].height = 24
    # Ligne 3 vide — le prochain append() tombera en ligne 4
    ws.append([])

# ==============================
# CONFIG
# ==============================
JIRA_SERVER = os.environ["JIRA_URL"]
JIRA_EMAIL  = os.environ["JIRA_EMAIL"]
JIRA_TOKEN  = os.environ["JIRA_API_TOKEN"]          # Jira API token (pas le mot de passe)
EXPORT_FILE    = f"jira_export_{datetime.now():%Y-%m-%d}.xlsx"
ABSENCES_FILE  = "absences.xlsx"   # Fichier Excel d'entrée pour les absences
OBJECTIFS_FILE     = "objectifs.xlsx"
CA_HISTORIQUE_FILE = "ca_historique.xlsx"
PERF_HIST_FILE     = "performance_historique.xlsx"
CHARGE_HIST_FILE      = "charge_historique.xlsx"
RAPPORT_GLOBAL_FILE = "rapport_global.xlsx"
BILLABILITY_FILE     = "Billability_France_PS.xlsx"   # Fichier Billability mensuel

# Valeurs par défaut — écrasées si l'onglet Projets clôturés s'exécute correctement
kpi_clotures_total  = 0
kpi_clotures_rework = 0
stock_rework_actuel = "—"


# ══════════════════════════════════════════════════════════════════
# PÉRIODE DE RÉFÉRENCE
# ══════════════════════════════════════════════════════════════════
# ┌─────────────────────────────────────────────────────────────┐
# │  MODE = "mois_courant"   → suivi en cours de mois          │
# │  MODE = "mois_precedent" → export définitif du 1er du mois │
# └─────────────────────────────────────────────────────────────┘
MODE = "mois_courant"   # ← changer ici selon le besoin

today           = datetime.now()
_first_of_month = today.replace(day=1)

if MODE == "mois_precedent":
    if _first_of_month.month == 1:
        REFERENCE_DATE = _first_of_month.replace(year=_first_of_month.year - 1, month=12)
    else:
        REFERENCE_DATE = _first_of_month.replace(month=_first_of_month.month - 1)
else:  # mois_courant
    REFERENCE_DATE = _first_of_month

week_start  = today - timedelta(days=today.weekday())
week_end    = week_start + timedelta(days=6)
month_start = REFERENCE_DATE
if month_start.month == 12:
    month_end = month_start.replace(year=month_start.year + 1, month=1) - timedelta(seconds=1)
else:
    month_end = month_start.replace(month=month_start.month + 1) - timedelta(seconds=1)

MOIS_FR = ["Janvier","Février","Mars","Avril","Mai","Juin",
           "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]
mois_courant_label = MOIS_FR[month_start.month - 1]

print(f"📅 Mode     : {MODE}")
print(f"📅 Période : {month_start:%Y-%m-%d} → {month_end:%Y-%m-%d}  |  Semaine : {week_start:%Y-%m-%d} → {week_end:%Y-%m-%d}")

# ══════════════════════════════════════════════════════════════════
# ÉQUIPE
# ══════════════════════════════════════════════════════════════════
SOLUTION_MAPPING = {
    "GEODP1"             : "GEODP",
    "GEODP2 new"         : "GEODP",
    "GEODP2 migration"   : "GEODP",
    "LITTERALIS"         : "LITTERALIS",
    "LITTERALIS STANDARD": "LITTERALIS",
    "SHERPA"             : "LITTERALIS",   # Sherpa est inclus dans Littéralis
}

ALLOWED_ACCOUNT_IDS = {
    "63fe26c90a4a47fb8d213c54",                        # QBO
    "712020:bb1c3fc2-7106-4c8b-9955-2e4917706722",     # FRE
    "63fe26c7f00d095406f2590c",                        # BBO
    "6405f3240d9b61193c263860",                        # TBO
    "712020:f59016f5-1528-45f5-a182-d91eb6f69cad",     # FBA
    "712020:e3996f09-39b9-4957-8f83-d925d9543c98",     # MPO
    "712020:30146e03-1c8d-42f4-a105-c6987f9b2e2c",     # MMA
    "712020:ab30e040-78a7-40e1-aef4-ff6781645c8e",     # DHA
    "712020:ba232bcb-61a8-4921-8448-c0792720536c",     # VCA
    "712020:dd9e6928-8a81-4d81-b640-0ae5f038d847",     # ALM
    "712020:a40869a4-1333-4240-ac8a-670ed2314e0b",     # RVI
}

ACCOUNT_TO_COLLAB = {
    "63fe26c90a4a47fb8d213c54"                    : "Quentin Bordillon",
    "712020:bb1c3fc2-7106-4c8b-9955-2e4917706722" : "Fabien Reutenauer",
    "63fe26c7f00d095406f2590c"                    : "Bérénice Bossard",
    "6405f3240d9b61193c263860"                    : "Timon Bos",
    "712020:f59016f5-1528-45f5-a182-d91eb6f69cad" : "Flavie Bardin",
    "712020:e3996f09-39b9-4957-8f83-d925d9543c98" : "Maxime Pontonnier",
    "712020:30146e03-1c8d-42f4-a105-c6987f9b2e2c" : "Marine Masingarbe",
    "712020:ab30e040-78a7-40e1-aef4-ff6781645c8e" : "Duncan Hamelin",
    "712020:ba232bcb-61a8-4921-8448-c0792720536c" : "Valentin Caujolle",
    "712020:dd9e6928-8a81-4d81-b640-0ae5f038d847" : "Alizée Margot",
    "712020:a40869a4-1333-4240-ac8a-670ed2314e0b" : "Rémy Vincent",
}

# Collaborateurs exclus des onglets Capacité / Charge / Performance
# (présents dans Suivi CA, mais pas dans l'analyse CD/CP)
EXCLUDED_FROM_ANALYSIS = {"Alizée Margot", "Valentin Caujolle", "Timon Bos"}

# ══════════════════════════════════════════════════════════════════
# ABSENCES
# ══════════════════════════════════════════════════════════════════
def load_absences(filepath):
    """
    Lit absences.xlsx. Colonnes : Collaborateur | Conges (jours) | Maladie (jours)
    Retourne (total_dict, conges_dict, maladie_dict).
    """
    empty = {}, {}, {}
    if not os.path.exists(filepath):
        print(f"\u26a0\ufe0f  Fichier absences introuvable : {filepath} -> absences a 0.")
        return empty
    try:
        df_abs = pd.read_excel(filepath, header=3).fillna(0)  # en-tête en ligne 4
        col_conges  = next((c for c in df_abs.columns if "cong" in c.lower()), None)
        col_maladie = next((c for c in df_abs.columns if "malad" in c.lower()), None)
        col_collab  = next((c for c in df_abs.columns if "collab" in c.lower()), None)
        if not all([col_collab, col_conges, col_maladie]):
            print(f"\u26a0\ufe0f  Colonnes attendues : Collaborateur | Conges (jours) | Maladie (jours)")
            print(f"   Colonnes trouvees : {list(df_abs.columns)}")
            return empty
        df_abs["_total"] = df_abs[col_conges] + df_abs[col_maladie]
        total   = dict(zip(df_abs[col_collab], df_abs["_total"]))
        conges  = dict(zip(df_abs[col_collab], df_abs[col_conges]))
        maladie = dict(zip(df_abs[col_collab], df_abs[col_maladie]))
        print(f"\u2705 Absences chargees pour {len(total)} collaborateurs.")
        return total, conges, maladie
    except Exception as e:
        print(f"\u26a0\ufe0f  Impossible de lire {filepath} : {e}. Absences a 0.")
        return empty

absences, absences_conges, absences_maladie = load_absences(ABSENCES_FILE)

# ══════════════════════════════════════════════════════════════════
# CONNEXION JIRA
# ══════════════════════════════════════════════════════════════════
try:
    jira = JIRA(options={"server": JIRA_SERVER}, basic_auth=(JIRA_EMAIL, JIRA_TOKEN))
    print("✅ Connexion Jira établie.")
except Exception as e:
    print(f"❌ Impossible de se connecter à Jira : {e}")
    sys.exit(1)

# ══════════════════════════════════════════════════════════════════
# RÉCUPÉRATION DES TICKETS
# ══════════════════════════════════════════════════════════════════
def get_all_issues(jql: str, expand: str = "worklog,subtasks") -> list:
    issues = []
    token  = None
    fields = (
        "summary,issuetype,parent,issuelinks,assignee,"
        "customfield_10221,customfield_24438,customfield_10220,customfield_22998,customfield_23610,"
        "customfield_23103,customfield_20514,customfield_10070,"
        "customfield_23608,customfield_24437,customfield_23955,"
        "customfield_22108,customfield_24529,customfield_24564,customfield_20701,status,created,timespent,worklog,subtasks"
    )
    try:
        while True:
            batch = jira.enhanced_search_issues(
                jql, expand=expand, maxResults=50,
                fields=fields, nextPageToken=token
            )
            issues.extend(batch)
            token = getattr(batch, "nextPageToken", None)
            if not token:
                break
    except Exception as e:
        print(f"❌ Erreur récupération tickets (JQL: {jql!r}) : {e}")
    return issues

print("📥 Récupération des tickets PDMDEP...")
all_issues   = get_all_issues("project = PDMDEP")
print(f"   → {len(all_issues)} tickets.")

print("📥 Récupération des tickets PSC...")
psc_epics    = get_all_issues('project = PSC AND issuetype = "New delivery"')
psc_children = get_all_issues('project = PSC AND issuetype != "New delivery"')
print(f"   → {len(psc_epics)} epics PSC, {len(psc_children)} enfants PSC.")

# ══════════════════════════════════════════════════════════════════
# MAPS GLOBALES
# ══════════════════════════════════════════════════════════════════
issue_map   : dict = {i.key: i for i in all_issues}
worklog_map : dict = {i.key: getattr(i.fields.worklog, "worklogs", []) for i in all_issues}

for issue in psc_epics + psc_children:
    issue_map[issue.key]   = issue
    worklog_map[issue.key] = getattr(issue.fields.worklog, "worklogs", [])

_all_issues_pool = list(all_issues) + list(psc_epics) + list(psc_children)

# ══════════════════════════════════════════════════════════════════
# FONCTIONS UTILITAIRES
# ══════════════════════════════════════════════════════════════════
def _make_worklog_obj(w_dict: dict):
    author = SimpleNamespace(
        accountId   = w_dict["author"]["accountId"],
        displayName = w_dict["author"]["displayName"],
    )
    return SimpleNamespace(
        author           = author,
        timeSpentSeconds = w_dict["timeSpentSeconds"],
        started          = w_dict["started"],
    )


def fetch_worklogs_for_ticket(ticket) -> list:
    """
    Récupère les worklogs d'un ticket.
    - Cas normal (< 20 worklogs embarqués et cohérent) : retour immédiat, aucun appel.
    - Cas ≥ 20 worklogs : appel paginé à /worklog (comportement d'origine).
    - Cas incohérent (0 worklogs embarqués MAIS timeSpent > 0 sur le ticket) :
      appel à /worklog pour récupérer les vrais worklogs.
      Cela couvre le cas Clockwork désactivé/réactivé où le champ embarqué
      revient vide alors que les worklogs existent toujours.
    """
    existing_wl = worklog_map.get(ticket.key, [])

    # Détection d'incohérence : le ticket a du temps logué mais 0 worklogs embarqués
    time_spent = getattr(ticket.fields, "timespent", None) or 0
    embedded_empty_but_has_time = len(existing_wl) == 0 and time_spent > 0

    if len(existing_wl) < 20 and not embedded_empty_but_has_time:
        return existing_wl

    # Appel à l'endpoint dédié /worklog (paginé)
    all_wl      = []
    start_at    = 0
    max_results = 50
    url         = f"{JIRA_SERVER}/rest/api/3/issue/{ticket.key}/worklog"

    while True:
        try:
            resp = jira._session.get(url, params={"startAt": start_at, "maxResults": max_results})
            resp.raise_for_status()
        except Exception as e:
            print(f"⚠️  Worklogs {ticket.key} : {e}")
            return existing_wl
        data  = resp.json()
        batch = data.get("worklogs", [])
        all_wl.extend(_make_worklog_obj(w) for w in batch)
        if len(batch) < max_results:
            break
        start_at += max_results
    return all_wl


def sum_worklogs(keys: list) -> tuple[float, float]:
    week_sec = month_sec = 0
    for k in keys:
        issue = issue_map.get(k)
        if not issue:
            continue
        try:
            wl_list = fetch_worklogs_for_ticket(issue)
        except Exception as e:
            print(f"⚠️  Worklogs ignorés {k} : {e}")
            continue
        for w in wl_list:
            if w.author.accountId not in ALLOWED_ACCOUNT_IDS:
                continue
            started = pd.to_datetime(w.started).date()
            if started < month_start.date() or started > month_end.date():
                continue
            if week_start.date() <= started <= week_end.date():
                week_sec += w.timeSpentSeconds
            month_sec += w.timeSpentSeconds
    return week_sec / 3600, month_sec / 3600


def get_all_children(epic_key: str) -> list:
    children = [
        t.key for t in _all_issues_pool
        if getattr(t.fields, "parent", None) and t.fields.parent.key == epic_key
    ]
    all_keys = set(children)
    for c in children:
        issue = issue_map.get(c)
        if issue:
            all_keys.update(s.key for s in getattr(issue.fields, "subtasks", []))
    return list(all_keys)


def find_linked_epic(issue) -> str | None:
    for link in getattr(issue.fields, "issuelinks", []):
        linked = getattr(link, "outwardIssue", None) or getattr(link, "inwardIssue", None)
        if linked and linked.key.startswith("PSC-") and getattr(link.type, "name", "") == "Relates":
            return linked.key
    return None


def get_category(issue) -> str:
    current = issue
    while True:
        if current.fields.issuetype.name.lower() == "epic":
            return getattr(getattr(current.fields, "customfield_23103", None), "value", "") or ""
        parent = getattr(current.fields, "parent", None)
        if not parent:
            return ""
        current = issue_map.get(parent.key)
        if not current:
            return ""


def get_solution(issue) -> str:
    current = issue
    while True:
        if current.fields.issuetype.name.lower() == "epic":
            return getattr(getattr(current.fields, "customfield_20514", None), "value", "") or ""
        parent = getattr(current.fields, "parent", None)
        if not parent:
            return ""
        current = issue_map.get(parent.key)
        if not current:
            return ""


def _read_client(raw_fields: dict) -> str:
    """
    Lit le client depuis customfield_10070.
    Ce champ peut être un dict {accountId, displayName}, une str, ou None.
    """
    val = raw_fields.get("customfield_10070")
    if not val:
        return ""
    if isinstance(val, dict):
        return val.get("displayName", val.get("name", ""))
    return str(val)


def _make_link(ws, row, col, key):
    """Rend la cellule d'une clé Epic cliquable vers Jira."""
    if not key or not str(key).strip():
        return
    cell = ws.cell(row=row, column=col)
    cell.hyperlink = f"{JIRA_SERVER}/browse/{key}"
    cell.style     = "Hyperlink"


def _read_montant(ca) -> float:
    """
    Lit le montant déclaré selon le type de ticket :
    - "Matériel GEODP" → customfield_10220
    - Autres           → customfield_22998
    """
    if getattr(ca.fields, "issuetype", None) and ca.fields.issuetype.name == "Matériel GEODP":
        return getattr(ca.fields, "customfield_10220", 0) or 0
    return getattr(ca.fields, "customfield_22998", 0) or 0


def _read_bdc(ca_fields_obj, ca_raw_fields: dict) -> str:
    """
    Lit le numéro BDC depuis customfield_10221 en priorité,
    sinon fallback sur customfield_24438 (ancien champ).
    """
    bdc = str(getattr(ca_fields_obj, "customfield_10221", "") or "").strip()
    if not bdc:
        bdc = str(ca_raw_fields.get("customfield_24438", "") or "").strip()
    return bdc


def _read_prestation(raw_fields: dict) -> str:
    """
    Lit le type de prestation depuis customfield_23955 (champ select : Hardware, Services…).
    Ce champ remonte comme un dict {"value": "Hardware", ...} ou une str.
    """
    val = raw_fields.get("customfield_23955")
    if not val:
        return ""
    if isinstance(val, dict):
        return val.get("value", "")
    return str(val)


DEAL_TYPE_MAPPING = {
    "new business"       : "Nouveau déploiement",
    "upsell"             : "Vente additionnelle",
    "additional setup"   : "Vente additionnelle",
}

def _read_deal_type(raw_fields: dict) -> str:
    """
    Détermine le type de deal :
    - "Migration"            si customfield_20514 (Solution cible) = "GEODP2 migration"
    - "Nouveau déploiement"  si customfield_22108 (Deal Type) = "New Business"
    - "Vente additionnelle"            si customfield_22108 = "Upsell" ou "Additional setup"
    - ""                     si non renseigné
    """
    solution_cible = (raw_fields.get("customfield_20514") or {}).get("value", "")
    if solution_cible.lower() == "geodp2 migration":
        return "Migration"
    deal_type_raw = (raw_fields.get("customfield_22108") or {}).get("value", "")
    if isinstance(raw_fields.get("customfield_22108"), str):
        deal_type_raw = raw_fields.get("customfield_22108", "")
    return DEAL_TYPE_MAPPING.get(deal_type_raw.lower().strip(), "")


# ══════════════════════════════════════════════════════════════════
# GÉNÉRATION DES LIGNES — PDMDEP
# ══════════════════════════════════════════════════════════════════
rows = []

for issue in all_issues:
    if issue.fields.issuetype.name != "Epic":
        continue

    epic_key  = issue.key
    epic_name = issue.fields.summary
    cf_obj    = getattr(issue.fields, "customfield_23103", None)

    raw_fields    = issue.raw.get("fields", {})
    raw_solution  = (raw_fields.get("customfield_20514") or {}).get("value", "")
    epic_solution = SOLUTION_MAPPING.get(raw_solution, raw_solution)
    projet        = raw_fields.get("customfield_23608", "")
    epic_category = cf_obj.value if hasattr(cf_obj, "value") else ""
    deal_type     = _read_deal_type(raw_fields)
    assignee_obj  = getattr(issue.fields, "assignee", None)
    assignee      = assignee_obj.displayName if assignee_obj else ""

    # Date de commande (= date de création de l'Epic)
    _created_raw = raw_fields.get("created") or getattr(issue.fields, "created", None)
    try:
        _created_dt = pd.to_datetime(_created_raw) if _created_raw else None
    except Exception:
        _created_dt = None
    date_commande = _created_dt.strftime("%d/%m/%Y") if _created_dt else ""
    ecart_mois    = (
        (month_start.year - _created_dt.year) * 12 + (month_start.month - _created_dt.month)
        if _created_dt else None
    )

    children_keys           = get_all_children(epic_key)
    total_week, total_month = sum_worklogs([epic_key] + children_keys)

    tickets_ca = [
        t for t in all_issues
        if getattr(t.fields, "parent", None)
        and t.fields.parent.key == epic_key
        and t.fields.issuetype.name in ("COORDIN : Suivi CA", "Matériel GEODP")
    ]

    has_amount = any(
        _read_montant(ca) > 0
        or (getattr(ca.fields, "customfield_23610", 0) or 0) != 0
        for ca in tickets_ca
    )
    if total_month <= 0 and not has_amount:
        continue

    nb_ca                  = max(1, len(tickets_ca))
    temps_mois_par_ligne   = total_month / nb_ca

    if tickets_ca:
        for ca in tickets_ca:
            montant                = _read_montant(ca)
            revenu_planifie_ticket = getattr(ca.fields, "customfield_23610", 0) or 0
            bdc                    = _read_bdc(ca.fields, ca.raw.get("fields", {}))
            rentabilite            = round(montant / temps_mois_par_ligne, 2) if temps_mois_par_ligne > 0 else 0
            fields_ca              = ca.raw.get("fields", {})
            prestation             = _read_prestation(fields_ca)
            type_presta            = fields_ca.get("customfield_24437", "") or ""
            client_ca              = _read_client(fields_ca)

            rows.append({
                "Epic"                   : epic_key,
                "Epic Nom"               : epic_name,
                "Assigné"                : assignee,
                "Client"                 : client_ca,
                "Projet"                 : projet,
                "Solution"               : epic_solution,
                "Catégorie"              : epic_category,
                "Type de deal"           : deal_type,
                "Date commande"          : date_commande,
                "Ancienneté (mois)"      : ecart_mois,
                "Temps mois (h)"         : round(temps_mois_par_ligne, 2),
                "Ticket CA"              : ca.key,
                "Type"                   : type_presta,
                "Numéro BDC"             : bdc,
                "Prestation"             : prestation,
                "Prévision du mois"      : revenu_planifie_ticket,
                "Montant déclaré ce mois": montant,
                "Rentabilité (€ / h)"    : rentabilite,
            })
    else:
        rows.append({
            "Epic"                   : epic_key,
            "Epic Nom"               : epic_name,
            "Assigné"                : assignee,
            "Client"                 : _read_client(raw_fields),
            "Projet"                 : projet,
            "Solution"               : epic_solution,
            "Catégorie"              : epic_category,
            "Type de deal"           : deal_type,
            "Date commande"          : date_commande,
            "Ancienneté (mois)"      : ecart_mois,
            "Temps mois (h)"         : round(temps_mois_par_ligne, 2),
            "Ticket CA"              : "",
            "Type"                   : "",
            "Numéro BDC"             : "",
            "Prestation"             : "",
            "Prévision du mois"      : 0,
            "Montant déclaré ce mois": 0,
            "Rentabilité (€ / h)"    : 0,
        })

# ══════════════════════════════════════════════════════════════════
# GÉNÉRATION DES LIGNES — PSC
# ══════════════════════════════════════════════════════════════════
for issue in psc_epics:
    children_keys           = get_all_children(issue.key)
    total_week, total_month = sum_worklogs([issue.key] + children_keys)

    # ✅ CORRECTIF : client lu depuis l'Epic PSC, pas depuis le ticket CA
    psc_raw_fields = issue.raw.get("fields", {})
    client_psc     = _read_client(psc_raw_fields)

    # Date de commande (= date de création de l'Epic PSC)
    _psc_created_raw = psc_raw_fields.get("created") or getattr(issue.fields, "created", None)
    try:
        _psc_created_dt = pd.to_datetime(_psc_created_raw) if _psc_created_raw else None
    except Exception:
        _psc_created_dt = None
    psc_date_commande = _psc_created_dt.strftime("%d/%m/%Y") if _psc_created_dt else ""
    psc_ecart_mois    = (
        (month_start.year - _psc_created_dt.year) * 12 + (month_start.month - _psc_created_dt.month)
        if _psc_created_dt else None
    )

    tickets_ca = [
        t for t in all_issues
        if t.fields.issuetype.name in ("COORDIN : Suivi CA", "Matériel GEODP")
        and find_linked_epic(t) == issue.key
    ]

    has_ca = any(
        _read_montant(ca) > 0
        or (getattr(ca.fields, "customfield_23610", 0) or 0) != 0
        for ca in tickets_ca
    )
    if total_month <= 0 and not has_ca:
        continue

    nb_ca                = max(1, len(tickets_ca))
    temps_mois_par_ligne = total_month / nb_ca

    if tickets_ca:
        for ca in tickets_ca:
            montant                = _read_montant(ca)
            revenu_planifie_ticket = getattr(ca.fields, "customfield_23610", 0) or 0
            bdc                    = _read_bdc(ca.fields, ca.raw.get("fields", {}))
            fields_ca              = ca.raw.get("fields", {})
            prestation             = _read_prestation(fields_ca)
            type_presta            = fields_ca.get("customfield_24437", "") or ""
            rentabilite            = round(montant / temps_mois_par_ligne, 2) if temps_mois_par_ligne > 0 else 0

            rows.append({
                "Epic"                   : issue.key,
                "Epic Nom"               : issue.fields.summary,
                "Assigné"                : assignee,
                "Client"                 : client_psc,
                "Projet"                 : "",
                "Solution"               : "LITTERALIS",
                "Catégorie"              : "Projet",
                "Type de deal"           : "Nouveau déploiement",
                "Date commande"          : psc_date_commande,
                "Ancienneté (mois)"      : psc_ecart_mois,
                "Temps mois (h)"         : round(temps_mois_par_ligne, 2),
                "Ticket CA"              : ca.key,
                "Type"                   : type_presta,
                "Numéro BDC"             : bdc,
                "Prestation"             : prestation,
                "Prévision du mois"      : revenu_planifie_ticket,
                "Montant déclaré ce mois": montant,
                "Rentabilité (€ / h)"    : rentabilite,
            })
    else:
        rows.append({
            "Epic"                   : issue.key,
            "Epic Nom"               : issue.fields.summary,
            "Assigné"                : assignee,
            "Client"                 : client_psc,
            "Projet"                 : "",
            "Solution"               : "LITTERALIS",
            "Catégorie"              : "Projet",
            "Type de deal"           : "Nouveau déploiement",
            "Date commande"          : psc_date_commande,
            "Ancienneté (mois)"      : psc_ecart_mois,
            "Temps mois (h)"         : round(temps_mois_par_ligne, 2),
            "Ticket CA"              : "",
            "Type"                   : "Services",
            "Numéro BDC"             : "",
            "Prestation"             : "",
            "Prévision du mois"      : 0,
            "Montant déclaré ce mois": 0,
            "Rentabilité (€ / h)"    : 0,
        })

# ══════════════════════════════════════════════════════════════════
# WORKLOGS BRUTS (df_wl) — tous collaborateurs hors exclus
# ══════════════════════════════════════════════════════════════════
print("📊 Construction du df_wl...")

worklog_rows = []
for issue in all_issues:
    try:
        wl_list = fetch_worklogs_for_ticket(issue)
    except Exception as e:
        print(f"⚠️  Worklogs ignorés {issue.key} : {e}")
        continue

    for w in wl_list:
        if w.author.accountId not in ALLOWED_ACCOUNT_IDS:
            continue
        started = pd.to_datetime(w.started).date()
        if started < month_start.date() or started > month_end.date():
            continue

        collab          = ACCOUNT_TO_COLLAB.get(w.author.accountId, w.author.displayName)
        ticket_cat      = get_category(issue)
        solution_jira   = get_solution(issue)
        solution_metier = SOLUTION_MAPPING.get(solution_jira, "Autre")

        ticket_status = getattr(getattr(issue.fields, "status", None), "name", "")
        worklog_rows.append({
            "Collaborateur": collab,
            "Ticket"       : issue.key,
            "Epic"         : epic_key,
            "Catégorie"    : ticket_cat,
            "Solution"     : solution_metier,
            "Temps (h)"    : w.timeSpentSeconds / 3600,
            "Date"         : started,
            "Exclu analyse": collab in EXCLUDED_FROM_ANALYSIS,
            "Bloqué"       : ticket_status.lower() == "bloqué",
        })

SUPPORT_LITT_EPICS  = {"PDMDEP-2171", "PDMDEP-12707"}
SUPPORT_GEODP_EPICS = {"PDMDEP-2172"}
DIVERS_LITT_EPIC    = "PDMDEP-31498"
DIVERS_GEODP_EPIC   = "PDMDEP-31499"
def _get_support_solution(ticket_key: str, all_issues_list) -> str:
    """
    Remonte la hiérarchie jusqu'à 4 niveaux pour trouver si un ticket
    appartient au support Littéralis ou GEODP.
    """
    issues_map = {i.key: i for i in all_issues_list}
    key = ticket_key
    for _ in range(4):
        issue = issues_map.get(key)
        if not issue:
            break
        if key in SUPPORT_LITT_EPICS:
            return "LITTERALIS"
        if key in SUPPORT_GEODP_EPICS:
            return "GEODP"
        parent = getattr(issue.fields, "parent", None)
        if not parent:
            break
        key = parent.key
    return ""

df_wl     = pd.DataFrame(worklog_rows)

# Corriger la solution des tickets Support selon PDMDEP-22897
if not df_wl.empty:
    support_mask = df_wl["Catégorie"] == "Support N2"
    for idx in df_wl[support_mask].index:
        ticket = df_wl.at[idx, "Ticket"]
        sol = _get_support_solution(ticket, all_issues)
        if sol:
            df_wl.at[idx, "Solution"] = sol

# df_wl pour les onglets d'analyse (hors Alizée, Valentin & Timon)
df_wl_ana = df_wl[~df_wl["Exclu analyse"]] if not df_wl.empty else df_wl

# ══════════════════════════════════════════════════════════════════
# EXPORT XLSX
# ══════════════════════════════════════════════════════════════════
df = pd.DataFrame(rows)

# Séparation Hardware : Prestation = "Hardware" OU issuetype = "Matériel GEODP"
if not df.empty and "Prestation" in df.columns:
    mask_hw = (
        df["Prestation"].str.strip().str.lower() == "hardware"
    ) | (
        df["Ticket CA"].isin([
            t.key for t in all_issues
            if t.fields.issuetype.name == "Matériel GEODP"
        ])
    )

    # Tickets CA clôturés (status -> "Terminé") durant le mois en cours
    print("📥 Récupération des tickets CA Hardware clôturés ce mois...")
    try:
        jql_hw_clos = (
            f'project = PDMDEP AND issuetype in ("Ticket CA", "Matériel GEODP") '
            f'AND status changed TO "Terminé" '
            f'DURING ("{month_start:%Y-%m-%d}", "{month_end:%Y-%m-%d}")'
        )
        hw_closed_keys = {t.key for t in get_all_issues(jql_hw_clos)}
        print(f"   → {len(hw_closed_keys)} ticket(s) CA Hardware clôturés ce mois.")
    except Exception as e:
        print(f"⚠️  Impossible de récupérer les clôtures Hardware : {e}")
        hw_closed_keys = set()

    # Un ticket Hardware figure dans le mois si son Ticket CA est clôturé ce mois
    mask_hw_clos = df["Ticket CA"].isin(hw_closed_keys)
    df_hardware = df[mask_hw & mask_hw_clos].copy()
    df_services  = df[~mask_hw & ~df["Catégorie"].isin(["Interne", "Support N2"])].copy()
else:
    df_hardware = pd.DataFrame(columns=df.columns)
    df_services  = df.copy()

# df_brut = services uniquement — Hardware exclu de tous les calculs CA
df_brut = df_services.copy()

# Exclure les lignes sans temps, sans prévision et sans montant
df_services = df_services[
    (df_services["Temps mois (h)"] > 0) |
    (df_services["Prévision du mois"] > 0) |
    (df_services["Montant déclaré ce mois"] != 0)
]

wb      = Workbook()

# ──────────────────────────────────────────────────────────────────
# ONGLET 1 : Suivi CA
# ──────────────────────────────────────────────────────────────────
def _write_ca_sheet(ws, df_source, title_str, subtitle_str):
    """Écrit un onglet de type Suivi CA : données + ligne de total."""
    write_sheet_title(ws, title_str, subtitle_str)

    data_start_row = 4
    rows_to_write = list(dataframe_to_rows(df_source, index=False, header=True))
    for r_offset, row_data in enumerate(rows_to_write):
        for c_offset, value in enumerate(row_data):
            ws.cell(row=data_start_row + r_offset, column=c_offset + 1, value=value)

    ncols = len(df_source.columns)
    col_indices = {name: idx for idx, name in enumerate(df_source.columns, start=1)}
    style_header_row(ws, data_start_row, ncols, dark=True)

    epic_col = col_indices.get("Epic")
    # Filtre automatique sur l'en-tête
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(ncols)
    ws.auto_filter.ref = f"A{data_start_row}:{last_col_letter}{data_start_row}"
    for i, row_cells in enumerate(ws.iter_rows(min_row=data_start_row + 1, max_row=ws.max_row)):
        row_i = row_cells[0].row
        style_row(ws, row_i, ncols, even=(i % 2 == 0))
        if epic_col:
            _make_link(ws, row_i, epic_col, ws.cell(row=row_i, column=epic_col).value)
            # Réappliquer le fill car _make_link écrase le style
            epic_cell = ws.cell(row=row_i, column=epic_col)
            epic_cell.fill = _fill(BLUE_LIGHT if (i % 2 == 0) else "FFFFFF")

    # Formatage monétaire sur les données
    money_cols = ["Prévision du mois", "Montant déclaré ce mois", "Rentabilité (€ / h)"]
    for col_name in money_cols:
        if col_name in col_indices:
            col_idx = col_indices[col_name]
            for row in ws.iter_rows(min_row=data_start_row + 1, max_row=ws.max_row,
                                    min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = '#,##0.00 "€"'

    # ── Ligne de total ────────────────────────────────────────────
    if not df_source.empty:
        total_row_idx = ws.max_row + 1
        cols = df_source.columns
        total_previ = df_source["Prévision du mois"].sum()      if "Prévision du mois"      in cols else None
        total_ca    = df_source["Montant déclaré ce mois"].sum() if "Montant déclaré ce mois" in cols else 0
        h_sum       = df_source["Temps mois (h)"].sum()          if "Temps mois (h)"          in cols else 0
        total_rent  = round(total_ca / h_sum, 2) if h_sum > 0 else None

        # Construire la ligne total (valeurs vides sur les colonnes non sommées)
        total_values = []
        for col_name in cols:
            if col_name == "Epic":
                total_values.append("TOTAL")
            elif col_name == "Prévision du mois":
                total_values.append(total_previ)
            elif col_name == "Montant déclaré ce mois":
                total_values.append(total_ca)
            elif col_name == "Rentabilité (€ / h)":
                total_values.append(total_rent)
            else:
                total_values.append("")

        for c_offset, value in enumerate(total_values):
            ws.cell(row=total_row_idx, column=c_offset + 1, value=value)
        style_row(ws, total_row_idx, ncols, total=True)
        for col_name in ["Prévision du mois", "Montant déclaré ce mois", "Rentabilité (€ / h)"]:
            if col_name in col_indices:
                ws.cell(row=total_row_idx, column=col_indices[col_name]).number_format = '#,##0.00 "€"'

    # ── Mise en rouge : Montant déclaré < Prévision (si les deux colonnes existent) ──
    if "Prévision du mois" in col_indices and "Montant déclaré ce mois" in col_indices:
        col_previ = col_indices["Prévision du mois"]
        col_ca    = col_indices["Montant déclaré ce mois"]
        RED_FILL  = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")
        RED_FONT  = _font(bold=False, color="C00000")
        for row_idx in range(data_start_row + 1, ws.max_row):  # hors ligne total
            previ = ws.cell(row=row_idx, column=col_previ).value or 0
            ca    = ws.cell(row=row_idx, column=col_ca).value   or 0
            if isinstance(previ, (int, float)) and isinstance(ca, (int, float)) and previ > 0 and ca < previ:
                cell = ws.cell(row=row_idx, column=col_ca)
                cell.fill = RED_FILL
                cell.font = RED_FONT

    ws.freeze_panes = f"A{data_start_row + 1}"
    autofit(ws)


# Sélection des colonnes pour chaque onglet
COLS_SUIVI_CA = [
    "Epic", "Epic Nom", "Assigné", "Client", "Projet", "Solution", "Catégorie",
    "Type de deal", "Date commande", "Ancienneté (mois)", "Temps mois (h)", "Ticket CA", "Numéro BDC",
    "Prévision du mois", "Montant déclaré ce mois", "Rentabilité (€ / h)",
]
COLS_SUIVI_HW = [
    "Epic", "Epic Nom", "Client", "Projet", "Solution", "Catégorie",
    "Type de deal", "Temps mois (h)", "Ticket CA", "Numéro BDC",
    "Montant déclaré ce mois",
]

def _filter_cols(df, cols):
    return df[[c for c in cols if c in df.columns]].copy()

ws = wb.active
ws.title = "Suivi de production"
_write_ca_sheet(ws, _filter_cols(df_services, COLS_SUIVI_CA),
    f"Suivi de production — {mois_courant_label} {month_start.year}",
    "Prestations de services — Hardware exclu")

# ──────────────────────────────────────────────────────────────────
# ONGLET 2 : Suivi Hardware
# ──────────────────────────────────────────────────────────────────
ws_hw = wb.create_sheet("Suivi Hardware")
_write_ca_sheet(ws_hw, _filter_cols(df_hardware, COLS_SUIVI_HW),
    f"Suivi Hardware — {mois_courant_label} {month_start.year}",
    f"Prestation = Hardware ou type 'Matériel GEODP' — {len(df_hardware)} ligne(s)")

# ──────────────────────────────────────────────────────────────────
# JOURS OUVRÉS
# ──────────────────────────────────────────────────────────────────
start_date       = month_start
end_date         = month_end
all_days         = pd.date_range(start_date, end_date, freq="D")
fr_holidays      = holidays.France(years=start_date.year)
holidays_in_period = [d for d in all_days if d in fr_holidays]
working_days     = [d for d in all_days if d.weekday() < 5 and d not in holidays_in_period]
num_working_days = len(working_days)

# ──────────────────────────────────────────────────────────────────
# ONGLET 2 : Capacité productive
# Inclut CD + CP, exclut Alizée, Valentin & Timon
# ──────────────────────────────────────────────────────────────────
ws_cap = wb.create_sheet("Capacité productive")
write_sheet_title(ws_cap, "Capacité productive — Chargés de déploiement & Chefs de projet",
                  f"Mois de {mois_courant_label} {month_start.year} — {num_working_days} jours ouvrés")

# Collaborateurs analysés : CD + CP (sans Alizée, Valentin & Timon)
collaborateurs_ana = [
    "Bérénice Bossard", "Marine Masingarbe", "Duncan Hamelin",
    "Maxime Pontonnier", "Quentin Bordillon",
    "Fabien Reutenauer", "Flavie Bardin", "Rémy Vincent",
]
taux_global_ana    = [1.00, 1.00, 1.00, 0.80, 1.00, 1.00, 1.00, 1.00]
heures_par_jour    = 7

# Ventilation CP par solution (ce mois — à supprimer le mois prochain quand tous 2 solutions)
CP_SPLIT = {
    "Flavie Bardin"     : {"LITTERALIS": 0.40, "GEODP": 0.60},
    "Fabien Reutenauer" : {"LITTERALIS": 1.00, "GEODP": 0.00},
    "Rémy Vincent"      : {"LITTERALIS": 1.00, "GEODP": 0.00},
}

headers_cap = ["Collaborateur", "Disponibilité", "Nb jours ouvrés",
               "Congés (j)", "Maladie (j)", "Total absences (j)", "Jours effectifs", "Heures totales dispo"]

cap_header_row = 4
ws_cap.append(headers_cap)
style_header_row(ws_cap, cap_header_row, len(headers_cap))

commentaires_t1 = {
    "Collaborateur"        : "Chargés de déploiement et Chefs de projet (hors Alizée Margot, Valentin Caujolle et Timon Bos)",
    "Disponibilité"        : "Temps de travail théorique (100% ou temps partiel)",
    "Nb jours ouvrés"      : "Nombre de jours ouvrés dans le mois en cours",
    "Congés (j)"           : "Jours de congés posés ce mois — à renseigner dans absences.xlsx",
    "Maladie (j)"          : "Jours d'arrêt maladie ce mois — à renseigner dans absences.xlsx",
    "Total absences (j)"   : "Congés + Maladie",
    "Jours effectifs"      : "Nb jours ouvrés - total absences",
    "Heures totales dispo" : "Jours effectifs × heures/jour × disponibilité",
}
for cell in ws_cap[cap_header_row]:
    texte = commentaires_t1.get(cell.value)
    if texte:
        cell.comment = Comment(texte, "Info")

capacites = []
for i, (name, taux) in enumerate(zip(collaborateurs_ana, taux_global_ana)):
    jours_conges    = absences_conges.get(name, 0)
    jours_maladie   = absences_maladie.get(name, 0)
    jours_abs       = absences.get(name, 0)
    jours_effectifs = max(0, num_working_days - jours_abs)
    heures_total    = jours_effectifs * heures_par_jour * taux
    capacites.append((name, taux, jours_effectifs, heures_total))
    ws_cap.append([name, f"{taux*100:.0f}%", num_working_days,
                   jours_conges, jours_maladie, jours_abs,
                   jours_effectifs, round(heures_total, 2)])
    style_row(ws_cap, cap_header_row + 1 + i, len(headers_cap), even=(i % 2 == 0))

# Ligne total
total_row = cap_header_row + 1 + len(collaborateurs_ana)
ws_cap.append(["TOTAL ÉQUIPE", "", "",
               sum(absences_conges.get(n, 0) for n in collaborateurs_ana),
               sum(absences_maladie.get(n, 0) for n in collaborateurs_ana),
               sum(absences.get(n, 0) for n in collaborateurs_ana),
               sum(c[2] for c in capacites),
               round(sum(c[3] for c in capacites), 2)])
style_row(ws_cap, total_row, len(headers_cap), total=True)

# Répartition théorique
repartition_litteralis = {"Productif": 0.75, "Support": 0.20, "Interne": 0.05}
repartition_geodp      = {"Productif": 0.95, "Support": 0.00, "Interne": 0.05}
repartition_cp         = {"Productif": 0.90, "Support": 0.00, "Interne": 0.10}

equipe_litteralis = ["Bérénice Bossard", "Marine Masingarbe", "Duncan Hamelin"]
equipe_geodp      = ["Maxime Pontonnier", "Quentin Bordillon"]
equipe_cp         = ["Fabien Reutenauer", "Flavie Bardin", "Rémy Vincent"]  # inchangé

heures_litteralis = sum(c[3] for c in capacites if c[0] in equipe_litteralis)
heures_geodp      = sum(c[3] for c in capacites if c[0] in equipe_geodp)
heures_cp         = sum(c[3] for c in capacites if c[0] in equipe_cp)

def _write_repartition(ws, label, repartition, heures_equipe):
    ws.append([])
    ws.append([f"Répartition théorique — {label}"])
    title_row = ws.max_row
    ws[title_row][0].font      = _font(bold=True, size=12, color=BLUE_MID)
    ws[title_row][0].alignment = _align()
    ws.append(["Catégorie", "% théorique", "Heures théoriques"])
    hdr_row = ws.max_row
    style_header_row(ws, hdr_row, 3, dark=False)
    for j, (cat, pct) in enumerate(repartition.items()):
        ws.append([cat, f"{int(pct*100)}%", round(heures_equipe * pct, 2)])
        style_row(ws, ws.max_row, 3, even=(j % 2 == 0))

_write_repartition(ws_cap, "Littéralis (CD)", repartition_litteralis, heures_litteralis)
_write_repartition(ws_cap, "GEODP (CD)",      repartition_geodp,      heures_geodp)
_write_repartition(ws_cap, "Chefs de projet", repartition_cp,         heures_cp)

# ── Tableau global équipe ────────────────────────────────────
heures_total_equipe = heures_litteralis + heures_geodp + heures_cp
if heures_total_equipe > 0:
    prod_global = (
        heures_litteralis * repartition_litteralis.get("Productif", 0) +
        heures_geodp      * repartition_geodp.get("Productif", 0) +
        heures_cp         * repartition_cp.get("Productif", 0)
    )
    sup_global = (
        heures_litteralis * repartition_litteralis.get("Support", 0) +
        heures_geodp      * repartition_geodp.get("Support", 0) +
        heures_cp         * repartition_cp.get("Support", 0)
    )
    int_global = heures_total_equipe - prod_global - sup_global
    _write_repartition(ws_cap, "Équipe globale",
        {"Productif": prod_global / heures_total_equipe,
         "Support":   sup_global  / heures_total_equipe,
         "Interne":   int_global  / heures_total_equipe},
        heures_total_equipe)

autofit(ws_cap)

# ══════════════════════════════════════════════════════════════════
# CONSTANTES PARTAGÉES POUR LES ONGLETS DE CHARGE
# ══════════════════════════════════════════════════════════════════
CATS_PRODUCTIF = ["Projet", "Commande sans prestation", "Maintenance", "Prestation offerte"]

CATEGORIES_DET = {
    "Projet"    : ["Projet", "Commande sans prestation", "Prestation offerte", "Maintenance"],
    "Rework"   : ["Rework"],
    "Support"  : ["Support N2"],
    "Interne"  : ["Interne"],
}
REPARTITION_THEO = {
    "Littéralis"    : {"Projet": 0.75, "Support": 0.20, "Interne": 0.05},
    "GEODP"         : {"Projet": 0.95, "Support": 0,    "Interne": 0.05},
    "CP"            : {"Projet": 0.90, "Support": 0,    "Interne": 0.10},
    "CP Littéralis" : {"Projet": 0.90, "Support": 0,    "Interne": 0.10},
    "CP GEODP"      : {"Projet": 0.90, "Support": 0,    "Interne": 0.10},
}

ajustements = {
    "Bérénice"         : 0.25,
    "Marine"           : 0.25,
    "Duncan"           : 0.25,
    "Maxime"           : 0.05,
    "Quentin"          : 0.05,
    "Fabien Reutenauer": 0.10,
    "Flavie Bardin"    : 0.10,
    "Rémy Vincent"     : 0.10,
}

# IDs stables des sous-epics de support (enfants de PDMDEP-22897)


# Charge globale : 2 lignes (Littéralis + GEODP), CP fusionnés dans chaque
# Capacités par solution (CD + CP ventilés selon CP_SPLIT)


capacites_dict = {
    name: heures_total
    for name, taux, jours_effectifs, heures_total in capacites
}

capacites_dict_theo = {
    name: heures_total * (1 - ajustements.get(name, 0))
    for name, taux, jours_effectifs, heures_total in capacites
}

def _cap_solution(sol_key: str) -> float:
    # CD de l'équipe concernée selon la solution
    equipe_cd = equipe_litteralis if sol_key == "LITTERALIS" else equipe_geodp
    cap = sum(capacites_dict_theo.get(m, 0) for m in equipe_cd)
    # CP ventilés selon leur split
    for m in equipe_cp:
        cap += capacites_dict_theo.get(m, 0) * CP_SPLIT.get(m, {}).get(sol_key, 0)
    return cap

cap_litt    = _cap_solution("LITTERALIS")
cap_geodp   = _cap_solution("GEODP")
cap_interne = sum(capacites_dict.get(m, 0) for m in equipe_litteralis + equipe_geodp + equipe_cp)

LIGNES_GLOBAL = {
    "Littéralis": {"solution": "LITTERALIS", "capacite": cap_litt,    "repartition": REPARTITION_THEO["Littéralis"]},
    "GEODP"     : {"solution": "GEODP",      "capacite": cap_geodp,   "repartition": REPARTITION_THEO["GEODP"]},
    "Interne"   : {"solution": None,          "capacite": 0, "repartition": {}},  # Pas de théorique pour l'Interne
}
LIGNES_LITT = {
    "Littéralis": {"solution": "LITTERALIS", "capacite": cap_litt, "repartition": REPARTITION_THEO["Littéralis"]},
}
LIGNES_GEODP = {
    "GEODP": {"solution": "GEODP", "capacite": cap_geodp, "repartition": REPARTITION_THEO["GEODP"]},
}


def pct_vs(reel, theo):
    return round(reel / theo * 100, 1) if theo > 0 else 0



def _charge_tableau(ws, df_source, lignes: dict, title: str, subtitle: str):
    """
    Génère un onglet de charge complet.
    lignes = {
        "Littéralis": {"solution": "LITTERALIS", "capacite": 420, "repartition": {...}},
        "GEODP"     : {"solution": "GEODP",      "capacite": 252, "repartition": {...}},
        "Interne"   : {"solution": None,          "capacite": 0,   "repartition": {}},
    }
    Le filtrage se fait par Solution du ticket, pas par collaborateur.
    """
    write_sheet_title(ws, title, subtitle)

    # ── Tableau principal ─────────────────────────────────────────
    headers_cd = [
        "Ligne",
        "Productif (h)", "Théorique (h)", "% Productif",
        "Support (h)",   "Théorique (h)", "% Support",
        "Interne (h)",   "Théorique (h)", "% Interne",
        "Total réalisé (h)", "Total théorique (h)", "% Saisie",
    ]
    has_interne = any(cfg.get("solution") is None for cfg in lignes.values())
    if not has_interne:
        headers_cd = [
            "Ligne",
            "Productif (h)", "Théorique (h)", "% Productif",
            "Support (h)",   "Théorique (h)", "% Support",
            "Total réalisé (h)", "Total théorique (h)", "% Saisie",
        ]
    hdr_row = 4
    ws.append(headers_cd)
    style_header_row(ws, hdr_row, len(headers_cd))

    for i, (label, cfg) in enumerate(lignes.items()):
        sol = cfg.get("solution")
        cap = cfg.get("capacite", 0)
        rep = cfg.get("repartition", {})

        if sol is None:
            # Interne : heures sans solution assignée (toutes catégories Interne)
            df_team = df_source[df_source["Catégorie"] == "Interne"] if not df_source.empty else pd.DataFrame()
        else:
            df_team = df_source[df_source["Solution"] == sol] if not df_source.empty else pd.DataFrame()

        heures_cat = {
            cat: df_team[df_team["Catégorie"].isin(cats)]["Temps (h)"].sum() if not df_team.empty else 0
            for cat, cats in CATEGORIES_DET.items()
        }
        # Productif inclut le Rework
        heures_cat["Projet"] = heures_cat["Projet"] + heures_cat.get("Rework", 0)
        total_reel  = heures_cat["Projet"] + heures_cat["Support"] + heures_cat["Interne"]
        heures_theo = {cat: cap * pct for cat, pct in rep.items()}
        # Théorique Interne = 0 pour les lignes par solution (non ventilable)
        if sol is not None:
            heures_theo["Interne"] = 0
        cap_theo    = sum(heures_theo.values())
        pct_saisie  = round(total_reel / cap * 100, 1) if cap > 0 else 0

        if has_interne:
            ws.append([
                label,
                round(heures_cat["Projet"],  2), round(heures_theo.get("Projet",  0), 2), pct_vs(heures_cat["Projet"],  heures_theo.get("Projet",  0)),
                round(heures_cat["Support"], 2), round(heures_theo.get("Support", 0), 2), pct_vs(heures_cat["Support"], heures_theo.get("Support", 0)),
                round(heures_cat["Interne"], 2), round(heures_theo.get("Interne", 0), 2), pct_vs(heures_cat["Interne"], heures_theo.get("Interne", 0)),
                round(total_reel, 2), round(cap_theo, 2), pct_saisie,
            ])
        else:
            ws.append([
                label,
                round(heures_cat["Projet"],  2), round(heures_theo.get("Projet",  0), 2), pct_vs(heures_cat["Projet"],  heures_theo.get("Projet",  0)),
                round(heures_cat["Support"], 2), round(heures_theo.get("Support", 0), 2), pct_vs(heures_cat["Support"], heures_theo.get("Support", 0)),
                round(total_reel, 2), round(cap_theo, 2), pct_saisie,
            ])
        style_row(ws, hdr_row + 1 + i, len(headers_cd), even=(i % 2 == 0))

    # Mise en forme conditionnelle % (colonnes D, G, J et % Saisie = L)
    n_equipes = len(lignes)
    pct_cols   = ["D", "G", "J"] if has_interne else ["D", "G"]
    saisie_col = "L" if has_interne else "I"
    for col_letter in pct_cols:
        rule = ColorScaleRule(
            start_type="num", start_value=0,   start_color="FFCCCC",
            mid_type="num",   mid_value=80,    mid_color="FFFF99",
            end_type="num",   end_value=100,   end_color="C6EFCE",
        )
        ws.conditional_formatting.add(
            f"{col_letter}{hdr_row + 1}:{col_letter}{hdr_row + n_equipes}", rule
        )
    from openpyxl.styles import PatternFill as _PF
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f"{saisie_col}{hdr_row + 1}:{saisie_col}{hdr_row + n_equipes}",
        CellIsRule(operator="lessThan", formula=["100"], fill=_PF("solid", start_color="FFCCCC", end_color="FFCCCC"))
    )

    # ── Données agrégées pour les graphiques ─────────────────────
    # Filtrer par les solutions de la config lignes
    solutions_scope = [cfg["solution"] for cfg in lignes.values() if cfg.get("solution") is not None]
    if solutions_scope:
        df_scope = df_source[df_source["Solution"].isin(solutions_scope)] if not df_source.empty else pd.DataFrame()
        # Ajouter l'Interne si présent dans les lignes
        if any(cfg.get("solution") is None for cfg in lignes.values()):
            df_int = df_source[df_source["Catégorie"] == "Interne"] if not df_source.empty else pd.DataFrame()
            df_scope = pd.concat([df_scope, df_int], ignore_index=True)
    else:
        df_scope = df_source.copy() if not df_source.empty else pd.DataFrame()

    ch_prod        = df_scope[df_scope["Catégorie"].isin(CATS_PRODUCTIF + ["Rework"])]["Temps (h)"].sum()
    ch_rework      = df_scope[df_scope["Catégorie"] == "Rework"]["Temps (h)"].sum()
    ch_projet      = df_scope[df_scope["Catégorie"] == "Projet"]["Temps (h)"].sum()
    ch_maintenance = df_scope[df_scope["Catégorie"] == "Maintenance"]["Temps (h)"].sum()
    ch_offerte     = df_scope[df_scope["Catégorie"] == "Prestation offerte"]["Temps (h)"].sum()
    ch_sans_presta = df_scope[df_scope["Catégorie"] == "Commande sans prestation"]["Temps (h)"].sum()
    ch_sup    = df_scope[df_scope["Catégorie"] == "Support N2"]["Temps (h)"].sum()
    ch_int    = df_scope[df_scope["Catégorie"] == "Interne"]["Temps (h)"].sum()
    ch_total  = df_scope["Temps (h)"].sum() if not df_scope.empty else 0
    ch_bloque = df_scope[df_scope.get("Bloqué", pd.Series(False, index=df_scope.index)) == True]["Temps (h)"].sum() if not df_scope.empty and "Bloqué" in df_scope.columns else 0

    autofit(ws)


# ──────────────────────────────────────────────────────────────────
# ONGLET 3 : Charge globale (toutes équipes)
# ──────────────────────────────────────────────────────────────────

def _ventile_cp(df, cp_split):
    if df.empty:
        return df
    rows_extra = []
    mask_cp = df["Collaborateur"].isin(cp_split.keys())
    df_cp   = df[mask_cp].copy()
    df_out  = df[~mask_cp].copy()
    for collab, splits in cp_split.items():
        df_c = df_cp[df_cp["Collaborateur"] == collab].copy()
        for sol, pct in splits.items():
            if pct > 0:
                df_c_sol = df_c.copy()
                df_c_sol["Solution"]  = sol
                df_c_sol["Temps (h)"] = df_c_sol["Temps (h)"] * pct
                rows_extra.append(df_c_sol)
    if rows_extra:
        df_out = pd.concat([df_out] + rows_extra, ignore_index=True)
    return df_out

df_wl_cp = _ventile_cp(df_wl_ana, CP_SPLIT)

ws_global = wb.create_sheet("Charge globale")
write_sheet_title(ws_global,
    f"Charge globale — {mois_courant_label} {month_start.year}",
    "CD + CP")

df_global = df_wl_ana  # toutes heures, hors exclus

# ── Tableau 1 : Répartition globale Productif / Support / Interne ─
ws_global.append(["Répartition globale"])
ws_global[ws_global.max_row][0].font      = _font(bold=True, size=12, color=BLUE_MID)
ws_global[ws_global.max_row][0].fill      = _fill("FFFFFF")
ws_global[ws_global.max_row][0].alignment = _align()
style_header_row(ws_global, ws_global.max_row, 1)

t1_headers = ["Catégorie", "Réalisé (h)"]
ws_global.append(t1_headers)
t1_hdr = ws_global.max_row
style_header_row(ws_global, t1_hdr, 2)

h_productif = df_global[df_global["Catégorie"].isin(CATS_PRODUCTIF + ["Rework"])]["Temps (h)"].sum() if not df_global.empty else 0
h_support   = df_global[df_global["Catégorie"] == "Support N2"]["Temps (h)"].sum()                   if not df_global.empty else 0
h_interne   = df_global[df_global["Catégorie"] == "Interne"]["Temps (h)"].sum()                      if not df_global.empty else 0
h_bloque    = df_global[df_global["Bloqué"] == True]["Temps (h)"].sum() if not df_global.empty and "Bloqué" in df_global.columns else 0

h_projet      = df_global[df_global["Catégorie"] == "Projet"]["Temps (h)"].sum()             if not df_global.empty else 0
h_maintenance = df_global[df_global["Catégorie"] == "Maintenance"]["Temps (h)"].sum()         if not df_global.empty else 0
h_offerte     = df_global[df_global["Catégorie"] == "Prestation offerte"]["Temps (h)"].sum()  if not df_global.empty else 0
h_sans_presta = df_global[df_global["Catégorie"] == "Commande sans prestation"]["Temps (h)"].sum() if not df_global.empty else 0
h_rework_gl   = df_global[df_global["Catégorie"] == "Rework"]["Temps (h)"].sum()              if not df_global.empty else 0

t1_data = [
    ("Productif",                    round(h_productif,   2)),
    ("  dont Projet",                round(h_projet,       2)),
    ("  dont Maintenance",           round(h_maintenance,  2)),
    ("  dont Prestation offerte",    round(h_offerte,      2)),
    ("  dont Commande sans presta.", round(h_sans_presta,  2)),
    ("  dont Rework",                round(h_rework_gl,    2)),
    ("  dont Bloqué",                round(h_bloque,       2)),
    ("Support",                      round(h_support,      2)),
    ("Interne",                      round(h_interne,      2)),
]
t1_anchor_row = ws_global.max_row  # ligne du titre tableau 1
t1_start = ws_global.max_row + 1
for j, (cat, val) in enumerate(t1_data):
    ws_global.append([cat, val])
    style_row(ws_global, ws_global.max_row, 2, even=(j % 2 == 0))
t1_end = ws_global.max_row







# ── Tableau heures par collaborateur ─────────────────────────────
ws_global.append(["Heures par collaborateur"])
ws_global[ws_global.max_row][0].font = _font(bold=True, size=12, color=BLUE_MID)

t4_headers = ["Collaborateur", "Réalisé (h)", "Théorique (h)", "Delta (h)"]
ws_global.append(t4_headers)
style_header_row(ws_global, ws_global.max_row, 4)

all_membres_global = equipe_litteralis + equipe_geodp + equipe_cp
if not df_wl_ana.empty:
    df_collab_g = (
        df_wl_ana[df_wl_ana["Collaborateur"].isin(all_membres_global)]
        .groupby("Collaborateur", as_index=False)["Temps (h)"]
        .sum().sort_values("Temps (h)", ascending=False)
    )
    for j, (_, r) in enumerate(df_collab_g.iterrows()):
        collab_name = r["Collaborateur"]
        h_reel = round(r["Temps (h)"], 2)
        h_theo = round(capacites_dict.get(collab_name, 0), 2)
        delta  = round(h_reel - h_theo, 2)
        ws_global.append([collab_name, h_reel, h_theo, delta])
        row_i = ws_global.max_row
        style_row(ws_global, row_i, 4, even=(j % 2 == 0))
        if delta < 0:
            ws_global.cell(row=row_i, column=4).font = _font(bold=False, color="C00000")

autofit(ws_global)

# ──────────────────────────────────────────────────────────────────
# CHARGE HISTORIQUE — mise à jour charge_historique.xlsx
# ──────────────────────────────────────────────────────────────────
_cats_prod = CATS_PRODUCTIF + ["Rework"]

def _h(sol, cats):
    if df_wl_ana.empty:
        return 0.0
    return round(df_wl_ana[
        (df_wl_ana["Solution"] == sol) &
        (df_wl_ana["Catégorie"].isin(cats))
    ]["Temps (h)"].sum(), 2)

ch_mois = {
    "Mois"              : mois_courant_label,
    "Année"             : month_start.year,
    "Productif Litt"    : _h("LITTERALIS", _cats_prod),
    "Rework Litt"       : _h("LITTERALIS", ["Rework"]),
    "Productif GEODP"   : _h("GEODP",      _cats_prod),
    "Rework GEODP"      : _h("GEODP",      ["Rework"]),
    "Support Litt"      : _h("LITTERALIS", ["Support N2"]),
    "Support GEODP"     : _h("GEODP",      ["Support N2"]),
}

try:
    if os.path.exists(CHARGE_HIST_FILE):
        df_ch_hist = pd.read_excel(CHARGE_HIST_FILE)
    else:
        df_ch_hist = pd.DataFrame(columns=list(ch_mois.keys()))

    mask_ch = (df_ch_hist["Mois"] == ch_mois["Mois"]) & (df_ch_hist["Année"] == ch_mois["Année"])
    if mask_ch.any():
        for k, v in ch_mois.items():
            df_ch_hist.loc[mask_ch, k] = v
        print(f"🔄 Charge historique mis à jour : {mois_courant_label} {month_start.year}")
    else:
        df_ch_hist = pd.concat([df_ch_hist, pd.DataFrame([ch_mois])], ignore_index=True)
        print(f"➕ Charge historique : nouvelle ligne {mois_courant_label} {month_start.year}")

    df_ch_hist.to_excel(CHARGE_HIST_FILE, index=False)
except Exception as e:
    print(f"⚠️  Impossible de sauvegarder {CHARGE_HIST_FILE} : {e}")



# Chargement historique charge pour les courbes
try:
    df_ch_hist_graph = pd.read_excel(CHARGE_HIST_FILE) if os.path.exists(CHARGE_HIST_FILE) else None
except Exception:
    df_ch_hist_graph = None


# ──────────────────────────────────────────────────────────────────
# ONGLET 4 : Charge Littéralis
# ──────────────────────────────────────────────────────────────────
# CD Littéralis : toutes leurs heures (support/interne sans solution assignée inclus)
ws_litt = wb.create_sheet("Charge Littéralis")
_charge_tableau(
    ws_litt, df_wl_ana, LIGNES_LITT,
    title    = f"Charge Littéralis — {mois_courant_label} {month_start.year}",
    subtitle = "Filtrage par solution Epic LITTERALIS",
)


# ──────────────────────────────────────────────────────────────────
# ONGLET 5 : Charge GEODP
# ──────────────────────────────────────────────────────────────────
ws_geo = wb.create_sheet("Charge GEODP")
_charge_tableau(
    ws_geo, df_wl_ana, LIGNES_GEODP,
    title    = f"Charge GEODP — {mois_courant_label} {month_start.year}",
    subtitle = "Filtrage par solution Epic GEODP",
)


# Courbes historiques Charge GEODP
# ──────────────────────────────────────────────────────────────────
# ONGLET 6 : Performance CA du mois — version enrichie
# ──────────────────────────────────────────────────────────────────

# ── Chargement des fichiers de référence ─────────────────────────
def _load_ref_file(filepath: str, label: str) -> dict:
    """
    Lit objectifs.xlsx ou ca_historique.xlsx.
    Retourne un dict { "Janvier": {"Global": x, "LITTERALIS": y, ...}, ... }
    """
    if not os.path.exists(filepath):
        print(f"⚠️  {label} introuvable : {filepath} → données à 0.")
        return {}
    try:
        df_ref = pd.read_excel(filepath, sheet_name=0, header=3)
        # Renommer la première colonne en "Mois"
        df_ref.columns = ["Mois"] + list(df_ref.columns[1:])
        df_ref = df_ref.dropna(subset=["Mois"])
        df_ref = df_ref[df_ref["Mois"] != "TOTAL"]
        result = {}
        for _, row in df_ref.iterrows():
            mois = str(row["Mois"]).strip()
            result[mois] = {
                col.replace(" (€)", "").strip(): (row[col] if pd.notna(row[col]) else 0)
                for col in df_ref.columns[1:]
            }
        return result
    except Exception as e:
        print(f"⚠️  Impossible de lire {label} ({filepath}) : {e}")
        return {}


objectifs_data  = _load_ref_file(OBJECTIFS_FILE,     "Fichier objectifs")
historique_data = _load_ref_file(CA_HISTORIQUE_FILE, "Fichier CA N-1")

obj_mois  = objectifs_data.get(mois_courant_label, {})
hist_mois = historique_data.get(mois_courant_label, {})

# ── Calculs de base ───────────────────────────────────────────────
total_realise        = df_brut["Montant déclaré ce mois"].sum()
aujourd_hui          = datetime.now().date()
days_passed          = len([d for d in working_days if d.date() <= aujourd_hui])
pct_avancement_jours = days_passed / num_working_days * 100 if num_working_days > 0 else 0

def _safe_evol_db(val, ref):
    if ref and ref > 0:
        return (val / ref - 1) * 100
    return 0

def _pct(val, ref):
    return round(val / ref * 100, 1) if ref and ref > 0 else None

def _delta(val, ref):
    return round(val - ref, 2) if ref is not None else None

def _fmt_pct(p):
    """Pourcentage sans signe + (pour atteinte objectif, couverture)."""
    if p is None: return "—"
    return f"{p:.1f}%"

def _fmt_pct_evol(p):
    """Pourcentage avec signe +/- (pour évolution N/N-1)."""
    if p is None: return "—"
    sign = "+" if p >= 0 else ""
    return f"{sign}{p:.1f}%"

def _fmt_eur(v):
    if v is None or v == 0: return "—"
    return f"{v:,.2f} €"

# ── Couverture projets ────────────────────────────────────────────
# Epics travaillés ce mois = epics avec temps > 0
epics_travailles = set(
    df_brut[
        (df_brut["Temps mois (h)"] > 0) | (df_brut["Montant déclaré ce mois"] != 0)
    ]["Epic"].unique()
)
# Epics avec CA déclaré > 0
epics_avec_ca    = set(df_brut[df_brut["Montant déclaré ce mois"] != 0]["Epic"].unique())
nb_travailles    = len(epics_travailles)
nb_avec_ca       = len(epics_avec_ca)

# ── Répartition par solution ──────────────────────────────────────
SOLUTIONS_PERF = ["LITTERALIS", "GEODP"]  # SHERPA inclus dans LITTERALIS
ca_par_solution = (
    df_brut.groupby("Solution")["Montant déclaré ce mois"].sum().to_dict()
    if not df_brut.empty else {}
)
heures_par_solution = (
    df_brut.groupby("Solution")["Temps mois (h)"].sum().to_dict()
    if not df_brut.empty else {}
)

# ── Rentabilité ───────────────────────────────────────────────────
total_heures_mois  = df_brut["Temps mois (h)"].sum() if not df_brut.empty else 0
rentabilite_globale = round(total_realise / total_heures_mois, 2) if total_heures_mois > 0 else 0

# ── Répartition CA par type de deal ──────────────────────────────
ca_par_deal_type = (
    df_brut[df_brut["Montant déclaré ce mois"] != 0]
    .groupby("Type de deal")["Montant déclaré ce mois"].sum()
    .sort_values(ascending=False)
    .to_dict()
    if not df_brut.empty and "Type de deal" in df_brut.columns else {}
)

# ── Top 5 projets par CA déclaré ─────────────────────────────────
top5 = (
    df_brut[df_brut["Montant déclaré ce mois"] > 0]
    .groupby(["Epic", "Epic Nom"])["Montant déclaré ce mois"].sum()
    .reset_index()
    .sort_values("Montant déclaré ce mois", ascending=False)
    .head(5)
    if not df_brut.empty else pd.DataFrame()
)

# ════════════════════════════════════════════════════════════
# ÉCRITURE DE L'ONGLET
# ════════════════════════════════════════════════════════════
ws_perf = wb.create_sheet("Performance CA du mois")
write_sheet_title(ws_perf,
    f"Performance CA — {mois_courant_label} {month_start.year}",
    f"Export du {aujourd_hui:%d/%m/%Y}  |  Avancement mois : {days_passed}/{num_working_days} jours ouvrés ({pct_avancement_jours:.0f}%)")

def _section_title(ws, txt):
    ws.append([])
    ws.append([txt])
    r = ws.max_row
    ws[r][0].font      = _font(bold=True, size=12, color=BLUE_MID)
    ws[r][0].alignment = _align()
    ws.row_dimensions[r].height = 20

def _write_table(ws, headers, rows_data, ncols=None):
    """Écrit un en-tête + lignes stylées, retourne la ligne de début des données."""
    if ncols is None: ncols = len(headers)
    ws.append(headers)
    hdr_r = ws.max_row
    style_header_row(ws, hdr_r, ncols)
    for i, row in enumerate(rows_data):
        ws.append(row)
        style_row(ws, ws.max_row, ncols, even=(i % 2 == 0))
    return hdr_r

# ── 1. Avancement du mois ─────────────────────────────────────────
_section_title(ws_perf, "📅 Avancement du mois")
_write_table(ws_perf,
    ["Indicateur", "Valeur"],
    [
        ["Jours ouvrés écoulés",  f"{days_passed} / {num_working_days}"],
        ["% avancement calendaire", f"{pct_avancement_jours:.0f}%"],
    ], ncols=2
)

# ── 2. Couverture projets ─────────────────────────────────────────
_section_title(ws_perf, "📂 Couverture projets")
taux_couv = _pct(nb_avec_ca, nb_travailles)
_write_table(ws_perf,
    ["Indicateur", "Nombre", "% couverture"],
    [
        ["Epics travaillés ce mois",       nb_travailles, "—"],
        ["Epics avec CA déclaré > 0",      nb_avec_ca,    _fmt_pct(taux_couv) if taux_couv else "—"],
        ["Epics sans CA déclaré (à vérifier)", nb_travailles - nb_avec_ca, ""],
    ], ncols=3
)

# ── 3. CA global : réalisé / objectif / N-1 ──────────────────────
_section_title(ws_perf, "💶 CA réalisé — Global & par solution")

obj_global  = obj_mois.get("Global", 0)
hist_global = hist_mois.get("Global", 0)
pct_obj_g   = _pct(total_realise, obj_global)
pct_nmo1_g  = _pct(total_realise, hist_global) - 100 if hist_global else None

ca_rows = []
# Ligne globale
ca_rows.append([
    "GLOBAL",
    _fmt_eur(total_realise),
    _fmt_eur(obj_global)  if obj_global  else "—",
    _fmt_pct(pct_obj_g)   if pct_obj_g   else "—",
    _fmt_eur(hist_global) if hist_global else "—",
    _fmt_pct_evol(pct_nmo1_g)  if pct_nmo1_g  is not None else "—",
])
# Lignes par solution
for sol in SOLUTIONS_PERF:
    ca_sol  = ca_par_solution.get(sol, 0)
    obj_sol = obj_mois.get(sol, 0)
    h_sol   = hist_mois.get(sol, 0)
    pct_o   = _pct(ca_sol, obj_sol)
    pct_n1  = (_pct(ca_sol, h_sol) - 100) if h_sol else None
    ca_rows.append([
        sol,
        _fmt_eur(ca_sol),
        _fmt_eur(obj_sol) if obj_sol else "—",
        _fmt_pct(pct_o)   if pct_o   else "—",
        _fmt_eur(h_sol)   if h_sol   else "—",
        _fmt_pct_evol(pct_n1)  if pct_n1  is not None else "—",
    ])

hdr_ca = _write_table(ws_perf,
    ["Solution", "Réalisé N (€)", "Objectif N (€)", "% Atteinte objectif",
     "Réalisé N-1 (€)", "Évolution N/N-1"],
    ca_rows, ncols=6
)

# Coloration alerte sur % atteinte objectif (col D = 4)
for r_idx in range(hdr_ca + 1, ws_perf.max_row + 1):
    # Colonne 5 = Évolution N/N-1 : rouge si négatif
    cell_evol = ws_perf.cell(row=r_idx, column=6)
    if isinstance(cell_evol.value, str):
        if cell_evol.value.startswith("-"):
            cell_evol.font = _font(bold=True, color="C00000")
        elif cell_evol.value.startswith("+"):
            cell_evol.font = _font(bold=True, color="375623")
    cell_pct = ws_perf.cell(row=r_idx, column=4)
    val_str  = str(cell_pct.value or "")
    try:
        pct_val = float(val_str.replace("%", "").replace("+", ""))
        if pct_val < pct_avancement_jours:
            cell_pct.fill = _fill(RED_LIGHT)
            cell_pct.font = _font(bold=True, color="C00000")
        elif pct_val >= 100:
            cell_pct.fill = _fill("C6EFCE")
            cell_pct.font = _font(bold=True, color="375623")
    except ValueError:
        pass


# ── 4. Rentabilité ────────────────────────────────────────────────
_section_title(ws_perf, "⚙️ Rentabilité moyenne (€/h)")
rent_rows = [["GLOBAL", _fmt_eur(rentabilite_globale), total_heures_mois]]
for sol in SOLUTIONS_PERF:
    h_sol    = heures_par_solution.get(sol, 0)
    ca_sol   = ca_par_solution.get(sol, 0)
    rent_sol = round(ca_sol / h_sol, 2) if h_sol > 0 else 0
    rent_rows.append([sol, _fmt_eur(rent_sol), round(h_sol, 1)])

_write_table(ws_perf,
    ["Solution", "Rentabilité (€/h)", "Heures loguées"],
    rent_rows, ncols=3
)

# ── 5. Répartition CA par type de deal — 3 tableaux ─────────────
_section_title(ws_perf, "🏷️ Répartition du CA par type de deal")

# Calculs par solution + type
def _ca_by_deal(df, solution=None):
    d = df[df["Montant déclaré ce mois"] > 0].copy() if not df.empty else df
    if solution:
        d = d[d["Solution"] == solution]
    if d.empty:
        return {}
    return d.groupby("Type de deal")["Montant déclaré ce mois"].sum().to_dict()

ca_deal_global  = _ca_by_deal(df_brut)
ca_deal_litt    = _ca_by_deal(df_brut, "LITTERALIS")
ca_deal_geodp   = _ca_by_deal(df_brut, "GEODP")

# Agrégation : Migration comptabilisée dans Upsell pour le global et Littéralis
def _agrege_migration(d):
    d2 = dict(d)
    mig = d2.pop("Migration", 0)
    d2["Vente additionnelle"] = d2.get("Vente additionnelle", 0) + mig
    return d2

ca_deal_global_agg = _agrege_migration(ca_deal_global)
ca_deal_litt_agg   = _agrege_migration(ca_deal_litt)

# Tableau Global (Upsell+Migration vs New)
_section_title(ws_perf, "Global")
global_rows = [
    [k, _fmt_eur(v), _fmt_pct(_pct(v, total_realise))]
    for k, v in sorted(ca_deal_global_agg.items(), key=lambda x: -x[1])
    if v > 0
] or [["—", "—", "—"]]
_write_table(ws_perf, ["Type de deal", "CA (€)", "% total"], global_rows, ncols=3)

# Tableau GEODP (Upsell vs New vs Migration)
_section_title(ws_perf, "GEODP")
ca_total_geodp = sum(ca_deal_geodp.values()) or 1
geodp_rows = [
    [k, _fmt_eur(v), _fmt_pct(_pct(v, ca_total_geodp))]
    for k, v in sorted(ca_deal_geodp.items(), key=lambda x: -x[1])
    if v > 0
] or [["—", "—", "—"]]
_write_table(ws_perf, ["Type de deal", "CA (€)", "% GEODP"], geodp_rows, ncols=3)

# Tableau Littéralis (Upsell+Migration vs New)
_section_title(ws_perf, "Littéralis")
ca_total_litt = sum(ca_deal_litt_agg.values()) or 1
litt_rows = [
    [k, _fmt_eur(v), _fmt_pct(_pct(v, ca_total_litt))]
    for k, v in sorted(ca_deal_litt_agg.items(), key=lambda x: -x[1])
    if v > 0
] or [["—", "—", "—"]]
_write_table(ws_perf, ["Type de deal", "CA (€)", "% Littéralis"], litt_rows, ncols=3)

# ── 6. Top 5 projets ─────────────────────────────────────────────
_section_title(ws_perf, "🏆 Top 5 projets — CA déclaré ce mois")
top5_rows = []
if not top5.empty:
    for rank, (_, row) in enumerate(top5.iterrows(), start=1):
        ca_proj  = row["Montant déclaré ce mois"]
        pct_proj = _pct(ca_proj, total_realise)
        top5_rows.append([
            f"#{rank}",
            row["Epic"],
            row["Epic Nom"][:60] + ("…" if len(str(row["Epic Nom"])) > 60 else ""),
            _fmt_eur(ca_proj),
            _fmt_pct(pct_proj),
        ])
else:
    top5_rows = [["—", "—", "—", "—", "—"]]

_write_table(ws_perf,
    ["#", "Clé", "Epic / Projet", "CA déclaré (€)", "% du CA total"],
    top5_rows, ncols=5
)

autofit(ws_perf, min_w=8, max_w=60)


# ══════════════════════════════════════════════════════════════════
# MISE À JOUR DE L'HISTORIQUE DE PERFORMANCE
# ══════════════════════════════════════════════════════════════════

HIST_COLUMNS = [
    "Mois", "Année", "CA réalisé global (€)",
    "CA LITTERALIS (€)", "CA GEODP (€)",
    "Objectif global (€)", "% Atteinte objectif",
    "CA N-1 global (€)", "Évolution N/N-1 (%)",
    "Rentabilité globale (€/h)",
    "Nb projets travaillés", "Nb projets avec CA",
    "Nb jours ouvrés", "Jours écoulés", "% Avancement calendaire",
]

def _safe_pct(val, ref):
    if ref and ref > 0:
        return round(val / ref * 100, 1)
    return None

def _safe_evol(val, ref):
    if ref and ref > 0:
        return round((val / ref - 1) * 100, 1)
    return None

# Ligne du mois courant
new_row = {
    "Mois"                      : mois_courant_label,
    "Année"                     : month_start.year,
    "CA réalisé global (€)"     : round(total_realise, 2),
    "CA LITTERALIS (€)"         : round(ca_par_solution.get("LITTERALIS", 0), 2),
    "CA GEODP (€)"              : round(ca_par_solution.get("GEODP", 0), 2),
    "Objectif global (€)"       : round(obj_global, 2) if obj_global else 0,
    "% Atteinte objectif"       : _safe_pct(total_realise, obj_global),
    "CA N-1 global (€)"         : round(hist_global, 2) if hist_global else 0,
    "Évolution N/N-1 (%)"       : _safe_evol(total_realise, hist_global),
    "Rentabilité globale (€/h)" : rentabilite_globale,
    "Nb projets travaillés"     : nb_travailles,
    "Nb projets avec CA"        : nb_avec_ca,
    "Nb jours ouvrés"           : num_working_days,
    "Jours écoulés"             : days_passed,
    "% Avancement calendaire"   : round(pct_avancement_jours, 1),
}

# Charger ou créer l'historique
hist_key = f"{month_start.year}-{month_start.month:02d}"
if os.path.exists(PERF_HIST_FILE):
    try:
        df_hist = pd.read_excel(PERF_HIST_FILE)
        # Reconstruire la clé
        df_hist["_key"] = df_hist["Année"].astype(str) + "-" + df_hist.index.map(
            lambda i: f"{list(MOIS_FR).index(df_hist.loc[i, 'Mois']) + 1:02d}"
            if df_hist.loc[i, "Mois"] in MOIS_FR else "00"
        )
        # Écraser si même mois/année
        mask = (df_hist["Mois"] == mois_courant_label) & (df_hist["Année"] == month_start.year)
        if mask.any():
            for col, val in new_row.items():
                df_hist.loc[mask, col] = val
        else:
            df_hist = pd.concat([df_hist, pd.DataFrame([new_row])], ignore_index=True)
        df_hist = df_hist.drop(columns=["_key"], errors="ignore")
    except Exception as e:
        print(f"⚠️  Impossible de lire {PERF_HIST_FILE} : {e}. Recréation.")
        df_hist = pd.DataFrame([new_row])
else:
    df_hist = pd.DataFrame([new_row])

# Trier par année + mois
df_hist["_sort"] = df_hist.apply(
    lambda r: r["Année"] * 100 + (MOIS_FR.index(r["Mois"]) + 1 if r["Mois"] in MOIS_FR else 0),
    axis=1
)
df_hist = df_hist.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

# Sauvegarder l'historique
try:
    df_hist.to_excel(PERF_HIST_FILE, index=False)
    print(f"✅ Historique mis à jour : {PERF_HIST_FILE} ({len(df_hist)} mois)")
except Exception as e:
    print(f"⚠️  Impossible de sauvegarder {PERF_HIST_FILE} : {e}")

# ──────────────────────────────────────────────────────────────────
# ONGLET : Historique CA
# ──────────────────────────────────────────────────────────────────
ws_hist = wb.create_sheet("Historique CA")
write_sheet_title(ws_hist,
    f"Historique CA — {month_start.year}",
    f"Évolution mensuelle depuis janvier {month_start.year} — mis à jour le {aujourd_hui:%d/%m/%Y}")

# Filtrer sur l'année courante uniquement
df_hist_yr = df_hist[df_hist["Année"] == month_start.year].copy()

# ── Tableau historique ────────────────────────────────────────────
hist_headers = [
    "Mois",
    "CA réalisé (€)", "CA LITT. (€)", "CA GEODP (€)",
    "Objectif (€)", "% Objectif",
    "CA N-1 (€)", "Évol. N/N-1",
    "Rentab. (€/h)",
    "Projets travaillés", "Projets avec CA",
    "Avancement (%)",
]
hist_col_map = [
    "Mois",
    "CA réalisé global (€)", "CA LITTERALIS (€)", "CA GEODP (€)",
    "Objectif global (€)", "% Atteinte objectif",
    "CA N-1 global (€)", "Évolution N/N-1 (%)",
    "Rentabilité globale (€/h)",
    "Nb projets travaillés", "Nb projets avec CA",
    "% Avancement calendaire",
]

ws_hist.append(hist_headers)
hdr_hist_row = ws_hist.max_row
style_header_row(ws_hist, hdr_hist_row, len(hist_headers))

EUR_FMT = '#,##0.00 "€"'
PCT_FMT = '0.0"%"'

for i, (_, row) in enumerate(df_hist_yr.iterrows()):
    data_row = [row.get(col, "") for col in hist_col_map]
    # Formater les pourcentages / None
    for j, col in enumerate(hist_col_map):
        v = data_row[j]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            data_row[j] = "—"
    ws_hist.append(data_row)
    row_idx = ws_hist.max_row
    style_row(ws_hist, row_idx, len(hist_headers),
              even=(i % 2 == 0),
              bold=(row.get("Mois") == mois_courant_label))  # mois courant en gras

    # Formatage monétaire et %
    for j, col in enumerate(hist_col_map, start=1):
        cell = ws_hist.cell(row=row_idx, column=j)
        if "€" in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = EUR_FMT
        elif col in ("% Atteinte objectif", "Évolution N/N-1 (%)", "% Avancement calendaire"):
            if isinstance(cell.value, (int, float)):
                cell.number_format = '0.0"%"'

# Ligne total YTD
ws_hist.append([])
ytd_row_idx = ws_hist.max_row + 1
ytd_values = ["TOTAL YTD"]
for col in hist_col_map[1:]:
    if "€" in col and "%" not in col and "Évol" not in col:
        v = df_hist_yr[col].sum() if col in df_hist_yr.columns else 0
        ytd_values.append(round(v, 2))
    elif col == "Rentabilité globale (€/h)":
        ca_tot = df_hist_yr.get("CA réalisé global (€)", pd.Series([0])).sum()
        h_col  = "CA réalisé global (€)"  # approx — on n'a pas les heures totales en hist
        ytd_values.append("—")
    else:
        ytd_values.append("—")
ws_hist.append(ytd_values)
style_row(ws_hist, ws_hist.max_row, len(hist_headers), total=True)
for j, col in enumerate(hist_col_map[1:], start=2):
    cell = ws_hist.cell(row=ws_hist.max_row, column=j)
    if isinstance(cell.value, (int, float)):
        cell.number_format = EUR_FMT

autofit(ws_hist, min_w=10, max_w=25)


# ══════════════════════════════════════════════════════════════════
# ONGLET : Commandes du mois
# Epics PDMDEP créés pendant le mois en cours (date created)
# ══════════════════════════════════════════════════════════════════
print("📥 Récupération des commandes du mois...")
try:
    jql_commandes = (
        f'project = PDMDEP AND issuetype = Epic ' 
        f'AND created >= "{month_start:%Y-%m-%d}" ' 
        f'AND created <= "{month_end:%Y-%m-%d}"' 
    )
    epics_commandes = get_all_issues(jql_commandes)
    print(f"   → {len(epics_commandes)} épics créés ce mois.")
except Exception as e:
    print(f"⚠️  Impossible de récupérer les commandes : {e}")
    epics_commandes = []

commande_rows = []
DIVERS_EPICS_EXCLUS = {DIVERS_LITT_EPIC, DIVERS_GEODP_EPIC}
for epic in sorted(
    [e for e in epics_commandes if e.key not in DIVERS_EPICS_EXCLUS],
    key=lambda e: e.raw.get('fields', {}).get('created', '') or ''
):
    epic_key  = epic.key
    epic_name = epic.fields.summary
    raw_fields = epic.raw.get("fields", {})

    # Date de création — lire depuis raw_fields car fields.created peut être None
    created_raw = raw_fields.get("created") or getattr(epic.fields, "created", None)
    try:
        created_dt = pd.to_datetime(created_raw).strftime("%d/%m/%Y") if created_raw else ""
    except Exception:
        created_dt = str(created_raw)[:10] if created_raw else ""

    # Statut
    statut = getattr(epic.fields.status, "name", "") if hasattr(epic.fields, "status") else ""

    # Solution + client
    raw_solution  = (raw_fields.get("customfield_20514") or {}).get("value", "")
    epic_solution = SOLUTION_MAPPING.get(raw_solution, raw_solution)
    client_epic   = _read_client(raw_fields)

    # Tickets CA enfants
    tickets_ca_cmd = [
        t for t in all_issues
        if getattr(t.fields, "parent", None)
        and t.fields.parent.key == epic_key
        and t.fields.issuetype.name in ("COORDIN : Suivi CA", "Matériel GEODP")
    ]

    # Agrégation : montant total prestations + CA reconnu
    # Montant : customfield_20701 en priorité, sinon customfield_24564
    def _montant_presta(ca):
        return getattr(ca.fields, "customfield_20701", 0) or 0

    montant_total_presta = sum(_montant_presta(ca) for ca in tickets_ca_cmd)
    ca_reconnu = sum(
        _read_montant(ca)
        for ca in tickets_ca_cmd
    )

    # BDC : depuis les tickets CA, sinon fallback sur customfield_24529 de l'Epic
    bdcs = [
        _read_bdc(ca.fields, ca.raw.get("fields", {}))
        for ca in tickets_ca_cmd
    ]
    bdc_concat = " / ".join(b for b in bdcs if b)
    if not bdc_concat:
        bdc_epic = raw_fields.get("customfield_24529", "") or ""
        if isinstance(bdc_epic, list):
            bdc_concat = " / ".join(str(x) for x in bdc_epic if x)
        elif bdc_epic:
            bdc_concat = str(bdc_epic).strip()

    # Client depuis ticket CA si absent de l'Epic
    if not client_epic and tickets_ca_cmd:
        fields_first_ca = tickets_ca_cmd[0].raw.get("fields", {})
        client_epic = _read_client(fields_first_ca)

    commande_rows.append({
        "Date de création"          : created_dt,
        "Clé Epic"                  : epic_key,
        "Nom / Résumé"              : epic_name,
        "Client"                    : client_epic,
        "Solution"                  : epic_solution,
        "Type de deal"              : _read_deal_type(raw_fields),
        "Statut"                    : statut,
        "Numéro BDC"                : bdc_concat,
        "Montant total prestations" : montant_total_presta,
        "CA reconnu ce mois"        : ca_reconnu,
    })

df_cmd = pd.DataFrame(commande_rows) if commande_rows else pd.DataFrame(columns=[
    "Date de création", "Clé Epic", "Nom / Résumé", "Client", "Solution",
    "Statut", "Numéro BDC", "Montant total prestations", "CA reconnu ce mois",
])

# ── Écriture de l'onglet ─────────────────────────────────────────
ws_cmd = wb.create_sheet("Commandes du mois")
write_sheet_title(ws_cmd,
    f"Commandes du mois — {mois_courant_label} {month_start.year}",
    f"Epics PDMDEP créés entre le {month_start:%d/%m/%Y} et le {month_end:%d/%m/%Y} — {len(commande_rows)} commande(s)")

CMD_HEADERS = [
    "Date de création", "Clé Epic", "Nom / Résumé", "Client", "Solution",
    "Type de deal", "Statut", "Numéro BDC",
    "Montant total prestations (€)", "CA reconnu ce mois (€)",
]

ws_cmd.append(CMD_HEADERS)
hdr_cmd = ws_cmd.max_row
style_header_row(ws_cmd, hdr_cmd, len(CMD_HEADERS))

EUR_FMT = '#,##0.00 "€"'
money_cmd_cols = {"Montant total prestations (€)", "CA reconnu ce mois (€)"}
col_cmd_idx = {h: i for i, h in enumerate(CMD_HEADERS, start=1)}

for i, row in enumerate(commande_rows):
    ws_cmd.append([
        row["Date de création"],
        row["Clé Epic"],
        row["Nom / Résumé"],
        row["Client"],
        row["Solution"],
        row["Type de deal"],
        row["Statut"],
        row["Numéro BDC"],
        row["Montant total prestations"],
        row["CA reconnu ce mois"],
    ])
    row_idx = ws_cmd.max_row
    _make_link(ws_cmd, row_idx, CMD_HEADERS.index("Clé Epic") + 1, row.get("Clé Epic", ""))
    style_row(ws_cmd, row_idx, len(CMD_HEADERS), even=(i % 2 == 0))
    for h in money_cmd_cols:
        ws_cmd.cell(row=row_idx, column=col_cmd_idx[h]).number_format = EUR_FMT

# Ligne total
if commande_rows:
    # Construire la ligne total avec le bon nombre de colonnes
    _n_cmd_cols = len(CMD_HEADERS)
    _col_mont = CMD_HEADERS.index("Montant total prestations (€)")
    _col_ca   = CMD_HEADERS.index("CA reconnu ce mois (€)")
    _total_row_vals = [""] * _n_cmd_cols
    _total_row_vals[0]         = "TOTAL"
    _total_row_vals[_col_mont] = sum(r["Montant total prestations"] for r in commande_rows)
    _total_row_vals[_col_ca]   = sum(r["CA reconnu ce mois"] for r in commande_rows)
    ws_cmd.append(_total_row_vals)
    total_cmd_row = ws_cmd.max_row
    style_row(ws_cmd, total_cmd_row, len(CMD_HEADERS), total=True)
    for h in money_cmd_cols:
        ws_cmd.cell(row=total_cmd_row, column=col_cmd_idx[h]).number_format = EUR_FMT

ws_cmd.freeze_panes = f"A{hdr_cmd + 1}"
autofit(ws_cmd, min_w=10, max_w=55)


# ══════════════════════════════════════════════════════════════════
# ONGLET : Projets clôturés
# ══════════════════════════════════════════════════════════════════
try:
    CLOTURES_HIST_FILE = "clotures_historique.xlsx"
    CLOTURES_HIST_COLS = ["Mois", "Année", "Clôturés Projet", "Clôturés Rework", "Total clôturés", "Stock Rework fin de mois"]
    
    # ── 1. Requête JQL : Epics PDMDEP passés à "Terminé" ce mois ─────
    print("📥 Récupération des projets clôturés ce mois...")
    try:
        jql_clotures = (
            f'project = PDMDEP AND issuetype = Epic '
            f'AND status changed TO "Terminé" '
            f'DURING ("{month_start:%Y-%m-%d}", "{month_end:%Y-%m-%d}")'
        )
        epics_clotures = get_all_issues(jql_clotures)
        print(f"   → {len(epics_clotures)} épics clôturés ce mois.")
    except Exception as e:
        print(f"⚠️  Impossible de récupérer les clôtures : {e}")
        epics_clotures = []

    # Stock Rework actuel : Epics Rework encore ouverts
    print("📥 Calcul du stock Rework actuel...")
    try:
        _champ_rework = "Type d'Epic[Dropdown]"
        jql_stock_rework = (
            f'project = PDMDEP AND issuetype = Epic '
            f'AND "{_champ_rework}" = Rework AND statusCategory != Done'
        )
        stock_rework_epics = get_all_issues(jql_stock_rework)
        stock_rework_actuel = len(stock_rework_epics)
        print(f"   → Stock Rework : {stock_rework_actuel} épics ouverts.")
    except Exception as e:
        print(f"⚠️  Impossible de calculer le stock Rework : {e}")
        stock_rework_actuel = "—"
    
    # ── 2. Analyser chaque Epic clôturé ──────────────────────────────
    clotures_rows = []
    nb_clotures_projet = 0
    nb_clotures_rework = 0
    
    for epic in epics_clotures:
        raw_fields = epic.raw.get("fields", {})
    
        # Catégorie (Projet ou Rework)
        cf_cat = getattr(epic.fields, "customfield_23103", None)
        categorie = cf_cat.value if hasattr(cf_cat, "value") else str(cf_cat or "")
    
        date_cloture = month_start.strftime("%m/%Y")
    
        # Solution
        raw_solution  = (raw_fields.get("customfield_20514") or {}).get("value", "")
        epic_solution = SOLUTION_MAPPING.get(raw_solution, raw_solution)
    
        # Client
        client = _read_client(raw_fields)
    
        is_rework = "rework" in categorie.lower()
        if is_rework:
            nb_clotures_rework += 1
        else:
            nb_clotures_projet += 1
    
        clotures_rows.append({
            "Mois clôture"  : date_cloture,
            "Clé Epic"      : epic.key,
            "Nom / Résumé"  : epic.fields.summary,
            "Client"        : client,
            "Solution"      : epic_solution,
            "Catégorie"     : categorie,
        })
    
    nb_clotures_total = nb_clotures_projet + nb_clotures_rework
    
    # ── 3. Fichier historique clôtures ───────────────────────────────
    new_row_cl = {
        "Mois"                    : mois_courant_label,
        "Année"                   : month_start.year,
        "Clôturés Projet"         : nb_clotures_projet,
        "Clôturés Rework"         : nb_clotures_rework,
        "Total clôturés"          : nb_clotures_total,
        "Stock Rework fin de mois": stock_rework_actuel,
    }
    
    if os.path.exists(CLOTURES_HIST_FILE):
        try:
            df_cl_hist = pd.read_excel(CLOTURES_HIST_FILE)
            # Si le fichier a un titre en ligne 1, chercher la vraie ligne d'en-tête
            if "Mois" not in df_cl_hist.columns:
                df_cl_hist = pd.read_excel(CLOTURES_HIST_FILE, header=3)
            mask_cl = (df_cl_hist["Mois"] == mois_courant_label) & (df_cl_hist["Année"] == month_start.year)
            if mask_cl.any():
                for col in ["Clôturés Projet", "Clôturés Rework", "Total clôturés", "Stock Rework fin de mois"]:
                    df_cl_hist.loc[mask_cl, col] = new_row_cl[col]
                # Ne pas écraser le stock Rework s'il a été saisi manuellement
            else:
                df_cl_hist = pd.concat([df_cl_hist, pd.DataFrame([new_row_cl])], ignore_index=True)
        except Exception as e:
            print(f"⚠️  Impossible de lire {CLOTURES_HIST_FILE} : {e}. Recréation.")
            df_cl_hist = pd.DataFrame([new_row_cl])
    else:
        # Initialisation avec janvier 2026
        init_jan = {
            "Mois": "Janvier", "Année": 2026,
            "Clôturés Projet": 8, "Clôturés Rework": 26,
            "Total clôturés": 34, "Stock Rework fin de mois": 117,
        }
        if mois_courant_label == "Janvier" and month_start.year == 2026:
            df_cl_hist = pd.DataFrame([new_row_cl])
            df_cl_hist.loc[0, "Stock Rework fin de mois"] = 117
        else:
            df_cl_hist = pd.DataFrame([init_jan, new_row_cl])
    
    # Tri chronologique
    df_cl_hist["_sort"] = df_cl_hist.apply(
        lambda r: r["Année"] * 100 + (MOIS_FR.index(r["Mois"]) + 1 if r["Mois"] in MOIS_FR else 0), axis=1
    )
    df_cl_hist = df_cl_hist.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)
    
    try:
        df_cl_hist.to_excel(CLOTURES_HIST_FILE, index=False)
        print(f"✅ Historique clôtures mis à jour : {CLOTURES_HIST_FILE}")
    except Exception as e:
        print(f"⚠️  Impossible de sauvegarder {CLOTURES_HIST_FILE} : {e}")
    
    # ── 4. Onglet Excel ───────────────────────────────────────────────
    ws_cl = wb.create_sheet("Projets clôturés")
    write_sheet_title(ws_cl,
        f"Projets clôturés — {mois_courant_label} {month_start.year}",
        f"{nb_clotures_total} Epic(s) passés à 'Terminé' ce mois  |  dont {nb_clotures_rework} Rework")
    
    # Tableau du mois
    CL_HEADERS = ["Mois clôture", "Clé Epic", "Nom / Résumé", "Client", "Solution", "Catégorie"]
    ws_cl.append(CL_HEADERS)
    hdr_cl = ws_cl.max_row
    style_header_row(ws_cl, hdr_cl, len(CL_HEADERS))
    
    CL_EPIC_COL = CL_HEADERS.index("Clé Epic") + 1
    for i, row in enumerate(clotures_rows):
        ws_cl.append([row[h] for h in CL_HEADERS])
        row_i = ws_cl.max_row
        style_row(ws_cl, row_i, len(CL_HEADERS), even=(i % 2 == 0))
        _make_link(ws_cl, row_i, CL_EPIC_COL, row.get("Clé Epic", ""))
        # Rework en orange
        if "rework" in str(row.get("Catégorie", "")).lower():
            for c in range(1, len(CL_HEADERS) + 1):
                ws_cl.cell(row=row_i, column=c).fill = _fill("FFF2CC")
    
    # Ligne résumé
    ws_cl.append([])
    ws_cl.append(["RÉSUMÉ", "", f"{nb_clotures_projet} Projet  +  {nb_clotures_rework} Rework  =  {nb_clotures_total} total", "", "", ""])
    style_row(ws_cl, ws_cl.max_row, len(CL_HEADERS), total=True)
    
    # ── 5. Tableau historique ─────────────────────────────────────────
    ws_cl.append([])
    ws_cl.append([])
    ws_cl.append(["📊 Historique des clôtures — depuis janvier 2026"])
    ws_cl[ws_cl.max_row][0].font = _font(bold=True, size=12, color=BLUE_MID)
    
    HIST_CL_HEADERS = ["Mois", "Clôturés Projet", "Clôturés Rework", "Total clôturés", "Stock Rework fin de mois"]
    ws_cl.append(HIST_CL_HEADERS)
    hdr_hist_cl = ws_cl.max_row
    style_header_row(ws_cl, hdr_hist_cl, len(HIST_CL_HEADERS))
    
    df_cl_yr = df_cl_hist[df_cl_hist["Année"] == month_start.year]
    data_cl_start = hdr_hist_cl + 1
    
    for i, (_, row) in enumerate(df_cl_yr.iterrows()):
        ws_cl.append([
            row.get("Mois", ""),
            row.get("Clôturés Projet", 0),
            row.get("Clôturés Rework", 0),
            row.get("Total clôturés", 0),
            row.get("Stock Rework fin de mois", ""),
        ])
        style_row(ws_cl, ws_cl.max_row, len(HIST_CL_HEADERS), even=(i % 2 == 0))
    
    data_cl_end = ws_cl.max_row
    
    
    autofit(ws_cl, min_w=12, max_w=60)
    
    # ── 7. KPI dans le tableau de bord ───────────────────────────────
    # (sera affiché après création du TDB — on stocke pour y accéder)
    kpi_clotures_total  = nb_clotures_total
    kpi_clotures_rework = nb_clotures_rework
except Exception as _e_cl:
    print(f"⚠️  Onglet Projets clôturés ignoré : {_e_cl}")


# ══════════════════════════════════════════════════════════════════
# ONGLET 0 : Tableau de bord — KPI clés
# ══════════════════════════════════════════════════════════════════
# Valeurs par défaut si l'onglet clôtures a échoué
if 'kpi_clotures_total'  not in dir(): kpi_clotures_total  = 0
if 'kpi_clotures_rework' not in dir(): kpi_clotures_rework = 0
if 'stock_rework_actuel' not in dir(): stock_rework_actuel = '—'

ws_db = wb.create_sheet("Tableau de bord", 0)  # en premier
write_sheet_title(ws_db,
    f"Tableau de bord — {mois_courant_label} {month_start.year}",
    f"Export du {aujourd_hui:%d/%m/%Y}  |  {days_passed}/{num_working_days} jours ouvrés écoulés ({pct_avancement_jours:.0f}%)")

BLUE_DARK_C = "1F3864"; BLUE_MID_C = "2E75B6"; GREEN_C = "375623"; RED_C = "C00000"
ORANGE_C = "BF5700"; WHITE_C = "FFFFFF"; BLUE_XL_C = "EBF3FB"; BLUE_L_C = "D6E4F0"

def _kpi_block(ws, start_row, start_col, label, value, sub="", color=BLUE_MID_C):
    ws.merge_cells(start_row=start_row, start_column=start_col,
                   end_row=start_row, end_column=start_col + 2)
    c = ws.cell(row=start_row, column=start_col, value=label)
    c.font = _font(bold=True, size=10, color="595959")
    c.alignment = _align("left")

    ws.merge_cells(start_row=start_row+1, start_column=start_col,
                   end_row=start_row+1, end_column=start_col + 2)
    v = ws.cell(row=start_row+1, column=start_col, value=value)
    v.font = _font(bold=True, size=20, color=color)
    v.alignment = _align("left")
    v.fill = _fill(BLUE_XL_C)

    if sub:
        ws.merge_cells(start_row=start_row+2, start_column=start_col,
                       end_row=start_row+2, end_column=start_col + 2)
        s = ws.cell(row=start_row+2, column=start_col, value=sub)
        s.font = _font(bold=False, size=9, color="595959")
        s.alignment = _align("left")

def _section_db(ws, row, col, title):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 8)
    c = ws.cell(row=row, column=col, value=title)
    c.font      = _font(bold=True, size=12, color=WHITE_C)
    c.fill      = _fill(BLUE_DARK_C)
    c.alignment = _align("left")
    ws.row_dimensions[row].height = 22

# ── Section CA ────────────────────────────────────────────────────
_section_db(ws_db, 5, 1, "💶  Chiffre d'affaires")
ws_db.row_dimensions[5].height = 22

pct_obj_str  = f"{_pct(total_realise, obj_global):.1f}%" if obj_global else "—"
evol_str     = f"{_safe_evol_db(total_realise, hist_global):+.1f}%" if hist_global else "—"
evol_color   = RED_C if hist_global and total_realise < hist_global else GREEN_C
rent_str     = f"{rentabilite_globale:.0f} €/h"

_kpi_block(ws_db, 6, 1,  "CA réalisé (services)",  f"{total_realise:,.0f} €")
_kpi_block(ws_db, 6, 5,  "Objectif du mois",        f"{obj_global:,.0f} €" if obj_global else "—")
_kpi_block(ws_db, 6, 9,  "% Atteinte objectif",     pct_obj_str,
           color=GREEN_C if obj_global and total_realise >= obj_global else RED_C)
_kpi_block(ws_db, 10, 1, "CA Littéralis",           f"{ca_par_solution.get('LITTERALIS', 0):,.0f} €")
_kpi_block(ws_db, 10, 5, "CA GEODP",                f"{ca_par_solution.get('GEODP', 0):,.0f} €")
_kpi_block(ws_db, 10, 9, "Évolution vs N-1",        evol_str, color=evol_color)

# ── Section Activité ──────────────────────────────────────────────
_section_db(ws_db, 14, 1, "📂  Activité projets")

_kpi_block(ws_db, 15, 1, "Projets travaillés",  str(nb_travailles))
_kpi_block(ws_db, 15, 5, "Projets avec CA",     str(nb_avec_ca))
_kpi_block(ws_db, 15, 9, "Rentabilité moy.",    rent_str)
_kpi_block(ws_db, 19, 1, "Commandes ce mois",   str(len(commande_rows)) if "commande_rows" in dir() else "—")
_kpi_block(ws_db, 19, 9, "Clôturés ce mois",    str(kpi_clotures_total),
           f"dont {kpi_clotures_rework} Rework")

# ── Section Charge ────────────────────────────────────────────────
_section_db(ws_db, 23, 1, "⏱️  Charge équipe")

h_loguees  = df_wl_ana["Temps (h)"].sum() if not df_wl_ana.empty else 0
h_dispo    = sum(c[3] for c in capacites)
taux_occup = round(h_loguees / h_dispo * 100, 1) if h_dispo > 0 else 0
h_non_sais = round(h_dispo - h_loguees, 1)

_kpi_block(ws_db, 24, 1, "Heures loguées",       f"{h_loguees:.1f} h")
_kpi_block(ws_db, 24, 5, "Heures disponibles",   f"{h_dispo:.1f} h")
_kpi_block(ws_db, 24, 9, "Taux d'occupation",    f"{taux_occup:.1f}%",
           color=GREEN_C if taux_occup >= 80 else RED_C)
_kpi_block(ws_db, 28, 1, "Heures non saisies",   f"{h_non_sais:.1f} h",
           color=RED_C if h_non_sais > 0 else GREEN_C)

# ── Section Rework ──────────────────────────────────────────────
_section_db(ws_db, 32, 1, "🔁  Stock Rework")
_kpi_block(ws_db, 33, 1, "Stock Rework actuel",
           str(stock_rework_actuel),
           "Epics Rework encore ouverts (fin de mois précédent)",
           color=RED_C if isinstance(stock_rework_actuel, int) and stock_rework_actuel > 50 else ORANGE_C)

# Largeurs colonnes
for c in range(1, 14):
    ws_db.column_dimensions[get_column_letter(c)].width = 14
for r in [6, 7, 8, 10, 11, 12, 15, 16, 17, 19, 20, 21, 24, 25, 26, 28, 29, 30]:
    ws_db.row_dimensions[r].height = 20

"""
BLOC À AJOUTER dans Export_worklog_V3.py
—————————————————————————————————————————
Placer ce bloc JUSTE AVANT la section "SAUVEGARDE" (wb.save),
APRÈS les sections Backlog, Dashboard et Projets clôturés.

Ce bloc exporte _data_mois.xlsx qui sera lu par Rapport_global.py.

Il remplace tout le bloc "RAPPORT GLOBAL" (lignes ~2706 à ~3582) du script actuel.
"""

# ══════════════════════════════════════════════════════════════════
# EXPORT DES DONNÉES INTERMÉDIAIRES → _data_mois.xlsx
# ══════════════════════════════════════════════════════════════════
DATA_MOIS_FILE = "_data_mois.xlsx"

try:
    print(f"\n📤 Export des données intermédiaires → {DATA_MOIS_FILE}...")

    # ── 1. Onglet "ca" : CA global et par solution + objectifs + N-1 ──
    _ca_row = {
        "Mois":                     mois_courant_label,
        "Année":                    month_start.year,
        "CA réalisé (€)":          round(total_realise, 2),
        "CA LITTERALIS (€)":       round(ca_par_solution.get("LITTERALIS", 0), 2),
        "CA GEODP (€)":            round(ca_par_solution.get("GEODP", 0), 2),
        "Objectif global (€)":     round(obj_global, 2) if obj_global else 0,
        "Objectif LITTERALIS (€)": round(obj_mois.get("LITTERALIS", 0), 2),
        "Objectif GEODP (€)":      round(obj_mois.get("GEODP", 0), 2),
        "CA N-1 global (€)":       round(hist_global, 2) if hist_global else 0,
        "CA N-1 LITTERALIS (€)":   round(hist_mois.get("LITTERALIS", 0), 2),
        "CA N-1 GEODP (€)":        round(hist_mois.get("GEODP", 0), 2),
    }
    _df_ca_export = pd.DataFrame([_ca_row])

    # ── 2. Onglet "ca_deal" : CA par type de deal et solution ──
    _deal_rows = []
    for sol in [None, "LITTERALIS", "GEODP"]:
        d = df_brut[df_brut["Montant déclaré ce mois"] > 0].copy() if not df_brut.empty else df_brut
        if sol:
            d = d[d["Solution"] == sol]
        if d.empty:
            continue
        ca_by_type = d.groupby("Type de deal")["Montant déclaré ce mois"].sum().to_dict()
        for deal_type, montant in ca_by_type.items():
            if montant > 0:
                _deal_rows.append({
                    "Mois":         mois_courant_label,
                    "Année":        month_start.year,
                    "Type de deal": deal_type,
                    "Solution":     sol or "GLOBAL",
                    "CA (€)":       round(montant, 2),
                })
    _df_ca_deal = pd.DataFrame(_deal_rows) if _deal_rows else pd.DataFrame(
        columns=["Mois", "Année", "Type de deal", "Solution", "CA (€)"])

    # ── 3. Onglet "commandes" : commandes du mois ──
    _cmd_cols = ["Mois", "Année", "Date de création", "Clé Epic", "Nom / Résumé",
                 "Client", "Solution", "Type de deal", "Statut", "Numéro BDC",
                 "Montant total prestations", "CA reconnu ce mois"]
    if 'commande_rows' in dir() and commande_rows:
        _df_cmd_export = pd.DataFrame(commande_rows)
        _df_cmd_export.insert(0, "Année", month_start.year)
        _df_cmd_export.insert(0, "Mois", mois_courant_label)
    else:
        _df_cmd_export = pd.DataFrame(columns=_cmd_cols)

    # ── 4. Onglet "charge" : charge par solution ──
    #    (reprend le ch_mois déjà calculé dans la section CHARGE HISTORIQUE)
    _df_charge_export = pd.DataFrame([ch_mois])

    # ── 5. Onglet "backlog" : métriques backlog par périmètre ──
    _bl_row = {"Mois": mois_courant_label, "Année": month_start.year}
    if 'df_bl' in dir() and not df_bl.empty:
        for suffix, df_sub in [("TOTAL", df_bl),
                                ("LITT", df_bl[df_bl["solution_cible"].isin(_BL_LITT_SOL)] if 'df_bl' in dir() else pd.DataFrame()),
                                ("GEODP", df_bl[df_bl["solution_cible"].isin(_BL_GEODP_SOL)] if 'df_bl' in dir() else pd.DataFrame())]:
            m = _bl_compute_metrics(df_sub) if not df_sub.empty else {}
            for key in ["total_backlog", "total_bloque", "total_mobilisable",
                        "nb_epics", "nb_projets", "nb_rework", "nb_bloques"]:
                _bl_row[f"{key}_{suffix}"] = m.get(key, 0)
    _df_backlog_export = pd.DataFrame([_bl_row])

    # ── Écriture du fichier ──
    with pd.ExcelWriter(DATA_MOIS_FILE, engine="openpyxl") as writer:
        _df_ca_export.to_excel(writer, sheet_name="ca", index=False)
        _df_ca_deal.to_excel(writer, sheet_name="ca_deal", index=False)
        _df_cmd_export.to_excel(writer, sheet_name="commandes", index=False)
        _df_charge_export.to_excel(writer, sheet_name="charge", index=False)
        _df_backlog_export.to_excel(writer, sheet_name="backlog", index=False)

    print(f"✅ Données intermédiaires exportées : {DATA_MOIS_FILE}")
    print(f"   → Exécutez ensuite : python Rapport_global.py")

except Exception as _e_data_mois:
    import traceback
    print(f"⚠️  Export _data_mois.xlsx échoué : {_e_data_mois}")
    traceback.print_exc()

# ══════════════════════════════════════════════════════════════════
# SAUVEGARDE
# ══════════════════════════════════════════════════════════════════
try:
    wb.save(EXPORT_FILE)
    print(f"✅ Export XLSX terminé : {EXPORT_FILE} ({len(df)} lignes)")
except Exception as e:
    print(f"❌ Impossible de sauvegarder le fichier Excel : {e}")
    sys.exit(1)