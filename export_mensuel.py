"""
export_mensuel.py — Export KPI mensuel (8 onglets)
===================================================
Génère l'export Excel à partir de jira_cache.pkl (produit par
collect_jira_worklogs.py), puis accumule les chiffres du mois dans
_historique.xlsx (via maj_historique.py).

8 onglets : Tableau de bord · Suivi de production · Suivi Hardware ·
Capacité productive · Charge · Backlog · Commandes du mois · Projets clôturés

Prérequis : jira_cache.pkl, absences_2026.xlsx, objectifs.xlsx, _historique.xlsx, ca_2025.xlsx
Sortie     : jira_export_AAAA-MM-JJ.xlsx
"""

import os
import sys
import pickle
from datetime import datetime, timedelta, date
from types import SimpleNamespace

import pandas as pd
import holidays
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

JIRA_SERVER = os.environ.get("JIRA_URL", "https://sogelink.atlassian.net")

# ══════════════════════════════════════════════════════════════════
# PALETTE & STYLES GLOBAUX
# ══════════════════════════════════════════════════════════════════
BLUE_DARK   = "1F3864"   # entêtes principaux
BLUE_MID    = "2E75B6"   # entêtes secondaires / titres
BLUE_LIGHT  = "D6E4F0"   # alternance lignes paires
BLUE_XLIGHT = "EBF3FB"   # fond léger
GREY_LINE   = "BDD7EE"   # bordures
WHITE       = "FFFFFF"
ORANGE      = "C55A11"   # alerte / rework
GREEN       = "375623"   # positif
RED_LIGHT   = "FFCCCC"   # alerte rouge

FONT_DEFAULT = "Calibri"
FONT_SIZE    = 11

def _font(bold=False, size=FONT_SIZE, color="000000", italic=False):
    return Font(name=FONT_DEFAULT, bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _border(color=GREY_LINE):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

STYLE_HEADER_DARK = {
    "font"      : _font(bold=True, color=WHITE),
    "fill"      : _fill(BLUE_DARK),
    "alignment" : _align("center"),
    "border"    : _border(BLUE_DARK),
}
STYLE_HEADER_MID = {
    "font"      : _font(bold=True, color=WHITE),
    "fill"      : _fill(BLUE_MID),
    "alignment" : _align("center"),
    "border"    : _border(BLUE_MID),
}
STYLE_TOTAL = {
    "font"      : _font(bold=True),
    "fill"      : _fill(BLUE_LIGHT),
    "alignment" : _align("center"),
    "border"    : _border(),
}

def apply_style(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)

def style_row(ws, row_idx, ncols, even=False, bold=False, total=False):
    """Applique alternance + bordures sur une ligne entière."""
    fill_color = BLUE_XLIGHT if even else WHITE
    if total:
        fill_color = BLUE_LIGHT
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill      = _fill(fill_color)
        cell.border    = _border()
        cell.font      = _font(bold=bold or total)
        cell.alignment = _align("left")

def style_header_row(ws, row_idx, ncols, dark=True):
    style = STYLE_HEADER_DARK if dark else STYLE_HEADER_MID
    for col in range(1, ncols + 1):
        apply_style(ws.cell(row=row_idx, column=col), style)

def autofit(ws, min_w=10, max_w=50, padding=3):
    """Ajuste la largeur de chaque colonne au contenu."""
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + padding, min_w), max_w)

def write_sheet_title(ws, title, subtitle=None):
    """Écrit le titre en A1, le sous-titre en A2, et laisse A3 vide.
    Les sections qui suivent n'ont besoin que d'un ws.append([]) pour
    atterrir en ligne 4 (en-tête de tableau)."""
    ws["A1"] = title
    ws["A1"].font = _font(bold=True, size=14, color=BLUE_DARK)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = _font(italic=True, size=10, color="595959")
    ws.row_dimensions[1].height = 24
    # Ligne 3 vide — le prochain append() tombera en ligne 4
    ws.append([])

# ══════════════════════════════════════════════════════════════════
# MAPPINGS MÉTIER
# ══════════════════════════════════════════════════════════════════
import unicodedata

def _normalize_solution(raw):
    """Normalise une valeur de solution Jira vers LITTERALIS / GEODP / la valeur brute.
       Insensible à la casse et aux accents, par mot-clé."""
    if not raw:
        return ""
    s = str(raw)
    # retirer accents + minuscules
    s_norm = "".join(c for c in unicodedata.normalize("NFD", s)
                     if unicodedata.category(c) != "Mn").lower()
    if "geodp" in s_norm:
        return "GEODP"
    if "litt" in s_norm or "liess" in s_norm or "sherpa" in s_norm:
        return "LITTERALIS"
    return s  # valeur inconnue : renvoyée telle quelle

class _SolutionMap:
    """Compat avec SOLUTION_MAPPING.get(x, default) : applique la normalisation
       par mot-clé. Si la valeur n'est pas reconnue, renvoie le défaut fourni."""
    def get(self, raw, default=None):
        res = _normalize_solution(raw)
        if res in ("LITTERALIS", "GEODP"):
            return res
        # non reconnue → comportement du dict d'origine : défaut si fourni, sinon valeur
        return default if default is not None else res

SOLUTION_MAPPING = _SolutionMap()

ABSENCES_FILE = "absences_2026.xlsx"

OBJECTIFS_FILE   = "objectifs.xlsx"
# Ordre de priorité : la dernière estimation prime. Pour chaque mois et chaque
# solution, on prend LE2 si renseigné, sinon LE1, sinon Budget.
OBJECTIFS_PRIORITE = ["LE2", "LE1", "Budget"]

def _lire_onglet_objectifs(filepath, onglet, mois_label):
    """Lit (global, litt, geodp) pour un mois dans un onglet donné.
       Retourne None si l'onglet ou le mois est absent."""
    try:
        df_obj = pd.read_excel(filepath, sheet_name=onglet, header=3)
        ligne = df_obj[df_obj["Mois"] == mois_label]
        if ligne.empty:
            return None
        row = ligne.iloc[0]
        col_g  = next((c for c in df_obj.columns if "global" in c.lower()), None)
        col_l  = next((c for c in df_obj.columns if "litt" in c.lower()), None)
        col_ge = next((c for c in df_obj.columns if "geodp" in c.lower()), None)
        def _v(col):
            if not col:
                return 0.0
            try:
                x = float(row[col])
                return x if x == x else 0.0   # écarte NaN
            except (ValueError, TypeError):
                return 0.0
        return _v(col_g), _v(col_l), _v(col_ge)
    except Exception:
        return None

def load_objectifs(filepath, mois_label, priorite=OBJECTIFS_PRIORITE):
    """Objectif du mois = la DERNIÈRE estimation disponible.
       Pour chaque valeur (global/litt/geodp), on prend LE2 si renseigné (>0),
       sinon LE1, sinon Budget. Le choix se fait valeur par valeur pour être
       robuste aux onglets partiellement remplis."""
    if not os.path.exists(filepath):
        print(f"⚠️  Fichier objectifs introuvable : {filepath} → objectifs à 0.")
        return 0.0, 0.0, 0.0

    # Lire les 3 onglets pour ce mois
    lectures = {}
    for onglet in priorite:
        r = _lire_onglet_objectifs(filepath, onglet, mois_label)
        if r is not None:
            lectures[onglet] = r

    if not lectures:
        print(f"⚠️  Mois '{mois_label}' absent des onglets objectifs → objectifs à 0.")
        return 0.0, 0.0, 0.0

    # Pour chaque composante, prendre la première valeur > 0 dans l'ordre de priorité
    def _choisir(idx):
        for onglet in priorite:
            if onglet in lectures and lectures[onglet][idx] > 0:
                return lectures[onglet][idx], onglet
        return 0.0, "—"

    g, src_g   = _choisir(0)
    l, src_l   = _choisir(1)
    ge, src_ge = _choisir(2)
    print(f"✅ Objectifs (dernière estimation) pour {mois_label} : "
          f"global {g:,.0f}€ [{src_g}] · Litt {l:,.0f}€ [{src_l}] · GEODP {ge:,.0f}€ [{src_ge}]")
    return g, l, ge

def load_absences(filepath, mois_label):
    """Lit l'onglet du mois dans absences_2026.xlsx (un onglet par mois,
       en-tête ligne 4). Colonnes : Collaborateur | Congés | Maladie.
       Retourne (total_dict, conges_dict, maladie_dict)."""
    empty = {}, {}, {}
    if not os.path.exists(filepath):
        print(f"⚠️  Fichier absences introuvable : {filepath} → absences à 0.")
        return empty
    try:
        xls = pd.ExcelFile(filepath)
        if mois_label not in xls.sheet_names:
            print(f"⚠️  Onglet '{mois_label}' absent de {filepath} (onglets : {xls.sheet_names}) → absences à 0.")
            return empty
        df_abs = pd.read_excel(filepath, sheet_name=mois_label, header=3).fillna(0)
        col_conges  = next((c for c in df_abs.columns if "cong" in c.lower()), None)
        col_maladie = next((c for c in df_abs.columns if "malad" in c.lower()), None)
        col_collab  = next((c for c in df_abs.columns if "collab" in c.lower()), None)
        if not all([col_collab, col_conges, col_maladie]):
            print(f"⚠️  Colonnes attendues : Collaborateur | Congés | Maladie. Trouvé : {list(df_abs.columns)}")
            return empty
        df_abs["_total"] = df_abs[col_conges] + df_abs[col_maladie]
        print(f"✅ Absences chargées depuis l'onglet '{mois_label}'.")
        return (dict(zip(df_abs[col_collab], df_abs["_total"])),
                dict(zip(df_abs[col_collab], df_abs[col_conges])),
                dict(zip(df_abs[col_collab], df_abs[col_maladie])))
    except Exception as e:
        print(f"⚠️  Lecture {filepath} impossible : {e}. Absences à 0.")
        return empty


# ══════════════════════════════════════════════════════════════════
# SOCLE : chargement cache + helpers hiérarchie/temps/CA
# ══════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════
# CHARGEMENT DU CACHE (remplace connexion Jira + récupération tickets)
# ══════════════════════════════════════════════════════════════════
# Le cache est produit par collect_jira_worklogs.py :
#   - worklogs      : liste Tempo filtrée équipe (issue_key résolu, nom, heures, date)
#   - issues_by_key : {key -> {id, key, summary, issuetype, parent_key, raw_fields}}
# On reconstruit des objets "issue-like" pour que les helpers existants
# (qui font issue.fields.customfield_xxx, .parent, .subtasks) continuent de marcher.


CACHE_FILE = "jira_cache.pkl"

print(f"📥 Chargement du cache {CACHE_FILE}...")
try:
    with open(CACHE_FILE, "rb") as f:
        _cache = pickle.load(f)
except FileNotFoundError:
    print(f"❌ Cache introuvable. Lance d'abord : python collect_jira_worklogs.py --refresh")
    sys.exit(1)

print(f"   → {len(_cache['issues_by_key'])} issues, "
      f"{len(_cache['worklogs'])} worklogs équipe ({_cache['mois_label']})")


# ── Reconstruction d'objets issue-like depuis les dicts bruts ──────
class _CFValue(dict):
    """Customfield de type 'option' : se comporte comme le dict brut
       {'value': 'X', ...} ET expose .value comme un objet Jira.
       Ainsi getattr(cf, 'value') ET cf.get('value') fonctionnent."""
    @property
    def value(self):
        return self.get("value", "")
    @property
    def name(self):
        return self.get("name", "")
    @property
    def displayName(self):
        return self.get("displayName", self.get("name", ""))


class _FieldsView:
    """Expose raw_fields comme des attributs : fields.customfield_xxx, fields.summary…
       - dicts d'option {'value':...} → renvoyés comme _CFValue (compat .value)
       - scalaires / listes → tels quels"""
    def __init__(self, info, issue_map_ref):
        raw = info["raw_fields"]
        object.__setattr__(self, "_raw", raw)
        object.__setattr__(self, "_info", info)
        object.__setattr__(self, "_imap", issue_map_ref)

    def __getattr__(self, name):
        raw = object.__getattribute__(self, "_raw")
        info = object.__getattribute__(self, "_info")

        if name == "summary":
            return info.get("summary", "")
        if name == "issuetype":
            return SimpleNamespace(name=info.get("issuetype") or "")
        if name == "parent":
            pk = info.get("parent_key")
            return SimpleNamespace(key=pk) if pk else None
        if name == "subtasks":
            # Non utilisé : la hiérarchie se reconstruit via children_by_parent
            # et les helpers remonter_vers_epic / descendre_depuis_epic.
            return []
        if name == "status":
            st = raw.get("status")
            return SimpleNamespace(name=st.get("name", "")) if isinstance(st, dict) else SimpleNamespace(name="")
        if name == "assignee":
            a = raw.get("assignee")
            return SimpleNamespace(displayName=a.get("displayName", "")) if isinstance(a, dict) else None

        # customfields & autres champs bruts
        if name in raw:
            val = raw[name]
            # dict d'option Jira {'value': ...} → wrapper compatible .value
            if isinstance(val, dict) and ("value" in val or "name" in val or "displayName" in val):
                return _CFValue(val)
            return val
        return None


class _IssueLike:
    def __init__(self, info, issue_map_ref):
        self.key = info["key"]
        self.id  = info["id"]
        self.fields = _FieldsView(info, issue_map_ref)
        # compat : certains helpers font issue.raw.get("fields", {})
        self.raw = {"fields": info["raw_fields"]}


# Construire issue_map (objets issue-like)
issue_map = {}
for key, info in _cache["issues_by_key"].items():
    issue_map[key] = _IssueLike(info, issue_map)

_all_issues_pool = list(issue_map.values())

# Index worklogs par issue_key (depuis Tempo, déjà filtré équipe + mois)
# Chaque worklog : {issue_key, author_id, author_name, hours, date}
worklog_by_issue = {}
for wl in _cache["worklogs"]:
    worklog_by_issue.setdefault(wl["issue_key"], []).append(wl)

# Index enfants par parent (depuis le cache)
children_by_parent = _cache.get("children_by_parent", {})

# Période (depuis le cache, fait foi)
month_start_date = _cache["month_start"]   # date
month_end_date   = _cache["month_end"]     # date (exclusif)
mois_courant_label = _cache["mois_label"]

# Clôtures du mois (clés des tickets passés à "Terminé" ce mois)
hw_clotures    = _cache.get("hw_clotures", set())
epics_clotures = _cache.get("epics_clotures", set())

# Absences du mois (onglet correspondant dans absences_2026.xlsx)
absences, absences_conges, absences_maladie = load_absences(ABSENCES_FILE, mois_courant_label)


# ── Remontée hiérarchique vers l'épic (récursive) ──────────────────
# Gère les 3 niveaux observés : worklog sur épic (0), tâche (1), sous-tâche (2).
# Cas tâche isolée sans parent (ex. PDMDEP-25450) → elle est sa propre racine.
_EPIC_TYPES = {"epic", "new delivery"}

def remonter_vers_epic(key):
    """Remonte par parent_key jusqu'à une épic. Retourne la clé de l'épic,
       ou la clé du ticket lui-même s'il n'a pas de parent (racine isolée)."""
    cur = key
    seen = set()
    while cur and cur in _cache["issues_by_key"] and cur not in seen:
        seen.add(cur)
        info = _cache["issues_by_key"][cur]
        itype = (info.get("issuetype") or "").lower()
        if itype in _EPIC_TYPES:
            return cur
        parent = info.get("parent_key")
        if not parent:
            return cur          # racine isolée : se compte elle-même
        cur = parent
    return key                  # filet de sécurité

def descendre_depuis_epic(epic_key):
    """Retourne l'épic + tous ses descendants (tâches + sous-tâches), récursivement."""
    result = {epic_key}
    pile = list(children_by_parent.get(epic_key, []))
    while pile:
        k = pile.pop()
        if k in result:
            continue
        result.add(k)
        pile.extend(children_by_parent.get(k, []))
    return result


# ── Temps du mois par épic (remplace sum_worklogs) ─────────────────
# Les worklogs Tempo sont déjà filtrés équipe + mois par le collecteur.
# On somme les heures de toutes les issues rattachées à l'épic (récursif).
def temps_mois_epic(epic_key):
    """Heures loggées ce mois sur l'épic + tous ses descendants."""
    total = 0.0
    for k in descendre_depuis_epic(epic_key):
        for wl in worklog_by_issue.get(k, []):
            total += wl["hours"]
    return total

def temps_mois_keys(keys):
    """Heures loggées ce mois sur une liste de clés (sans descente)."""
    total = 0.0
    for k in keys:
        for wl in worklog_by_issue.get(k, []):
            total += wl["hours"]
    return total


# ── Tickets CA enfants d'une épic ──────────────────────────────────
_CA_TYPES = ("COORDIN : Suivi CA", "Matériel GEODP")

def tickets_ca_de(epic_key):
    """Retourne les objets issue-like des tickets CA enfants directs de l'épic."""
    out = []
    for child_key in children_by_parent.get(epic_key, []):
        iss = issue_map.get(child_key)
        if iss and iss.fields.issuetype.name in _CA_TYPES:
            out.append(iss)
    return out


# ── Rattachement CA → épic PSC via les liens "Relates" ─────────────
# Les tickets CA (PDMDEP) pointent vers leur épic PSC par un lien
# Relates/outward. On lit le format brut issuelinks du cache.
def find_linked_epic(issue):
    """Retourne la clé de l'épic PSC liée à un ticket CA, ou None."""
    links = issue.raw["fields"].get("issuelinks") or []
    for ln in links:
        if (ln.get("type", {}) or {}).get("name") != "Relates":
            continue
        tgt = ln.get("outwardIssue") or ln.get("inwardIssue")
        if tgt and str(tgt.get("key", "")).startswith("PSC-"):
            return tgt["key"]
    return None


# ── Rattachement CA → épic PSC via les liens "Relates" ─────────────
# Les tickets CA (COORDIN : Suivi CA / Matériel GEODP) pointent vers leur
# épic PSC par un lien type "Relates" (outward ou inward), confirmé sur données.
def find_linked_epic(issue):
    """Retourne la clé de l'épic PSC liée à un ticket CA, ou None."""
    raw = issue.raw["fields"]
    for ln in raw.get("issuelinks") or []:
        if (ln.get("type") or {}).get("name") != "Relates":
            continue
        tgt = ln.get("outwardIssue") or ln.get("inwardIssue")
        if tgt and str(tgt.get("key", "")).startswith("PSC-"):
            return tgt["key"]
    return None


# ── Tickets CA liés à une épic PSC (via Relates) ───────────────────
# Index pré-calculé UNE fois : épic_psc_key -> [tickets CA].
# Évite de reparcourir les 26755 issues pour chaque épic PSC.
_ca_psc_index = {}
for _key, _iss in issue_map.items():
    if _iss.fields.issuetype.name in _CA_TYPES:
        _epic_psc = find_linked_epic(_iss)
        if _epic_psc:
            _ca_psc_index.setdefault(_epic_psc, []).append(_iss)

def tickets_ca_psc(epic_psc_key):
    """Retourne les tickets CA (issue-like) liés à une épic PSC donnée."""
    return _ca_psc_index.get(epic_psc_key, [])

# ══════════════════════════════════════════════════════════════════
# HELPERS MÉTIER (lecture des champs)
# ══════════════════════════════════════════════════════════════════
def get_category(issue) -> str:
    current = issue
    while True:
        if current.fields.issuetype.name.lower() == "epic":
            return getattr(getattr(current.fields, "customfield_23103", None), "value", "") or ""
        parent = getattr(current.fields, "parent", None)
        if not parent:
            return ""
        current = issue_map.get(parent.key)
        if not current:
            return ""


def get_solution(issue) -> str:
    current = issue
    while True:
        if current.fields.issuetype.name.lower() == "epic":
            return getattr(getattr(current.fields, "customfield_20514", None), "value", "") or ""
        parent = getattr(current.fields, "parent", None)
        if not parent:
            return ""
        current = issue_map.get(parent.key)
        if not current:
            return ""

def _read_client(raw_fields: dict) -> str:
    """
    Lit le client depuis customfield_10070.
    Ce champ peut être un dict {accountId, displayName}, une str, ou None.
    """
    val = raw_fields.get("customfield_10070")
    if not val:
        return ""
    if isinstance(val, dict):
        return val.get("displayName", val.get("name", ""))
    return str(val)


def _make_link(ws, row, col, key):
    """Rend la cellule d'une clé Epic cliquable vers Jira."""
    if not key or not str(key).strip():
        return
    cell = ws.cell(row=row, column=col)
    cell.hyperlink = f"{JIRA_SERVER}/browse/{key}"
    cell.style     = "Hyperlink"


def _read_montant(ca) -> float:
    """
    Lit le montant déclaré selon le type de ticket :
    - "Matériel GEODP" → customfield_10220
    - Autres           → customfield_22998
    """
    if getattr(ca.fields, "issuetype", None) and ca.fields.issuetype.name == "Matériel GEODP":
        return getattr(ca.fields, "customfield_10220", 0) or 0
    return getattr(ca.fields, "customfield_22998", 0) or 0


def _read_bdc(ca_fields_obj, ca_raw_fields: dict) -> str:
    """
    Lit le numéro BDC depuis customfield_10221 en priorité,
    sinon fallback sur customfield_24438 (ancien champ).
    """
    bdc = str(getattr(ca_fields_obj, "customfield_10221", "") or "").strip()
    if not bdc:
        bdc = str(ca_raw_fields.get("customfield_24438", "") or "").strip()
    return bdc


def _read_prestation(raw_fields: dict) -> str:
    """
    Lit le type de prestation depuis customfield_23955 (champ select : Hardware, Services…).
    Ce champ remonte comme un dict {"value": "Hardware", ...} ou une str.
    """
    val = raw_fields.get("customfield_23955")
    if not val:
        return ""
    if isinstance(val, dict):
        return val.get("value", "")
    return str(val)


DEAL_TYPE_MAPPING = {
    "new business"       : "Nouveau déploiement",
    "upsell"             : "Vente additionnelle",
    "additional setup"   : "Vente additionnelle",
}

def _read_deal_type(raw_fields: dict) -> str:
    """
    Détermine le type de deal :
    - "Migration"            si customfield_20514 (Solution cible) = "GEODP2 migration"
    - "Nouveau déploiement"  si customfield_22108 (Deal Type) = "New Business"
    - "Vente additionnelle"            si customfield_22108 = "Upsell" ou "Additional setup"
    - ""                     si non renseigné
    """
    solution_cible = (raw_fields.get("customfield_20514") or {}).get("value", "")
    if solution_cible.lower() == "geodp2 migration":
        return "Migration"
    deal_type_raw = (raw_fields.get("customfield_22108") or {}).get("value", "")
    if isinstance(raw_fields.get("customfield_22108"), str):
        deal_type_raw = raw_fields.get("customfield_22108", "")
    return DEAL_TYPE_MAPPING.get(deal_type_raw.lower().strip(), "")


# ══════════════════════════════════════════════════════════════════
# month_start compat (certains helpers utilisent month_start.year/.month)
# ══════════════════════════════════════════════════════════════════
month_start = datetime(month_start_date.year, month_start_date.month, 1)

# ══════════════════════════════════════════════════════════════════
# GÉNÉRATION DES LIGNES — PDMDEP
# ══════════════════════════════════════════════════════════════════
rows = []

for issue in _all_issues_pool:
    if issue.fields.issuetype.name != "Epic":
        continue

    epic_key  = issue.key
    epic_name = issue.fields.summary
    cf_obj    = getattr(issue.fields, "customfield_23103", None)

    raw_fields    = issue.raw.get("fields", {})
    raw_solution  = (raw_fields.get("customfield_20514") or {}).get("value", "")
    epic_solution = SOLUTION_MAPPING.get(raw_solution, raw_solution)
    projet        = raw_fields.get("customfield_23608", "")
    epic_category = cf_obj.value if hasattr(cf_obj, "value") else ""
    deal_type     = _read_deal_type(raw_fields)
    assignee_obj  = getattr(issue.fields, "assignee", None)
    assignee      = assignee_obj.displayName if assignee_obj else ""

    _created_raw = raw_fields.get("created")
    try:
        _created_dt = pd.to_datetime(_created_raw) if _created_raw else None
    except Exception:
        _created_dt = None
    date_commande = _created_dt.strftime("%d/%m/%Y") if _created_dt else ""
    ecart_mois    = (
        (month_start.year - _created_dt.year) * 12 + (month_start.month - _created_dt.month)
        if _created_dt else None
    )

    # Temps du mois (Tempo) sur l'épic + tous descendants
    total_month = temps_mois_epic(epic_key)

    # Tickets CA enfants directs (PDMDEP) — séparer services / hardware
    tickets_ca_all = tickets_ca_de(epic_key)
    tickets_ca = [t for t in tickets_ca_all if t.fields.issuetype.name != "Matériel GEODP"]
    tickets_hw = [t for t in tickets_ca_all if t.fields.issuetype.name == "Matériel GEODP"]

    has_amount = any(
        _read_montant(ca) > 0
        or (getattr(ca.fields, "customfield_23610", 0) or 0) != 0
        for ca in tickets_ca_all
    )
    if total_month <= 0 and not has_amount:
        continue

    # Logique A : le temps se répartit uniquement sur les tickets SERVICES.
    # Le Matériel GEODP ne consomme pas d'heures (vente de matériel).
    nb_ca                = max(1, len(tickets_ca))
    temps_mois_par_ligne = total_month / nb_ca

    # Lignes Matériel GEODP : montant seul, sans temps (iront en Suivi Hardware)
    for ca in tickets_hw:
        montant     = _read_montant(ca)
        fields_ca   = ca.raw.get("fields", {})
        bdc         = _read_bdc(ca.fields, fields_ca)
        rows.append({
            "Epic": epic_key, "Epic Nom": epic_name, "Assigné": assignee,
            "Client": _read_client(fields_ca), "Projet": projet, "Solution": epic_solution,
            "Catégorie": epic_category, "Type de deal": deal_type,
            "Date commande": date_commande, "Ancienneté (mois)": ecart_mois,
            "Temps mois (h)": 0, "Ticket CA": ca.key,
            "Type": fields_ca.get("customfield_24437", "") or "", "Numéro BDC": bdc,
            "Prestation": _read_prestation(fields_ca),
            "Prévision du mois": getattr(ca.fields, "customfield_23610", 0) or 0,
            "Montant déclaré ce mois": montant, "Rentabilité (€ / h)": 0,
            "_type_ca": "Matériel GEODP",
        })

    if tickets_ca:
        for ca in tickets_ca:
            montant                = _read_montant(ca)
            revenu_planifie_ticket = getattr(ca.fields, "customfield_23610", 0) or 0
            bdc                    = _read_bdc(ca.fields, ca.raw.get("fields", {}))
            rentabilite            = round(montant / temps_mois_par_ligne, 2) if temps_mois_par_ligne > 0 else 0
            fields_ca              = ca.raw.get("fields", {})
            prestation             = _read_prestation(fields_ca)
            type_presta            = fields_ca.get("customfield_24437", "") or ""
            client_ca              = _read_client(fields_ca)

            rows.append({
                "Epic": epic_key, "Epic Nom": epic_name, "Assigné": assignee,
                "Client": client_ca, "Projet": projet, "Solution": epic_solution,
                "Catégorie": epic_category, "Type de deal": deal_type,
                "Date commande": date_commande, "Ancienneté (mois)": ecart_mois,
                "Temps mois (h)": round(temps_mois_par_ligne, 2), "Ticket CA": ca.key,
                "Type": type_presta, "Numéro BDC": bdc, "Prestation": prestation,
                "Prévision du mois": revenu_planifie_ticket,
                "Montant déclaré ce mois": montant, "Rentabilité (€ / h)": rentabilite,
                "_type_ca": ca.fields.issuetype.name,
            })
    else:
        rows.append({
            "Epic": epic_key, "Epic Nom": epic_name, "Assigné": assignee,
            "Client": _read_client(raw_fields), "Projet": projet, "Solution": epic_solution,
            "Catégorie": epic_category, "Type de deal": deal_type,
            "Date commande": date_commande, "Ancienneté (mois)": ecart_mois,
            "Temps mois (h)": round(temps_mois_par_ligne, 2), "Ticket CA": "",
            "Type": "", "Numéro BDC": "", "Prestation": "",
            "Prévision du mois": 0, "Montant déclaré ce mois": 0, "Rentabilité (€ / h)": 0,
            "_type_ca": "",
        })

# ══════════════════════════════════════════════════════════════════
# GÉNÉRATION DES LIGNES — PSC
# ══════════════════════════════════════════════════════════════════
psc_epics = [iss for iss in _all_issues_pool
             if iss.key.startswith("PSC-")
             and iss.fields.issuetype.name == "New delivery"]

for issue in psc_epics:
    total_month    = temps_mois_epic(issue.key)
    psc_raw_fields = issue.raw.get("fields", {})
    client_psc     = _read_client(psc_raw_fields)

    _psc_created_raw = psc_raw_fields.get("created")
    try:
        _psc_created_dt = pd.to_datetime(_psc_created_raw) if _psc_created_raw else None
    except Exception:
        _psc_created_dt = None
    psc_date_commande = _psc_created_dt.strftime("%d/%m/%Y") if _psc_created_dt else ""
    psc_ecart_mois    = (
        (month_start.year - _psc_created_dt.year) * 12 + (month_start.month - _psc_created_dt.month)
        if _psc_created_dt else None
    )

    tous_tickets_ca_psc = tickets_ca_psc(issue.key)
    tickets_ca = [t for t in tous_tickets_ca_psc if t.fields.issuetype.name != "Matériel GEODP"]
    tickets_hw_psc = [t for t in tous_tickets_ca_psc if t.fields.issuetype.name == "Matériel GEODP"]

    has_ca = any(
        _read_montant(ca) > 0
        or (getattr(ca.fields, "customfield_23610", 0) or 0) != 0
        for ca in tous_tickets_ca_psc
    )
    if total_month <= 0 and not has_ca:
        continue

    # Logique A : temps réparti sur les services uniquement
    nb_ca                = max(1, len(tickets_ca))
    temps_mois_par_ligne = total_month / nb_ca

    raw_sol_psc   = (psc_raw_fields.get("customfield_20514") or {}).get("value", "")
    solution_psc  = SOLUTION_MAPPING.get(raw_sol_psc, raw_sol_psc)
    # Règle PSC : Projet si ≥1 ticket CA services rattaché, sinon Commande sans prestation
    cat_psc = "Projet" if tickets_ca else "Commande sans prestation"

    # Lignes Matériel GEODP liées au PSC : montant seul, sans temps
    for ca in tickets_hw_psc:
        fields_ca = ca.raw.get("fields", {})
        rows.append({
            "Epic": issue.key, "Epic Nom": issue.fields.summary, "Assigné": "",
            "Client": client_psc, "Projet": psc_raw_fields.get("customfield_23608", ""),
            "Solution": solution_psc, "Catégorie": "",
            "Type de deal": _read_deal_type(psc_raw_fields),
            "Date commande": psc_date_commande, "Ancienneté (mois)": psc_ecart_mois,
            "Temps mois (h)": 0, "Ticket CA": ca.key,
            "Type": fields_ca.get("customfield_24437", "") or "",
            "Numéro BDC": _read_bdc(ca.fields, fields_ca),
            "Prestation": _read_prestation(fields_ca),
            "Prévision du mois": getattr(ca.fields, "customfield_23610", 0) or 0,
            "Montant déclaré ce mois": _read_montant(ca), "Rentabilité (€ / h)": 0,
            "_type_ca": "Matériel GEODP",
        })

    if tickets_ca:
        for ca in tickets_ca:
            montant                = _read_montant(ca)
            revenu_planifie_ticket = getattr(ca.fields, "customfield_23610", 0) or 0
            bdc                    = _read_bdc(ca.fields, ca.raw.get("fields", {}))
            rentabilite            = round(montant / temps_mois_par_ligne, 2) if temps_mois_par_ligne > 0 else 0
            fields_ca              = ca.raw.get("fields", {})
            prestation             = _read_prestation(fields_ca)
            type_presta            = fields_ca.get("customfield_24437", "") or ""

            rows.append({
                "Epic": issue.key, "Epic Nom": issue.fields.summary, "Assigné": "",
                "Client": client_psc, "Projet": psc_raw_fields.get("customfield_23608", ""),
                "Solution": solution_psc, "Catégorie": cat_psc,
                "Type de deal": _read_deal_type(psc_raw_fields),
                "Date commande": psc_date_commande, "Ancienneté (mois)": psc_ecart_mois,
                "Temps mois (h)": round(temps_mois_par_ligne, 2), "Ticket CA": ca.key,
                "Type": type_presta, "Numéro BDC": bdc, "Prestation": prestation,
                "Prévision du mois": revenu_planifie_ticket,
                "Montant déclaré ce mois": montant, "Rentabilité (€ / h)": rentabilite,
                "_type_ca": ca.fields.issuetype.name,
            })
    elif not tickets_hw_psc:
        # Épic PSC avec du temps mais SANS aucun ticket CA (ni services ni hardware).
        # On garde la ligne (décision : montrer tout le temps passé).
        rows.append({
            "Epic": issue.key, "Epic Nom": issue.fields.summary, "Assigné": "",
            "Client": client_psc, "Projet": psc_raw_fields.get("customfield_23608", ""),
            "Solution": solution_psc, "Catégorie": cat_psc,
            "Type de deal": _read_deal_type(psc_raw_fields),
            "Date commande": psc_date_commande, "Ancienneté (mois)": psc_ecart_mois,
            "Temps mois (h)": round(temps_mois_par_ligne, 2), "Ticket CA": "",
            "Type": "", "Numéro BDC": "", "Prestation": "",
            "Prévision du mois": 0, "Montant déclaré ce mois": 0, "Rentabilité (€ / h)": 0,
            "_type_ca": "",
        })

df = pd.DataFrame(rows)
print(f"   {len(df)} lignes générées")

# ══════════════════════════════════════════════════════════════════
# CONSTRUCTION DU df_wl — worklogs Tempo catégorisés (pour la Charge)
# Chaque worklog (déjà filtré équipe + mois) est enrichi de la catégorie
# et de la solution de son épic, via remontée hiérarchique.
# ══════════════════════════════════════════════════════════════════
SUPPORT_LITT_EPICS  = {"PDMDEP-2171", "PDMDEP-12707"}
SUPPORT_GEODP_EPICS = {"PDMDEP-2172"}

def _categorie_et_solution(ticket_key):
    """Remonte vers l'épic (Epic ou New delivery) et lit catégorie + solution.
       Règles spéciales :
       - 'COORDIN : Suivi hors projet' → toujours Maintenance
       - épic PSC (New delivery) → Projet si elle a ≥1 ticket CA rattaché, sinon
         Commande sans prestation (PSC n'a pas de customfield_23103)."""
    cur = ticket_key
    seen = set()
    while cur and cur in _cache["issues_by_key"] and cur not in seen:
        seen.add(cur)
        info = _cache["issues_by_key"][cur]
        itype = (info.get("issuetype") or "")
        itype_l = itype.lower()

        # Règle : Suivi hors projet → Maintenance (quel que soit le niveau)
        if itype == "COORDIN : Suivi hors projet":
            return "Maintenance", "", cur

        if itype_l == "epic":
            raw = info["raw_fields"]
            cat = (raw.get("customfield_23103") or {}).get("value", "") or ""
            sol_raw = (raw.get("customfield_20514") or {}).get("value", "") or ""
            sol = SOLUTION_MAPPING.get(sol_raw, "Autre" if sol_raw else "")
            return cat, sol, cur

        if itype_l == "new delivery":
            # Épic PSC : toujours LITTERALIS (déploiement Littéralis).
            # Catégorie selon présence d'un ticket CA rattaché.
            a_ca = len(tickets_ca_psc(cur)) > 0
            cat = "Projet" if a_ca else "Commande sans prestation"
            return cat, "LITTERALIS", cur

        parent = info.get("parent_key")
        if not parent:
            break
        cur = parent
    return "", "", None

def _support_solution(ticket_key):
    """Pour les heures Support N2 : rattache à Littéralis/GEODP selon épic dédiée."""
    cur = ticket_key
    for _ in range(6):
        if cur in SUPPORT_LITT_EPICS:
            return "LITTERALIS"
        if cur in SUPPORT_GEODP_EPICS:
            return "GEODP"
        info = _cache["issues_by_key"].get(cur)
        if not info or not info.get("parent_key"):
            break
        cur = info["parent_key"]
    return ""

wl_rows = []
for wl in _cache["worklogs"]:
    tk = wl["issue_key"]
    cat, sol, _epic = _categorie_et_solution(tk)
    # Correction Support : solution rattachée via épic de support dédiée
    if cat == "Support N2":
        sol_sup = _support_solution(tk)
        if sol_sup:
            sol = sol_sup
    wl_rows.append({
        "Collaborateur": wl["author_name"],
        "Ticket": tk,
        "Catégorie": cat,
        "Solution": sol,
        "Epic": _epic or remonter_vers_epic(tk),
        "Temps (h)": wl["hours"],
        "Date": wl["date"],
    })

df_wl = pd.DataFrame(wl_rows)


# ══════════════════════════════════════════════════════════════════
# ANALYSE DU TEMPS NON VALORISÉ (temps travaillé sans CA déclaré ce mois)
# ══════════════════════════════════════════════════════════════════
def _analyse_temps_non_valorise(df_wl):
    """Décompose le temps productif du mois selon qu'il a généré du CA ou non.

    Structure (heures, par solution LITTERALIS/GEODP) :
      - rework            : temps sur épics Rework
      - gratuit           : temps sur épics gratuits (presta offerte, etc.)
      - projet_sans_ca    : temps sur projets qui n'ont PAS déclaré de CA ce mois
          dont bloque / depasse / reste
      - projet_avec_ca_perdu : heures bloquées/dépassées sur projets qui ONT déclaré du CA
          dont bloque / depasse

    Règle bloqué/dépassé : chaque heure bloquée/dépassée est rattachée à son épic ;
    selon que l'épic a déclaré du CA ce mois ou non, elle va dans projet_sans_ca
    (retranchée du "reste") ou dans projet_avec_ca_perdu.
    """
    # Les champs 27820/27853 sont déjà exprimés EN HEURES (pas en secondes).
    SEC_PAR_H = 1.0

    def _vide():
        return {
            "rework": 0.0, "gratuit": 0.0,
            "projet_sans_ca_total": 0.0,
            "sans_ca_bloque": 0.0, "sans_ca_depasse": 0.0, "sans_ca_reste": 0.0,
            "avec_ca_bloque": 0.0, "avec_ca_depasse": 0.0,
        }
    res = {"LITTERALIS": _vide(), "GEODP": _vide()}

    # 1) CA déclaré ce mois par épic (via le ticket Suivi CA de l'épic)
    ca_par_epic = {}   # epic_key -> CA du mois
    for k, info in _cache["issues_by_key"].items():
        if info.get("issuetype") != "COORDIN : Suivi CA":
            continue
        epic = remonter_vers_epic(k)
        ca_mois = float(info["raw_fields"].get("customfield_22998") or 0)
        ca_par_epic[epic] = ca_par_epic.get(epic, 0) + ca_mois

    def _a_declare_ca(epic):
        return ca_par_epic.get(epic, 0) > 0

    # 2) Temps bloqué/dépassé par épic (somme des tickets COORDIN : Paramétrages)
    bloque_par_epic = {}
    depasse_par_epic = {}
    for k, info in _cache["issues_by_key"].items():
        if info.get("issuetype") != "COORDIN : Paramétrages":
            continue
        epic = remonter_vers_epic(k)
        raw = info["raw_fields"]
        b = float(raw.get("customfield_27820") or 0) / SEC_PAR_H
        d = float(raw.get("customfield_27853") or 0) / SEC_PAR_H
        bloque_par_epic[epic]  = bloque_par_epic.get(epic, 0) + b
        depasse_par_epic[epic] = depasse_par_epic.get(epic, 0) + d

    # 3) Répartir le temps productif (worklogs) par catégorie et valorisation
    if df_wl is not None and not df_wl.empty:
        for _, r in df_wl.iterrows():
            sol = r["Solution"]
            if sol not in res:
                continue
            cat = r["Catégorie"]
            h = float(r["Temps (h)"] or 0)
            if cat == "Rework":
                res[sol]["rework"] += h
            elif cat in ("Prestation offerte", "Commande sans prestation", "Maintenance", "Assistance Expert"):
                res[sol]["gratuit"] += h
            elif cat == "Projet":
                epic = r["Epic"]
                if not _a_declare_ca(epic):
                    # projet sans CA déclaré ce mois → temps non valorisé
                    res[sol]["projet_sans_ca_total"] += h

    # 4) Ventiler bloqué/dépassé selon valorisation de l'épic
    def _sol_epic(epic):
        return _solution_de_ticket(epic) or ""
    for epic, b in bloque_par_epic.items():
        sol = _sol_epic(epic)
        if sol not in res or b <= 0:
            continue
        if _a_declare_ca(epic):
            res[sol]["avec_ca_bloque"] += b
        else:
            res[sol]["sans_ca_bloque"] += b
    for epic, d in depasse_par_epic.items():
        sol = _sol_epic(epic)
        if sol not in res or d <= 0:
            continue
        if _a_declare_ca(epic):
            res[sol]["avec_ca_depasse"] += d
        else:
            res[sol]["sans_ca_depasse"] += d

    # 5) Calculer le "reste à expliquer" = projet_sans_ca - bloqué - dépassé (plancher 0)
    for sol in res:
        d = res[sol]
        reste = d["projet_sans_ca_total"] - d["sans_ca_bloque"] - d["sans_ca_depasse"]
        d["sans_ca_reste"] = max(0.0, round(reste, 2))
        # arrondis
        for key in d:
            d[key] = round(d[key], 2)
    return res


# (l'appel à _analyse_temps_non_valorise est fait plus bas, une fois toutes
#  les fonctions dépendantes définies)

# Diagnostic : totaux par catégorie et par solution (pour validation)
if not df_wl.empty:
    print("\n   ── Charge Tempo : totaux de contrôle ──")
    print(f"   Total heures équipe : {df_wl['Temps (h)'].sum():.1f}h")
    print("   Par catégorie :")
    for cat, h in df_wl.groupby('Catégorie')['Temps (h)'].sum().sort_values(ascending=False).items():
        print(f"     {cat or '(vide)':<28} {h:>7.1f}h")
    print("   Par solution :")
    for sol, h in df_wl.groupby('Solution')['Temps (h)'].sum().sort_values(ascending=False).items():
        print(f"     {sol or '(vide)':<28} {h:>7.1f}h")

def write_sheet_title(ws, title, subtitle=None):
    """Écrit le titre en A1, le sous-titre en A2, et laisse A3 vide.
    Les sections qui suivent n'ont besoin que d'un ws.append([]) pour
    atterrir en ligne 4 (en-tête de tableau)."""
    ws["A1"] = title
    ws["A1"].font = _font(bold=True, size=14, color=BLUE_DARK)
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = _font(italic=True, size=10, color="595959")
    ws.row_dimensions[1].height = 24
    # Ligne 3 vide — le prochain append() tombera en ligne 4
    ws.append([])

def _write_ca_sheet(ws, df_source, title_str, subtitle_str):
    """Écrit un onglet de type Suivi CA : données + ligne de total."""
    write_sheet_title(ws, title_str, subtitle_str)

    data_start_row = 4
    rows_to_write = list(dataframe_to_rows(df_source, index=False, header=True))
    for r_offset, row_data in enumerate(rows_to_write):
        for c_offset, value in enumerate(row_data):
            ws.cell(row=data_start_row + r_offset, column=c_offset + 1, value=value)

    ncols = len(df_source.columns)
    col_indices = {name: idx for idx, name in enumerate(df_source.columns, start=1)}
    style_header_row(ws, data_start_row, ncols, dark=True)

    epic_col = col_indices.get("Epic")
    # Filtre automatique sur l'en-tête
    from openpyxl.utils import get_column_letter
    last_col_letter = get_column_letter(ncols)
    ws.auto_filter.ref = f"A{data_start_row}:{last_col_letter}{data_start_row}"
    for i, row_cells in enumerate(ws.iter_rows(min_row=data_start_row + 1, max_row=ws.max_row)):
        row_i = row_cells[0].row
        style_row(ws, row_i, ncols, even=(i % 2 == 0))
        if epic_col:
            _make_link(ws, row_i, epic_col, ws.cell(row=row_i, column=epic_col).value)
            # Réappliquer le fill car _make_link écrase le style
            epic_cell = ws.cell(row=row_i, column=epic_col)
            epic_cell.fill = _fill(BLUE_LIGHT if (i % 2 == 0) else "FFFFFF")

    # Formatage monétaire sur les données
    money_cols = ["Prévision du mois", "Montant déclaré ce mois", "Rentabilité (€ / h)"]
    for col_name in money_cols:
        if col_name in col_indices:
            col_idx = col_indices[col_name]
            for row in ws.iter_rows(min_row=data_start_row + 1, max_row=ws.max_row,
                                    min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = '#,##0.00 "€"'

    # ── Ligne de total ────────────────────────────────────────────
    if not df_source.empty:
        total_row_idx = ws.max_row + 1
        cols = df_source.columns
        total_previ = df_source["Prévision du mois"].sum()      if "Prévision du mois"      in cols else None
        total_ca    = df_source["Montant déclaré ce mois"].sum() if "Montant déclaré ce mois" in cols else 0
        h_sum       = df_source["Temps mois (h)"].sum()          if "Temps mois (h)"          in cols else 0
        total_rent  = round(total_ca / h_sum, 2) if h_sum > 0 else None

        # Construire la ligne total (valeurs vides sur les colonnes non sommées)
        total_values = []
        for col_name in cols:
            if col_name == "Epic":
                total_values.append("TOTAL")
            elif col_name == "Prévision du mois":
                total_values.append(total_previ)
            elif col_name == "Montant déclaré ce mois":
                total_values.append(total_ca)
            elif col_name == "Rentabilité (€ / h)":
                total_values.append(total_rent)
            else:
                total_values.append("")

        for c_offset, value in enumerate(total_values):
            ws.cell(row=total_row_idx, column=c_offset + 1, value=value)
        style_row(ws, total_row_idx, ncols, total=True)
        for col_name in ["Prévision du mois", "Montant déclaré ce mois", "Rentabilité (€ / h)"]:
            if col_name in col_indices:
                ws.cell(row=total_row_idx, column=col_indices[col_name]).number_format = '#,##0.00 "€"'

    # ── Mise en rouge : Montant déclaré < Prévision (si les deux colonnes existent) ──
    if "Prévision du mois" in col_indices and "Montant déclaré ce mois" in col_indices:
        col_previ = col_indices["Prévision du mois"]
        col_ca    = col_indices["Montant déclaré ce mois"]
        RED_FILL  = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")
        RED_FONT  = _font(bold=False, color="C00000")
        for row_idx in range(data_start_row + 1, ws.max_row):  # hors ligne total
            previ = ws.cell(row=row_idx, column=col_previ).value or 0
            ca    = ws.cell(row=row_idx, column=col_ca).value   or 0
            if isinstance(previ, (int, float)) and isinstance(ca, (int, float)) and previ > 0 and ca < previ:
                cell = ws.cell(row=row_idx, column=col_ca)
                cell.fill = RED_FILL
                cell.font = RED_FONT

    ws.freeze_panes = f"A{data_start_row + 1}"
    autofit(ws)

COLS_SUIVI_CA = [
    "Epic", "Epic Nom", "Assigné", "Client", "Projet", "Solution", "Catégorie",
    "Type de deal", "Date commande", "Ancienneté (mois)", "Temps mois (h)", "Ticket CA", "Numéro BDC",
    "Prévision du mois", "Montant déclaré ce mois", "Rentabilité (€ / h)",
]
COLS_SUIVI_HW = [
    "Epic", "Epic Nom", "Client", "Projet", "Solution", "Catégorie",
    "Type de deal", "Temps mois (h)", "Ticket CA", "Numéro BDC",
    "Montant déclaré ce mois",
]

def _filter_cols(df, cols):
    return df[[c for c in cols if c in df.columns]].copy()


# ══════════════════════════════════════════════════════════════════
# FILTRAGE SERVICES (Hardware exclu, Interne/Support exclus)
# ══════════════════════════════════════════════════════════════════
if not df.empty:
    # Hardware = tickets de type "Matériel GEODP" → exclus du Suivi de production
    mask_hw = df["_type_ca"] == "Matériel GEODP"
    df_services = df[~mask_hw & ~df["Catégorie"].isin(["Interne", "Support N2"])].copy()
else:
    df_services = df.copy()

# Exclure les lignes sans temps, sans prévision et sans montant
if not df_services.empty:
    df_services = df_services[
        (df_services["Temps mois (h)"] > 0) |
        (df_services["Prévision du mois"] > 0) |
        (df_services["Montant déclaré ce mois"] != 0)
    ]

print(f"   {len(df_services)} lignes services (après filtrage)")

# ══════════════════════════════════════════════════════════════════
# FILTRAGE HARDWARE (tickets Matériel GEODP CLÔTURÉS ce mois)
# ══════════════════════════════════════════════════════════════════
if not df.empty:
    df_hardware = df[df["_type_ca"] == "Matériel GEODP"].copy()
    # Ne garder que les tickets Matériel GEODP clôturés (passés à "Terminé") ce mois.
    # Le ticket Matériel GEODP est dans la colonne "Ticket CA".
    if not df_hardware.empty:
        df_hardware = df_hardware[df_hardware["Ticket CA"].isin(hw_clotures)]
else:
    df_hardware = df.copy()

print(f"   {len(df_hardware)} lignes hardware (clôturées ce mois)")

# ══════════════════════════════════════════════════════════════════
# ÉCRITURE DU FICHIER
# ══════════════════════════════════════════════════════════════════
wb = Workbook()
ws = wb.active
ws.title = "Suivi de production"
_write_ca_sheet(ws, _filter_cols(df_services, COLS_SUIVI_CA),
    f"Suivi de production — {mois_courant_label} {month_start.year}",
    "Prestations de services — Hardware exclu")

# ── Onglet 2 : Suivi Hardware ──────────────────────────────────────
ws_hw = wb.create_sheet("Suivi Hardware")
_write_ca_sheet(ws_hw, _filter_cols(df_hardware, COLS_SUIVI_HW),
    f"Suivi Hardware — {mois_courant_label} {month_start.year}",
    f"Tickets 'Matériel GEODP' clôturés ce mois — {len(df_hardware)} ligne(s)")

# ══════════════════════════════════════════════════════════════════
# JOURS OUVRÉS DU MOIS
# ══════════════════════════════════════════════════════════════════
all_days           = pd.date_range(month_start_date, month_end_date, freq="D")
fr_holidays        = holidays.France(years=month_start_date.year)
holidays_in_period = [d for d in all_days if d in fr_holidays]
working_days       = [d for d in all_days if d.weekday() < 5 and d not in holidays_in_period]
num_working_days   = len(working_days)

# ══════════════════════════════════════════════════════════════════
# ONGLET 3 : Capacité productive
# ══════════════════════════════════════════════════════════════════
ws_cap = wb.create_sheet("Capacité productive")
write_sheet_title(ws_cap, "Capacité productive — Chargés de déploiement & Chefs de projet",
                  f"Mois de {mois_courant_label} {month_start.year} — {num_working_days} jours ouvrés")

collaborateurs_ana = [
    "Bérénice Bossard", "Marine Masingarbe", "Duncan Hamelin",
    "Maxime Pontonnier", "Quentin Bordillon",
    "Fabien Reutenauer", "Flavie Bardin", "Rémy Vincent",
]
# Flavie Bardin à 0% (départ fin mois précédent, remplacement en cours)
taux_global_ana = [1.00, 1.00, 1.00, 0.80, 1.00, 1.00, 0.00, 1.00]
heures_par_jour = 7

# Calculs PAR PERSONNE (gardés en mémoire pour la Charge et le dashboard,
# mais NON écrits dans l'export public — données RH sensibles).
capacites = []
for name, taux in zip(collaborateurs_ana, taux_global_ana):
    jours_abs       = absences.get(name, 0)
    jours_effectifs = max(0, num_working_days - jours_abs)
    heures_total    = jours_effectifs * heures_par_jour * taux
    capacites.append((name, taux, jours_effectifs, heures_total))

# ── Onglet public : TOTAUX d'équipe uniquement (aucun nominatif) ──
headers_cap = ["Indicateur", "Valeur"]
cap_header_row = 4
ws_cap.append(headers_cap)
style_header_row(ws_cap, cap_header_row, len(headers_cap))

nb_etp_actifs   = sum(1 for _, taux, _, _ in capacites if taux > 0)
total_jours_eff = sum(c[2] for c in capacites)
total_heures    = sum(c[3] for c in capacites)
total_absences  = sum(absences.get(n, 0) for n in collaborateurs_ana)

resume_lignes = [
    ("Effectif de l'équipe", f"{len(collaborateurs_ana)} personnes"),
    ("Collaborateurs actifs ce mois", f"{nb_etp_actifs}"),
    ("Jours ouvrés du mois", f"{num_working_days}"),
    ("Total jours travaillés (équipe)", f"{total_jours_eff:.0f} j"),
    ("Capacité totale disponible", f"{total_heures:.0f} h"),
]
for i, (lib, val) in enumerate(resume_lignes):
    ws_cap.append([lib, val])
    style_row(ws_cap, cap_header_row + 1 + i, len(headers_cap), even=(i % 2 == 0))

total_row = cap_header_row + 1 + len(resume_lignes)

# Répartition théorique
repartition_litteralis = {"Productif": 0.75, "Support": 0.20, "Interne": 0.05}
repartition_geodp      = {"Productif": 0.95, "Support": 0.00, "Interne": 0.05}
repartition_cp         = {"Productif": 0.90, "Support": 0.00, "Interne": 0.10}

equipe_litteralis = ["Bérénice Bossard", "Marine Masingarbe", "Duncan Hamelin"]
equipe_geodp      = ["Maxime Pontonnier", "Quentin Bordillon"]
equipe_cp         = ["Fabien Reutenauer", "Flavie Bardin", "Rémy Vincent"]

heures_litteralis = sum(c[3] for c in capacites if c[0] in equipe_litteralis)
heures_geodp      = sum(c[3] for c in capacites if c[0] in equipe_geodp)
heures_cp         = sum(c[3] for c in capacites if c[0] in equipe_cp)

def _write_repartition(ws, label, repartition, heures_equipe):
    ws.append([])
    ws.append([f"Répartition théorique — {label}"])
    title_row = ws.max_row
    ws[title_row][0].font      = _font(bold=True, size=12, color=BLUE_MID)
    ws[title_row][0].alignment = _align()
    ws.append(["Catégorie", "% théorique", "Heures théoriques"])
    style_header_row(ws, ws.max_row, 3, dark=False)
    for j, (cat, pct) in enumerate(repartition.items()):
        ws.append([cat, f"{int(pct*100)}%", round(heures_equipe * pct, 2)])
        style_row(ws, ws.max_row, 3, even=(j % 2 == 0))

_write_repartition(ws_cap, "Littéralis (CD)", repartition_litteralis, heures_litteralis)
_write_repartition(ws_cap, "GEODP (CD)",      repartition_geodp,      heures_geodp)
_write_repartition(ws_cap, "Chefs de projet", repartition_cp,         heures_cp)

heures_total_equipe = heures_litteralis + heures_geodp + heures_cp
if heures_total_equipe > 0:
    prod_global = (heures_litteralis * repartition_litteralis["Productif"]
                   + heures_geodp * repartition_geodp["Productif"]
                   + heures_cp * repartition_cp["Productif"])
    sup_global = (heures_litteralis * repartition_litteralis["Support"]
                  + heures_geodp * repartition_geodp["Support"]
                  + heures_cp * repartition_cp["Support"])
    int_global = heures_total_equipe - prod_global - sup_global
    _write_repartition(ws_cap, "Équipe globale",
        {"Productif": prod_global / heures_total_equipe,
         "Support":   sup_global  / heures_total_equipe,
         "Interne":   int_global  / heures_total_equipe},
        heures_total_equipe)

autofit(ws_cap)

# ══════════════════════════════════════════════════════════════════
# ONGLET 4 : Charge (fusionné, sans graphiques)
# ══════════════════════════════════════════════════════════════════
# Capacités PRODUCTIVES par solution (CD de la solution + CP ventilés)
CP_SPLIT = {
    "Flavie Bardin"     : {"LITTERALIS": 0.40, "GEODP": 0.60},
    "Fabien Reutenauer" : {"LITTERALIS": 1.00, "GEODP": 0.00},
    "Rémy Vincent"      : {"LITTERALIS": 1.00, "GEODP": 0.00},
}
# Ajustement : part non-productive (interne/support) retirée de la capacité théorique productive
ajustements = {
    "Bérénice Bossard": 0.25, "Marine Masingarbe": 0.25, "Duncan Hamelin": 0.25,
    "Maxime Pontonnier": 0.05, "Quentin Bordillon": 0.05,
    "Fabien Reutenauer": 0.10, "Flavie Bardin": 0.10, "Rémy Vincent": 0.10,
}
cap_theo_dict = {name: htot * (1 - ajustements.get(name, 0))
                 for name, taux, jeff, htot in capacites}

def _cap_solution(sol_key):
    equipe_cd = equipe_litteralis if sol_key == "LITTERALIS" else equipe_geodp
    cap = sum(cap_theo_dict.get(m, 0) for m in equipe_cd)
    for m in equipe_cp:
        cap += cap_theo_dict.get(m, 0) * CP_SPLIT.get(m, {}).get(sol_key, 0)
    return cap

cap_par_sol = {"LITTERALIS": _cap_solution("LITTERALIS"), "GEODP": _cap_solution("GEODP")}

# Catégories regroupées
CAT_PROJET  = {"Projet"}
CAT_REWORK  = {"Rework"}
CAT_GRATUIT = {"Prestation offerte", "Commande sans prestation", "Maintenance", "Assistance Expert"}
CAT_SUPPORT = {"Support N2"}

def _h(df, sol, cats):
    if df.empty:
        return 0.0
    return df[(df["Solution"] == sol) & (df["Catégorie"].isin(cats))]["Temps (h)"].sum()

ws_ch = wb.create_sheet("Charge")
write_sheet_title(ws_ch, f"Charge — {mois_courant_label} {month_start.year}",
                  "Heures réalisées par solution et par type (source : Tempo)")

headers_ch = ["Solution", "Projet avec CA", "Rework", "Gratuit",
              "Productif (total)", "Support", "Capacité productive", "% occupation"]
ch_hdr = 4
ws_ch.append(headers_ch)
style_header_row(ws_ch, ch_hdr, len(headers_ch))

def _pct_color(pct):
    if pct < 70:    return "FFD9A0"  # orange : sous-occupé
    if pct < 80:    return "FFF2A8"  # jaune
    if pct <= 100:  return "C6EFCE"  # vert
    return "FFB3B3"                  # rouge : surcharge

ligne_idx = ch_hdr
for sol_label, sol_key in [("Littéralis", "LITTERALIS"), ("GEODP", "GEODP")]:
    h_proj = _h(df_wl, sol_key, CAT_PROJET)
    h_rew  = _h(df_wl, sol_key, CAT_REWORK)
    h_grat = _h(df_wl, sol_key, CAT_GRATUIT)
    h_sup  = _h(df_wl, sol_key, CAT_SUPPORT)
    productif = h_proj + h_rew + h_grat
    cap = cap_par_sol.get(sol_key, 0)
    pct = round(productif / cap * 100, 1) if cap > 0 else 0
    ligne_idx += 1
    ws_ch.append([sol_label, round(h_proj, 2), round(h_rew, 2), round(h_grat, 2),
                  round(productif, 2), round(h_sup, 2), round(cap, 2), f"{pct:.1f}%"])
    style_row(ws_ch, ligne_idx, len(headers_ch), even=(ligne_idx % 2 == 0))
    # couleur % occupation (dernière colonne)
    pcell = ws_ch.cell(ligne_idx, len(headers_ch))
    pcell.fill = _fill(_pct_color(pct))
    pcell.font = _font(bold=True)
    pcell.alignment = _align(h="center")

# ── Bloc Interne (séparé, non ventilé par solution) ────────────────
h_interne = df_wl[df_wl["Catégorie"] == "Interne"]["Temps (h)"].sum() if not df_wl.empty else 0
ws_ch.append([])
ws_ch.append(["Interne (toutes équipes)", round(h_interne, 2)])
r_int = ws_ch.max_row
ws_ch[r_int][0].font = _font(bold=True, size=11, color=BLUE_MID)
ws_ch.cell(r_int, 2).font = _font(bold=True)

autofit(ws_ch)

# ══════════════════════════════════════════════════════════════════
# ONGLET 5 : Backlog à date
# ══════════════════════════════════════════════════════════════════
# Unité : tickets "COORDIN : Suivi CA" non clôturés.
#   customfield_22042 = montant restant à reconnaître (backlog en €)
#   customfield_24269 = type de blocage (Aucun / Blocage client / commerce / produit)
# Solution remontée depuis l'épic parent.
def _solution_de_ticket(ticket_key):
    _cat, sol, _epic = _categorie_et_solution(ticket_key)
    if sol in ("LITTERALIS", "GEODP"):
        return sol
    # Pas de solution via parent → peut-être un ticket CA lié à une épic PSC
    # (rattachement par lien Relates, pas par parent). Les PSC = LITTERALIS.
    iss = issue_map.get(ticket_key)
    if iss is not None and find_linked_epic(iss):
        return "LITTERALIS"
    return sol

def _phase_de_pct(pct):
    """Mappe un % d'avancement (0-100) vers l'une des 4 phases du modèle."""
    if pct is None:
        return "Lancement & prérequis"
    if pct < 10:   return "Lancement & prérequis"   # 0→10%
    if pct < 20:   return "Mise en service"          # 10→20%
    if pct < 90:   return "Paramétrage & recette"    # 20→90%
    return "PV signé"                                 # 90→100%


def _composante_de_categorie(cat):
    """Regroupe la catégorie d'épic en 3 composantes du backlog."""
    if cat in ("Rework",):
        return "Rework"
    if cat in ("Prestation offerte", "Commande sans prestation", "Maintenance", "Assistance Expert"):
        return "Prestations gratuites"
    return "Facturable"


backlog_rows = []
for k, info in _cache["issues_by_key"].items():
    if info.get("issuetype") != "COORDIN : Suivi CA":
        continue
    raw = info["raw_fields"]
    statut = (raw.get("status") or {}).get("name", "")
    if statut in ("Terminé", "Terminée", "Done", "Clôturé", "Fermé"):
        continue  # backlog = non clôturé
    montant = float(raw.get("customfield_22042") or 0)
    if montant == 0:
        continue
    blocage_raw = raw.get("customfield_24269")
    if isinstance(blocage_raw, dict):
        blocage = blocage_raw.get("value") or blocage_raw.get("name") or ""
    else:
        blocage = blocage_raw or ""
    if blocage.lower() in ("aucun", "none", ""):
        blocage = "Aucun"
    # "Blocage projet" est comptabilisé avec "Blocage client"
    if blocage == "Blocage projet":
        blocage = "Blocage client"
    sol = _solution_de_ticket(k) or "Autre"
    # Catégorie (→ composante facturable / rework / gratuit)
    cat, _, epic_key = _categorie_et_solution(k)
    composante = _composante_de_categorie(cat)
    # Jalon % (0-100) → phase du modèle
    jalon_pct_raw = raw.get("customfield_21883")
    try:
        jalon_pct = float(jalon_pct_raw) if jalon_pct_raw not in (None, "") else None
    except (ValueError, TypeError):
        jalon_pct = None
    phase = _phase_de_pct(jalon_pct)
    # Temps restant (somme sous-tâches, champ 10384)
    temps_restant = 0.0
    tr_raw = raw.get("customfield_10384")
    if tr_raw not in (None, ""):
        try:
            temps_restant = float(tr_raw) / 3600  # secondes → heures
        except (ValueError, TypeError):
            temps_restant = 0.0
    # Ancienneté du BDC = création de l'ÉPIC PARENT → aujourd'hui (en mois).
    if not epic_key:
        epic_key = remonter_vers_epic(k)
    epic_info = _cache["issues_by_key"].get(epic_key, {})
    epic_created = epic_info.get("raw_fields", {}).get("created")
    anc_mois = None
    if epic_created:
        try:
            cd = pd.to_datetime(epic_created).date()
            anc_mois = (date.today().year - cd.year) * 12 + (date.today().month - cd.month)
        except Exception:
            anc_mois = None
    # Date de déblocage estimée (champ 26845) → tranche relative
    deb_raw = raw.get("customfield_26845")
    deblocage_tranche = None   # None = pas bloqué ou sans date
    if blocage != "Aucun":
        deblocage_tranche = "Sans date"   # bloqué mais pas de date par défaut
        if deb_raw:
            try:
                dd = pd.to_datetime(deb_raw).date()
                delta_j = (dd - date.today()).days
                if delta_j <= 31:
                    deblocage_tranche = "Ce mois"
                elif delta_j <= 92:
                    deblocage_tranche = "1-3 mois"
                else:
                    deblocage_tranche = "Au-delà"
            except Exception:
                deblocage_tranche = "Sans date"
    backlog_rows.append({"ticket": k, "solution": sol,
                         "montant": montant, "blocage": blocage,
                         "composante": composante, "phase": phase,
                         "jalon_pct": jalon_pct, "temps_restant": temps_restant,
                         "anciennete_mois": anc_mois,
                         "deblocage_tranche": deblocage_tranche})

df_bl = pd.DataFrame(backlog_rows)


def _backlog_facturable_analyse(df):
    """Analyse du backlog facturable, par solution.
       Retourne, pour LITTERALIS et GEODP :
       - mobilisable / bloqué (montant + temps)
       - répartition des montants par phase (4 phases du modèle)
       - répartition par type de blocage."""
    phases = ["Lancement & prérequis", "Mise en service",
              "Paramétrage & recette", "PV signé"]
    blocages = ["Aucun", "Blocage client", "Blocage produit",
                "Blocage commerce", "Autre"]
    tranches_deb = ["Ce mois", "1-3 mois", "Au-delà", "Sans date"]

    def _vide():
        return {
            "total": 0.0, "mobilisable": 0.0, "bloque": 0.0,
            "temps_total": 0.0, "temps_mobilisable": 0.0, "temps_bloque": 0.0,
            "nb": 0, "nb_mobilisable": 0, "nb_bloque": 0,
            "par_phase": {p: 0.0 for p in phases},
            "nb_par_phase": {p: 0 for p in phases},
            # bloqué/mobilisable par phase (montant)
            "phase_mobilisable": {p: 0.0 for p in phases},
            "phase_bloque": {p: 0.0 for p in phases},
            "par_blocage": {b: 0.0 for b in blocages},
            "nb_par_blocage": {b: 0 for b in blocages},
            # échéancier de déblocage (montant + nb) du backlog bloqué
            "deblocage_montant": {t: 0.0 for t in tranches_deb},
            "deblocage_nb": {t: 0 for t in tranches_deb},
        }

    res = {"LITTERALIS": _vide(), "GEODP": _vide()}
    if df.empty:
        return res

    # Seul le facturable a un montant > 0 ; rework/gratuit traités plus tard.
    fact = df[df["composante"] == "Facturable"]
    for _, r in fact.iterrows():
        sol = r["solution"]
        if sol not in res:
            continue
        d = res[sol]
        m = float(r["montant"])
        t = float(r.get("temps_restant") or 0)
        bloque = r["blocage"] != "Aucun"
        d["total"] += m
        d["temps_total"] += t
        d["nb"] += 1
        if bloque:
            d["bloque"] += m; d["temps_bloque"] += t; d["nb_bloque"] += 1
        else:
            d["mobilisable"] += m; d["temps_mobilisable"] += t; d["nb_mobilisable"] += 1
        # par phase (+ split bloqué/mobilisable)
        ph = r["phase"]
        if ph in d["par_phase"]:
            d["par_phase"][ph] += m
            d["nb_par_phase"][ph] += 1
            if bloque:
                d["phase_bloque"][ph] += m
            else:
                d["phase_mobilisable"][ph] += m
        # par blocage
        bl = r["blocage"] if r["blocage"] in blocages else "Autre"
        d["par_blocage"][bl] += m
        d["nb_par_blocage"][bl] += 1
        # échéancier de déblocage (uniquement le bloqué)
        if bloque:
            tr = r.get("deblocage_tranche") or "Sans date"
            if tr in d["deblocage_montant"]:
                d["deblocage_montant"][tr] += m
                d["deblocage_nb"][tr] += 1
    return res


def _bucket_anc(m):
    """Range une ancienneté (en mois) dans une tranche."""
    if m is None:
        return "Inconnu"
    if m <= 0:  return "M+0"
    if m == 1:  return "M+1"
    if m == 2:  return "M+2"
    if m == 3:  return "M+3"
    if m <= 6:  return "M+4→6"
    if m <= 12: return "M+7→12"
    return "M+13+"


def _backlog_par_tranche(df):
    """Répartit le backlog par tranche d'ancienneté × solution.
       Retourne {solution: {tranche: montant}}."""
    tranches = ["M+0", "M+1", "M+2", "M+3", "M+4→6", "M+7→12", "M+13+"]
    res = {"LITTERALIS": {t: 0.0 for t in tranches},
           "GEODP":      {t: 0.0 for t in tranches}}
    if df.empty:
        return res
    for _, r in df.iterrows():
        sol = r["solution"]
        if sol not in res:
            continue
        b = _bucket_anc(r.get("anciennete_mois"))
        if b == "Inconnu":
            continue
        res[sol][b] += float(r["montant"])
    return res

ws_bl = wb.create_sheet("Backlog")
write_sheet_title(ws_bl, f"Backlog à date — {date.today().strftime('%d/%m/%Y')}",
                  "Tickets Suivi CA non clôturés — montant restant à reconnaître")

# Types de blocage rencontrés (hors Aucun)
TYPES_BLOCAGE = ["Blocage client", "Blocage commerce", "Blocage produit"]

def _bilan(df_sub):
    total = df_sub["montant"].sum() if not df_sub.empty else 0
    par_type = {}
    for t in TYPES_BLOCAGE:
        par_type[t] = df_sub[df_sub["blocage"] == t]["montant"].sum() if not df_sub.empty else 0
    # blocages éventuels hors liste connue
    autres_bloc = df_sub[~df_sub["blocage"].isin(TYPES_BLOCAGE + ["Aucun"])]["montant"].sum() if not df_sub.empty else 0
    total_bloque = sum(par_type.values()) + autres_bloc
    mobilisable = total - total_bloque
    return total, mobilisable, total_bloque, par_type, autres_bloc

headers_bl = ["Périmètre", "Backlog total", "Mobilisable", "Bloqué",
              "dont Blocage client", "dont Blocage commerce", "dont Blocage produit"]
bl_hdr = 4
ws_bl.append(headers_bl)
style_header_row(ws_bl, bl_hdr, len(headers_bl))

lignes_bl = [("TOTAL", df_bl),
             ("Littéralis", df_bl[df_bl["solution"] == "LITTERALIS"] if not df_bl.empty else df_bl),
             ("GEODP", df_bl[df_bl["solution"] == "GEODP"] if not df_bl.empty else df_bl)]

r = bl_hdr
for i, (label, sub) in enumerate(lignes_bl):
    total, mobilisable, bloque, par_type, _autres = _bilan(sub)
    if _autres > 0:
        print(f"   ⚠ Backlog {label}: {_autres:,.0f}€ de blocage non catégorisé "
              f"(valeur de blocage hors client/commerce/produit) — à vérifier")
    r += 1
    ws_bl.append([label, round(total, 0), round(mobilisable, 0), round(bloque, 0),
                  round(par_type["Blocage client"], 0),
                  round(par_type["Blocage commerce"], 0),
                  round(par_type["Blocage produit"], 0)])
    style_row(ws_bl, r, len(headers_bl), total=(label == "TOTAL"), even=(i % 2 == 0))
    # format montant €
    for c in range(2, len(headers_bl) + 1):
        ws_bl.cell(r, c).number_format = "#,##0 €"

autofit(ws_bl)

# ══════════════════════════════════════════════════════════════════
# ONGLET 6 : Commandes reçues ce mois
# ══════════════════════════════════════════════════════════════════
# Une commande = ticket "COORDIN : Suivi CA" dont l'ÉPIC PARENT a été créée
# ce mois. Montant = customfield_20701 (montant total commande). Hardware exclu.
def _epic_de_ticket_ca(ticket_key):
    """Épic parente d'un ticket CA : via parent (PDMDEP) ou lien Relates (PSC)."""
    info = _cache["issues_by_key"].get(ticket_key)
    if not info:
        return None
    pk = info.get("parent_key")
    if pk and pk in _cache["issues_by_key"]:
        return pk
    iss = issue_map.get(ticket_key)
    if iss is not None:
        lk = find_linked_epic(iss)
        if lk:
            return lk
    return None

cmd_rows = []
for k, info in _cache["issues_by_key"].items():
    if info.get("issuetype") != "COORDIN : Suivi CA":
        continue
    raw = info["raw_fields"]

    # Exclure Hardware (prestation)
    prest = _read_prestation(raw)
    if prest.strip().lower() == "hardware":
        continue

    # Épic parente + sa date de création
    epic_key = _epic_de_ticket_ca(k)
    if not epic_key:
        continue
    epic_info = _cache["issues_by_key"].get(epic_key, {})
    created_raw = epic_info.get("raw_fields", {}).get("created")
    if not created_raw:
        continue
    try:
        created_dt = pd.to_datetime(created_raw).date()
    except Exception:
        continue
    # Filtre : épic créée dans le mois courant
    if not (month_start_date <= created_dt < month_end_date):
        continue

    # Montant total de la commande (customfield_20701)
    iss = issue_map.get(k)
    montant = float(getattr(iss.fields, "customfield_20701", 0) or 0) if iss else 0

    # Solution (PDMDEP via parent, PSC → LITTERALIS)
    sol = _solution_de_ticket(k) or "Autre"
    client = _read_client(raw)
    bdc = _read_bdc(iss.fields, raw) if iss else ""
    deal = _read_deal_type(epic_info.get("raw_fields", {}))

    cmd_rows.append({
        "Ticket CA": k, "Epic": epic_key, "Date commande": created_dt.strftime("%d/%m/%Y"),
        "Client": client, "Solution": sol, "Type de deal": deal,
        "Prestation": prest, "Numéro BDC": bdc, "Montant (€)": montant,
    })

df_cmd = pd.DataFrame(cmd_rows)

ws_cmd = wb.create_sheet("Commandes du mois")
write_sheet_title(ws_cmd, f"Commandes reçues — {mois_courant_label} {month_start.year}",
                  f"Épics créées ce mois — {len(df_cmd)} commande(s)")

cmd_cols = ["Ticket CA", "Epic", "Date commande", "Client", "Solution",
            "Type de deal", "Prestation", "Numéro BDC", "Montant (€)"]
cmd_hdr = 4
ws_cmd.append(cmd_cols)
style_header_row(ws_cmd, cmd_hdr, len(cmd_cols))

r = cmd_hdr
if not df_cmd.empty:
    df_cmd = df_cmd.sort_values("Solution")
    for i, row in enumerate(df_cmd.itertuples(index=False), start=1):
        r += 1
        ws_cmd.append(list(row))
        style_row(ws_cmd, r, len(cmd_cols), even=(i % 2 == 0))
        ws_cmd.cell(r, len(cmd_cols)).number_format = '#,##0 €'

# Totaux : global puis par solution
r += 1
ws_cmd.append([])
def _ligne_total(label, sub):
    global r
    r = ws_cmd.max_row + 1
    nb = len(sub)
    mt = sub["Montant (€)"].sum() if not sub.empty else 0
    ws_cmd.cell(r, 1, label)
    ws_cmd.cell(r, 8, f"{nb} commande(s)")
    ws_cmd.cell(r, 9, round(mt, 0))
    ws_cmd.cell(r, 9).number_format = '#,##0 €'
    style_row(ws_cmd, r, len(cmd_cols), total=True)

if not df_cmd.empty:
    _ligne_total("TOTAL", df_cmd)
    _ligne_total("dont Littéralis", df_cmd[df_cmd["Solution"] == "LITTERALIS"])
    _ligne_total("dont GEODP", df_cmd[df_cmd["Solution"] == "GEODP"])

autofit(ws_cmd)

# ══════════════════════════════════════════════════════════════════
# ONGLET 7 : Projets clôturés ce mois
# ══════════════════════════════════════════════════════════════════
# Épics passées à "Terminé" ce mois (epics_clotures, via changelog).
# Rework identifié par customfield_23103 == "Rework".
clot_rows = []
for epic_key in sorted(epics_clotures):
    info = _cache["issues_by_key"].get(epic_key)
    if not info:
        continue
    raw = info["raw_fields"]
    type_epic = (raw.get("customfield_23103") or {}).get("value", "") or ""
    is_rework = (type_epic == "Rework")
    sol_raw = (raw.get("customfield_20514") or {}).get("value", "") or ""
    solution = SOLUTION_MAPPING.get(sol_raw, sol_raw or "Autre")
    client = _read_client(raw)
    # Ancienneté : création de l'épic → aujourd'hui
    created_raw = raw.get("created")
    anciennete_mois = ""
    if created_raw:
        try:
            cd = pd.to_datetime(created_raw).date()
            anciennete_mois = (date.today().year - cd.year) * 12 + (date.today().month - cd.month)
        except Exception:
            pass
    clot_rows.append({
        "Epic": epic_key, "Nom": info.get("summary", "")[:60],
        "Client": client, "Solution": solution,
        "Type": "Rework" if is_rework else "Projet",
        "Ancienneté (mois)": anciennete_mois,
        "_rework": is_rework,
    })

df_clot = pd.DataFrame(clot_rows)
nb_projets = int((~df_clot["_rework"]).sum()) if not df_clot.empty else 0
nb_rework  = int(df_clot["_rework"].sum()) if not df_clot.empty else 0

ws_cl = wb.create_sheet("Projets clôturés")
write_sheet_title(ws_cl, f"Projets clôturés — {mois_courant_label} {month_start.year}",
                  f"{len(df_clot)} clôture(s) : {nb_projets} projet(s), {nb_rework} rework")

CL_COLS = ["Epic", "Nom", "Client", "Solution", "Type", "Ancienneté (mois)"]
cl_hdr = 4
ws_cl.append(CL_COLS)
style_header_row(ws_cl, cl_hdr, len(CL_COLS))

REWORK_FILL = "FCE4D6"  # orangé léger pour les lignes Rework
r = cl_hdr
if not df_clot.empty:
    df_clot_sorted = df_clot.sort_values(["_rework", "Solution"])
    for i, (_, row) in enumerate(df_clot_sorted.iterrows(), start=1):
        r += 1
        ws_cl.append([row["Epic"], row["Nom"], row["Client"], row["Solution"],
                      row["Type"], row["Ancienneté (mois)"]])
        style_row(ws_cl, r, len(CL_COLS), even=(i % 2 == 0))
        if row["_rework"]:
            for c in range(1, len(CL_COLS) + 1):
                ws_cl.cell(r, c).fill = _fill(REWORK_FILL)

# Totaux
r += 2
ws_cl.cell(r, 1, "TOTAL projets clôturés")
ws_cl.cell(r, 6, nb_projets)
style_row(ws_cl, r, len(CL_COLS), total=True)
r += 1
ws_cl.cell(r, 1, "TOTAL rework clôturés")
ws_cl.cell(r, 6, nb_rework)
style_row(ws_cl, r, len(CL_COLS), total=True)
for c in range(1, len(CL_COLS) + 1):
    ws_cl.cell(r, c).fill = _fill(REWORK_FILL)

autofit(ws_cl)

# ══════════════════════════════════════════════════════════════════
# ONGLET 0 : Tableau de bord (récapitulatif, inséré en première position)
# ══════════════════════════════════════════════════════════════════
# 1) CA du mois (global + par solution) depuis df_services
if not df_services.empty:
    ca_global = df_services["Montant déclaré ce mois"].sum()
    ca_litt   = df_services[df_services["Solution"] == "LITTERALIS"]["Montant déclaré ce mois"].sum()
    ca_geodp  = df_services[df_services["Solution"] == "GEODP"]["Montant déclaré ce mois"].sum()
else:
    ca_global = ca_litt = ca_geodp = 0

# 2) Objectifs du mois (dernière estimation : LE2 > LE1 > Budget)
obj_global, obj_litt, obj_geodp = load_objectifs(OBJECTIFS_FILE, mois_courant_label)

# 3) Charge à date au prorata des jours ouvrés écoulés
today_d = date.today()
jours_ecoules = len([d for d in working_days if d.date() <= today_d])
ratio_temps = (jours_ecoules / num_working_days) if num_working_days else 0
cap_mensuelle_totale = sum(c[3] for c in capacites)          # capacité dispo du mois
cap_attendue_a_date  = cap_mensuelle_totale * ratio_temps     # ce qu'on devrait avoir saisi
charge_reelle = df_wl["Temps (h)"].sum() if not df_wl.empty else 0
pct_charge = round(charge_reelle / cap_attendue_a_date * 100, 1) if cap_attendue_a_date > 0 else 0

# 4) Projets / Rework ouverts à date (comptage local depuis le cache)
TYPES_PRODUCTIFS = {"Projet", "Prestation offerte", "Commande sans prestation", "Maintenance", "Rework", "Assistance Expert"}
STATUTS_CLOS = {"Terminé", "Terminée", "Done", "Clôturé", "Fermé", "Annulé", "Abandonné"}
nb_proj_ouverts = nb_rew_ouverts = 0
for k, info in _cache["issues_by_key"].items():
    if info.get("issuetype") != "Epic":
        continue
    if not str(k).startswith("PDMDEP-"):
        continue
    raw = info["raw_fields"]
    statut = (raw.get("status") or {}).get("name", "")
    if statut in STATUTS_CLOS:
        continue
    type_epic = (raw.get("customfield_23103") or {}).get("value", "") or ""
    if type_epic not in TYPES_PRODUCTIFS:
        continue
    nb_proj_ouverts += 1
    if type_epic == "Rework":
        nb_rew_ouverts += 1

# ── Construction de l'onglet (design : bandeaux + KPI fond bleu clair) ──
ws_db = wb.create_sheet("Tableau de bord", 0)  # position 0
write_sheet_title(ws_db, f"Tableau de bord — {mois_courant_label} {month_start.year}",
                  f"Synthèse au {today_d.strftime('%d/%m/%Y')} "
                  f"({jours_ecoules}/{num_working_days} jours ouvrés écoulés)")

RED_C   = "C00000"
GREEN_C = "375623"

def _section(row, label):
    """Bandeau de section bleu foncé sur toute la largeur."""
    for c in range(1, 10):
        cell = ws_db.cell(row, c)
        cell.fill = _fill(BLUE_DARK)
    sc = ws_db.cell(row, 1, label)
    sc.font = _font(bold=True, size=12, color="FFFFFF")
    sc.alignment = _align(h="left")
    ws_db.row_dimensions[row].height = 22

def _kpi(row, col, label, value, color=None):
    """KPI : libellé en haut, gros chiffre dessous sur fond bleu clair (3 colonnes)."""
    lc = ws_db.cell(row, col, label)
    lc.font = _font(bold=False, size=10, color="404040")
    vc = ws_db.cell(row + 1, col, value)
    vc.font = _font(bold=True, size=18, color=color or BLUE_DARK)
    vc.alignment = _align(h="left")
    # fond bleu clair sur les 3 colonnes du bloc
    for c in range(col, col + 3):
        ws_db.cell(row + 1, c).fill = _fill(BLUE_LIGHT)
    ws_db.row_dimensions[row + 1].height = 26

def _pct_av(real, obj):
    return round(real / obj * 100, 1) if obj > 0 else 0

# ═══ Section Chiffre d'affaires ═══
_section(5, "  📊  Chiffre d'affaires")
_kpi(6, 1, "CA réalisé (global)", f"{ca_global:,.0f} €")
_kpi(6, 4, "Objectif du mois",    f"{obj_global:,.0f} €")
pct_g = _pct_av(ca_global, obj_global)
_kpi(6, 7, "% Atteinte objectif", f"{pct_g:.1f}%",
     color=GREEN_C if pct_g >= 100 else RED_C)
_kpi(10, 1, "CA Littéralis", f"{ca_litt:,.0f} €")
_kpi(10, 4, "CA GEODP",      f"{ca_geodp:,.0f} €")
_kpi(10, 7, "Objectifs Litt / GEODP", f"{obj_litt:,.0f} / {obj_geodp:,.0f} €")

# ═══ Section Charge ═══
_section(14, "  ⏱️  Charge & capacité")
_kpi(15, 1, "Charge à date", f"{charge_reelle:.0f} h")
_kpi(15, 4, "Attendu à date", f"{cap_attendue_a_date:.0f} h")
charge_color = RED_C if (pct_charge > 100 or pct_charge < 70) else GREEN_C
_kpi(15, 7, "% Complétion (prorata)", f"{pct_charge:.1f}%", color=charge_color)
_kpi(19, 1, "Capacité du mois", f"{cap_mensuelle_totale:.0f} h")

# ═══ Section Activité projets ═══
_section(23, "  📂  Activité projets")
_kpi(24, 1, "Projets clôturés", str(nb_projets))
_kpi(24, 4, "dont Rework clôturés", str(nb_rework),
     color=RED_C if nb_rework else BLUE_DARK)
_kpi(24, 7, "Projets ouverts (à date)", str(nb_proj_ouverts))
_kpi(28, 1, "Rework ouverts (à date)", str(nb_rew_ouverts),
     color=RED_C if nb_rew_ouverts else BLUE_DARK)

for c in range(1, 10):
    ws_db.column_dimensions[get_column_letter(c)].width = 17

output = f"jira_export_{date.today().strftime('%Y-%m-%d')}.xlsx"
wb.save(output)
print(f"\n✅ {output}")

# Petit résumé pour comparaison
if not df_services.empty:
    ca_total = df_services["Montant déclaré ce mois"].sum()
    h_total  = df_services["Temps mois (h)"].sum()
    print(f"   Suivi prod  : {len(df_services)} lignes | {ca_total:,.0f} € | {h_total:.1f} h")
if not df_hardware.empty:
    print(f"   Hardware    : {len(df_hardware)} lignes | {df_hardware['Montant déclaré ce mois'].sum():,.0f} €")

# ══════════════════════════════════════════════════════════════════
# ACCUMULATION DANS L'HISTORIQUE (option A : écrase le mois courant)
# ══════════════════════════════════════════════════════════════════
try:
    from maj_historique import maj_historique

    # Backlog par solution
    def _bl_sol(sol):
        if df_bl.empty:
            return 0, 0
        s = df_bl[df_bl["solution"] == sol]
        tot = s["montant"].sum()
        blo = s[s["blocage"] != "Aucun"]["montant"].sum()
        return tot, blo

    bl_litt_tot, bl_litt_blo = _bl_sol("LITTERALIS")
    bl_geodp_tot, bl_geodp_blo = _bl_sol("GEODP")

    # CA par type de deal × solution (onglet ca_deal)
    ca_deal_lignes = []
    if not df_services.empty:
        grp = df_services.groupby(["Type de deal", "Solution"])["Montant déclaré ce mois"].sum()
        for (deal, sol), montant in grp.items():
            if montant != 0:
                ca_deal_lignes.append({
                    "Mois": mois_courant_label, "Année": month_start.year,
                    "Type de deal": deal or "(non renseigné)", "Solution": sol,
                    "CA (€)": round(float(montant), 2),
                })

    # Ancienneté (depuis le Suivi de production : une ligne par épic AVEC CA)
    anciennete_lignes = []
    if not df_services.empty:
        # CA total par épic (pour ne garder que celles qui ont du CA)
        ca_par_epic = df_services.groupby("Epic")["Montant déclaré ce mois"].sum()
        vus = set()
        for _, r in df_services.iterrows():
            ep = r["Epic"]
            if ep in vus:
                continue
            vus.add(ep)
            ca_epic = float(ca_par_epic.get(ep, 0) or 0)
            if ca_epic == 0:
                continue  # on ne garde que les épics avec du CA
            anciennete_lignes.append({
                "Mois": mois_courant_label, "Année": month_start.year,
                "Epic": ep, "Nom": r.get("Epic Nom", ""),
                "Solution": r.get("Solution", ""), "Catégorie": r.get("Catégorie", ""),
                "Type de deal": r.get("Type de deal", ""),
                "Date commande": r.get("Date commande", ""),
                "Ancienneté": r.get("Ancienneté (mois)", ""),
                "CA": round(ca_epic, 2),
            })

    # Commandes (liste, format historique)
    commandes_lignes = []
    if not df_cmd.empty:
        for _, r in df_cmd.iterrows():
            epic_k = r.get("Epic", "")
            # statut + résumé de l'épic depuis le cache
            epic_info = _cache["issues_by_key"].get(epic_k, {})
            epic_raw = epic_info.get("raw_fields", {})
            statut_epic = (epic_raw.get("status") or {}).get("name", "") if isinstance(epic_raw.get("status"), dict) else ""
            resume_epic = epic_info.get("summary", "")
            commandes_lignes.append({
                "Mois": mois_courant_label, "Année": month_start.year,
                "Clé Epic": epic_k, "Numéro BDC": r.get("Numéro BDC", ""),
                "Client": r.get("Client", ""), "Nom / Résumé": resume_epic,
                "Solution": r.get("Solution", ""), "Statut": statut_epic,
                "Type de deal": r.get("Type de deal", ""),
                "Date de création": r.get("Date commande", ""),
                "Avancement projet": "", "Champs personnalisés (Temps passé (h))": "",
                "Montant total prestations (€)": round(float(r.get("Montant (€)", 0) or 0), 2),
                "CA reconnu ce mois (€)": "",
            })

    # Analyse du temps non valorisé (toutes les fonctions dépendantes sont
    # maintenant définies : _solution_de_ticket, remonter_vers_epic...).
    analyse_nv = _analyse_temps_non_valorise(df_wl)

    contexte = {
        "mois_label": mois_courant_label, "annee": month_start.year,
        "ca_global": float(ca_global), "ca_litt": float(ca_litt), "ca_geodp": float(ca_geodp),
        "obj_global": float(obj_global), "obj_litt": float(obj_litt), "obj_geodp": float(obj_geodp),
        "charge": {
            "prod_litt": _h(df_wl, "LITTERALIS", CAT_PROJET) + _h(df_wl, "LITTERALIS", CAT_REWORK) + _h(df_wl, "LITTERALIS", CAT_GRATUIT),
            "prod_geodp": _h(df_wl, "GEODP", CAT_PROJET) + _h(df_wl, "GEODP", CAT_REWORK) + _h(df_wl, "GEODP", CAT_GRATUIT),
            "sup_litt": _h(df_wl, "LITTERALIS", CAT_SUPPORT), "sup_geodp": _h(df_wl, "GEODP", CAT_SUPPORT),
            "rew_litt": _h(df_wl, "LITTERALIS", CAT_REWORK), "rew_geodp": _h(df_wl, "GEODP", CAT_REWORK),
            "proj_litt": _h(df_wl, "LITTERALIS", CAT_PROJET), "proj_geodp": _h(df_wl, "GEODP", CAT_PROJET),
            "gratuit_litt": _h(df_wl, "LITTERALIS", CAT_GRATUIT), "gratuit_geodp": _h(df_wl, "GEODP", CAT_GRATUIT),
            "interne": df_wl[df_wl["Catégorie"] == "Interne"]["Temps (h)"].sum() if not df_wl.empty else 0,
        },
        "backlog": {
            "total": df_bl["montant"].sum() if not df_bl.empty else 0,
            "total_litt": bl_litt_tot, "total_geodp": bl_geodp_tot,
            "bloque": df_bl[df_bl["blocage"] != "Aucun"]["montant"].sum() if not df_bl.empty else 0,
            "mobilisable": df_bl[df_bl["blocage"] == "Aucun"]["montant"].sum() if not df_bl.empty else 0,
            "bloque_litt": bl_litt_blo, "mob_litt": bl_litt_tot - bl_litt_blo,
            "bloque_geodp": bl_geodp_blo, "mob_geodp": bl_geodp_tot - bl_geodp_blo,
            "nb_projets": nb_proj_ouverts, "nb_rework": nb_rew_ouverts,
        },
        "backlog_anciennete": _backlog_par_tranche(df_bl),
        "backlog_facturable": _backlog_facturable_analyse(df_bl),
        "temps_non_valorise": analyse_nv,
        "commandes_lignes": commandes_lignes,
        "ca_deal_lignes": ca_deal_lignes,
        "anciennete_lignes": anciennete_lignes,
        "clotures": {
            "nb_projets_clotures": nb_projets,   # flux : clôturés CE MOIS
            "nb_rework_clotures": nb_rework,
        },
        "heures": {
            "heures_saisies": float(charge_reelle),
            "capacite_attendue": float(cap_attendue_a_date),
            "capacite_totale": float(cap_mensuelle_totale),
        },
    }
    maj_historique(contexte)
except Exception as e:
    import traceback
    print(f"   ⚠ Accumulation historique échouée : {e}")
    print("   ── Trace complète ──")
    traceback.print_exc()
    print("   ────────────────────")
