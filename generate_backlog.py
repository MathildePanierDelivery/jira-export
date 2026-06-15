"""
generate_backlog.py
-------------------
Lit le dernier fichier Backlog_projets_by_CA_*.xlsx et génère backlog.html
"""
import glob, os, sys
from datetime import datetime
from openpyxl import load_workbook
import design_system as ds

# ── Fichier source ────────────────────────────────────────────────
def find_file(path=None):
    if path and os.path.exists(path): return path
    files = sorted(glob.glob("Backlog_projets_by_CA_*.xlsx"), reverse=True)
    if not files:
        print("❌ Aucun fichier Backlog_projets_by_CA_*.xlsx trouvé"); sys.exit(1)
    print(f"📂 {files[0]}"); return files[0]

# ── Lecture des données ───────────────────────────────────────────
def load_data(fp):
    wb = load_workbook(fp, read_only=True, data_only=True)
    data = {}
    for scope in ["TOTAL", "LITTERALIS", "GEODP"]:
        sheet = next((s for s in wb.sheetnames if s.startswith(scope)), None)
        if not sheet: continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        d = {}
        for i, row in enumerate(rows):
            vals = [v for v in row if v is not None]
            if not vals: continue
            label = str(vals[0]).strip()
            # Structure : label en-tête sur une ligne, valeurs sur la suivante
            if label == "Total backlog" and i+1 < len(rows):
                nxt = [v for v in rows[i+1] if v is not None]
                if nxt and isinstance(nxt[0], (int, float)):
                    d["total"]       = float(nxt[0] or 0)
                    d["bloque"]      = float(nxt[1] or 0) if len(nxt) > 1 else 0
                    d["mobilisable"] = float(nxt[2] or 0) if len(nxt) > 2 else 0
            elif label == "Blocage client" and i+1 < len(rows) and "total" in d and "blocages" not in d:
                nxt = [v for v in rows[i+1] if v is not None]
                if nxt and isinstance(nxt[0], (int, float)):
                    d["blocages"] = {
                        "client":   float(nxt[0] or 0),
                        "commerce": float(nxt[1] or 0) if len(nxt) > 1 else 0,
                        "produit":  float(nxt[2] or 0) if len(nxt) > 2 else 0,
                    }
            elif label == "Total BDC ouverts" and i+1 < len(rows):
                nxt = [v for v in rows[i+1] if v is not None]
                if nxt and isinstance(nxt[0], (int, float)):
                    d["nb_bdc"]     = int(nxt[0] or 0)
                    d["nb_bloques"] = int(nxt[1] or 0) if len(nxt) > 1 else 0
        data[scope] = d

    # Lecture onglet DETAIL pour le tableau des projets bloqués
    detail_sheet = next((s for s in wb.sheetnames if s.startswith("DETAIL")), None)
    projets_bloques = []
    if detail_sheet:
        ws_d = wb[detail_sheet]
        rows_d = list(ws_d.iter_rows(values_only=True))
        # Trouver l'en-tête
        hdr = None
        for i, row in enumerate(rows_d):
            if row and "BDC / Order" in str(row[0] or ""):
                hdr = {str(v).strip(): j for j, v in enumerate(row) if v}
                start = i + 1; break
        if hdr:
            for row in rows_d[start:]:
                if not row or not row[0]: continue
                blocage = str(row[hdr.get("Blocage", 18)] or "").strip()
                if blocage and blocage not in ("Aucun", "NC", ""):
                    projets_bloques.append({
                        "bdc":      str(row[hdr.get("BDC / Order", 0)] or "").strip(),
                        "client":   str(row[hdr.get("Client", 5)] or "").strip(),
                        "solution": str(row[hdr.get("Solution", 4)] or "").strip(),
                        "cp":       str(row[hdr.get("Chef de projet", 6)] or "").strip(),
                        "montant":  float(row[hdr.get("Montant restant", 8)] or 0),
                        "blocage":  blocage,
                        "detail":   str(row[hdr.get("Détail blocage", 19)] or "").strip(),
                        "date_deb": str(row[hdr.get("Date de déblocage estimée", 20)] or "").strip(),
                    })
    projets_bloques.sort(key=lambda x: x["montant"], reverse=True)
    wb.close()
    return data, projets_bloques

# ── Génération HTML ───────────────────────────────────────────────
def kpi_section(d, label):
    total      = d.get("total", 0)
    bloque     = d.get("bloque", 0)
    mobilisable = d.get("mobilisable", 0)
    nb_bdc     = d.get("nb_bdc", 0)
    nb_bloq    = d.get("nb_bloques", 0)
    pct_bloq   = (bloque / total * 100) if total else 0
    pct_mob    = (mobilisable / total * 100) if total else 0

    blocages   = d.get("blocages", {})
    bl_client  = blocages.get("client", 0)
    bl_comm    = blocages.get("commerce", 0)
    bl_prod    = blocages.get("produit", 0)
    bl_autre   = max(0, bloque - bl_client - bl_comm - bl_prod)

    bar_items = [
        ("Mobilisable", pct_mob,  mobilisable, "#3B6D11"),
        ("Bloqué",      pct_bloq, bloque,      "#A32D2D"),
    ]
    bars = "".join(f"""
        <div class="bar-row">
          <div class="bar-label">{lbl}</div>
          <div class="bar-track"><div class="bar-fill" style="width:{min(pct,100):.1f}%;background:{col}"></div></div>
          <div class="bar-value">{ds.fmt_eur(val)}</div>
        </div>""" for lbl, pct, val, col in bar_items)

    blocage_rows = ""
    for lbl, val, css in [
        ("Blocage client",   bl_client,  "pill-red"),
        ("Blocage commerce", bl_comm,    "pill-amber"),
        ("Blocage produit",  bl_prod,    "pill-amber"),
        ("Autre / divers",   bl_autre,   "pill-blue"),
    ]:
        if val > 0:
            pct_b = (val / bloque * 100) if bloque else 0
            blocage_rows += f"""
            <tr>
              <td><span class="pill {css}">{lbl}</span></td>
              <td class="r">{ds.fmt_eur(val)}</td>
              <td class="r">{pct_b:.1f} %</td>
            </tr>"""

    return f"""
    <div class="kpi-row">
      <div class="kpi">
        <div class="kpi-label">Backlog total</div>
        <div class="kpi-value">{ds.fmt_eur(total)}</div>
        <div class="kpi-trend flat">{nb_bdc} BDC ouverts</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">Dont bloqué</div>
        <div class="kpi-value red">{ds.fmt_eur(bloque)}</div>
        <div class="kpi-trend flat">{pct_bloq:.1f}% du backlog · {nb_bloq} BDC</div>
      </div>
      <div class="kpi">
        <div class="kpi-label">Mobilisable</div>
        <div class="kpi-value green">{ds.fmt_eur(mobilisable)}</div>
        <div class="kpi-trend flat">{pct_mob:.1f}% du backlog</div>
      </div>
    </div>

    <div class="section-label">Répartition</div>
    <div class="card">
      <div class="card-body">{bars}</div>
    </div>

    <div class="section-label">Détail des blocages par motif</div>
    <div class="card">
      <div class="card-head">
        <span class="card-title">Montants bloqués</span>
        <span class="card-sub">{ds.fmt_eur(bloque)} total</span>
      </div>
      <div class="card-body" style="padding:0">
        <table class="data-table">
          <thead><tr><th>Motif</th><th class="r">Montant</th><th class="r">% du bloqué</th></tr></thead>
          <tbody>{blocage_rows}</tbody>
        </table>
      </div>
    </div>"""

def generate_html(data, projets_bloques, today_str):
    scopes = [("TOTAL","Global"), ("LITTERALIS","Littéralis"), ("GEODP","GEODP")]
    tabs_js = []
    panels = []
    for i, (key, label) in enumerate(scopes):
        d = data.get(key, {})
        active = "active" if i == 0 else ""
        tabs_js.append(f'"{key}"')
        panels.append(f'<div id="panel-{key}" class="scope-panel" style="display:{"block" if i==0 else "none"}">{kpi_section(d, label)}</div>')

    # Tableau projets bloqués
    rows_bloq = ""
    for p in projets_bloques[:30]:
        sol_css = "pill-litt" if "LITT" in p["solution"].upper() else "pill-geodp"
        rows_bloq += f"""
        <tr>
          <td><span class="pill {sol_css}">{p['solution']}</span></td>
          <td>{p['client'][:40]}</td>
          <td>{p['cp'].split()[0] if p['cp'] else '—'}</td>
          <td class="r">{ds.fmt_eur(p['montant'])}</td>
          <td><span class="pill pill-red">{p['blocage']}</span></td>
          <td style="color:var(--text-2);max-width:200px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis">{p['detail'][:60]}</td>
          <td style="font-family:'DM Mono',monospace;font-size:11px;color:var(--text-3)">{p['date_deb'] or '—'}</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Backlog projets — PS France</title>
{ds.DESIGN_CSS}
</head>
<body>
{ds.nav("backlog", today_str)}
<div class="page">
  <div class="page-header">
    <div class="page-title">Backlog projets</div>
    <div class="page-sub">Vision à date — données extraites depuis Jira</div>
  </div>

  <div class="scope-wrap">
    <div class="scope-tabs">
      <button class="scope-tab active" onclick="setScope('TOTAL',this)">Global</button>
      <button class="scope-tab" onclick="setScope('LITTERALIS',this)">Littéralis</button>
      <button class="scope-tab" onclick="setScope('GEODP',this)">GEODP</button>
    </div>
  </div>

  {"".join(panels)}

  <div class="section-label" style="margin-top:8px">Détail des projets bloqués</div>
  <div class="card">
    <div class="card-head">
      <span class="card-title">Projets bloqués — top {min(30,len(projets_bloques))} par montant</span>
      <span class="card-sub">{len(projets_bloques)} projets bloqués au total</span>
    </div>
    <div class="card-body" style="padding:0;overflow-x:auto">
      <table class="data-table">
        <thead>
          <tr>
            <th>Solution</th><th>Client</th><th>CP</th>
            <th class="r">Montant</th><th>Blocage</th>
            <th>Détail</th><th>Déblocage prévu</th>
          </tr>
        </thead>
        <tbody>{rows_bloq}</tbody>
      </table>
    </div>
  </div>
</div>

<script>
function setScope(key, btn) {{
  document.querySelectorAll(".scope-panel").forEach(p => p.style.display="none");
  document.querySelectorAll(".scope-tab").forEach(b => b.classList.remove("active"));
  document.getElementById("panel-"+key).style.display="block";
  btn.classList.add("active");
}}
</script>
</body></html>"""

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", help="Fichier backlog xlsx")
    parser.add_argument("--output", default="backlog.html")
    args = parser.parse_args()

    fp = find_file(args.file)
    today_str = datetime.now().strftime("%d/%m/%Y")
    print("📊 Lecture des données...")
    data, projets_bloques = load_data(fp)
    print(f"   {len(projets_bloques)} projets bloqués trouvés")
    print("🖊️  Génération HTML...")
    html = generate_html(data, projets_bloques, today_str)
    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅ {args.output}")
