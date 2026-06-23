"""
generate_previsionnel.py
========================
Capture le prévisionnel pondéré du mois (snapshot du 1er vendredi) à partir
de jira_cache.pkl, et l'accumule dans previsionnel_2026.xlsx.

Prévisionnel pondéré d'un ticket = Prévision du mois (23610) × taux(confiance 23595)
  Haute → 0.85 · Moyenne → 0.50 · Basse → 0.15 · vide/autre → 0 (non planifié)

Une ligne par ticket CA (ticket, CP, solution, prévision brute, niveau, taux,
prévision pondérée), avec sous-totaux par CP et par solution, et total général.

Le fichier est NON sensible (publiable). Un onglet par mois + un onglet récap.

Usage : python generate_previsionnel.py [--mois Juin]
"""

import os
import sys
import pickle
import argparse
import unicodedata
from datetime import date
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

CACHE_FILE = "jira_cache.pkl"
OUTPUT     = "previsionnel_2026.xlsx"

CF_PREVISION  = "customfield_23610"   # Revenu planifié / prévision du mois
CF_CONFIANCE  = "customfield_23595"   # Niveau de confiance
CF_SOLUTION   = "customfield_20514"
CF_CLIENT     = "customfield_10070"
CF_PROJET     = "customfield_23608"

MOIS_FR = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]

# En-tête / styles
BLUE_DARK = "1F3864"
BLUE_MID  = "2E75B6"
GREY      = "D9D9D9"
TOTAL_BG  = "FFF2CC"


def _norm(s):
    """minuscule + sans accents."""
    if not s:
        return ""
    return "".join(c for c in unicodedata.normalize("NFD", str(s))
                   if unicodedata.category(c) != "Mn").lower().strip()


def taux_confiance(valeur):
    """Mappe le niveau de confiance vers son taux (insensible casse/accents)."""
    n = _norm(valeur)
    if n == "haute":
        return 0.85
    if n == "moyenne":
        return 0.50
    if n == "basse":
        return 0.15
    return 0.0   # vide ou inconnu = non planifié


def _val(raw):
    if isinstance(raw, dict):
        return raw.get("value") or raw.get("name") or ""
    return raw or ""


def _solution(raw):
    s = _norm(_val(raw.get(CF_SOLUTION)))
    if "geodp" in s:
        return "GEODP"
    if "litt" in s or "liess" in s or "sherpa" in s:
        return "LITTERALIS"
    return "Autre"


def _assignee(raw):
    a = raw.get("assignee")
    if isinstance(a, dict):
        return a.get("displayName") or a.get("name") or "(non assigné)"
    return "(non assigné)"


def construire_previsionnel(cache):
    """Retourne la liste des lignes de prévisionnel (tickets CA avec prévision)."""
    issues = cache["issues_by_key"]
    lignes = []
    for key, info in issues.items():
        if info.get("issuetype") != "COORDIN : Suivi CA":
            continue
        raw = info["raw_fields"]
        prevision = raw.get(CF_PREVISION)
        try:
            prevision = float(prevision) if prevision is not None else 0.0
        except (TypeError, ValueError):
            prevision = 0.0
        if prevision == 0:
            continue  # pas de prévision → pas dans le prévisionnel

        niveau = _val(raw.get(CF_CONFIANCE)) or "Non planifié"
        taux   = taux_confiance(niveau)
        client = _val(raw.get(CF_CLIENT))
        if isinstance(raw.get(CF_CLIENT), dict):
            client = raw[CF_CLIENT].get("displayName") or raw[CF_CLIENT].get("name") or ""

        lignes.append({
            "Ticket CA": key,
            "CP": _assignee(raw),
            "Client": client,
            "Projet": raw.get(CF_PROJET) or "",
            "Solution": _solution(raw),
            "Prévision brute (€)": round(prevision, 2),
            "Niveau de confiance": niveau,
            "Taux": taux,
            "Prévision pondérée (€)": round(prevision * taux, 2),
        })
    return lignes


# ── Styles ──
def _hdr(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row, c)
        cell.fill = PatternFill("solid", fgColor=BLUE_DARK)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _ecrire_onglet_mois(wb, mois, annee, lignes):
    nom = mois
    if nom in wb.sheetnames:
        del wb[nom]
    ws = wb.create_sheet(nom)

    titre = f"Prévisionnel pondéré — {mois} {annee}"
    ws.cell(1, 1, titre).font = Font(bold=True, size=13, color=BLUE_DARK)
    ws.cell(2, 1, f"Capturé le {date.today().strftime('%d/%m/%Y')}").font = Font(italic=True, size=9, color="808080")

    cols = ["Ticket CA", "CP", "Client", "Projet", "Solution",
            "Prévision brute (€)", "Niveau de confiance", "Taux", "Prévision pondérée (€)"]
    hdr_row = 4
    for c, col in enumerate(cols, 1):
        ws.cell(hdr_row, c, col)
    _hdr(ws, hdr_row, len(cols))

    # tri par CP puis solution
    lignes_tri = sorted(lignes, key=lambda x: (x["CP"], x["Solution"], -x["Prévision pondérée (€)"]))
    r = hdr_row
    for ln in lignes_tri:
        r += 1
        for c, col in enumerate(cols, 1):
            v = ln[col]
            cell = ws.cell(r, c, v)
            if "€" in col:
                cell.number_format = '#,##0 "€"'
            if col == "Taux":
                cell.number_format = '0%'

    # Total général
    r += 2
    ws.cell(r, 5, "TOTAL GÉNÉRAL").font = Font(bold=True)
    tot_brut = sum(l["Prévision brute (€)"] for l in lignes)
    tot_pond = sum(l["Prévision pondérée (€)"] for l in lignes)
    ws.cell(r, 6, round(tot_brut, 2)).number_format = '#,##0 "€"'
    ws.cell(r, 9, round(tot_pond, 2)).number_format = '#,##0 "€"'
    for c in range(5, 10):
        ws.cell(r, c).fill = PatternFill("solid", fgColor=TOTAL_BG)
        ws.cell(r, c).font = Font(bold=True)

    # Sous-totaux par CP
    r += 2
    ws.cell(r, 1, "Sous-totaux par CP").font = Font(bold=True, color=BLUE_MID, size=11)
    r += 1
    for col, lib in [(1, "CP"), (2, "Prévision brute (€)"), (3, "Prévision pondérée (€)")]:
        ws.cell(r, col, lib)
    _hdr(ws, r, 3)
    par_cp = defaultdict(lambda: [0.0, 0.0])
    for l in lignes:
        par_cp[l["CP"]][0] += l["Prévision brute (€)"]
        par_cp[l["CP"]][1] += l["Prévision pondérée (€)"]
    for cp, (brut, pond) in sorted(par_cp.items(), key=lambda x: -x[1][1]):
        r += 1
        ws.cell(r, 1, cp)
        ws.cell(r, 2, round(brut, 2)).number_format = '#,##0 "€"'
        ws.cell(r, 3, round(pond, 2)).number_format = '#,##0 "€"'

    # Sous-totaux par solution
    r += 2
    ws.cell(r, 1, "Sous-totaux par solution").font = Font(bold=True, color=BLUE_MID, size=11)
    r += 1
    for col, lib in [(1, "Solution"), (2, "Prévision brute (€)"), (3, "Prévision pondérée (€)")]:
        ws.cell(r, col, lib)
    _hdr(ws, r, 3)
    par_sol = defaultdict(lambda: [0.0, 0.0])
    for l in lignes:
        par_sol[l["Solution"]][0] += l["Prévision brute (€)"]
        par_sol[l["Solution"]][1] += l["Prévision pondérée (€)"]
    for sol, (brut, pond) in sorted(par_sol.items()):
        r += 1
        ws.cell(r, 1, sol)
        ws.cell(r, 2, round(brut, 2)).number_format = '#,##0 "€"'
        ws.cell(r, 3, round(pond, 2)).number_format = '#,##0 "€"'

    # largeurs
    for c, w in zip(range(1, 10), [14, 22, 26, 30, 12, 16, 18, 8, 18]):
        ws.column_dimensions[chr(64 + c)].width = w
    return tot_brut, tot_pond


def _maj_recap(wb, mois, annee, tot_brut, tot_pond, par_cp_sol):
    """Onglet récap : totaux par mois + ventilation CP×solution."""
    nom = "Récap"
    ws = wb[nom] if nom in wb.sheetnames else wb.create_sheet(nom, 0)
    if ws.max_row < 1 or ws.cell(1, 1).value is None:
        ws.cell(1, 1, "Récapitulatif prévisionnel pondéré")
        ws.cell(1, 1).font = Font(bold=True, size=13, color=BLUE_DARK)
        headers = ["Mois", "Année", "Prévision brute (€)", "Prévision pondérée (€)"]
        for c, hh in enumerate(headers, 1):
            ws.cell(3, c, hh)
        _hdr(ws, 3, len(headers))

    # upsert ligne du mois
    cible = None
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == mois and ws.cell(r, 2).value == annee:
            cible = r
            break
    if cible is None:
        cible = ws.max_row + 1
        ws.cell(cible, 1, mois)
        ws.cell(cible, 2, annee)
    ws.cell(cible, 3, round(tot_brut, 2)).number_format = '#,##0 "€"'
    ws.cell(cible, 4, round(tot_pond, 2)).number_format = '#,##0 "€"'
    for c, w in zip(range(1, 5), [14, 8, 18, 18]):
        ws.column_dimensions[chr(64 + c)].width = w


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--mois", default=None, help="Mois cible (défaut : mois courant du cache)")
    args = ap.parse_args()

    if not os.path.exists(CACHE_FILE):
        print(f"❌ {CACHE_FILE} introuvable. Lancer d'abord collect_jira_worklogs.py")
        sys.exit(1)

    with open(CACHE_FILE, "rb") as f:
        cache = pickle.load(f)

    mois  = args.mois or cache.get("mois_label", MOIS_FR[date.today().month - 1])
    annee = cache.get("annee", date.today().year)

    print(f"📋 Construction du prévisionnel pondéré — {mois} {annee}")
    lignes = construire_previsionnel(cache)
    print(f"   {len(lignes)} ticket(s) CA avec prévision")

    if os.path.exists(OUTPUT):
        wb = load_workbook(OUTPUT)
    else:
        wb = Workbook()
        wb.remove(wb.active)  # retirer la feuille par défaut

    tot_brut, tot_pond = _ecrire_onglet_mois(wb, mois, annee, lignes)
    _maj_recap(wb, mois, annee, tot_brut, tot_pond, None)

    # Récap en première position
    if "Récap" in wb.sheetnames:
        wb.move_sheet("Récap", -(wb.sheetnames.index("Récap")))

    wb.save(OUTPUT)
    print(f"   Total brut     : {tot_brut:,.0f} €")
    print(f"   Total pondéré  : {tot_pond:,.0f} €")
    print(f"✅ {OUTPUT}")


if __name__ == "__main__":
    main()
