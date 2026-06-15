"""
maj_reconnaissance_ca.py
========================
Script de mise à jour du tableau de reconnaissance CA.

Usage :
    python maj_reconnaissance_ca.py --mois Juin --fichier "Juin26.xlsx"

Arguments :
    --mois      Nom du mois à ajouter (ex: Juin, Juillet…)
    --fichier   Chemin vers le fichier export mensuel
    --tableau   (optionnel) Chemin vers le tableau existant
                Par défaut : Reconnaissance_CA_2026.xlsx dans le même dossier

Exemple :
    python maj_reconnaissance_ca.py --mois Juin --fichier exports/Juin26.xlsx
"""

import argparse
import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# ⚙️  CONFIGURATION JIRA — à compléter
# ---------------------------------------------------------------------------
JIRA_URL = "https://sogelink.atlassian.net"
JIRA_USER = "mathilde.panier@sogelink.com"
JIRA_API_TOKEN = os.environ["JIRA_API_TOKEN"]
JIRA_FIELD_ID         = "customfield_20701"   # Montant total commande
JIRA_FIELD_AVANT_2026 = "customfield_26980"   # Montant reconnu avant 2026
JIRA_FIELD_AVANCEMENT = "customfield_21883"   # % d'avancement

# ---------------------------------------------------------------------------
# ⚙️  CORRESPONDANCE MANUELLE Epic (tickets sans Epic dans l'export)
#     À compléter si de nouveaux cas apparaissent
# ---------------------------------------------------------------------------
EPIC_FALLBACK = {
    'PDMDEP-29031': 'PSC-1101',
    'PDMDEP-28860': 'PSC-1095',
    'PDMDEP-28859': 'PSC-1079',
}

# ---------------------------------------------------------------------------
# Constantes de style
# ---------------------------------------------------------------------------
HEADER_FILL      = PatternFill('solid', start_color='1F4E79')
SUBHEADER_FILL   = PatternFill('solid', start_color='2E75B6')
ALT_FILL         = PatternFill('solid', start_color='EBF3FB')
TOTAL_COL_FILL   = PatternFill('solid', start_color='D6E4F0')
GRAND_TOTAL_FILL = PatternFill('solid', start_color='1F4E79')
OVERRUN_FILL      = PatternFill('solid', start_color='FFCCCC')
OVERRUN_FONT_COLOR = 'C00000'

BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)

def apply_border(cell):
    cell.border = BORDER

def latest(series):
    return series.dropna().iloc[-1] if series.dropna().any() else ''


# ---------------------------------------------------------------------------
# Récupération du montant total commande via l'API Jira
# ---------------------------------------------------------------------------
def fetch_jira_budget(ticket_ids):
    if not ticket_ids:
        return {}

    if "VOTRE" in JIRA_API_TOKEN:
        print("⚠️  Configuration Jira incomplète (JIRA_API_TOKEN non renseigné).")
        print("   Les colonnes Jira resteront vides.")
        return {}, {}, {}

    try:
        from jira import JIRA
    except ImportError:
        print("⚠️  Librairie 'jira' manquante. Installez-la avec : pip install jira")
        return {}, {}, {}

    try:
        jira = JIRA(server=JIRA_URL, basic_auth=(JIRA_USER, JIRA_API_TOKEN))
    except Exception as e:
        print(f"⚠️  Connexion Jira impossible : {e}")
        return {}, {}, {}

    budgets = {}
    avant_2026_cache = {}
    avancement_cache = {}
    batch_size = 50
    ticket_list = list(ticket_ids)

    for i in range(0, len(ticket_list), batch_size):
        batch = ticket_list[i:i + batch_size]
        jql = f"issueKey in ({', '.join(batch)})"
        try:
            issues = jira.search_issues(
                jql, maxResults=batch_size,
                fields=f"{JIRA_FIELD_ID},{JIRA_FIELD_AVANT_2026},{JIRA_FIELD_AVANCEMENT}"
            )
            for issue in issues:
                key = issue.key
                fields = issue.raw["fields"]

                def get_float(f):
                    raw = fields.get(f)
                    if isinstance(raw, dict):
                        raw = raw.get("value") or raw.get("name")
                    try:
                        return float(raw) if raw is not None else None
                    except (TypeError, ValueError):
                        return None

                budgets[key] = get_float(JIRA_FIELD_ID)
                avant_2026_cache[key] = get_float(JIRA_FIELD_AVANT_2026)
                avancement_cache[key] = get_float(JIRA_FIELD_AVANCEMENT)
        except Exception as e:
            print(f"⚠️  Erreur API Jira (batch {i//batch_size + 1}) : {e}")
            for t in batch:
                budgets[t] = None
                avant_2026_cache[t] = None
                avancement_cache[t] = None

    for t in ticket_list:
        if t not in budgets:
            budgets[t] = None
        if t not in avant_2026_cache:
            avant_2026_cache[t] = None
        if t not in avancement_cache:
            avancement_cache[t] = None

    return budgets, avant_2026_cache, avancement_cache


# ---------------------------------------------------------------------------
# Lecture d'un export mensuel
# ---------------------------------------------------------------------------
def load_monthly_export(filepath, mois):
    xl = pd.ExcelFile(filepath)

    if 'Suivi CA' in xl.sheet_names:
        df = pd.read_excel(filepath, sheet_name='Suivi CA', header=0)
        df = df[df['Type'] != 'Hardware']
        df['Numéro BDC'] = df['Numéro BDC'].apply(
            lambda x: f"{int(x):08d}"
            if pd.notna(x) and str(x).replace('.0', '').isdigit()
            else x
        )
    elif 'Suivi de production' in xl.sheet_names:
        df = pd.read_excel(filepath, sheet_name='Suivi de production', header=3)
    else:
        raise ValueError(
            f"Onglet 'Suivi CA' ou 'Suivi de production' introuvable dans {filepath}.\n"
            f"Onglets disponibles : {xl.sheet_names}"
        )

    df = df[df['Montant déclaré ce mois'] != 0]
    df = df[df['Ticket CA'].notna() & ~df['Ticket CA'].astype(str).str.upper().str.contains('TOTAL')]
    df = df[['Ticket CA', 'Epic', 'Projet', 'Client', 'Solution', 'Numéro BDC', 'Montant déclaré ce mois']].copy()

    # En janvier, corriger les lignes où le Ticket CA est erroné (erreur de saisie dans l'export)
    # Format : {ticket_ca_errone: (ticket_ca_correct, epic_correct)}
    CORRECTIONS_TICKET_CA_JANVIER = {
        'PDMDEP-29505': ('PDMDEP-29510', 'PDMDEP-29505'),  # Erreur saisie : Epic Pont l'Abbé saisi au lieu du vrai Ticket CA
    }
    if 'Suivi CA' in pd.ExcelFile(filepath).sheet_names:
        for tc_errone, (tc_correct, epic_correct) in CORRECTIONS_TICKET_CA_JANVIER.items():
            mask = df['Ticket CA'].astype(str) == tc_errone
            df.loc[mask, 'Ticket CA'] = tc_correct
            df.loc[mask, 'Epic'] = epic_correct

    # Fallback Epic (tickets sans Epic ou avec collision Epic/Ticket CA)
    def resolve_epic(r):
        tc = str(r['Ticket CA'])
        epic = r['Epic']
        # Si ce Ticket CA est dans EPIC_FALLBACK, forcer l'Epic
        if tc in EPIC_FALLBACK:
            return EPIC_FALLBACK[tc]
        # Si Epic est vide, utiliser le Ticket CA comme Epic
        if pd.isna(epic):
            return tc
        return epic
    df['Epic'] = df.apply(resolve_epic, axis=1)
    df['Mois'] = mois
    return df


# ---------------------------------------------------------------------------
# Lecture du tableau existant
# ---------------------------------------------------------------------------
def load_existing_tableau(tableau_path):
    wb = openpyxl.load_workbook(tableau_path, data_only=True)

    # Lire l'onglet "Par Ticket CA" comme source de vérité
    if 'Par Ticket CA' in wb.sheetnames:
        ws = wb['Par Ticket CA']
    else:
        ws = wb.active

    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]

    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0] or str(row[0]).upper() == 'TOTAL':
            continue
        rows.append(dict(zip(headers, row)))

    df = pd.DataFrame(rows)

    fixed = {'Ticket CA', 'Epic', 'Numéro BDC', 'Solution', 'Client', 'Projet',
             'Montant commande', 'Total', 'Total reconnu', 'Avant 2026', '% Avancement', 'Vérification', None}
    existing_months = [h for h in headers if h not in fixed]

    # Caches déjà présents dans le tableau
    existing_budgets = {}
    existing_avant_2026 = {}
    existing_avancement = {}
    for _, row in df.iterrows():
        tc = row.get('Ticket CA')
        if not pd.notna(tc):
            continue
        tc = str(tc)
        if pd.notna(row.get('Montant commande')):
            existing_budgets[tc] = float(row['Montant commande'])
        if pd.notna(row.get('Avant 2026')):
            existing_avant_2026[tc] = float(row['Avant 2026'])
        if pd.notna(row.get('% Avancement')):
            # Excel stocke 80% comme 0.8 — on remet à l'échelle Jira (80)
            existing_avancement[tc] = float(row['% Avancement']) * 100

    id_cols = ['Ticket CA', 'Epic', 'Numéro BDC', 'Solution', 'Client', 'Projet']
    id_cols_present = [c for c in id_cols if c in df.columns]
    melted = df.melt(id_vars=id_cols_present, value_vars=existing_months,
                     var_name='Mois', value_name='Montant déclaré ce mois')
    melted = melted[melted['Montant déclaré ce mois'].notna()]
    melted = melted[melted['Montant déclaré ce mois'] != 0]
    melted = melted.rename(columns={'Numéro BDC': 'Numéro BDC'})

    return melted, existing_months, existing_budgets, existing_avant_2026, existing_avancement


# ---------------------------------------------------------------------------
# Construction des deux onglets Excel
# ---------------------------------------------------------------------------
def write_sheet(ws, pivot_wide, cols, display_headers, months, title,
                budgets=None, overrun_col=None):
    ncols = len(cols)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f'A1:{last_col}1')
    c = ws['A1']
    c.value = title
    c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    for col_idx, h in enumerate(display_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        apply_border(cell)
    ws.row_dimensions[2].height = 22

    for row_idx, row in pivot_wide.iterrows():
        excel_row = row_idx + 3
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill()

        total_val = row.get('Total', 0) or 0
        ticket_ca = str(row.get('Ticket CA', '') or row.get('Ticket_CA', ''))
        budget_val = budgets.get(ticket_ca) if budgets else None
        is_overrun = (
            budget_val is not None and pd.notna(budget_val) and round(total_val, 2) > round(float(budget_val), 2)
        )

        for col_idx, col_name in enumerate(cols, 1):
            val = row[col_name] if col_name in row.index else None
            cell = ws.cell(row=excel_row, column=col_idx)

            if col_name == 'Montant_commande':
                cell.value = budget_val if budget_val is not None else None
                if cell.value is not None:
                    cell.number_format = '#,##0.00 €'
                cell.alignment = Alignment(horizontal='right')
                cell.font = Font(name='Arial', size=9)
                cell.fill = fill
            elif col_name == 'Avant_2026':
                cell.value = val if pd.notna(val) else None
                if cell.value is not None:
                    cell.number_format = '#,##0.00 €'
                cell.alignment = Alignment(horizontal='right')
                cell.font = Font(name='Arial', size=9)
                cell.fill = fill
            elif col_name == 'Avancement':
                cell.value = (float(val) / 100) if pd.notna(val) else None
                if cell.value is not None:
                    cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal='right')
                cell.font = Font(name='Arial', size=9)
                cell.fill = fill
            elif col_name == 'Verification':
                cell.value = val if pd.notna(val) and val != '' else None
                cell.alignment = Alignment(horizontal='center')
                if val == 'OK':
                    cell.font = Font(name='Arial', size=9, bold=True, color='375623')
                    cell.fill = PatternFill('solid', start_color='E2EFDA')
                elif val and val != '':
                    cell.font = Font(name='Arial', size=9, bold=True, color=OVERRUN_FONT_COLOR)
                    cell.fill = OVERRUN_FILL
                else:
                    cell.font = Font(name='Arial', size=9)
                    cell.fill = fill
            elif col_name in months or col_name == 'Total':
                if pd.notna(val) and val != 0:
                    cell.value = val
                    cell.number_format = '#,##0.00 €'
                else:
                    cell.value = None
                cell.alignment = Alignment(horizontal='right')
                if col_name == 'Total' and is_overrun:
                    cell.fill = OVERRUN_FILL
                    cell.font = Font(name='Arial', size=9, bold=True, color=OVERRUN_FONT_COLOR)
                else:
                    cell.fill = TOTAL_COL_FILL if col_name == 'Total' else fill
                    cell.font = Font(name='Arial', size=9, bold=(col_name == 'Total'))
            else:
                str_val = val if pd.notna(val) else ''
                # Rendre Ticket CA et Epic cliquables
                if col_name in ('Ticket CA', 'Ticket_CA', 'Epic') and str_val:
                    cell.value = str_val
                    cell.hyperlink = f'https://sogelink.atlassian.net/browse/{str_val}'
                    cell.font = Font(name='Arial', size=9, color='0563C1', underline='single')
                else:
                    cell.value = str_val
                    cell.font = Font(name='Arial', size=9)
                cell.alignment = Alignment(horizontal='left', wrap_text=(col_name == 'Projet'))
                cell.fill = fill

            apply_border(cell)

    total_row = len(pivot_wide) + 3
    for c in range(1, ncols + 1):
        cell = ws.cell(row=total_row, column=c)
        cell.fill = GRAND_TOTAL_FILL
        cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.value = ''
        apply_border(cell)
    ws.cell(row=total_row, column=1).value = 'TOTAL'

    month_start_col = next((i+1 for i, cn in enumerate(cols) if cn == 'Avant_2026'), None) or next(i+1 for i, cn in enumerate(cols) if cn == months[0])
    total_col_idx = next((i+1 for i, cn in enumerate(cols) if cn == 'Total'), None)

    for col_idx in range(month_start_col, ncols + 1):
        col_name = cols[col_idx - 1]
        if col_name in ('Montant_commande', 'Avancement', 'Verification'):
            continue
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=total_row, column=col_idx)
        cell.value = f'=SUM({col_letter}3:{col_letter}{total_row - 1})'
        cell.number_format = '#,##0.00 €'
        cell.alignment = Alignment(horizontal='right')

    ws.row_dimensions[total_row].height = 20
    ws.freeze_panes = f'{get_column_letter(month_start_col)}3'


def build_excel(combined, all_months, budgets, avant_2026, avancement, output_path):
    wb = openpyxl.Workbook()

    # --- Onglet Par Epic ---
    epic_info = combined.groupby('Epic').agg(
        Client=('Client', latest),
        Solution=('Solution', latest),
    ).reset_index()

    pivot_epic = combined.groupby(['Epic', 'Mois'])['Montant déclaré ce mois'].sum().reset_index()
    pw_epic = pivot_epic.pivot_table(
        index='Epic', columns='Mois', values='Montant déclaré ce mois', aggfunc='sum'
    ).reset_index()
    for m in all_months:
        if m not in pw_epic.columns:
            pw_epic[m] = None
    pw_epic = pw_epic.merge(epic_info, on='Epic', how='left')
    pw_epic['Total'] = pw_epic[all_months].sum(axis=1)
    pw_epic = pw_epic.sort_values('Total', ascending=False).reset_index(drop=True)
    cols_epic = ['Epic', 'Solution', 'Client'] + all_months + ['Total']
    pw_epic = pw_epic[cols_epic]

    ws_epic = wb.active
    ws_epic.title = 'Par Epic'
    write_sheet(ws_epic, pw_epic, cols_epic,
                ['Epic', 'Solution', 'Client'] + all_months + ['Total'],
                all_months, 'Suivi de Reconnaissance CA 2026 — par Epic')
    ws_epic.column_dimensions['A'].width = 16
    ws_epic.column_dimensions['B'].width = 12
    ws_epic.column_dimensions['C'].width = 32
    for i in range(len(all_months)):
        ws_epic.column_dimensions[get_column_letter(4 + i)].width = 13 if i < 2 else 12
    ws_epic.column_dimensions[get_column_letter(4 + len(all_months))].width = 16

    # --- Onglet Par Ticket CA ---
    ticket_info = combined.groupby('Ticket CA').agg(
        Epic=('Epic', latest),
        Projet=('Projet', latest),
        Client=('Client', latest),
        Solution=('Solution', latest),
        BDC=('Numéro BDC', latest),
    ).reset_index()

    pivot_ticket = combined.groupby(['Ticket CA', 'Mois'])['Montant déclaré ce mois'].sum().reset_index()
    pw_ticket = pivot_ticket.pivot_table(
        index='Ticket CA', columns='Mois', values='Montant déclaré ce mois', aggfunc='sum'
    ).reset_index()
    for m in all_months:
        if m not in pw_ticket.columns:
            pw_ticket[m] = None
    pw_ticket = pw_ticket.merge(ticket_info, on='Ticket CA', how='left')

    # Champs Jira complémentaires
    pw_ticket['Avant_2026']  = pw_ticket['Ticket CA'].astype(str).map(avant_2026)
    pw_ticket['Avancement']  = pw_ticket['Ticket CA'].astype(str).map(avancement)
    pw_ticket['Montant_commande'] = pw_ticket['Ticket CA'].astype(str).map(budgets)

    # Total = Avant 2026 + tous les mois 2026
    pw_ticket['Total'] = pw_ticket[all_months].sum(axis=1).add(pw_ticket['Avant_2026'].fillna(0))

    # Vérification : % avancement × montant commande vs total reconnu (tolérance 1€)
    def verif(row):
        pct  = row.get('Avancement')
        cmd  = row.get('Montant_commande')
        tot  = row.get('Total')
        if pd.isna(pct) or pd.isna(cmd) or pd.isna(tot):
            return ''
        attendu = round(float(pct) / 100 * float(cmd), 2)
        diff    = round(float(tot) - attendu, 2)
        return 'OK' if abs(diff) <= 1 else f'{diff:+.2f} €'

    pw_ticket['Verification'] = pw_ticket.apply(verif, axis=1)

    pw_ticket = pw_ticket.sort_values('Total', ascending=False).reset_index(drop=True)
    cols_ticket = ['Ticket CA', 'Epic', 'BDC', 'Solution', 'Client', 'Projet', 'Avant_2026'] + all_months + ['Total', 'Montant_commande', 'Avancement', 'Verification']
    for col in cols_ticket:
        if col not in pw_ticket.columns:
            pw_ticket[col] = None
    pw_ticket = pw_ticket[cols_ticket]

    ws_ticket = wb.create_sheet('Par Ticket CA')
    display_ticket = ['Ticket CA', 'Epic', 'Numéro BDC', 'Solution', 'Client', 'Projet', 'Avant 2026'] + all_months + ['Total reconnu', 'Montant commande', '% Avancement', 'Vérification']
    write_sheet(ws_ticket, pw_ticket, cols_ticket, display_ticket,
                all_months, 'Suivi de Reconnaissance CA 2026 — par Ticket CA',
                budgets=budgets)
    ws_ticket.column_dimensions['A'].width = 16
    ws_ticket.column_dimensions['B'].width = 16
    ws_ticket.column_dimensions['C'].width = 14
    ws_ticket.column_dimensions['D'].width = 12
    ws_ticket.column_dimensions['E'].width = 28
    ws_ticket.column_dimensions['F'].width = 40
    for i in range(len(all_months)):
        ws_ticket.column_dimensions[get_column_letter(7 + i)].width = 13 if i < 2 else 12
    ws_ticket.column_dimensions[get_column_letter(7)].width = 14             # Avant 2026
    for i in range(len(all_months)):
        ws_ticket.column_dimensions[get_column_letter(8 + i)].width = 13 if i < 2 else 12
    ws_ticket.column_dimensions[get_column_letter(8 + len(all_months))].width = 16     # Total
    ws_ticket.column_dimensions[get_column_letter(8 + len(all_months) + 1)].width = 16 # Montant commande
    ws_ticket.column_dimensions[get_column_letter(8 + len(all_months) + 2)].width = 12 # % Avancement
    ws_ticket.column_dimensions[get_column_letter(8 + len(all_months) + 3)].width = 14 # Vérification

    wb.save(output_path)
    return pw_epic, pw_ticket


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description='Mise à jour du tableau de reconnaissance CA.')
    parser.add_argument('--mois',    required=True, help='Nom du mois (ex: Juin)')
    parser.add_argument('--fichier', required=True, help='Chemin vers le fichier export mensuel (.xlsx)')
    parser.add_argument('--tableau', default=None,  help='Chemin vers le tableau existant (optionnel)')
    parser.add_argument('--refresh-jira', action='store_true', help='Forcer le rechargement de tous les champs Jira (ignore le cache)')
    args = parser.parse_args()

    script_dir   = os.path.dirname(os.path.abspath(__file__))
    tableau_path = args.tableau or os.path.join(script_dir, 'Reconnaissance_CA_2026.xlsx')

    if not os.path.exists(args.fichier):
        print(f"Erreur : fichier introuvable → {args.fichier}")
        sys.exit(1)
    if not os.path.exists(tableau_path):
        print(f"Erreur : tableau introuvable → {tableau_path}")
        sys.exit(1)

    print(f"→ Lecture du tableau existant : {tableau_path}")
    existing_df, existing_months, existing_budgets, existing_avant_2026, existing_avancement = load_existing_tableau(tableau_path)

    # Comparaison insensible à la casse
    mois_match = next((m for m in existing_months if m.lower() == args.mois.lower()), None)
    args.mois = mois_match if mois_match else args.mois.capitalize()

    if mois_match:
        print(f"Attention : le mois '{mois_match}' est déjà présent dans le tableau.")
        confirm = input("Écraser les données existantes pour ce mois ? (o/n) : ").strip().lower()
        if confirm != 'o':
            print("Annulé.")
            sys.exit(0)
        existing_df     = existing_df[existing_df['Mois'] != mois_match]
        existing_months = [m for m in existing_months if m != mois_match]

    print(f"→ Lecture du nouvel export : {args.fichier} (mois : {args.mois})")
    new_df = load_monthly_export(args.fichier, args.mois)
    print(f"   {len(new_df)} lignes chargées.")

    combined = pd.concat([existing_df, new_df], ignore_index=True)

    ORDRE_MOIS = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                  'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    all_months = existing_months + ([args.mois] if args.mois not in existing_months else [])
    all_months = sorted(set(all_months), key=lambda m: ORDRE_MOIS.index(m) if m in ORDRE_MOIS else 99)

    # Récupération des budgets Jira
    all_tickets = set(combined['Ticket CA'].dropna().astype(str).unique())
    # Un ticket est "connu" seulement s'il est en cache pour TOUS les champs Jira
    # --refresh-jira force le rappel de tous les tickets
    if args.refresh_jira:
        known_tickets = set()
    else:
        known_tickets = set(existing_budgets.keys()) & set(existing_avancement.keys())
    new_tickets = all_tickets - known_tickets

    if new_tickets:
        print(f"→ Interrogation Jira pour {len(new_tickets)} ticket(s)…")
        new_budgets, new_avant_2026, new_avancement = fetch_jira_budget(new_tickets)
        budgets    = {**existing_budgets,    **new_budgets}
        avant_2026 = {**existing_avant_2026, **new_avant_2026}
        avancement = {**existing_avancement, **new_avancement}
    else:
        print("→ Tous les budgets sont déjà en cache, pas d'appel Jira nécessaire.")
        budgets    = existing_budgets
        avant_2026 = existing_avant_2026
        avancement = existing_avancement

    found = sum(1 for v in budgets.values() if v is not None)
    print(f"   Montants commande récupérés : {found}/{len(all_tickets)}")

    pw_epic, pw_ticket = build_excel(combined, all_months, budgets, avant_2026, avancement, tableau_path)

    print(f"\nTableau mis à jour avec succès : {tableau_path}")

    mapped_budgets = pw_ticket['Ticket CA'].astype(str).map(budgets)
    overruns = pw_ticket[
        mapped_budgets.notna()
        & pw_ticket['Total'].notna()
        & (pw_ticket['Total'].astype(float) > mapped_budgets.astype(float))
    ]
    if len(overruns) > 0:
        print(f"\n⚠️  {len(overruns)} ticket(s) avec total reconnu > montant commande :")
        for _, r in overruns.iterrows():
            budget = budgets.get(str(r['Ticket CA']))
            print(f"   {r['Ticket CA']} — {r.get('Projet','')} : reconnu {r['Total']:,.2f} € / commande {budget:,.2f} €")
    else:
        print("\n✅ Aucun dépassement détecté.")

    print("\nTotaux par mois (Par Epic) :")
    for m in all_months:
        print(f"  {m} : {pw_epic[m].sum():,.2f} €")
    print(f"  TOTAL : {pw_epic['Total'].sum():,.2f} €")
    print(f"  Epics : {len(pw_epic)} | Tickets CA : {len(pw_ticket)}")


if __name__ == '__main__':
    main()
