import os
"""
Extraction des temps Clockwork depuis Jira
Genere un fichier Excel avec les temps passes par mois, par ligne Epic,
avec distinction Chef de projet / Charges de deploiement.

Prerequis :
  pip install jira openpyxl pandas
"""

import sys
from collections import defaultdict
from types import SimpleNamespace

import pandas as pd
from jira import JIRA
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -----------------------------------------------------------------------------
#  CONFIGURATION
# -----------------------------------------------------------------------------
JIRA_SERVER = "https://sogelink.atlassian.net"
JIRA_EMAIL = "mathilde.panier@sogelink.com"
JIRA_TOKEN = os.environ["JIRA_API_TOKEN"]

SOURCE_EXCEL = 'Reconnaissance_CA_2026.xlsx'
OUTPUT_EXCEL = 'Temps_Clockwork_2026.xlsx'
ANNEE        = 2026

# -----------------------------------------------------------------------------
#  COLLABORATEURS
# -----------------------------------------------------------------------------
ACCOUNT_TO_COLLAB = {
    '63fe26c90a4a47fb8d213c54'                    : 'Quentin Bordillon',
    '712020:bb1c3fc2-7106-4c8b-9955-2e4917706722' : 'Fabien Reutenauer',
    '63fe26c7f00d095406f2590c'                    : 'Berenice Bossard',
    '712020:f59016f5-1528-45f5-a182-d91eb6f69cad' : 'Flavie Bardin',
    '712020:e3996f09-39b9-4957-8f83-d925d9543c98' : 'Maxime Pontonnier',
    '712020:30146e03-1c8d-42f4-a105-c6987f9b2e2c' : 'Marine Masingarbe',
    '712020:ab30e040-78a7-40e1-aef4-ff6781645c8e' : 'Duncan Hamelin',
    '712020:a40869a4-1333-4240-ac8a-670ed2314e0b' : 'Remy Vincent',
}
CHARGES_DE_DEPLOIEMENT = {
    'Quentin Bordillon', 'Berenice Bossard',
    'Maxime Pontonnier', 'Marine Masingarbe', 'Duncan Hamelin',
}
ALLOWED_ACCOUNT_IDS = set(ACCOUNT_TO_COLLAB.keys())

MOIS_LABELS = ['Janvier', 'Fevrier', 'Mars', 'Avril', 'Mai', 'Juin',
               'Juillet', 'Aout', 'Septembre', 'Octobre', 'Novembre', 'Decembre']

# -----------------------------------------------------------------------------
#  CONNEXION JIRA
# -----------------------------------------------------------------------------
try:
    jira = JIRA(options={'server': JIRA_SERVER}, basic_auth=(JIRA_EMAIL, JIRA_TOKEN))
    print('Connexion Jira etablie.')
except Exception as e:
    print(f'Impossible de se connecter a Jira : {e}')
    sys.exit(1)

# -----------------------------------------------------------------------------
#  HELPERS WORKLOGS
# -----------------------------------------------------------------------------
def _make_worklog_obj(w_dict):
    author = SimpleNamespace(
        accountId   = w_dict['author']['accountId'],
        displayName = w_dict['author']['displayName'],
    )
    return SimpleNamespace(
        author           = author,
        timeSpentSeconds = w_dict['timeSpentSeconds'],
        started          = w_dict['started'],
    )


def fetch_worklogs_for_issue(issue):
    """
    Recupere tous les worklogs d'un ticket.
    - Si < 20 worklogs embarques et coherent : retour direct, 0 appel API.
    - Sinon : appel pagine a /worklog.
    """
    existing_wl = getattr(issue.fields.worklog, 'worklogs', [])
    time_spent  = getattr(issue.fields, 'timespent', None) or 0
    need_api    = len(existing_wl) >= 20 or (len(existing_wl) == 0 and time_spent > 0)

    if not need_api:
        return existing_wl

    all_wl   = []
    start_at = 0
    url      = f'{JIRA_SERVER}/rest/api/3/issue/{issue.key}/worklog'
    while True:
        try:
            resp = jira._session.get(url, params={'startAt': start_at, 'maxResults': 50})
            resp.raise_for_status()
        except Exception as e:
            print(f'  [!] Worklogs {issue.key} : {e}')
            return existing_wl
        data  = resp.json()
        batch = data.get('worklogs', [])
        all_wl.extend(_make_worklog_obj(w) for w in batch)
        if len(batch) < 50:
            break
        start_at += 50
    return all_wl


def parse_worklog(wl):
    """Retourne (nom, mois, heures) ou None si hors perimetre."""
    if wl.author.accountId not in ALLOWED_ACCOUNT_IDS:
        return None
    try:
        started = pd.to_datetime(wl.started).date()
    except Exception:
        return None
    if started.year != ANNEE:
        return None
    return {
        'nom'    : ACCOUNT_TO_COLLAB[wl.author.accountId],
        'mois'   : started.month,
        'heures' : wl.timeSpentSeconds / 3600.0,
    }

# -----------------------------------------------------------------------------
#  RECUPERATION DE TOUTES LES ISSUES PDMDEP (comme Export_worklog_V3)
# -----------------------------------------------------------------------------
print('Recuperation de toutes les issues PDMDEP...')

all_issues = jira.search_issues(
    'project = PDMDEP',
    expand='worklog,subtasks',
    maxResults=False,
    fields='summary,issuetype,parent,subtasks,timespent,worklog',
)

print(f'  -> {len(all_issues)} issues recuperees.')
issue_map = {i.key: i for i in all_issues}

# Index enfants par parent
children_by_parent = defaultdict(list)
for issue in all_issues:
    parent = getattr(issue.fields, 'parent', None)
    if parent:
        children_by_parent[parent.key].append(issue.key)


def get_all_children(epic_key):
    """Retourne tous les tickets enfants + leurs sous-taches."""
    children = children_by_parent.get(epic_key, [])
    all_keys = set(children)
    for c in children:
        issue = issue_map.get(c)
        if issue:
            for sub in getattr(issue.fields, 'subtasks', []):
                all_keys.add(sub.key)
    return list(all_keys)

# -----------------------------------------------------------------------------
#  LECTURE DU FICHIER SOURCE
# -----------------------------------------------------------------------------
print(f'Lecture de {SOURCE_EXCEL}...')
df_source = pd.read_excel(SOURCE_EXCEL, header=1)
df_source = df_source[df_source['Epic'] != 'TOTAL'].copy()
df_source = df_source[df_source['Epic'].notna()].copy()
df_source['Epic'] = df_source['Epic'].astype(str).str.strip()
lignes = df_source.to_dict('records')
print(f'  -> {len(lignes)} lignes dans le fichier source.')

# -----------------------------------------------------------------------------
#  CALCUL DES TEMPS PAR EPIC
#  Structure : {epic -> {mois -> {CP, CD, total}}}
# -----------------------------------------------------------------------------
data      = defaultdict(lambda: defaultdict(lambda: {'CP': 0.0, 'CD': 0.0, 'total': 0.0}))
epic_done = {}   # cache epic -> liste de toutes les issues

total_lignes = len(lignes)
for i, ligne in enumerate(lignes, 1):
    epic = ligne['Epic']
    print(f'  [{i:3}/{total_lignes}] Epic={epic}', end='  ', flush=True)

    if epic not in epic_done:
        epic_done[epic] = [epic] + get_all_children(epic)
    all_keys = epic_done[epic]

    nb_wl = 0
    for key in all_keys:
        issue = issue_map.get(key)
        if not issue:
            continue
        for wl in fetch_worklogs_for_issue(issue):
            parsed = parse_worklog(wl)
            if not parsed:
                continue
            mois   = parsed['mois']
            heures = parsed['heures']
            nom    = parsed['nom']
            data[epic][mois]['total'] += heures
            if nom in CHARGES_DE_DEPLOIEMENT:
                data[epic][mois]['CD'] += heures
            else:
                data[epic][mois]['CP'] += heures
            nb_wl += 1

    print(f'-> {len(all_keys)} issues, {nb_wl} worklog(s) retenus')

print('Recuperation terminee.')

# -----------------------------------------------------------------------------
#  CONSTRUCTION DU DATAFRAME DE SORTIE
# -----------------------------------------------------------------------------
rows = []
for ligne in lignes:
    epic       = ligne['Epic']
    epic_data  = data.get(epic, {})

    out_row = {
        'Epic'     : epic,
        'Solution' : str(ligne.get('Solution', '')),
        'Client'   : str(ligne.get('Client',   '')),
    }

    total_cp = total_cd = total_all = 0.0
    for m in range(1, 13):
        ml  = MOIS_LABELS[m - 1]
        md  = epic_data.get(m, {})
        cp  = md.get('CP',    0.0)
        cd  = md.get('CD',    0.0)
        tot = md.get('total', 0.0)
        out_row[f'{ml}_CP']    = round(cp,  2) if cp  else None
        out_row[f'{ml}_CD']    = round(cd,  2) if cd  else None
        out_row[f'{ml}_Total'] = round(tot, 2) if tot else None
        total_cp  += cp
        total_cd  += cd
        total_all += tot

    out_row['Total_CP']    = round(total_cp,  2)
    out_row['Total_CD']    = round(total_cd,  2)
    out_row['Total_Temps'] = round(total_all, 2)
    rows.append(out_row)

df_out = pd.DataFrame(rows)

# Ligne totaux
totals = {c: '' for c in df_out.columns}
totals['Epic'] = 'TOTAL'
for col in df_out.columns:
    if pd.api.types.is_numeric_dtype(df_out[col]):
        totals[col] = round(df_out[col].sum(), 2)
df_out = pd.concat([df_out, pd.DataFrame([totals])], ignore_index=True)

# -----------------------------------------------------------------------------
#  ECRITURE EXCEL FORMATE
# -----------------------------------------------------------------------------
print(f'Generation de {OUTPUT_EXCEL}...')
df_out.to_excel(OUTPUT_EXCEL, index=False, sheet_name='Temps Clockwork 2026')

wb = load_workbook(OUTPUT_EXCEL)
ws = wb.active

COULEUR_ENTETE    = '1F3864'
COULEUR_CP        = 'D9E1F2'
COULEUR_CD        = 'E2EFDA'
COULEUR_TOTAL_COL = 'FCE4D6'
COULEUR_TOTAL_ROW = 'FFF2CC'

font_entete  = Font(bold=True, color='FFFFFF', name='Arial', size=9)
font_total   = Font(bold=True, name='Arial', size=9)
font_normal  = Font(name='Arial', size=9)
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

def make_fill(h):
    return PatternFill('solid', start_color=h, fgColor=h)

thin   = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

col_names = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

def col_type(name):
    if not name: return 'other'
    if name.endswith('_CP'):    return 'CP'
    if name.endswith('_CD'):    return 'CD'
    if name.endswith('_Total') or name in ('Total_CP', 'Total_CD', 'Total_Temps'):
        return 'total'
    return 'info'

ws.row_dimensions[1].height = 32
for c in range(1, ws.max_column + 1):
    cell = ws.cell(1, c)
    cell.font = font_entete; cell.fill = make_fill(COULEUR_ENTETE)
    cell.alignment = align_center; cell.border = border

for r in range(2, ws.max_row + 1):
    is_total_row = str(ws.cell(r, 1).value) == 'TOTAL'
    ws.row_dimensions[r].height = 14
    for c in range(1, ws.max_column + 1):
        cell  = ws.cell(r, c)
        ctype = col_type(col_names[c - 1])
        cell.border = border
        if is_total_row:
            cell.font = font_total; cell.fill = make_fill(COULEUR_TOTAL_ROW)
        elif ctype == 'CP':
            cell.font = font_normal; cell.fill = make_fill(COULEUR_CP)
        elif ctype == 'CD':
            cell.font = font_normal; cell.fill = make_fill(COULEUR_CD)
        elif ctype == 'total':
            cell.font = font_total; cell.fill = make_fill(COULEUR_TOTAL_COL)
        else:
            cell.font = font_normal
        cell.alignment = align_left if ctype == 'info' else align_center

col_widths = {'Epic': 15, 'Solution': 12, 'Client': 32}
for c in range(1, ws.max_column + 1):
    name = col_names[c - 1]
    if name in col_widths:
        ws.column_dimensions[get_column_letter(c)].width = col_widths[name]
    elif col_type(name) == 'total':
        ws.column_dimensions[get_column_letter(c)].width = 9
    else:
        ws.column_dimensions[get_column_letter(c)].width = 7

ws.freeze_panes = 'D2'

# Feuille recap par mois
ws2 = wb.create_sheet('Recap par mois')
ws2.append(['Mois', 'Heures CP', 'Heures CD', 'Heures Total'])
for m in range(1, 13):
    ml  = MOIS_LABELS[m - 1]
    cp  = df_out.iloc[:-1][f'{ml}_CP'].sum()
    cd  = df_out.iloc[:-1][f'{ml}_CD'].sum()
    tot = df_out.iloc[:-1][f'{ml}_Total'].sum()
    ws2.append([ml, round(cp, 2), round(cd, 2), round(tot, 2)])
ws2.append([
    'TOTAL',
    round(df_out.iloc[:-1]['Total_CP'].sum(),    2),
    round(df_out.iloc[:-1]['Total_CD'].sum(),    2),
    round(df_out.iloc[:-1]['Total_Temps'].sum(), 2),
])
for r in range(1, ws2.max_row + 1):
    for c in range(1, 5):
        cell = ws2.cell(r, c)
        cell.border = border; cell.alignment = align_center
        if r == 1:
            cell.font = font_entete; cell.fill = make_fill(COULEUR_ENTETE)
        elif r == ws2.max_row:
            cell.font = font_total; cell.fill = make_fill(COULEUR_TOTAL_ROW)
        else:
            cell.font = font_normal
            if c == 2:   cell.fill = make_fill(COULEUR_CP)
            elif c == 3: cell.fill = make_fill(COULEUR_CD)
            elif c == 4: cell.fill = make_fill(COULEUR_TOTAL_COL)
for col_letter, width in [('A', 14), ('B', 12), ('C', 12), ('D', 14)]:
    ws2.column_dimensions[col_letter].width = width

wb.save(OUTPUT_EXCEL)
print(f'Fichier genere : {OUTPUT_EXCEL}')
print(f'  -> {len(df_out) - 1} lignes de projet')
print(f'  -> 2 feuilles : Temps Clockwork 2026 + Recap par mois')
