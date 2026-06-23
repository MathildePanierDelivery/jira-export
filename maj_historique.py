"""
maj_historique.py
=================
Met à jour _historique.xlsx avec les données du mois courant (option A :
écrase la ligne du mois si elle existe, sinon l'ajoute).

Appelé par export_v4_palier.py en fin d'exécution.
Ne touche QUE les onglets/colonnes qu'on sait remplir ; préserve le reste
(notamment les colonnes N-1 saisies à la main, sauf le CA N-1 qu'on
récupère automatiquement depuis ca_2025.xlsx).

6 onglets alimentés : ca, charge, backlog, commandes, anciennete, ca_deal
"""

import os
import pandas as pd
from openpyxl import load_workbook

HISTORIQUE_FILE = "_historique.xlsx"
CA_N1_FILE      = "ca_2025.xlsx"


def _lire_ca_n1():
    """Lit le CA 2025 (année N-1) par mois depuis ca_2025.xlsx.
       Retourne {mois_label: (global, litt, geodp)}."""
    if not os.path.exists(CA_N1_FILE):
        return {}
    try:
        df = pd.read_excel(CA_N1_FILE, sheet_name=0, header=3)
        col_m = next((c for c in df.columns if "mois" in str(c).lower()), None)
        col_g = next((c for c in df.columns if "global" in str(c).lower()), None)
        col_l = next((c for c in df.columns if "litt" in str(c).lower()), None)
        col_ge = next((c for c in df.columns if "geodp" in str(c).lower()), None)
        out = {}
        for _, r in df.iterrows():
            m = r.get(col_m)
            if not isinstance(m, str) or m.upper() == "TOTAL":
                continue
            out[m] = (float(r.get(col_g) or 0), float(r.get(col_l) or 0), float(r.get(col_ge) or 0))
        return out
    except Exception as e:
        print(f"   ⚠ CA N-1 illisible : {e}")
        return {}


def _upsert_ligne(ws, cle_cols, valeurs, headers):
    """Insère ou met à jour une ligne identifiée par (Mois, Année).
       cle_cols = {'Mois': 'Juin', 'Année': 2026}
       valeurs  = {nom_colonne: valeur} pour les colonnes à écrire.
       Préserve les colonnes non fournies si la ligne existe déjà."""
    # index des colonnes par nom
    col_idx = {h: i+1 for i, h in enumerate(headers) if h}

    # chercher la ligne existante
    ligne_cible = None
    for r in range(2, ws.max_row + 1):
        match = all(
            str(ws.cell(r, col_idx[k]).value) == str(v)
            for k, v in cle_cols.items() if k in col_idx
        )
        if match:
            ligne_cible = r
            break

    if ligne_cible is None:
        ligne_cible = ws.max_row + 1
        # écrire la clé
        for k, v in cle_cols.items():
            if k in col_idx:
                ws.cell(ligne_cible, col_idx[k], v)

    # écrire les valeurs (seulement les colonnes fournies)
    for nom, val in valeurs.items():
        if nom in col_idx:
            ws.cell(ligne_cible, col_idx[nom], val)
    return ligne_cible


def _remplacer_lignes_mois(ws, cle_cols, lignes_data, headers):
    """Pour les onglets liste (commandes, anciennete, ca_deal) :
       supprime toutes les lignes du mois/année, puis réécrit la liste fournie."""
    col_idx = {h: i+1 for i, h in enumerate(headers) if h}

    # collecter les lignes à supprimer (du mois courant)
    a_supprimer = []
    for r in range(2, ws.max_row + 1):
        match = all(
            str(ws.cell(r, col_idx[k]).value) == str(v)
            for k, v in cle_cols.items() if k in col_idx
        )
        if match:
            a_supprimer.append(r)

    # supprimer de bas en haut
    for r in sorted(a_supprimer, reverse=True):
        ws.delete_rows(r, 1)

    # ajouter les nouvelles lignes à la fin
    for data in lignes_data:
        row_vals = [data.get(h, "") for h in headers]
        ws.append(row_vals)


def maj_historique(contexte):
    """Point d'entrée. `contexte` est un dict fourni par l'export avec
       toutes les données calculées du mois courant."""
    if not os.path.exists(HISTORIQUE_FILE):
        print(f"   ⚠ {HISTORIQUE_FILE} introuvable — accumulation ignorée.")
        return

    mois  = contexte["mois_label"]
    annee = contexte["annee"]
    cle   = {"Mois": mois, "Année": annee}

    wb = load_workbook(HISTORIQUE_FILE)

    def headers_de(ws):
        return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # ── Onglet CA (avec N-1 auto) ──
    if "ca" in wb.sheetnames:
        ws = wb["ca"]
        h = headers_de(ws)
        ca_n1 = _lire_ca_n1().get(mois, (0, 0, 0))
        vals = {
            "CA réalisé (€)": contexte["ca_global"],
            "CA LITTERALIS (€)": contexte["ca_litt"],
            "CA GEODP (€)": contexte["ca_geodp"],
            "Objectif global (€)": contexte["obj_global"],
            "Objectif LITTERALIS (€)": contexte["obj_litt"],
            "Objectif GEODP (€)": contexte["obj_geodp"],
            "CA N-1 global (€)": ca_n1[0],
            "CA N-1 LITTERALIS (€)": ca_n1[1],
            "CA N-1 GEODP (€)": ca_n1[2],
        }
        _upsert_ligne(ws, cle, vals, h)

    # ── Onglet charge ──
    if "charge" in wb.sheetnames:
        ws = wb["charge"]
        h = headers_de(ws)
        c = contexte["charge"]
        vals = {
            "Productif Litt": c["prod_litt"], "Productif GEODP": c["prod_geodp"],
            "Support Litt": c["sup_litt"], "Support GEODP": c["sup_geodp"],
            "Rework Litt": c["rew_litt"], "Rework GEODP": c["rew_geodp"],
            "Rework Total": c["rew_litt"] + c["rew_geodp"],
            "Interne": c["interne"],
            "Support Total": c["sup_litt"] + c["sup_geodp"],
            "Productif Total": c["prod_litt"] + c["prod_geodp"],
        }
        _upsert_ligne(ws, cle, vals, h)

    # ── Onglet backlog ──
    if "backlog" in wb.sheetnames:
        ws = wb["backlog"]
        h = headers_de(ws)
        b = contexte["backlog"]
        # Valeurs COURANTES (mises à jour à chaque run)
        vals = {
            "total_backlog_TOTAL": b["total"], "total_backlog_LITT": b["total_litt"],
            "total_backlog_GEODP": b["total_geodp"],
            "total_bloque_TOTAL": b["bloque"], "total_mobilisable_TOTAL": b["mobilisable"],
            "nb_projets_TOTAL": b["nb_projets"], "nb_rework_TOTAL": b["nb_rework"],
            "total_bloque_LITT": b["bloque_litt"], "total_mobilisable_LITT": b["mob_litt"],
            "total_bloque_GEODP": b["bloque_geodp"], "total_mobilisable_GEODP": b["mob_geodp"],
        }
        ligne = _upsert_ligne(ws, cle, vals, h)

        # PHOTO DÉBUT DE MOIS : figée au premier run du mois, jamais écrasée.
        # On (re)lit les headers car _upsert a pu ajouter des colonnes.
        h = headers_de(ws)
        col_idx = {hh: i + 1 for i, hh in enumerate(h) if hh}
        photo = {
            "backlog_debut_mois_TOTAL": b["total"],
            "backlog_debut_mois_LITT":  b["total_litt"],
            "backlog_debut_mois_GEODP": b["total_geodp"],
        }
        for nom, valeur in photo.items():
            # créer la colonne si elle n'existe pas encore
            if nom not in col_idx:
                new_col = ws.max_column + 1
                ws.cell(1, new_col, nom)
                col_idx[nom] = new_col
            cell = ws.cell(ligne, col_idx[nom])
            # n'écrire QUE si la case est vide (photo non encore prise ce mois)
            if cell.value in (None, ""):
                cell.value = valeur

    # ── Onglet commandes (liste) ──
    if "commandes" in wb.sheetnames and contexte.get("commandes_lignes") is not None:
        ws = wb["commandes"]
        h = headers_de(ws)
        _remplacer_lignes_mois(ws, cle, contexte["commandes_lignes"], h)

    # ── Onglet ca_deal (liste) ──
    if "ca_deal" in wb.sheetnames and contexte.get("ca_deal_lignes") is not None:
        ws = wb["ca_deal"]
        h = headers_de(ws)
        _remplacer_lignes_mois(ws, cle, contexte["ca_deal_lignes"], h)

    # ── Onglet anciennete (liste) ──
    if "anciennete" in wb.sheetnames and contexte.get("anciennete_lignes") is not None:
        ws = wb["anciennete"]
        h = headers_de(ws)
        _remplacer_lignes_mois(ws, cle, contexte["anciennete_lignes"], h)

    wb.save(HISTORIQUE_FILE)
    print(f"   ✅ Historique mis à jour pour {mois} {annee}")
