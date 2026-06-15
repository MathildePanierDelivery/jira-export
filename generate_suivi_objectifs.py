"""
generate_dashboard.py
---------------------
Lit le dernier fichier jira_export_*.xlsx produit par Export_worklog_V3.py
et génère un fichier index.html prêt à être publié sur GitHub Pages.

Usage :
    python generate_dashboard.py
    python generate_dashboard.py --file jira_export_2026-06-12.xlsx  # fichier spécifique
"""

import argparse
import glob
import os
import sys
from datetime import datetime

from openpyxl import load_workbook

# ══════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════
OBJECTIF_ANNUEL_CA   = 1_300_000   # €
DECLENCHEUR_CA       = 1_000_000   # €
STOCK_REWORK_DEPART  = 117         # au 1er janvier 2026
PALIERS_REWORK       = [           # (pct_label, nb_clotures_requises)
    ("80 %", 50),
    ("90 %", 60),
    ("100 %", 70),
]
MOIS_FR = ["Janvier","Février","Mars","Avril","Mai","Juin",
           "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

# ══════════════════════════════════════════════════════════════════
# LECTURE DU FICHIER EXCEL
# ══════════════════════════════════════════════════════════════════
def find_export_file(path=None):
    if path:
        if not os.path.exists(path):
            print(f"❌ Fichier introuvable : {path}")
            sys.exit(1)
        return path
    files = sorted(glob.glob("jira_export_*.xlsx"), reverse=True)
    if not files:
        print("❌ Aucun fichier jira_export_*.xlsx trouvé dans le répertoire courant.")
        sys.exit(1)
    print(f"📂 Fichier utilisé : {files[0]}")
    return files[0]


def load_data(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)

    # ── Historique CA ──────────────────────────────────────────────
    ws_ca = wb["Historique CA"]
    rows_ca = list(ws_ca.iter_rows(min_row=5, max_col=6, values_only=True))

    monthly = []
    ytd_ca = ytd_litt = ytd_geodp = ytd_obj = None

    for row in rows_ca:
        if not row[0]:
            continue
        label = str(row[0])
        if label == "TOTAL YTD":
            ytd_ca    = float(row[1] or 0)
            ytd_litt  = float(row[2] or 0)
            ytd_geodp = float(row[3] or 0)
            ytd_obj   = float(row[4] or 0)
        elif label in MOIS_FR:
            monthly.append({
                "mois":    label,
                "ca":      float(row[1] or 0),
                "litt":    float(row[2] or 0),
                "geodp":   float(row[3] or 0),
                "obj":     float(row[4] or 0),
                "pct":     float(row[5]) if row[5] and row[5] != "—" else None,
            })

    # ── Projets clôturés ───────────────────────────────────────────
    ws_clos = wb["Projets clôturés"]
    rows_clos = list(ws_clos.iter_rows(min_row=1, max_col=6, values_only=True))

    historique_clotures = []
    in_historique = False
    for row in rows_clos:
        if not row[0]:
            continue
        label = str(row[0])
        if "Historique des clôtures" in label:
            in_historique = True
            continue
        if in_historique and label == "Mois":
            continue
        if in_historique and label in MOIS_FR:
            historique_clotures.append({
                "mois":          label,
                "clos_projet":   int(row[1] or 0),
                "clos_rework":   int(row[2] or 0),
                "total_clos":    int(row[3] or 0),
                "stock_rework":  int(row[4] or 0),
            })

    # Cumul rework clôturés
    total_rework_clos = sum(m["clos_rework"] for m in historique_clotures)
    stock_actuel      = historique_clotures[-1]["stock_rework"] if historique_clotures else None
    mois_actuel       = historique_clotures[-1]["mois"] if historique_clotures else "—"

    wb.close()

    return {
        "monthly":            monthly,
        "ytd_ca":             ytd_ca,
        "ytd_litt":           ytd_litt,
        "ytd_geodp":          ytd_geodp,
        "ytd_obj":            ytd_obj,
        "historique_clotures": historique_clotures,
        "total_rework_clos":  total_rework_clos,
        "stock_actuel":       stock_actuel,
        "mois_actuel":        mois_actuel,
    }


# ══════════════════════════════════════════════════════════════════
# HELPERS HTML
# ══════════════════════════════════════════════════════════════════
def fmt_eur(val):
    """123456.78 → '123 457 €'"""
    return f"{round(val):,} €".replace(",", "\u202f")

def fmt_pct(val):
    return f"{val:.1f}\u202f%"

def pct_pill(pct):
    if pct is None:
        return '<span class="pct-pill pct-warn">—</span>'
    if pct >= 85:
        css = "pct-good"
    elif pct >= 65:
        css = "pct-warn"
    else:
        css = "pct-bad"
    return f'<span class="pct-pill {css}">{fmt_pct(pct)}</span>'


def build_monthly_rows(monthly):
    last_idx = len(monthly) - 1
    rows = []
    for i, m in enumerate(monthly):
        cls = ' class="current-month"' if i == last_idx else ""
        suffix = " ← en cours" if i == last_idx else ""
        rows.append(f"""
            <tr{cls}>
              <td>{m['mois']}{suffix}</td>
              <td>{fmt_eur(m['ca'])}</td>
              <td>{fmt_eur(m['obj'])}</td>
              <td>{pct_pill(m['pct'])}</td>
            </tr>""")
    return "\n".join(rows)


def build_stock_blocks(historique):
    blocks = []
    last_idx = len(historique) - 1
    for i, h in enumerate(historique):
        if i == last_idx:
            style = 'style="background:var(--green-light); border: 1px solid #6EE7B7"'
            color = 'style="color:var(--green)"'
        else:
            style = ""
            color = 'style="color:var(--orange)"'
        blocks.append(f"""
            <div class="stock-block" {style}>
              <div class="stock-num" {color}>{h['stock_rework']}</div>
              <div class="stock-label">{h['mois']}{'&nbsp;(actuel)' if i == last_idx else ''}</div>
            </div>""")
    # Bloc de départ
    depart = f"""
            <div class="stock-block">
              <div class="stock-num">{STOCK_REWORK_DEPART}</div>
              <div class="stock-label">Stock jan. 2026</div>
            </div>"""
    return depart + "\n".join(blocks)


def build_rework_steps(total_clos):
    steps = []
    for label, target in PALIERS_REWORK:
        achieved = total_clos >= target
        css = "achieved" if achieved else ("current" if total_clos >= target * 0.8 else "")
        check = '<div class="check-icon">✓</div>' if achieved else ""
        steps.append(f"""
          <div class="rework-step {css}">
            {check}
            <div class="step-pct">{label}</div>
            <div class="step-target">{target}</div>
            <div class="step-label">clôtures requises</div>
          </div>""")
    return "\n".join(steps)


def rework_status_text(total_clos):
    max_target = PALIERS_REWORK[-1][1]
    achieved = [label for label, t in PALIERS_REWORK if total_clos >= t]
    if len(achieved) == len(PALIERS_REWORK):
        return f"<strong>Les {len(PALIERS_REWORK)} paliers atteints ✓</strong> · Objectif 100\u202f% dépassé ({total_clos}\u202f/\u202f{max_target})"
    elif achieved:
        next_label, next_target = [(l, t) for l, t in PALIERS_REWORK if total_clos < t][0]
        return f"<strong>Palier {achieved[-1]} atteint ✓</strong> · Prochain objectif\u202f: {next_target} clôtures ({next_target - total_clos} restantes)"
    else:
        first_label, first_target = PALIERS_REWORK[0]
        return f"Objectif {first_label}\u202f: {first_target} clôtures · {first_target - total_clos} restantes"


# ══════════════════════════════════════════════════════════════════
# GÉNÉRATION HTML
# ══════════════════════════════════════════════════════════════════
def generate_html(data):
    ytd         = data["ytd_ca"]
    ytd_litt    = data["ytd_litt"]
    ytd_geodp   = data["ytd_geodp"]
    total_rework = data["total_rework_clos"]
    today_str   = datetime.now().strftime("%d/%m/%Y").lstrip("0")

    pct_ca      = ytd / OBJECTIF_ANNUEL_CA * 100
    pct_litt    = ytd_litt / ytd * 100 if ytd else 0
    pct_geodp   = ytd_geodp / ytd * 100 if ytd else 0
    declencheur_pct = DECLENCHEUR_CA / OBJECTIF_ANNUEL_CA * 100

    fill_ca = min(pct_ca, 100)

    # Rework bar : ramené à 100 si dépassé
    max_rework = PALIERS_REWORK[-1][1]
    fill_rework = min(total_rework / max_rework * 100, 100)
    palier_positions = [(t / max_rework * 100, str(t)) for _, t in PALIERS_REWORK]
    palier_marks = "\n".join(
        f'<div class="progress-milestone" style="left:{p:.1f}%"><div class="progress-milestone-label">{lbl}</div></div>'
        for p, lbl in palier_positions
    )

    monthly_rows   = build_monthly_rows(data["monthly"])
    stock_blocks   = build_stock_blocks(data["historique_clotures"])
    rework_steps   = build_rework_steps(total_rework)
    rework_status  = rework_status_text(total_rework)

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Suivi Objectifs TEAM — PS France 2026</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  :root {{
    --bg: #F0F2F5; --surface: #FFFFFF; --border: #E2E6EA;
    --text-primary: #1A1D23; --text-secondary: #6B7280; --text-muted: #9CA3AF;
    --blue: #2563EB; --blue-light: #EFF6FF; --blue-mid: #BFDBFE;
    --green: #059669; --green-light: #ECFDF5;
    --orange: #D97706; --orange-light: #FFFBEB;
    --red: #DC2626;
    --litt: #7C3AED; --litt-light: #F5F3FF;
    --geodp: #0891B2; --geodp-light: #ECFEFF;
    --radius: 12px;
    --shadow: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
  }}
  body {{ font-family: 'Inter', system-ui, sans-serif; background: var(--bg); color: var(--text-primary); min-height: 100vh; }}
  header {{ background: var(--text-primary); color: white; padding: 28px 40px; display: flex; align-items: center; justify-content: space-between; gap: 16px; }}
  header .title-block h1 {{ font-size: 18px; font-weight: 600; letter-spacing: -0.01em; }}
  header .title-block p {{ font-size: 13px; color: rgba(255,255,255,0.5); margin-top: 3px; font-family: 'DM Mono', monospace; }}
  header .update-badge {{ font-family: 'DM Mono', monospace; font-size: 11px; color: rgba(255,255,255,0.45); background: rgba(255,255,255,0.07); border: 1px solid rgba(255,255,255,0.1); padding: 6px 12px; border-radius: 6px; white-space: nowrap; }}
  main {{ max-width: 1100px; margin: 0 auto; padding: 36px 24px 60px; display: flex; flex-direction: column; gap: 28px; }}
  .section-label {{ font-size: 11px; font-weight: 600; letter-spacing: 0.08em; text-transform: uppercase; color: var(--text-muted); margin-bottom: 12px; }}
  .card {{ background: var(--surface); border-radius: var(--radius); border: 1px solid var(--border); box-shadow: var(--shadow); overflow: hidden; }}
  .card-header {{ padding: 20px 24px 16px; border-bottom: 1px solid var(--border); display: flex; align-items: flex-start; justify-content: space-between; gap: 12px; }}
  .card-header h2 {{ font-size: 15px; font-weight: 600; }}
  .card-header p {{ font-size: 12px; color: var(--text-secondary); margin-top: 2px; }}
  .card-body {{ padding: 24px; }}
  .big-stat {{ font-family: 'DM Mono', monospace; font-size: 38px; font-weight: 500; letter-spacing: -0.02em; line-height: 1; }}
  .big-stat-label {{ font-size: 12px; color: var(--text-secondary); margin-top: 6px; }}
  .progress-wrap {{ margin-top: 20px; }}
  .progress-meta {{ display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 8px; }}
  .progress-pct {{ font-family: 'DM Mono', monospace; font-size: 13px; font-weight: 500; }}
  .progress-target {{ font-size: 12px; color: var(--text-secondary); }}
  .progress-track {{ height: 10px; background: var(--border); border-radius: 99px; position: relative; overflow: visible; }}
  .progress-fill {{ height: 100%; border-radius: 99px; }}
  .progress-milestone {{ position: absolute; top: 50%; transform: translate(-50%, -50%); width: 2px; height: 18px; background: var(--text-primary); opacity: 0.2; border-radius: 1px; }}
  .progress-milestone-label {{ position: absolute; top: 16px; transform: translateX(-50%); font-size: 10px; font-family: 'DM Mono', monospace; color: var(--text-muted); white-space: nowrap; }}
  .ca-split {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-top: 20px; }}
  .split-block {{ padding: 14px 16px; border-radius: 8px; border: 1px solid var(--border); }}
  .split-block .split-label {{ font-size: 11px; font-weight: 600; letter-spacing: 0.05em; text-transform: uppercase; margin-bottom: 6px; }}
  .split-block .split-amount {{ font-family: 'DM Mono', monospace; font-size: 20px; font-weight: 500; }}
  .split-block .split-pct {{ font-size: 11px; color: var(--text-secondary); margin-top: 2px; }}
  .split-litt {{ background: var(--litt-light); border-color: #DDD6FE; }}
  .split-litt .split-label, .split-litt .split-amount {{ color: var(--litt); }}
  .split-geodp {{ background: var(--geodp-light); border-color: #A5F3FC; }}
  .split-geodp .split-label, .split-geodp .split-amount {{ color: var(--geodp); }}
  .monthly-table {{ width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 4px; }}
  .monthly-table th {{ text-align: left; font-size: 11px; font-weight: 600; letter-spacing: 0.04em; text-transform: uppercase; color: var(--text-muted); padding: 8px 12px; border-bottom: 1px solid var(--border); }}
  .monthly-table th:not(:first-child) {{ text-align: right; }}
  .monthly-table td {{ padding: 10px 12px; border-bottom: 1px solid #F3F4F6; font-family: 'DM Mono', monospace; }}
  .monthly-table td:first-child {{ font-family: 'Inter', sans-serif; font-weight: 500; }}
  .monthly-table td:not(:first-child) {{ text-align: right; }}
  .monthly-table tr:last-child td {{ border-bottom: none; }}
  .monthly-table tr.current-month td {{ background: var(--blue-light); color: var(--blue); }}
  .monthly-table tr.current-month td:first-child {{ font-weight: 600; }}
  .pct-pill {{ display: inline-block; padding: 2px 8px; border-radius: 99px; font-size: 11px; font-weight: 500; }}
  .pct-good {{ background: var(--green-light); color: var(--green); }}
  .pct-warn {{ background: var(--orange-light); color: var(--orange); }}
  .pct-bad  {{ background: #FEF2F2; color: var(--red); }}
  .rework-steps {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-top: 20px; }}
  .rework-step {{ border: 1px solid var(--border); border-radius: 8px; padding: 14px; text-align: center; position: relative; }}
  .rework-step.achieved {{ background: var(--green-light); border-color: #6EE7B7; }}
  .rework-step.current  {{ background: var(--orange-light); border-color: #FCD34D; }}
  .rework-step .step-pct {{ font-size: 11px; font-weight: 700; letter-spacing: 0.05em; text-transform: uppercase; margin-bottom: 6px; }}
  .rework-step.achieved .step-pct {{ color: var(--green); }}
  .rework-step.current  .step-pct {{ color: var(--orange); }}
  .rework-step .step-target {{ font-family: 'DM Mono', monospace; font-size: 22px; font-weight: 500; }}
  .rework-step .step-label {{ font-size: 11px; color: var(--text-secondary); margin-top: 3px; }}
  .check-icon {{ position: absolute; top: 10px; right: 10px; color: var(--green); font-size: 14px; }}
  .rework-current {{ margin-top: 16px; display: flex; align-items: center; gap: 12px; padding: 14px 16px; background: var(--bg); border-radius: 8px; }}
  .rework-current .big-num {{ font-family: 'DM Mono', monospace; font-size: 32px; font-weight: 500; line-height: 1; }}
  .rework-current .big-context {{ font-size: 13px; color: var(--text-secondary); line-height: 1.5; }}
  .rework-current .big-context strong {{ color: var(--text-primary); font-weight: 600; }}
  .stock-row {{ display: flex; gap: 10px; margin-top: 16px; flex-wrap: wrap; }}
  .stock-block {{ flex: 1; min-width: 90px; background: var(--bg); border-radius: 8px; padding: 12px 14px; text-align: center; }}
  .stock-block .stock-num {{ font-family: 'DM Mono', monospace; font-size: 24px; font-weight: 500; }}
  .stock-block .stock-label {{ font-size: 11px; color: var(--text-secondary); margin-top: 3px; }}
  @media (max-width: 640px) {{
    header {{ padding: 20px; flex-direction: column; align-items: flex-start; }}
    .big-stat {{ font-size: 28px; }}
    .ca-split, .rework-steps {{ grid-template-columns: 1fr; }}
  }}
</style>
</head>
<body>

<header>
  <div class="title-block">
    <h1>Suivi Objectifs TEAM — Professional Services France</h1>
    <p>Année 2026 · Mis à jour le {today_str}</p>
  </div>
  <div class="update-badge">⟳ Automatique · chaque matin</div>
</header>

<main>

  <!-- CA ANNUEL -->
  <section>
    <div class="section-label">Chiffre d'affaires non récurrent</div>
    <div class="card">
      <div class="card-header">
        <div>
          <h2>CA cumulé 2026</h2>
          <p>Objectif annuel : {fmt_eur(OBJECTIF_ANNUEL_CA)} · Déclencheur : {fmt_eur(DECLENCHEUR_CA)}</p>
        </div>
        <div style="text-align:right">
          <div class="big-stat" style="color:var(--blue)">{fmt_eur(ytd)}</div>
          <div class="big-stat-label">réalisés sur {fmt_eur(OBJECTIF_ANNUEL_CA)}</div>
        </div>
      </div>
      <div class="card-body">

        <div class="progress-wrap">
          <div class="progress-meta">
            <span class="progress-pct" style="color:var(--blue)">{fmt_pct(pct_ca)}</span>
            <span class="progress-target">Objectif : {fmt_eur(OBJECTIF_ANNUEL_CA)}</span>
          </div>
          <div class="progress-track" style="height:14px">
            <div class="progress-fill" style="width:{fill_ca:.1f}%; background: linear-gradient(90deg, #2563EB, #60A5FA)"></div>
            <div class="progress-milestone" style="left:{declencheur_pct:.1f}%">
              <div class="progress-milestone-label">{fmt_eur(DECLENCHEUR_CA)}</div>
            </div>
          </div>
        </div>

        <div class="ca-split">
          <div class="split-block split-litt">
            <div class="split-label">Littéralis</div>
            <div class="split-amount">{fmt_eur(ytd_litt)}</div>
            <div class="split-pct">{fmt_pct(pct_litt)} du CA total</div>
          </div>
          <div class="split-block split-geodp">
            <div class="split-label">GEODP</div>
            <div class="split-amount">{fmt_eur(ytd_geodp)}</div>
            <div class="split-pct">{fmt_pct(pct_geodp)} du CA total</div>
          </div>
        </div>

        <div style="margin-top:24px">
          <div class="section-label" style="margin-bottom:0">Détail mensuel</div>
        </div>
        <table class="monthly-table">
          <thead>
            <tr>
              <th>Mois</th><th>CA réalisé</th><th>Objectif</th><th>Atteinte</th>
            </tr>
          </thead>
          <tbody>{monthly_rows}</tbody>
        </table>

      </div>
    </div>
  </section>

  <!-- REWORK -->
  <section>
    <div class="section-label">Réduction du backlog Rework</div>
    <div class="card">
      <div class="card-header">
        <div>
          <h2>Projets Rework clôturés — cumul 2026</h2>
          <p>Départ : {STOCK_REWORK_DEPART} Rework au 1er janvier 2026 · Stock actuel : {data['stock_actuel']}</p>
        </div>
      </div>
      <div class="card-body">

        <div class="rework-current">
          <div class="big-num">{total_rework}</div>
          <div class="big-context">
            Rework clôturés depuis janvier<br>
            {rework_status}
          </div>
        </div>

        <div class="rework-steps">{rework_steps}</div>

        <div class="progress-wrap" style="margin-top:20px">
          <div class="progress-meta">
            <span class="progress-pct" style="color:var(--green)">{total_rework} clôturés</span>
            <span class="progress-target">Objectif max : {max_rework}</span>
          </div>
          <div class="progress-track">
            <div class="progress-fill" style="width:{fill_rework:.1f}%; background: linear-gradient(90deg, #059669, #34D399)"></div>
            {palier_marks}
          </div>
        </div>

        <div style="margin-top:24px">
          <div class="section-label" style="margin-bottom:8px">Évolution du stock Rework</div>
          <div class="stock-row">{stock_blocks}</div>
        </div>

      </div>
    </div>
  </section>


  <!-- SATISFACTION CLIENT -->
  <section>
    <div class="section-label">Satisfaction client</div>
    <div class="card">
      <div class="card-header">
        <div>
          <h2>Suivi de la satisfaction client</h2>
          <p>Suivi disponible depuis le canal Teams dédié · Mis à jour en continu par l'équipe</p>
        </div>
      </div>
      <div class="card-body">
        <p style="font-size:13px; color:var(--text-secondary); line-height:1.6; margin-bottom:16px">
          Le suivi de la satisfaction client est géré séparément depuis le canal Teams
          <strong style="color:var(--text-primary)">Suivi de la satisfaction client</strong>.
          Le fichier de réponses est accessible directement via le lien ci-dessous.
        </p>
        <a href="https://sogelink.sharepoint.com/:x:/t/EquipeDeliveryLittralisGeODP-Suividelasatisfactionclient/IQCO8d4z_n7GQLJXQTADrXO3AewqcvG6amiuGealjJiWioI?e=2fZZx9"
           target="_blank"
           style="display:inline-flex; align-items:center; gap:8px; padding:10px 18px;
                  background:var(--blue-light); color:var(--blue);
                  border:1px solid #BFDBFE; border-radius:8px;
                  text-decoration:none; font-size:13px; font-weight:500;">
          📊 Ouvrir Réponses.xlsx sur SharePoint ↗
        </a>
      </div>
    </div>
  </section>

</main>
</body>
</html>"""
    return html


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Génère le dashboard HTML depuis l'export Jira.")
    parser.add_argument("--file", help="Chemin vers le fichier jira_export_*.xlsx (optionnel)")
    parser.add_argument("--output", default="objectifs.html", help="Fichier HTML de sortie (défaut : objectifs.html)")
    args = parser.parse_args()

    filepath = find_export_file(args.file)
    print("📊 Lecture des données...")
    data = load_data(filepath)

    print("🖊️  Génération du HTML...")
    html = generate_html(data)

    with open(args.output, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"✅ Dashboard généré : {args.output}")
    print(f"   CA YTD       : {fmt_eur(data['ytd_ca'])} / {fmt_eur(OBJECTIF_ANNUEL_CA)}")
    print(f"   Rework clos  : {data['total_rework_clos']} / {PALIERS_REWORK[-1][1]} (objectif max)")
    print(f"   Stock rework : {data['stock_actuel']} (était {STOCK_REWORK_DEPART} en janv.)")
