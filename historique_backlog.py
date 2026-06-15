import os
"""
Extraction du "Reste à Reconnaître" (customfield_22042) pour les tickets
"COORDIN : Suivi CA", reconstitué au 1er de chaque mois depuis janvier.

Pré-requis :
    pip install jira openpyxl pandas python-dateutil
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from jira import JIRA
import requests
from requests.auth import HTTPBasicAuth
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, date
from dateutil.relativedelta import relativedelta

# ─────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────
JIRA_URL = "https://sogelink.atlassian.net"
JIRA_EMAIL = "mathilde.panier@sogelink.com"
JIRA_TOKEN = os.environ["JIRA_API_TOKEN"]

JIRA_PROJECT    = "PDMDEP"
ISSUE_TYPE      = "COORDIN : Suivi CA"
CUSTOM_FIELD_ID = "customfield_22042"   # Reste à Reconnaître

ANNEE_DEBUT = 2026
MOIS_DEBUT  = 1

# Snapshots ponctuels supplémentaires (onglet dédié)
SNAPSHOTS_PONCTUELS = [
    datetime(2026, 1, 12, tzinfo=timezone.utc),
    datetime(2026, 2, 18, tzinfo=timezone.utc),
]

today       = date.today().strftime("%Y%m%d")
OUTPUT_FILE = f"RAR_Mensuel_{today}.xlsx"

# ─────────────────────────────────────────────
#  SNAPSHOTS : 1er de chaque mois
# ─────────────────────────────────────────────
def build_snapshots() -> list[datetime]:
    snapshots = []
    current   = datetime(ANNEE_DEBUT, MOIS_DEBUT, 1, tzinfo=timezone.utc)
    now_utc   = datetime.now(timezone.utc)
    while current <= now_utc:
        snapshots.append(current)
        current += relativedelta(months=1)
    return snapshots


# ─────────────────────────────────────────────
#  CHANGELOG COMPLET VIA ENDPOINT DÉDIÉ
# ─────────────────────────────────────────────
def get_changelog_complet(issue_key: str, http_auth) -> list[dict]:
    """Récupère toutes les entrées du changelog via pagination (pas de limite à 100)."""
    histories, start = [], 0
    while True:
        r = requests.get(
            f"{JIRA_URL}/rest/api/3/issue/{issue_key}/changelog",
            auth=http_auth,
            headers={"Accept": "application/json"},
            params={"startAt": start, "maxResults": 100},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        histories.extend(data.get("values", []))
        start += len(data.get("values", []))
        if start >= data["total"]:
            break
    return histories


# ─────────────────────────────────────────────
#  RECONSTITUTION DE LA VALEUR À UNE DATE
# ─────────────────────────────────────────────
def parse_dt(s: str) -> datetime:
    return datetime.fromisoformat(s.replace("Z", "+00:00")).astimezone(timezone.utc)


def valeur_au_snapshot(histories: list[dict], created_str: str, valeur_actuelle, snapshot_dt: datetime):
    """
    Retrouve la valeur de CUSTOM_FIELD_ID au moment du snapshot_dt.

    Stratégie : on reconstitue la timeline complète des changements du champ,
    triée du plus ancien au plus récent. On cherche le dernier changement dont
    la date est <= snapshot_dt et on retourne sa toString.
    Si aucun changement n'est antérieur au snapshot, le champ n'avait pas encore
    été modifié → on retourne le fromString du tout premier changement connu
    (c'est-à-dire la valeur à la création).
    """
    if created_str and parse_dt(created_str) > snapshot_dt:
        return None   # Ticket pas encore créé à cette date

    changements = []
    for history in histories:
        h_dt = parse_dt(history["created"])
        for item in history.get("items", []):
            if item.get("fieldId") == CUSTOM_FIELD_ID:
                changements.append((h_dt, item))

    if not changements:
        # Champ jamais modifié : valeur actuelle = valeur depuis toujours
        return valeur_actuelle

    # Trier du plus ancien au plus récent
    changements.sort(key=lambda x: x[0])

    # Valeur initiale = fromString du premier changement.
    # Si None (champ vide avant la première saisie), on initialise à None.
    valeur = changements[0][1].get("fromString")

    # Avancer dans la timeline jusqu'au snapshot
    for h_dt, item in changements:
        if h_dt <= snapshot_dt:
            valeur = item.get("toString")
        else:
            break

    # Si aucun changement n'est antérieur au snapshot mais le ticket existait,
    # cela signifie que le champ n'avait pas encore été renseigné → None
    # Convertir en float si possible
    try:
        return float(valeur) if valeur is not None else None
    except (ValueError, TypeError):
        return valeur


# ─────────────────────────────────────────────
#  PROGRAMME PRINCIPAL
# ─────────────────────────────────────────────
def main():
    snapshots = build_snapshots()
    labels    = [s.strftime("%Y-%m-%d") for s in snapshots]

    print("Connexion à Jira...")
    jira = JIRA(server=JIRA_URL, basic_auth=(JIRA_EMAIL, JIRA_TOKEN))
    print("Connecté ✓")

    print("\nRécupération des tickets Suivi CA...")
    jql = (
        f'project = "{JIRA_PROJECT}" '
        f'AND issuetype = "{ISSUE_TYPE}" '
        f'ORDER BY created DESC'
    )
    issues = jira.search_issues(
        jql,
        maxResults=False,
        fields=f"summary,created,{CUSTOM_FIELD_ID}",
    )
    print(f"  → {len(issues)} tickets récupérés.")

    http_auth = HTTPBasicAuth(JIRA_EMAIL, JIRA_TOKEN)

    # Cache des changelogs pour éviter de les récupérer deux fois
    changelog_cache = {}

    rows = []
    for i, issue in enumerate(issues, 1):
        key     = issue.key
        fields  = issue.raw["fields"]
        summary = fields.get("summary", "")
        created = fields.get("created", "")
        valeur_actuelle = fields.get(CUSTOM_FIELD_ID)

        print(f"  [{i}/{len(issues)}] {key} — changelog complet…")
        histories = get_changelog_complet(key, http_auth)
        changelog_cache[key] = (histories, created, valeur_actuelle, summary)

        row = {"Ticket": key, "Résumé": summary}
        for snap_dt, label in zip(snapshots, labels):
            row[label] = valeur_au_snapshot(histories, created, valeur_actuelle, snap_dt)

        rows.append(row)

    # ── Construction du DataFrame ──────────────────────────────
    df = pd.DataFrame(rows)
    for label in labels:
        df[label] = pd.to_numeric(df[label], errors="ignore")

    # ── Export Excel avec mise en forme ───────────────────────
    print("\nGénération du fichier Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RAR Mensuel"
    ws.sheet_view.showGridLines = True

    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    HDR_FONT  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BORDER    = Border(
        right=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    # En-têtes
    headers = list(df.columns)
    ws.row_dimensions[1].height = 28
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border    = BORDER

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    for c_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 14

    # Données
    for r_idx, (_, row_data) in enumerate(df.iterrows(), start=2):
        bg = "EBF3FB" if r_idx % 2 == 0 else "FFFFFF"
        ws.row_dimensions[r_idx].height = 16
        for c_idx, col in enumerate(headers, 1):
            val  = row_data[col]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = BORDER
            cell.alignment = Alignment(
                horizontal="right" if c_idx >= 3 else "left",
                vertical="center"
            )
            cell.font = Font(name="Calibri", size=9, color="595959")
            if c_idx >= 3 and pd.notna(val):
                cell.number_format = "#,##0.00"

    # ── Ligne TOTAL ───────────────────────────────────────────
    total_row = len(df) + 2
    ws.row_dimensions[total_row].height = 18
    TOTAL_FILL = PatternFill("solid", fgColor="1F3864")
    TOTAL_FONT = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

    cell = ws.cell(row=total_row, column=1, value="TOTAL")
    cell.fill      = TOTAL_FILL
    cell.font      = TOTAL_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = BORDER

    cell = ws.cell(row=total_row, column=2, value="")
    cell.fill   = TOTAL_FILL
    cell.border = BORDER

    for c_idx, col in enumerate(headers[2:], start=3):
        col_letter = get_column_letter(c_idx)
        # Somme uniquement des cellules numériques (ignore les None)
        total_val = df[col].apply(pd.to_numeric, errors="coerce").sum()
        cell = ws.cell(row=total_row, column=c_idx, value=total_val if total_val else None)
        cell.fill          = TOTAL_FILL
        cell.font          = TOTAL_FONT
        cell.alignment     = Alignment(horizontal="right", vertical="center")
        cell.border        = BORDER
        cell.number_format = "#,##0.00"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes    = "A2"

    # ── Onglet snapshots ponctuels ────────────────────────────
    print("\nGénération de l'onglet snapshots ponctuels...")
    labels_ponc = [s.strftime("%Y-%m-%d") for s in SNAPSHOTS_PONCTUELS]
    rows_ponc = []
    for i, issue in enumerate(issues, 1):
        key = issue.key
        histories, created, valeur_actuelle, summary = changelog_cache[key]
        row = {"Ticket": key, "Résumé": summary}
        for snap_dt, label in zip(SNAPSHOTS_PONCTUELS, labels_ponc):
            row[label] = valeur_au_snapshot(histories, created, valeur_actuelle, snap_dt)
        rows_ponc.append(row)

    df_ponc = pd.DataFrame(rows_ponc)
    for label in labels_ponc:
        df_ponc[label] = pd.to_numeric(df_ponc[label], errors="ignore")

    ws2 = wb.create_sheet("Snapshots ponctuels")
    ws2.sheet_view.showGridLines = True
    headers_ponc = list(df_ponc.columns)
    ws2.row_dimensions[1].height = 28
    for c_idx, h in enumerate(headers_ponc, 1):
        cell = ws2.cell(row=1, column=c_idx, value=h)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = HDR_ALIGN
        cell.border    = BORDER
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 45
    for c_idx in range(3, len(headers_ponc) + 1):
        ws2.column_dimensions[get_column_letter(c_idx)].width = 14

    for r_idx, (_, row_data) in enumerate(df_ponc.iterrows(), start=2):
        bg = "EBF3FB" if r_idx % 2 == 0 else "FFFFFF"
        ws2.row_dimensions[r_idx].height = 16
        for c_idx, col in enumerate(headers_ponc, 1):
            val  = row_data[col]
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = BORDER
            cell.alignment = Alignment(
                horizontal="right" if c_idx >= 3 else "left",
                vertical="center"
            )
            cell.font = Font(name="Calibri", size=9, color="595959")
            if c_idx >= 3 and pd.notna(val):
                cell.number_format = "#,##0.00"

    # Ligne TOTAL onglet ponctuels
    total_row_p = len(df_ponc) + 2
    ws2.row_dimensions[total_row_p].height = 18
    cell = ws2.cell(row=total_row_p, column=1, value="TOTAL")
    cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER
    cell = ws2.cell(row=total_row_p, column=2, value="")
    cell.fill = TOTAL_FILL; cell.border = BORDER
    for c_idx, col in enumerate(headers_ponc[2:], start=3):
        total_val = df_ponc[col].apply(pd.to_numeric, errors="coerce").sum()
        cell = ws2.cell(row=total_row_p, column=c_idx, value=total_val if total_val else None)
        cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = BORDER; cell.number_format = "#,##0.00"

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers_ponc))}1"
    ws2.freeze_panes    = "A2"

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Fichier généré : {OUTPUT_FILE}")


if __name__ == "__main__":
    main()