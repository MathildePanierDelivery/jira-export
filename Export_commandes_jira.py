"""
Export CA Année — tickets "COORDIN : Suivi CA" et "Matériel GEODP"
-------------------------------------------------------------------
Récupère tous les tickets CA créés depuis le 1er janvier de l'année
en cours, avec pour chaque ticket :
  - Date de création
  - Montant (customfield_22998 ou customfield_20701 pour Matériel GEODP)
  - Epic parent, Client, Solution, BDC, Prestation...

Produit un fichier Excel avec :
  - 1 onglet de données brutes (tous les tickets)
  - 1 onglet de synthèse avec total par mois + total général
"""

import os
import sys
from datetime import datetime

import pandas as pd
from jira import JIRA
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# CONFIG — reprendre les mêmes valeurs que Export_worklog_V3.py
# ══════════════════════════════════════════════════════════════════
JIRA_SERVER = "https://sogelink.atlassian.net"
JIRA_EMAIL = "mathilde.panier@sogelink.com"
JIRA_TOKEN = os.environ["JIRA_API_TOKEN"]

today      = datetime.now()
YEAR_START = datetime(today.year, 1, 1)  # 1er janvier de l'année courante

EXPORT_FILE = f"export_ca_{today.year}_{today:%Y%m%d}.xlsx"

MOIS_FR = ["Janvier","Février","Mars","Avril","Mai","Juin",
           "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

SOLUTION_MAPPING = {
    "GEODP1"             : "GEODP",
    "GEODP2 new"         : "GEODP",
    "GEODP2 migration"   : "GEODP",
    "LITTERALIS"         : "LITTERALIS",
    "LITTERALIS STANDARD": "LITTERALIS",
    "SHERPA"             : "LITTERALIS",
}

# ── Styles ────────────────────────────────────────────────────────
BLUE_DARK   = "1F3864"
BLUE_MID    = "2E75B6"
BLUE_LIGHT  = "D6E4F0"
BLUE_XLIGHT = "EBF3FB"
GREY_LINE   = "BDD7EE"
WHITE       = "FFFFFF"
GREEN_DARK  = "375623"
GREEN_LIGHT = "E2EFDA"

FONT_DEFAULT = "Calibri"

def _font(bold=False, size=11, color="000000", italic=False):
    return Font(name=FONT_DEFAULT, bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _border(color=GREY_LINE):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def apply_style(cell, font=None, fill=None, border=None, alignment=None):
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if border:    cell.border    = border
    if alignment: cell.alignment = alignment

def style_header(ws, row, ncols, dark=True):
    bg    = BLUE_DARK if dark else BLUE_MID
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font      = _font(bold=True, color=WHITE)
        c.fill      = _fill(bg)
        c.border    = _border(bg)
        c.alignment = _align("center")

def style_data_row(ws, row, ncols, even=False, bold=False, total=False, green=False):
    if green:
        bg = GREEN_LIGHT
    elif total:
        bg = BLUE_LIGHT
    elif even:
        bg = BLUE_XLIGHT
    else:
        bg = WHITE
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill      = _fill(bg)
        c.border    = _border()
        c.font      = _font(bold=bold or total)
        c.alignment = _align("left")

def autofit(ws, min_w=10, max_w=55, padding=3):
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + padding, min_w), max_w)

# ══════════════════════════════════════════════════════════════════
# CONNEXION JIRA
# ══════════════════════════════════════════════════════════════════
try:
    jira = JIRA(options={"server": JIRA_SERVER}, basic_auth=(JIRA_EMAIL, JIRA_TOKEN))
    print("✅ Connexion Jira établie.")
except Exception as e:
    print(f"❌ Impossible de se connecter à Jira : {e}")
    sys.exit(1)

# ══════════════════════════════════════════════════════════════════
# RÉCUPÉRATION DES TICKETS CA
# ══════════════════════════════════════════════════════════════════
def get_all_issues(jql: str) -> list:
    issues = []
    token  = None
    fields = (
        "summary,issuetype,parent,assignee,created,status,"
        "customfield_10221,customfield_24438,"   # BDC
        "customfield_22998,"                     # Montant Services
        "customfield_20701,"                     # Montant CA
        "customfield_23610,"                     # Revenu planifié
        "customfield_10070,"                     # Client
        "customfield_20514,"                     # Solution cible
        "customfield_23955,"                     # Prestation
        "customfield_24437,"                     # Type prestation
        "customfield_22108,"                     # Deal type
        "customfield_23608"                      # Projet
    )
    try:
        while True:
            batch = jira.enhanced_search_issues(
                jql, maxResults=100, fields=fields, nextPageToken=token
            )
            issues.extend(batch)
            token = getattr(batch, "nextPageToken", None)
            if not token:
                break
    except Exception as e:
        print(f"❌ Erreur JQL ({jql!r}) : {e}")
    return issues


# JQL 1 : tous les tickets CA — sans filtre de date (c'est la date de l'Epic parent qui compte)
jql_ca = (
    f'project in (PDMDEP, PSC) '
    f'AND issuetype = "COORDIN : Suivi CA" '
    f'ORDER BY created ASC'
)

print(f"\n📥 Récupération des tickets CA...")
ca_tickets = get_all_issues(jql_ca)
print(f"   → {len(ca_tickets)} ticket(s) CA trouvé(s).")

# JQL 2 : Epics dont la date de création est dans l'année en cours
jql_epics = (
    f'project in (PDMDEP, PSC) '
    f'AND issuetype in (Epic, "New delivery") '
    f'AND created >= "{YEAR_START:%Y-%m-%d}" '
    f'ORDER BY created ASC'
)

print(f"📥 Récupération des Epics créés depuis le {YEAR_START:%d/%m/%Y}...")
epics = get_all_issues(jql_epics)
print(f"   → {len(epics)} Epic(s) trouvé(s).")

# Map epic_key → date de création de l'Epic
epic_created_map: dict = {}
for ep in epics:
    raw_ep = ep.raw.get("fields", {})
    created_raw_ep = raw_ep.get("created") or getattr(ep.fields, "created", None)
    try:
        epic_created_map[ep.key] = pd.to_datetime(created_raw_ep) if created_raw_ep else None
    except Exception:
        epic_created_map[ep.key] = None

# ══════════════════════════════════════════════════════════════════
# FONCTIONS UTILITAIRES
# ══════════════════════════════════════════════════════════════════
def _read_montant(ticket) -> float:
    return float(getattr(ticket.fields, "customfield_20701", 0) or 0)

def _read_bdc(fields_obj, raw_fields: dict) -> str:
    bdc = str(getattr(fields_obj, "customfield_10221", "") or "").strip()
    if not bdc:
        bdc = str(raw_fields.get("customfield_24438", "") or "").strip()
    return bdc

def _read_client(raw_fields: dict) -> str:
    val = raw_fields.get("customfield_10070")
    if not val:
        return ""
    if isinstance(val, dict):
        return val.get("displayName", val.get("name", ""))
    return str(val)

def _read_prestation(raw_fields: dict) -> str:
    val = raw_fields.get("customfield_23955")
    if not val:
        return ""
    return val.get("value", "") if isinstance(val, dict) else str(val)

def _read_solution(raw_fields: dict) -> str:
    val = (raw_fields.get("customfield_20514") or {}).get("value", "")
    return SOLUTION_MAPPING.get(val, val)

DEAL_MAPPING = {
    "new business"     : "Nouveau déploiement",
    "upsell"           : "Vente additionnelle",
    "additional setup" : "Vente additionnelle",
}

def _read_deal_type(raw_fields: dict) -> str:
    sol = (raw_fields.get("customfield_20514") or {}).get("value", "")
    if sol.lower() == "geodp2 migration":
        return "Migration"
    dt = (raw_fields.get("customfield_22108") or {}).get("value", "")
    if isinstance(raw_fields.get("customfield_22108"), str):
        dt = raw_fields.get("customfield_22108", "")
    return DEAL_MAPPING.get(dt.lower().strip(), "")

# ══════════════════════════════════════════════════════════════════
# CONSTRUCTION DU DATAFRAME
# ══════════════════════════════════════════════════════════════════
rows = []
for ticket in ca_tickets:
    raw   = ticket.raw.get("fields", {})
    f     = ticket.fields

    parent_key = getattr(getattr(f, "parent", None), "key", "")

    # Date de référence = date de création de l'Epic parent
    created_dt = epic_created_map.get(parent_key) if parent_key else None
    # Si l'Epic n'est pas dans notre map (créé avant l'année), on ignore ce ticket
    if created_dt is None:
        continue

    assignee   = getattr(getattr(f, "assignee", None), "displayName", "")
    status     = getattr(getattr(f, "status", None), "name", "")
    montant    = _read_montant(ticket)
    bdc        = _read_bdc(f, raw)
    client     = _read_client(raw)
    prestation = _read_prestation(raw)
    solution   = _read_solution(raw)
    deal_type  = _read_deal_type(raw)
    projet     = raw.get("customfield_23608", "") or ""

    rows.append({
        "Ticket CA"     : ticket.key,
        "Résumé"        : getattr(f, "summary", ""),
        "Type"          : getattr(getattr(f, "issuetype", None), "name", ""),
        "Epic parent"   : parent_key,
        "Projet"        : projet,
        "Client"        : client,
        "Solution"      : solution,
        "Type de deal"  : deal_type,
        "Prestation"    : prestation,
        "Numéro BDC"    : bdc,
        "Assigné"       : assignee,
        "Statut"        : status,
        "Date création" : created_dt.strftime("%d/%m/%Y"),
        "Mois"          : MOIS_FR[created_dt.month - 1],
        "Num mois"      : created_dt.month,
        "Montant (€)"   : montant,
    })

df = pd.DataFrame(rows)

# Exclure les tickets dont la Prestation est explicitement "Hardware"
if not df.empty:
    nb_avant = len(df)
    df = df[df["Prestation"].str.strip().str.lower() != "hardware"]
    nb_exclus = nb_avant - len(df)
    if nb_exclus:
        print(f"   → {nb_exclus} ticket(s) Hardware exclus.")

if df.empty:
    print("⚠️  Aucun ticket CA trouvé — fichier vide généré.")
else:
    print(f"   → {len(df)} lignes retenues / Montant total : {df['Montant (\u20ac)'].sum():,.2f} €")

# ══════════════════════════════════════════════════════════════════
# CONSTRUCTION DU WORKBOOK
# ══════════════════════════════════════════════════════════════════
wb = Workbook()

# ──────────────────────────────────────────────────────────────────
# ONGLET 1 : Données brutes
# ──────────────────────────────────────────────────────────────────
ws_raw = wb.active
ws_raw.title = "📋 Tickets CA bruts"
ws_raw.sheet_view.showGridLines = False

# Titre
ws_raw["A1"] = f"Tickets CA — {today.year} (depuis le 01/01/{today.year})"
ws_raw["A1"].font      = _font(bold=True, size=14, color=BLUE_DARK)
ws_raw["A1"].alignment = _align("left")
ws_raw.row_dimensions[1].height = 24
ws_raw["A2"] = f"Export du {today:%d/%m/%Y} — {len(df)} tickets"
ws_raw["A2"].font = _font(italic=True, size=10, color="595959")
ws_raw.append([])  # ligne 3 vide

# Colonnes à afficher (sans les colonnes internes Num mois)
COLS_DISPLAY = [
    "Ticket CA", "Date création", "Mois", "Client", "Projet",
    "Solution", "Type de deal", "Prestation", "Type",
    "Numéro BDC", "Epic parent", "Résumé", "Assigné", "Statut",
    "Montant (€)"
]
df_display = df[COLS_DISPLAY] if not df.empty else pd.DataFrame(columns=COLS_DISPLAY)

# En-tête
header_row = 4
for col_i, col_name in enumerate(df_display.columns, start=1):
    ws_raw.cell(row=header_row, column=col_i, value=col_name)
style_header(ws_raw, header_row, len(df_display.columns), dark=True)

# Filtre auto
last_col = get_column_letter(len(df_display.columns))
ws_raw.auto_filter.ref = f"A{header_row}:{last_col}{header_row}"

# Données
MONEY_COL_IDX  = df_display.columns.tolist().index("Montant (€)") + 1
TICKET_COL_IDX = df_display.columns.tolist().index("Ticket CA") + 1
for i, row_data in enumerate(df_display.itertuples(index=False), start=1):
    r = header_row + i
    for col_i, val in enumerate(row_data, start=1):
        ws_raw.cell(row=r, column=col_i, value=val)
    style_data_row(ws_raw, r, len(df_display.columns), even=(i % 2 == 0))
    ws_raw.cell(row=r, column=MONEY_COL_IDX).number_format = '#,##0.00 "€"'
    # Lien cliquable sur le Ticket CA
    ticket_key = ws_raw.cell(row=r, column=TICKET_COL_IDX).value
    if ticket_key:
        cell = ws_raw.cell(row=r, column=TICKET_COL_IDX)
        cell.hyperlink = f"{JIRA_SERVER}/browse/{ticket_key}"
        cell.font = _font(color="0563C1", bold=False)
        cell.fill = _fill(BLUE_XLIGHT if (i % 2 == 0) else WHITE)

# Ligne total
if not df_display.empty:
    total_r = header_row + len(df_display) + 1
    ws_raw.cell(row=total_r, column=1, value="TOTAL")
    ws_raw.cell(row=total_r, column=MONEY_COL_IDX, value=df["Montant (€)"].sum())
    ws_raw.cell(row=total_r, column=MONEY_COL_IDX).number_format = '#,##0.00 "€"'
    style_data_row(ws_raw, total_r, len(df_display.columns), total=True)

autofit(ws_raw)

# ──────────────────────────────────────────────────────────────────
# ONGLET 2 : Synthèse par mois
# ──────────────────────────────────────────────────────────────────
ws_syn = wb.create_sheet("📊 Synthèse par mois")
ws_syn.sheet_view.showGridLines = False

ws_syn["A1"] = f"Synthèse CA par mois — {today.year}"
ws_syn["A1"].font      = _font(bold=True, size=14, color=BLUE_DARK)
ws_syn["A1"].alignment = _align("left")
ws_syn.row_dimensions[1].height = 24
ws_syn["A2"] = f"Export du {today:%d/%m/%Y}"
ws_syn["A2"].font = _font(italic=True, size=10, color="595959")
ws_syn.append([])  # ligne 3 vide

# Colonnes synthèse
SYN_COLS = [
    "Mois",
    "Nb tickets CA", "Montant total (€)",
    "dont Nb GEODP",  "dont Montant GEODP (€)",
    "dont Nb LITTERALIS", "dont Montant LITTERALIS (€)",
]

# En-tête principal (ligne 4)
syn_header = 4
for col_i, col_name in enumerate(SYN_COLS, start=1):
    ws_syn.cell(row=syn_header, column=col_i, value=col_name)
style_header(ws_syn, syn_header, len(SYN_COLS), dark=True)
# Sous-entêtes visuels : colonnes GEODP en bleu mid, LITTERALIS en vert
for col_i in (4, 5):
    c = ws_syn.cell(row=syn_header, column=col_i)
    c.fill = _fill(BLUE_MID)
for col_i in (6, 7):
    c = ws_syn.cell(row=syn_header, column=col_i)
    c.fill      = _fill("375623")
    c.font      = _font(bold=True, color=WHITE)
    c.alignment = _align("center")

# Agrégation par mois (dans l'ordre chronologique)
if not df.empty:
    df_geo  = df[df["Solution"] == "GEODP"]
    df_litt = df[df["Solution"] == "LITTERALIS"]

    def _grp(df_sub):
        if df_sub.empty:
            return {}
        g = (
            df_sub[df_sub["Num mois"] > 0]
            .groupby("Num mois", as_index=False)
            .agg(nb=("Ticket CA", "count"), montant=("Montant (€)", "sum"))
        )
        return {int(row["Num mois"]): (int(row["nb"]), row["montant"]) for _, row in g.iterrows()}

    grp_total = (
        df[df["Num mois"] > 0]
        .groupby(["Num mois", "Mois"], as_index=False)
        .agg(nb=("Ticket CA", "count"), montant=("Montant (€)", "sum"))
        .sort_values("Num mois")
    )
    geo_map  = _grp(df_geo)
    litt_map = _grp(df_litt)

    month_rows = []
    for _, grp_row in grp_total.iterrows():
        num = int(grp_row["Num mois"])
        nb_geo,   mt_geo   = geo_map.get(num,  (0, 0.0))
        nb_litt,  mt_litt  = litt_map.get(num, (0, 0.0))
        month_rows.append((
            grp_row["Mois"], int(grp_row["nb"]), grp_row["montant"],
            nb_geo, mt_geo, nb_litt, mt_litt,
        ))
else:
    month_rows = []

for i, (mois_label, nb, montant_sum, nb_geo, mt_geo, nb_litt, mt_litt) in enumerate(month_rows, start=1):
    r = syn_header + i
    vals = [mois_label, nb, montant_sum, nb_geo, mt_geo, nb_litt, mt_litt]
    for col_i, val in enumerate(vals, start=1):
        ws_syn.cell(row=r, column=col_i, value=val)
    style_data_row(ws_syn, r, len(SYN_COLS), even=(i % 2 == 0))
    for money_col in (3, 5, 7):
        ws_syn.cell(row=r, column=money_col).number_format = '#,##0.00 "€"'

# Ligne total général
total_syn_r   = syn_header + len(month_rows) + 1
total_nb      = sum(r[1] for r in month_rows)
total_montant = sum(r[2] for r in month_rows)
total_nb_geo  = sum(r[3] for r in month_rows)
total_mt_geo  = sum(r[4] for r in month_rows)
total_nb_litt = sum(r[5] for r in month_rows)
total_mt_litt = sum(r[6] for r in month_rows)
totals = ["TOTAL", total_nb, total_montant, total_nb_geo, total_mt_geo, total_nb_litt, total_mt_litt]
for col_i, val in enumerate(totals, start=1):
    ws_syn.cell(row=total_syn_r, column=col_i, value=val)
for money_col in (3, 5, 7):
    ws_syn.cell(row=total_syn_r, column=money_col).number_format = '#,##0.00 "€"'
style_data_row(ws_syn, total_syn_r, len(SYN_COLS), total=True)

# Largeurs fixes synthèse
ws_syn.column_dimensions["A"].width = 18
ws_syn.column_dimensions["B"].width = 16
ws_syn.column_dimensions["C"].width = 22
ws_syn.column_dimensions["D"].width = 16
ws_syn.column_dimensions["E"].width = 26
ws_syn.column_dimensions["F"].width = 20
ws_syn.column_dimensions["G"].width = 28

# ══════════════════════════════════════════════════════════════════
# SAUVEGARDE
# ══════════════════════════════════════════════════════════════════
try:
    wb.save(EXPORT_FILE)
    print(f"\n✅ Export terminé : {EXPORT_FILE}")
    print(f"   → {len(df)} tickets CA / {len(month_rows)} mois / Total : {total_montant:,.2f} €")
except Exception as e:
    print(f"❌ Impossible de sauvegarder : {e}")
    sys.exit(1)